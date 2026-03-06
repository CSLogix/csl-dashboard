import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Colors
HEADER_FILL = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HUB_FILL = PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid")
HUB_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=13)
OK_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
WRONG_FILL = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
MISSING_FILL = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
NEW_FILL = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
WRAP = Alignment(wrap_text=True, vertical="top")

HEADERS = ["Col Letter", "Index", "Sheet Header", "Purpose", "Code Variable", "Current Value", "Correct Value", "Status", "Notes"]

def style_header(ws, row):
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP
        cell.border = THIN_BORDER

def style_hub_row(ws, row, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(HEADERS))
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill = HUB_FILL
    cell.font = HUB_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")

def add_row(ws, row, data):
    status = data[7] if len(data) > 7 else ""
    for col, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.border = THIN_BORDER
        cell.alignment = WRAP
        cell.font = Font(name="Arial", size=10)
        if status == "OK":
            cell.fill = OK_FILL
        elif status == "WRONG":
            cell.fill = WRONG_FILL
            cell.font = Font(name="Arial", size=10, bold=True)
        elif status == "MISSING":
            cell.fill = MISSING_FILL
        elif status == "NEW":
            cell.fill = NEW_FILL

# ═══════════════════════════════════════════════════════════════
# Sheet 1: Tolead Hubs
# ═══════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Tolead Hubs"

r = 1

# ── ORD ──
style_hub_row(ws, r, "ORD — Schedule tab — 802 rows"); r += 1
for col, h in enumerate(HEADERS, 1):
    ws.cell(row=r, column=col, value=h)
style_header(ws, r); r += 1

ord_data = [
    ["A", 0, "Column 1", "(blank row#)", "", "", "", "", ""],
    ["B", 1, "Column 1", "LINE # (customer load)", "col_load_id", 1, 1, "OK", ""],
    ["C", 2, "APT ID", "USPS Appointment ID", "", "", "col_appt_id: 2", "MISSING", "Add if needed"],
    ["D", 3, "APPT DATE & TIME", "Delivery appointment", "", "", "col_delivery: 3", "MISSING", "Delivery date/time"],
    ["E", 4, "Date", "Pickup Date", "col_date", 4, 4, "OK", ""],
    ["F", 5, "Time", "Pickup Time", "col_time", 5, 5, "OK", ""],
    ["G", 6, "Pickup Address", "Origin", "col_origin", 6, 6, "OK", ""],
    ["H", 7, "Destination", "Destination", "col_dest", 7, 7, "OK", ""],
    ["I", 8, "Loads", "Scheduling status (Ready/Picked)", "", "", "col_loads: 8", "MISSING", "Same pattern as DFW Col J"],
    ["J", 9, "Status", "Status (New/Scheduled/Delivered)", "col_status", 9, 9, "OK", ""],
    ["K", 10, "Size", "Equipment type", "", "", "", "", ""],
    ["L", 11, "Dock #", "", "", "", "", "", ""],
    ["M", 12, "ETA Tolead", "", "", "", "", "", ""],
    ["N", 13, "ATD Tolead", "", "", "", "", "", ""],
    ["O", 14, "ATA Destination", "", "", "", "", "", ""],
    ["P", 15, "Tracking", "EFJ Pro # + Macropoint hyperlink", "col_efj", 15, 15, "OK", ""],
    ["Q", 16, "Trailer #", "Driver Trailer #", "col_trailer", 16, 16, "OK", ""],
    ["R", 17, "Driver Contact", "Driver Phone #", "", "", "col_phone: 17", "MISSING", "Need for POD reminder emails"],
    ["S", 18, "Others", "Notes", "", "", "", "", ""],
]
for d in ord_data:
    add_row(ws, r, d); r += 1

r += 1  # spacer

# Lifecycle
style_hub_row(ws, r, "ORD Lifecycle"); r += 1
ws.cell(row=r, column=1, value="Stage").font = Font(bold=True)
ws.cell(row=r, column=2, value="Col I (Loads)").font = Font(bold=True)
ws.cell(row=r, column=3, value="Col J (Status)").font = Font(bold=True)
ws.cell(row=r, column=4, value="Col P (EFJ)").font = Font(bold=True)
ws.cell(row=r, column=5, value="Dashboard Status").font = Font(bold=True)
ws.cell(row=r, column=6, value="Active Loads (now)").font = Font(bold=True)
r += 1
for stage in [
    ["New load from customer", "Ready", "New", "(empty)", "Needs to Cover", "3"],
    ["Dispatcher covers it", "Ready→Picked", "Scheduled", "EFJ + MP link", "Scheduled", "~5"],
    ["Driver picks up", "Picked", "Scheduled", "EFJ + MP link", "Active (MP tracking)", ""],
    ["MP: Tracking Completed", "Picked", "Scheduled→Delivered", "EFJ + MP link", "Delivered → POD reminder", ""],
    ["Final close", "Picked", "Delivered", "EFJ + MP link", "Removed from active", "696"],
]:
    for col, val in enumerate(stage, 1):
        ws.cell(row=r, column=col, value=val).border = THIN_BORDER
    r += 1

r += 2

# ── JFK ──
style_hub_row(ws, r, "JFK — Schedule tab — 185 rows"); r += 1
for col, h in enumerate(HEADERS, 1):
    ws.cell(row=r, column=col, value=h)
style_header(ws, r); r += 1

jfk_data = [
    ["A", 0, "LINE #", "Customer load ID", "col_load_id", 0, 0, "OK", ""],
    ["B", 1, "appt #", "Appointment #", "", "", "", "", ""],
    ["C", 2, "Rate", "", "", "", "", "", ""],
    ["D", 3, "Pickup Date", "Pickup Date", "col_date", 3, 3, "OK", ""],
    ["E", 4, "Pickup Time", "Pickup Time", "col_time", 4, 4, "OK", ""],
    ["F", 5, "delivery date/time", "Delivery appointment", "", "", "col_delivery: 5", "MISSING", "Delivery date/time"],
    ["G", 6, "Pickup Address", "Origin", "col_origin", 6, 6, "OK", ""],
    ["H", 7, "Destination", "Destination", "col_dest", 7, 7, "OK", ""],
    ["I", 8, "Loads", "Scheduling status (Picked/empty)", "", "", "col_loads: 8", "MISSING", "Same pattern as DFW Col J"],
    ["J", 9, "Status", "Status (New/Assigned/In Transit/Delivered)", "col_status", 9, 9, "OK", ""],
    ["K", 10, "Size", "Equipment type", "", "", "", "", ""],
    ["L", 11, "ETA Tolead", "", "", "", "", "", ""],
    ["M", 12, "ATD Tolead", "", "", "", "", "", ""],
    ["N", 13, "ATA Destination", "", "", "", "", "", ""],
    ["O", 14, "Tracking", "EFJ Pro # + Macropoint hyperlink", "col_efj", 14, 14, "OK", ""],
    ["P", 15, "License Plate", "Driver Trailer / Plate #", "col_trailer", 15, 15, "OK", ""],
    ["Q", 16, "Driver Contact", "Driver Phone #", "", "", "col_phone: 16", "MISSING", "Need for POD reminder emails"],
    ["R", 17, "Others", "Notes", "", "", "", "", ""],
]
for d in jfk_data:
    add_row(ws, r, d); r += 1

r += 1
style_hub_row(ws, r, "JFK Lifecycle"); r += 1
ws.cell(row=r, column=1, value="Stage").font = Font(bold=True)
ws.cell(row=r, column=2, value="Col I (Loads)").font = Font(bold=True)
ws.cell(row=r, column=3, value="Col J (Status)").font = Font(bold=True)
ws.cell(row=r, column=4, value="Col O (EFJ)").font = Font(bold=True)
ws.cell(row=r, column=5, value="Dashboard Status").font = Font(bold=True)
ws.cell(row=r, column=6, value="Active Loads (now)").font = Font(bold=True)
r += 1
for stage in [
    ["New load", "(empty)", "New", "EFJ (no MP yet?)", "Needs to Cover", "1"],
    ["Assigned", "(empty)", "Assigned", "EFJ + MP link", "Assigned", "1"],
    ["In Transit", "(empty)", "In Transit", "EFJ + MP link", "Active (MP tracking)", "1"],
    ["MP: Tracking Completed", "Picked", "Delivered", "EFJ + MP link", "Delivered → POD reminder", ""],
    ["Final close", "Picked", "Delivered", "EFJ + MP link", "Removed from active", "180"],
]:
    for col, val in enumerate(stage, 1):
        ws.cell(row=r, column=col, value=val).border = THIN_BORDER
    r += 1

r += 2

# ── LAX ──
style_hub_row(ws, r, "LAX — LAX tab — 1284 rows — ⚠ 3 WRONG COLUMNS IN CODE"); r += 1
for col, h in enumerate(HEADERS, 1):
    ws.cell(row=r, column=col, value=h)
style_header(ws, r); r += 1

lax_data = [
    ["A", 0, "EFJ", "EFJ Pro # + Macropoint hyperlink", "col_efj", 0, 0, "OK", ""],
    ["B", 1, "POD Status", "TRUE/FALSE", "", "", "", "", ""],
    ["C", 2, "Rate", "", "", "", "", "", ""],
    ["D", 3, "Container / Load ID #", "LINE # / Load ID", "col_load_id", 3, 3, "OK", ""],
    ["E", 4, "PU Date", "Pickup Date", "col_date", 4, 4, "OK", ""],
    ["F", 5, "Time", "Pickup Time", "col_time", 5, 5, "OK", ""],
    ["G", 6, "Pickup Location", "Origin (pickup address)", "col_dest", 6, "col_origin: 6", "WRONG", "Code says col_dest but this is PICKUP location"],
    ["H", 7, "Delivery Location", "Destination", "", "", "col_dest: 7", "WRONG", "This should be col_dest, not G"],
    ["I", 8, "Delivery Date/Time", "Delivery appointment date", "col_status", 8, "col_delivery: 8", "WRONG", "Code thinks this is STATUS — it's delivery date!"],
    ["J", 9, "Loads", "Status (Delivered/Assigned/INTRANSIT/Unassigned)", "", "", "col_status: 9", "WRONG", "THIS is the real status column"],
    ["K", 10, "Size", "Equipment type (53ft, Flatbed)", "col_trailer", 10, "", "WRONG", "Code says trailer but this is Size"],
    ["L", 11, "Trailer #", "Driver Trailer #", "", "", "col_trailer: 11", "WRONG", "Real trailer column"],
    ["M", 12, "Driver Phone #", "Driver Phone", "", "", "col_phone: 12", "MISSING", "Need for POD reminder emails"],
    ["N", 13, "Carrier Email", "", "", "", "", "", ""],
    ["O", 14, "RATINGS", "", "", "", "", "", ""],
    ["P", 15, "BILLING", "", "", "", "", "", ""],
    ["Q", 16, "Rate agreed", "", "", "", "", "", ""],
    ["R", 17, "Notes", "", "", "", "", "", ""],
]
for d in lax_data:
    add_row(ws, r, d); r += 1

r += 1
style_hub_row(ws, r, "LAX Lifecycle"); r += 1
ws.cell(row=r, column=1, value="Stage").font = Font(bold=True)
ws.cell(row=r, column=2, value="Col J (Loads/Status)").font = Font(bold=True)
ws.cell(row=r, column=3, value="Col A (EFJ)").font = Font(bold=True)
ws.cell(row=r, column=4, value="Dashboard Status").font = Font(bold=True)
ws.cell(row=r, column=5, value="Active Loads (now)").font = Font(bold=True)
r += 1
for stage in [
    ["New load", "Unassigned", "EFJ (no MP?)", "Needs to Cover", "1"],
    ["Assigned", "Assigned", "EFJ + MP link", "Assigned / Tracking", "2"],
    ["In Transit", "INTRANSIT", "EFJ + MP link", "Active (MP tracking)", "2"],
    ["MP: Tracking Completed", "Delivered", "EFJ + MP link", "Delivered → POD reminder", ""],
    ["Cargo Claim", "Cargo Claim", "EFJ", "Special handling", "2"],
    ["Cancelled", "CANCELLED", "EFJ", "Skip", "1"],
    ["Final close", "Delivered", "EFJ", "Removed from active", "750"],
]:
    for col, val in enumerate(stage, 1):
        ws.cell(row=r, column=col, value=val).border = THIN_BORDER
    r += 1

r += 2

# ── DFW (reference) ──
style_hub_row(ws, r, "DFW — DFW tab — 956 rows — ✅ ALREADY FIXED"); r += 1
for col, h in enumerate(HEADERS, 1):
    ws.cell(row=r, column=col, value=h)
style_header(ws, r); r += 1

dfw_data = [
    ["A", 0, "Column 1", "(blank)", "", "", "", "", ""],
    ["B", 1, "USPS APP ID", "USPS Appointment ID", "col_appt_id", 1, 1, "OK", ""],
    ["C", 2, "USPS APPT TIME", "Delivery Date", "col_delivery_date", 2, 2, "OK", ""],
    ["D", 3, "Location", "Destination", "col_dest", 3, 3, "OK", ""],
    ["E", 4, "LINE #", "Customer load ID", "col_load_id", 4, 4, "OK", ""],
    ["F", 5, "Date", "Pickup Date", "col_date", 5, 5, "OK", ""],
    ["G", 6, "Pick Up time", "Pickup Time", "col_time", 6, 6, "OK", ""],
    ["H", 7, "Trailer", "Equipment trailer", "", "", "", "", "Not driver trailer"],
    ["I", 8, "Size", "Equipment type", "col_equipment", 8, 8, "OK", ""],
    ["J", 9, "Loads", "Scheduling (Scheduled/Picked)", "col_loads_j", 9, 9, "OK", "Key lifecycle column"],
    ["K", 10, "EFJ Pro #", "EFJ + Macropoint hyperlink", "col_efj", 10, 10, "OK", ""],
    ["L", 11, "Status", "Final status (Delivered)", "col_status", 11, 11, "OK", "Manual — set after POD received"],
    ["M", 12, "Trailer", "Driver Trailer #", "col_trailer", 12, 12, "OK", ""],
    ["N", 13, "Phone No.", "Driver Phone #", "col_phone", 13, 13, "OK", ""],
    ["O", 14, "Other", "Notes", "", "", "", "", ""],
]
for d in dfw_data:
    add_row(ws, r, d); r += 1

# ═══════════════════════════════════════════════════════════════
# Sheet 2: Boviet Tabs
# ═══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Boviet")
r2 = 1

boviet_tabs = [
    ("DTE Fresh/Stock", [
        ["A", 0, "EFJ Pro #", "EFJ + MP hyperlink", "efj_col", 0, 0, "OK", ""],
        ["B", 1, "Load ID", "Load ID", "load_id_col", 1, 1, "OK", ""],
        ["C", 2, "PU Location", "Origin", "", "", "", "", ""],
        ["D", 3, "Pickup Date/Time", "Pickup", "pickup_col", 3, 3, "OK", ""],
        ["E", 4, "Delivery Date/Time", "Delivery", "delivery_col", 4, 4, "OK", ""],
        ["F", 5, "Status", "Status", "status_col", 5, 5, "OK", ""],
        ["G", 6, "ETA", "", "", "", "", "", ""],
        ["H", 7, "Carrier Email", "", "", "", "", "", ""],
        ["I", 8, "Driver Phone #", "Driver Phone", "", "", "phone_col: 8", "MISSING", ""],
        ["J", 9, "Trailer #", "Driver Trailer", "", "", "trailer_col: 9", "MISSING", ""],
        ["K", 10, "NOTE", "", "", "", "", "", ""],
        ["L", 11, "Macropoint Status", "", "", "", "", "", ""],
    ]),
    ("Sundance", [
        ["A", 0, "EFJ Pro #", "EFJ + MP hyperlink", "efj_col", 0, 0, "OK", ""],
        ["B", 1, "Load ID", "Load ID", "load_id_col", 1, 1, "OK", ""],
        ["C", 2, "Rate", "", "", "", "", "", ""],
        ["D", 3, "PU Location", "Origin", "", "", "", "", ""],
        ["E", 4, "Pickup Date/Time", "Pickup", "pickup_col", 4, 4, "OK", ""],
        ["F", 5, "Delivery Date/Time", "Delivery", "delivery_col", 5, 5, "OK", ""],
        ["G", 6, "Status", "Status", "status_col", 6, 6, "OK", ""],
        ["H", 7, "Carrier Email", "", "", "", "", "", ""],
        ["I", 8, "Driver Phone #", "Driver Phone", "", "", "phone_col: 8", "MISSING", ""],
        ["J", 9, "Trailer #", "Driver Trailer", "", "", "trailer_col: 9", "MISSING", ""],
        ["K", 10, "Driver Name", "", "", "", "", "", ""],
        ["L", 11, "Notes", "", "", "", "", "", ""],
        ["M", 12, "Macropoint Status", "", "", "", "", "", ""],
    ]),
    ("Piedra", [
        ["A", 0, "EFJ Pro #", "EFJ + MP hyperlink", "efj_col", 0, 0, "OK", ""],
        ["B", 1, "Rate", "", "", "", "", "", ""],
        ["C", 2, "Load ID", "Load ID", "load_id_col", 2, 2, "OK", ""],
        ["D", 3, "BV #", "Boviet ref #", "", "", "", "", ""],
        ["E", 4, "PU Location", "Origin", "", "", "", "", ""],
        ["F", 5, "DEL Location", "Destination", "", "", "", "", ""],
        ["G", 6, "Pickup Date/Time", "Pickup", "pickup_col", "5 ← WRONG", 6, "WRONG", "Code says 5 (F=DEL Location)"],
        ["H", 7, "Delivery Date/Time", "Delivery", "delivery_col", "6 ← WRONG", 7, "WRONG", "Code says 6 (G=Pickup)"],
        ["I", 8, "Status", "Status", "status_col", "7 ← WRONG", 8, "WRONG", "Code says 7 (H=Delivery)"],
        ["J", 9, "ETA", "", "", "", "", "", ""],
        ["K", 10, "Carrier Email", "", "", "", "", "", ""],
        ["L", 11, "Driver Phone #", "Driver Phone", "", "", "phone_col: 11", "MISSING", ""],
        ["M", 12, "Trailer #", "Driver Trailer", "", "", "trailer_col: 12", "MISSING", ""],
        ["N", 13, "Driver Name", "", "", "", "", "", ""],
        ["O", 14, "NOTE", "", "", "", "", "", ""],
        ["P", 15, "Macropoint Status", "", "", "", "", "", ""],
    ]),
    ("Hanson", [
        ["A", 0, "EFJ Pro #", "EFJ + MP hyperlink", "efj_col", 0, 0, "OK", ""],
        ["B", 1, "Load ID / Container No.", "Load ID", "load_id_col", 1, 1, "OK", ""],
        ["C", 2, "BV #", "Boviet ref #", "", "", "", "", ""],
        ["D", 3, "Pickup Location", "Origin", "", "", "", "", ""],
        ["E", 4, "Pickup Date/Time", "Pickup", "pickup_col", 4, 4, "OK", ""],
        ["F", 5, "Del Date/Time", "Delivery", "delivery_col", 5, 5, "OK", ""],
        ["G", 6, "Status", "Status", "status_col", 6, 6, "OK", ""],
        ["H", 7, "Carrier Email", "", "", "", "", "", ""],
        ["I", 8, "Driver Phone #", "Driver Phone", "", "", "phone_col: 8", "MISSING", ""],
        ["J", 9, "Driver Name", "", "", "", "", "", ""],
        ["K", 10, "Trailer #", "Driver Trailer", "", "", "trailer_col: 10", "MISSING", ""],
        ["L", 11, "Note", "", "", "", "", "", ""],
        ["M", 12, "Macropoint Status", "", "", "", "", "", ""],
    ]),
    ("Renewable Energy", [
        ["A", 0, "EFJ Pro #", "EFJ + MP hyperlink", "efj_col", 0, 0, "OK", ""],
        ["B", 1, "Load ID", "Load ID", "load_id_col", 1, 1, "OK", ""],
        ["C", 2, "Rate", "", "", "", "", "", ""],
        ["D", 3, "Pickup Date/Time", "Pickup", "pickup_col", 3, 3, "OK", ""],
        ["E", 4, "Delivery Date/Time", "Delivery", "delivery_col", 4, 4, "OK", ""],
        ["F", 5, "Status", "Status", "status_col", 5, 5, "OK", ""],
        ["G", 6, "Carrier Email", "", "", "", "", "", ""],
        ["H", 7, "Notes", "", "", "", "", "", ""],
        ["I", 8, "Macropoint Status", "", "", "", "", "", ""],
    ]),
    ("Radiance Solar", [
        ["A", 0, "EFJ Pro #", "EFJ + MP hyperlink", "efj_col", 0, 0, "OK", ""],
        ["B", 1, "Load ID", "Load ID", "load_id_col", 1, 1, "OK", ""],
        ["C", 2, "Rate", "", "", "", "", "", ""],
        ["D", 3, "Pickup Date/Time", "Pickup", "pickup_col", 3, 3, "OK", ""],
        ["E", 4, "Delivery Date/Time", "Delivery", "delivery_col", 4, 4, "OK", ""],
        ["F", 5, "Status", "Status", "status_col", 5, 5, "OK", ""],
        ["G", 6, "Carrier Email", "", "", "", "", "", ""],
        ["H", 7, "Notes", "", "", "", "", "", ""],
        ["I", 8, "Macropoint Status", "", "", "", "", "", ""],
    ]),
]

for tab_name, rows_data in boviet_tabs:
    wrong_count = sum(1 for d in rows_data if d[7] == "WRONG")
    missing_count = sum(1 for d in rows_data if d[7] == "MISSING")
    flag = ""
    if wrong_count:
        flag = f" — ⚠ {wrong_count} WRONG"
    if missing_count:
        flag += f" + {missing_count} MISSING"
    style_hub_row(ws2, r2, f"Boviet: {tab_name}{flag}"); r2 += 1
    for col, h in enumerate(HEADERS, 1):
        ws2.cell(row=r2, column=col, value=h)
    style_header(ws2, r2); r2 += 1
    for d in rows_data:
        add_row(ws2, r2, d); r2 += 1
    r2 += 1

# ═══════════════════════════════════════════════════════════════
# Column widths
# ═══════════════════════════════════════════════════════════════
for sheet in [ws, ws2]:
    sheet.column_dimensions["A"].width = 10
    sheet.column_dimensions["B"].width = 8
    sheet.column_dimensions["C"].width = 25
    sheet.column_dimensions["D"].width = 30
    sheet.column_dimensions["E"].width = 18
    sheet.column_dimensions["F"].width = 14
    sheet.column_dimensions["G"].width = 18
    sheet.column_dimensions["H"].width = 10
    sheet.column_dimensions["I"].width = 35

# Save
OUTPUT = r"C:\Users\jsfel\Downloads\CSL_Column_Mapping.xlsx"
wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
