"""
FastAPI dashboard for CSL AI Dispatch — Operations Dashboard.
Pulls real-time data from Google Sheet + systemd journal + PostgreSQL.
"""

import json
from decimal import Decimal
import logging
import mimetypes
import os
import re
import shutil
import subprocess
import sys
import threading
import time as _time
from datetime import datetime, timedelta
from pathlib import Path

from fastapi import FastAPI, HTTPException, Query, Form, Request, UploadFile, File, BackgroundTasks
from fastapi.responses import HTMLResponse, FileResponse, RedirectResponse, JSONResponse, Response
from fastapi.staticfiles import StaticFiles
from starlette.middleware.base import BaseHTTPMiddleware
import uvicorn
import gspread
from google.oauth2.service_account import Credentials

import auth
import config
from crypto import encrypt_data, decrypt_data
import database as db
# ── Macropoint tracking constants ──
TRACKING_PHONE = os.getenv("MACROPOINT_TRACKING_PHONE", "4437614954")
DISPATCH_EMAIL = os.getenv("EMAIL_CC", "efj-operations@evansdelivery.com")

# ── Delivery Email Notification ──────────────────────────────────────────
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import threading

JANICE_EMAIL = "Janice.Cortes@evansdelivery.com"
SMTP_HOST = os.getenv("SMTP_HOST", "smtp.gmail.com")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER = os.getenv("SMTP_USER", "")
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD", "")

def _has_delivery_email_been_sent(efj: str) -> bool:
    """Check if delivery email was already sent for this EFJ."""
    try:
        with db.get_cursor() as cur:
            cur.execute("SELECT efj FROM delivery_emails_sent WHERE efj = %s", (efj,))
            return cur.fetchone() is not None
    except Exception:
        return False

def _record_delivery_email(efj: str, account: str):
    """Record that delivery email was sent for this EFJ."""
    try:
        with db.get_conn() as conn:
            with db.get_cursor(conn) as cur:
                cur.execute(
                    "INSERT INTO delivery_emails_sent (efj, account) VALUES (%s, %s) ON CONFLICT (efj) DO NOTHING",
                    (efj, account)
                )
    except Exception as e:
        log.error("Failed to record delivery email for %s: %s", efj, e)

def send_delivery_email(shipment: dict):
    """Send delivery notification email to Janice for Master loads."""
    efj = shipment.get("efj", "")
    account = shipment.get("account", "")

    # Only Master loads (not Tolead/Boviet)
    if account in ("Tolead", "Boviet"):
        return
    # Dedup check
    if _has_delivery_email_been_sent(efj):
        log.info("Delivery email already sent for %s — skipping", efj)
        return
    if not SMTP_USER or not SMTP_PASSWORD:
        log.warning("SMTP credentials not configured — skipping delivery email for %s", efj)
        return

    container = shipment.get("container", "") or shipment.get("loadNumber", "")
    carrier = shipment.get("carrier", "")
    origin = shipment.get("origin", "")
    destination = shipment.get("destination", "")
    delivery_date = shipment.get("delivery", "")
    rep = shipment.get("rep", "")

    subject = f"CSL — Delivered: {efj} | {account} | {container}"
    deep_link = f"https://cslogixdispatch.com/app?view=billing&load={efj}"

    body = f"""
    <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
      <div style="background: linear-gradient(135deg, #22C55E, #4ADE80); padding: 16px 24px; border-radius: 12px 12px 0 0;">
        <h2 style="color: white; margin: 0; font-size: 18px;">&#10022; Load Delivered</h2>
      </div>
      <div style="background: #141A28; padding: 24px; border: 1px solid #1E293B; border-top: none; border-radius: 0 0 12px 12px;">
        <table style="width: 100%; border-collapse: collapse; color: #F0F2F5; font-size: 14px;">
          <tr><td style="padding: 8px 0; color: #8B95A8; width: 120px;">EFJ #</td><td style="padding: 8px 0; font-weight: 600;">{efj}</td></tr>
          <tr><td style="padding: 8px 0; color: #8B95A8;">Account</td><td style="padding: 8px 0; font-weight: 600;">{account}</td></tr>
          <tr><td style="padding: 8px 0; color: #8B95A8;">Container/Load</td><td style="padding: 8px 0;">{container}</td></tr>
          <tr><td style="padding: 8px 0; color: #8B95A8;">Carrier</td><td style="padding: 8px 0;">{carrier}</td></tr>
          <tr><td style="padding: 8px 0; color: #8B95A8;">Route</td><td style="padding: 8px 0;">{origin} &#8594; {destination}</td></tr>
          <tr><td style="padding: 8px 0; color: #8B95A8;">Delivery Date</td><td style="padding: 8px 0;">{delivery_date}</td></tr>
          <tr><td style="padding: 8px 0; color: #8B95A8;">Rep</td><td style="padding: 8px 0;">{rep}</td></tr>
        </table>
        <div style="margin-top: 20px; text-align: center;">
          <a href="{deep_link}" style="display: inline-block; background: #00D4AA; color: #0A0E17; padding: 10px 28px; border-radius: 8px; text-decoration: none; font-weight: 700; font-size: 14px;">
            Open in Billing Dashboard
          </a>
        </div>
      </div>
      <p style="color: #64748B; font-size: 11px; margin-top: 12px; text-align: center;">
        CSLogix Dispatch &mdash; Automated delivery notification
      </p>
    </div>
    """

    def _send():
        try:
            msg = MIMEMultipart("alternative")
            msg["Subject"] = subject
            msg["From"] = SMTP_USER
            msg["To"] = JANICE_EMAIL
            msg["Cc"] = DISPATCH_EMAIL
            msg.attach(MIMEText(body, "html"))

            recipients = [JANICE_EMAIL, DISPATCH_EMAIL]
            with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as smtp:
                smtp.ehlo()
                smtp.starttls()
                smtp.login(SMTP_USER, SMTP_PASSWORD)
                smtp.sendmail(SMTP_USER, recipients, msg.as_string())

            _record_delivery_email(efj, account)
            log.info("Delivery email sent for %s → %s (cc: %s)", efj, JANICE_EMAIL, DISPATCH_EMAIL)
        except Exception as exc:
            log.error("Failed to send delivery email for %s: %s", efj, exc)

    threading.Thread(target=_send, daemon=True).start()

TRACKING_CACHE_FILE = "/root/csl-bot/ftl_tracking_cache.json"

# Status → progress step mapping (cumulative: each status implies all prior steps done)
_MP_PROGRESS_ORDER = [
    "Driver Assigned",
    "Ready To Track",
    "Arrived At Origin",
    "Departed Origin",
    "At Delivery",
    "Delivered",
]

_STATUS_TO_STEP = {
    # Sheet statuses
    "pending": 0,
    "booked": 0,
    "assigned": 1,
    "ready to track": 1,
    "tracking now": 1,
    "tracking waiting for update": 1,
    "at pickup": 2,
    "driver arrived at pickup": 2,
    "arrived at pickup": 2,
    "arrived at origin": 2,
    "loading": 2,
    "in transit": 3,
    "departed pickup": 3,
    "departed pickup - en route": 3,
    "en route": 3,
    "running late": 3,
    "tracking behind schedule": 3,
    "driver phone unresponsive": 3,
    "at delivery": 4,
    "arrived at delivery": 4,
    "unloading": 4,
    "out for delivery": 4,
    "delivered": 5,
    "departed delivery": 5,
    "completed": 5,
    "pod received": 5,
}

def _read_tracking_cache() -> dict:
    """Read the FTL tracking cache written by ftl_monitor.py."""
    try:
        with open(TRACKING_CACHE_FILE) as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}



def _find_tracking_entry(cache: dict, efj: str) -> dict:
    """Find tracking cache entry, handling EFJ prefix mismatch."""
    if efj in cache:
        return cache[efj]
    stripped = efj.replace("EFJ", "").strip()
    if stripped in cache:
        return cache[stripped]
    prefixed = f"EFJ{efj}" if not efj.startswith("EFJ") else None
    if prefixed and prefixed in cache:
        return cache[prefixed]
    for entry in cache.values():
        e = entry.get("efj", "")
        if e == efj or e == stripped or e.replace("EFJ", "").strip() == stripped:
            return entry
    return {}

def _get_driver_contact(efj: str) -> dict:
    """Get driver contact info from PostgreSQL."""
    try:
        with db.get_cursor() as cur:
            cur.execute(
                "SELECT driver_name, driver_phone, driver_email, notes, "
                "carrier_email, trailer_number, macropoint_url "
                "FROM driver_contacts WHERE efj = %s",
                (efj,),
            )
            row = cur.fetchone()
            if row:
                return dict(row)
    except Exception as e:
        log.warning("Failed to read driver contact for %s: %s", efj, e)
    return {}


def _upsert_driver_contact(efj: str, name: str = None, phone: str = None,
                           email: str = None, notes: str = None,
                           carrier_email: str = None, trailer_number: str = None,
                           macropoint_url: str = None):
    """Insert or update driver contact info."""
    with db.get_conn() as conn:
        with db.get_cursor(conn) as cur:
            cur.execute("""
                INSERT INTO driver_contacts (efj, driver_name, driver_phone, driver_email, notes, carrier_email, trailer_number, macropoint_url)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (efj) DO UPDATE SET
                    driver_name    = COALESCE(EXCLUDED.driver_name,    driver_contacts.driver_name),
                    driver_phone   = COALESCE(EXCLUDED.driver_phone,   driver_contacts.driver_phone),
                    driver_email   = COALESCE(EXCLUDED.driver_email,   driver_contacts.driver_email),
                    notes          = COALESCE(EXCLUDED.notes,          driver_contacts.notes),
                    carrier_email  = COALESCE(EXCLUDED.carrier_email,  driver_contacts.carrier_email),
                    trailer_number = COALESCE(EXCLUDED.trailer_number, driver_contacts.trailer_number),
                    macropoint_url = COALESCE(EXCLUDED.macropoint_url, driver_contacts.macropoint_url),
                    updated_at     = NOW()
            """, (efj, name or None, phone or None, email or None, notes or None, carrier_email or None, trailer_number or None, macropoint_url or None))


def _build_macropoint_progress(status_str: str):
    """Build progress array from a status string."""
    key = (status_str or "").strip().lower()
    step = _STATUS_TO_STEP.get(key, -1)
    # Fuzzy match: check if any known key is contained in the status
    if step == -1:
        for k, v in _STATUS_TO_STEP.items():
            if k in key or key in k:
                step = v
                break
    if step == -1:
        step = 0  # default: only first step done
    return [
        {"label": lbl, "done": i <= step}
        for i, lbl in enumerate(_MP_PROGRESS_ORDER)
    ]


log = logging.getLogger(__name__)

app = FastAPI(title="CSL AI Dispatch")

# ---------------------------------------------------------------------------
# Authentication middleware
# ---------------------------------------------------------------------------
PUBLIC_PATHS = {"/login", "/setup", "/health", "/logo.svg", "/app", "/assets", "/", "/macropoint-webhook", "/webhook-test"}

class AuthMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        path = request.url.path
        # Public paths: static assets, login/setup pages, React app shell
        if path in PUBLIC_PATHS or path.startswith("/static") or path.startswith("/app") or path.startswith("/assets") or path.endswith((".png", ".svg", ".ico", ".jpg", ".webp")):
            return await call_next(request)
        # API routes require session auth (return 401 JSON, not redirect)
        if path.startswith("/api/"):
            # Dev API key bypass for local development (IP-restricted)
            dev_key = os.environ.get("CSL_DEV_KEY", "")
            dev_ips = os.environ.get("CSL_DEV_IPS", "").split(",")
            if dev_key and request.headers.get("x-dev-key") == dev_key:
                client_ip = request.headers.get("x-real-ip", "")
                if client_ip in dev_ips:
                    return await call_next(request)
            token = request.cookies.get("csl_session")
            user = auth.verify_session_token(token)
            if not user:
                from starlette.responses import JSONResponse
                return JSONResponse({"error": "unauthorized"}, status_code=401)
            return await call_next(request)
        # Legacy HTML pages: redirect to login
        if not auth.is_configured():
            return RedirectResponse("/setup", status_code=302)
        token = request.cookies.get("csl_session")
        user = auth.verify_session_token(token)
        if not user:
            return RedirectResponse("/login", status_code=302)
        return await call_next(request)

app.add_middleware(AuthMiddleware)

# ── CORS middleware for React dev ──
from starlette.middleware.cors import CORSMiddleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://localhost:5174", "http://localhost:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── Serve React production build assets ──
_react_dist = Path(__file__).parent / "static" / "dist"
if _react_dist.exists():
    app.mount("/assets", StaticFiles(directory=str(_react_dist / "assets")), name="react-assets")


# ── Serve public assets (images in dist root) ──
from fastapi.responses import FileResponse as _FileResponse

@app.get("/rateiq-bot.png")
async def _serve_rateiq_bot():
    f = _react_dist / "rateiq-bot.png"
    if f.exists():
        return _FileResponse(str(f), media_type="image/png")
    return JSONResponse({"error": "not found"}, 404)

@app.get("/astrobot.png")
async def _serve_astrobot():
    f = _react_dist / "astrobot.png"
    if f.exists():
        return _FileResponse(str(f), media_type="image/png")
    return JSONResponse({"error": "not found"}, 404)

@app.get("/logo.svg")
async def _serve_logo():
    f = _react_dist / "logo.svg"
    if f.exists():
        return _FileResponse(str(f), media_type="image/svg+xml")
    return JSONResponse({"error": "not found"}, 404)




# ---------------------------------------------------------------------------
# Google Sheet configuration
# ---------------------------------------------------------------------------
SHEET_ID = "19MB5HmmWwsVXY_nADCYYLJL-zWXYt8yWrfeRBSfB2S0"
CREDS_FILE = "/root/csl-credentials.json"
SKIP_TABS = {
    "Sheet 4", "DTCELNJW", "Account Rep", "SSL Links",
    "Completed Eli", "Completed Radka", "Completed John F", "Boviet",
}
CACHE_TTL = 600  # 10 minutes
# Boviet + Tolead separate sheets
BOVIET_SHEET_ID = "1OP-ZDaMCOsPxcxezHSPfN5ftUXlUcOjFgsfCQgDp3wI"
import re as _re


def _classify_mp_display_status(cache_entry, shipment=None):
    """Compute a user-friendly MP display status from tracking data.

    Returns (display_status: str, detail: str).
    display_status is the short badge label, detail is optional context.
    """
    import re as _re
    from datetime import datetime as _dt, timezone as _tz
    from zoneinfo import ZoneInfo as _ZI

    if not cache_entry:
        return ("No MP", "")

    status = (cache_entry.get("status") or "").strip()
    schedule_alert = (cache_entry.get("schedule_alert") or "").strip()
    last_loc = cache_entry.get("last_location") or {}
    last_loc_ts = (last_loc.get("timestamp") or "").strip()

    status_lower = status.lower()
    alert_upper = schedule_alert.upper()

    # ── Compute staleness from last_event_at / last_scraped ──
    _last_ev = (cache_entry.get("last_event_at") or cache_entry.get("last_scraped") or "").strip()
    _stale_hours = None
    if _last_ev:
        try:
            _now_s = _dt.now(_ZI("America/New_York"))
            _ts_s = _last_ev.replace(" ET", "").replace(" EST", "").replace(" EDT", "")
            for _fmt_s in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%SZ"):
                try:
                    _ev_dt = _dt.strptime(_ts_s, _fmt_s)
                    _ev_dt = _ev_dt.replace(tzinfo=_ZI("America/New_York"))
                    _stale_hours = (_now_s - _ev_dt).total_seconds() / 3600
                    break
                except ValueError:
                    continue
        except Exception:
            pass

    # ── Cross-ref PG shipment status for stale/empty cache ──
    _ship_status = ""
    if shipment:
        _ship_status = (shipment.get("status") or "").strip().lower()
    _TERMINAL_SHIP = {"delivered", "ready to close", "completed", "billed_closed",
                      "billed/closed", "canceled", "cancelled"}

    # If cache never got real data (empty stop_times + no last_event_at),
    # defer to shipment status
    _has_mp_data = bool(_last_ev) or bool(cache_entry.get("stop_times", {}))
    if not _has_mp_data and _ship_status in _TERMINAL_SHIP:
        return ("Delivered", "Per shipment status")

    # If cache is very stale (>48h) and shipment is terminal, auto-classify
    if _stale_hours is not None and _stale_hours > 48 and _ship_status in _TERMINAL_SHIP:
        return ("Delivered", f"Per shipment status")

    # ── No status at all ──
    if not status:
        # Even with no MP status, if cache has stop2_departed, it's delivered
        if cache_entry.get("stop_times", {}).get("stop2_departed"):
            return ("Delivered", "Departed delivery")
        return ("No MP", "")

    if status_lower == "unassigned":
        return ("Unassigned", "")

    # ── Terminal: Delivered / Completed ──
    if "delivered" in status_lower or "completed" in status_lower:
        return ("Delivered", "")

    # ── Driver phone unresponsive ──
    if "unresponsive" in status_lower:
        return ("No Signal", "Driver phone unresponsive")

    # ── Parse hours from schedule alert ──
    hours_offset = None  # positive = ahead, negative = behind
    _m = _re.search(r'([\d.]+)\s*Hours?\s*(BEHIND|AHEAD)', alert_upper)
    if _m:
        hours_offset = float(_m.group(1))
        if _m.group(2) == "BEHIND":
            hours_offset = -hours_offset

    is_behind = hours_offset is not None and hours_offset < 0
    is_ahead = hours_offset is not None and hours_offset >= 0

    detail = ""
    if hours_offset is not None:
        abs_h = abs(hours_offset)
        if abs_h >= 1:
            detail = f"{abs_h:.1f}h {'behind' if is_behind else 'ahead'}"
        else:
            mins = int(abs_h * 60)
            detail = f"{mins}m {'behind' if is_behind else 'ahead'}"

    # ── GPS staleness ──
    gps_stale = True  # assume stale unless proven otherwise
    if last_loc_ts:
        try:
            now = _dt.now(_ZI("America/New_York"))
            ts = last_loc_ts.replace(" ET", "").replace(" EST", "").replace(" EDT", "")
            for fmt in ("%Y-%m-%dT%H:%M:%SZ", "%Y-%m-%d %H:%M:%S",
                        "%m/%d/%Y %I:%M:%S %p", "%m/%d/%Y %H:%M:%S"):
                try:
                    loc_dt = _dt.strptime(ts, fmt)
                    if fmt.endswith("Z"):
                        loc_dt = loc_dt.replace(tzinfo=_tz.utc).astimezone(_ZI("America/New_York"))
                    else:
                        loc_dt = loc_dt.replace(tzinfo=_ZI("America/New_York"))
                    age_hours = (now - loc_dt).total_seconds() / 3600
                    gps_stale = age_hours > 2.0
                    break
                except ValueError:
                    continue
        except Exception:
            pass

    # ── At specific stops ──
    if "arrived" in status_lower and "pickup" in status_lower:
        return ("At Pickup", detail)
    if "at pickup" in status_lower:
        return ("At Pickup", detail)
    if "arrived" in status_lower and "delivery" in status_lower:
        return ("At Delivery", detail)
    if "at delivery" in status_lower:
        return ("At Delivery", detail)
    if "departed delivery" in status_lower:
        return ("Delivered", "Departed delivery")

    # ── In Transit ──
    if any(x in status_lower for x in ("transit", "departed pickup", "en route")):
        if is_behind:
            return ("Behind Schedule", detail)
        if is_ahead:
            return ("On Time", detail)
        if gps_stale:
            return ("Awaiting Update", "GPS stale")
        return ("In Transit", "")

    # ── Running late ──
    if "behind" in status_lower or "late" in status_lower:
        return ("Behind Schedule", detail or "Running late")

    # ── Waiting for update ──
    if "waiting" in status_lower:
        return ("Awaiting Update", "")

    # ── Tracking Started (the key improvement) ──
    if "tracking started" in status_lower:
        if is_behind:
            return ("Behind Schedule", detail)
        if is_ahead:
            return ("On Time", detail)
        if not last_loc_ts:
            return ("Assigned", "Awaiting first GPS ping")
        if gps_stale:
            return ("Awaiting Update", "GPS stale")
        return ("On Time", "Tracking active")

    # ── On-Site / stale active statuses ──
    if "on-site" in status_lower or "on site" in status_lower:
        if cache_entry.get("stop_times", {}).get("stop2_departed"):
            return ("Delivered", "Departed delivery")
        if _stale_hours is not None and _stale_hours > 24 and _ship_status in _TERMINAL_SHIP:
            return ("Delivered", "Per shipment status")
        return ("At Delivery", detail or "On site")

    # ── Fallback ──
    if is_behind:
        return ("Behind Schedule", detail)
    if is_ahead:
        return ("On Time", detail)
    if _stale_hours is not None and _stale_hours > 48 and _ship_status in _TERMINAL_SHIP:
        return ("Delivered", "Per shipment status")
    return (status, "")


def _shorten_address(addr):
    """Shorten full address to city/state/zip format.
    '640 N Central Ave, Wood Dale, IL 60191' -> 'Wood Dale, IL 60191'
    '(66Z) Kansas City, MO (14hrs)' -> 'Kansas City, MO'
    '(ATL) 495 Horizon Dr Ste 300 Suwanee GA 30024' -> 'Suwanee, GA 30024'
    """
    if not addr:
        return addr
    # Strip leading parenthetical codes like (66Z), (ATL), (LAX)
    addr = _re.sub(r'^\(\w+\)\s*', '', addr).strip()
    # Strip trailing parenthetical like (14hrs)
    addr = _re.sub(r'\s*\([^)]*\)\s*$', '', addr).strip()
    # Try to match "City, ST ZIP" at end of address
    m = _re.search(r'([A-Za-z][A-Za-z .]+),\s*([A-Z]{2})\s+(\d{5}(?:-\d{4})?)\s*$', addr)
    if m:
        return f"{m.group(1).strip()}, {m.group(2)} {m.group(3)}"
    # Try "City, ST" without zip
    m = _re.search(r'([A-Za-z][A-Za-z .]+),\s*([A-Z]{2})\s*$', addr)
    if m:
        return f"{m.group(1).strip()}, {m.group(2)}"
    # Try "City ST ZIP" without comma (common in some formats)
    m = _re.search(r'([A-Za-z][A-Za-z .]+)\s+([A-Z]{2})\s+(\d{5}(?:-\d{4})?)\s*$', addr)
    if m:
        return f"{m.group(1).strip()}, {m.group(2)} {m.group(3)}"
    return addr


BOVIET_SKIP_TABS = {"POCs", "Boviet Master"}
BOVIET_TAB_CONFIGS = {
    "DTE Fresh/Stock":  {"efj_col": 0, "load_id_col": 1, "status_col": 5},
    "Sundance":         {"efj_col": 0, "load_id_col": 1, "status_col": 6},
    "Renewable Energy": {"efj_col": 0, "load_id_col": 1, "status_col": 5},
    "Radiance Solar":   {"efj_col": 0, "load_id_col": 1, "status_col": 5},
    "Piedra":           {"efj_col": 0, "load_id_col": 2, "status_col": 8,
                         "pickup_col": 6, "delivery_col": 7,
                         "phone_col": 11, "trailer_col": 12,
                         "carrier_email_col": 10, "driver_name_col": 13,
                         "default_origin": "Greenville, NC", "default_dest": "Mexia, TX",
                         "start_row": 45},
    "Hanson":           {"efj_col": 0, "load_id_col": 1, "status_col": 6,
                         "pickup_col": 4, "delivery_col": 5,
                         "phone_col": 8, "trailer_col": 10,
                         "carrier_email_col": 7, "driver_name_col": 9},
}
BOVIET_DONE_STATUSES = {"delivered", "completed", "canceled", "cancelled", "ready to close"}

TOLEAD_SHEET_ID = "1-zl7CCFdy2bWRTm1FsGDDjU-KVwqPiThQuvJc2ZU2ac"
TOLEAD_TAB = "Schedule"
TOLEAD_COL_EFJ = 15     # P
TOLEAD_COL_ORD = 1      # B
TOLEAD_COL_STATUS = 9   # J
TOLEAD_COL_ORIGIN = 6   # G
TOLEAD_COL_DEST = 7     # H
TOLEAD_COL_DATE = 4     # E
TOLEAD_SKIP_STATUSES = {"delivered", "canceled", "cancelled"}

# --- Tolead JFK ---
TOLEAD_JFK_SHEET_ID = "1mfhEsK2zg6GWTqMDTo5VwCgY-QC1tkNPknurHMxtBhs"
TOLEAD_JFK_TAB = "Schedule"
TOLEAD_ORD_COLS = {
    "efj": 15, "load_id": 1, "status": 9, "origin": 6,
    "destination": 7, "pickup_date": 4, "pickup_time": 5,
    "delivery": 3, "driver": 16, "phone": 17, "appt_id": 2,
}

TOLEAD_JFK_COLS = {
    "efj": 14, "load_id": 0, "status": 9, "origin": 6,
    "destination": 7, "pickup_date": 3, "pickup_time": 4,
    "delivery": 5, "driver": 15, "phone": 16,
}

# --- Tolead LAX ---
TOLEAD_LAX_SHEET_ID = "1YLB6z5LdL0kFYTcfq_H5e6acU8-VLf8nN0XfZrJ-bXo"
TOLEAD_LAX_TAB = "LAX"
TOLEAD_LAX_COLS = {
    "efj": 0, "load_id": 3, "status": 9, "origin": 6,
    "destination": 7, "pickup_date": 4, "pickup_time": 5,
    "delivery": 8, "driver": 11, "phone": 12,
}

# --- Tolead DFW ---
TOLEAD_DFW_SHEET_ID = "1RfGcq25x4qshlBDWDJD-xBJDlJm6fvDM_w2RTHhn9oI"
TOLEAD_DFW_TAB = "DFW"
TOLEAD_DFW_COLS = {
    "efj": 10, "load_id": 4, "status": 11, "origin": None,
    "destination": 3, "pickup_date": 5, "pickup_time": 6,
    "delivery": 2, "driver": 12, "phone": 13,
    "appt_id": 1, "equipment": 8,
}

TOLEAD_HUB_CONFIGS = {
    "ORD": {"sheet_id": TOLEAD_SHEET_ID, "tab": TOLEAD_TAB, "cols": TOLEAD_ORD_COLS, "default_origin": "ORD", "start_row": 790},
    "JFK": {"sheet_id": TOLEAD_JFK_SHEET_ID, "tab": TOLEAD_JFK_TAB, "cols": TOLEAD_JFK_COLS, "default_origin": "Garden City, NY", "start_row": 184},
    "LAX": {"sheet_id": TOLEAD_LAX_SHEET_ID, "tab": TOLEAD_LAX_TAB, "cols": TOLEAD_LAX_COLS, "default_origin": "Vernon, CA", "start_row": 755},
    "DFW": {"sheet_id": TOLEAD_DFW_SHEET_ID, "tab": TOLEAD_DFW_TAB, "cols": TOLEAD_DFW_COLS, "default_origin": "Irving, TX", "start_row": 172},
}

def _get_sheet_hyperlinks(creds, sheet_id, tab_name):
    """Fetch hyperlink URLs from a sheet tab via Sheets API v4."""
    import requests as _requests
    from google.auth.transport.requests import Request as GoogleRequest
    try:
        creds.refresh(GoogleRequest())
        api_url = (
            f"https://sheets.googleapis.com/v4/spreadsheets/{sheet_id}"
            f"?ranges={_requests.utils.quote(tab_name)}"
            f"&fields=sheets.data.rowData.values.hyperlink"
            f"&includeGridData=true"
        )
        resp = _requests.get(api_url, headers={"Authorization": f"Bearer {creds.token}"})
        resp.raise_for_status()
        data = resp.json()
        result = []
        for row_data in data["sheets"][0]["data"][0].get("rowData", []):
            result.append([cell.get("hyperlink") for cell in row_data.get("values", [])])
        return result
    except Exception as e:
        log.warning("Hyperlink fetch failed for %s/%s: %s", sheet_id[:8], tab_name, e)
        return []





COL = {
    "efj": 0, "move_type": 1, "container": 2, "bol": 3,
    "ssl": 4, "carrier": 5, "origin": 6, "destination": 7,
    "eta": 8, "lfd": 9, "pickup": 10, "delivery": 11,
    "status": 12, "notes": 13, "bot_alert": 14, "return_port": 15,
}

BOT_SERVICES = [
    # Import/Export are cron jobs (7:30 AM & 1:30 PM M-F), not systemd services
    # {"unit": "csl-import", "name": "Dray Import Scanner", "poll_min": 180},
    # {"unit": "csl-export", "name": "Dray Export Scanner", "poll_min": 60},
    # {"unit": "csl-ftl", "name": "FTL / MacroPoint Tracker", "poll_min": 30},  # now cron job
    {"unit": "csl-boviet", "name": "Boviet Monitor", "poll_min": 20},
    {"unit": "csl-tolead", "name": "Tolead Monitor", "poll_min": 20},
    {"unit": "csl-upload", "name": "Upload Server", "poll_min": 0},
]

REP_STYLES = {
    "Eli": {"color": "var(--accent-blue)", "bg": "linear-gradient(135deg,#3b82f6,#2563eb)", "initials": "EL"},
    "Radka": {"color": "var(--accent-green)", "bg": "linear-gradient(135deg,#22c55e,#16a34a)", "initials": "RK"},
    "John F": {"color": "var(--accent-purple)", "bg": "linear-gradient(135deg,#8b5cf6,#7c3aed)", "initials": "JF"},
    "Janice": {"color": "var(--accent-cyan)", "bg": "linear-gradient(135deg,#06b6d4,#0891b2)", "initials": "JC"},
    "Boviet": {"color": "var(--accent-amber)", "bg": "linear-gradient(135deg,#f59e0b,#d97706)", "initials": "BV"},
    "Tolead": {"color": "var(--accent-red)", "bg": "linear-gradient(135deg,#ef4444,#dc2626)", "initials": "TL"},
}

# ---------------------------------------------------------------------------
# File upload validation
# ---------------------------------------------------------------------------
ALLOWED_EXTENSIONS = {".pdf", ".png", ".jpg", ".jpeg", ".xlsx", ".xls", ".doc", ".docx"}
MAX_UPLOAD_SIZE = 25 * 1024 * 1024  # 25 MB


def _sanitize_filename(filename: str) -> str:
    """Strip path components and dangerous characters from a filename."""
    name = Path(filename).name
    name = re.sub(r'[^\w\-.]', '_', name)
    name = re.sub(r'_{2,}', '_', name)
    name = re.sub(r'\.{2,}', '.', name)
    if not name or name.startswith('.'):
        name = f"upload_{int(_time.time())}"
    return name

# ---------------------------------------------------------------------------
# Sheet cache — reads all tabs, caches in memory
# ---------------------------------------------------------------------------

class SheetCache:
    def __init__(self):
        self.shipments = []
        self.rep_map = {}
        self.stats = {"active": 0, "on_schedule": 0, "eta_changed": 0, "at_risk": 0, "completed_today": 0}
        self.accounts = []
        self.team = {}
        self._lock = threading.Lock()
        self._last = 0

    def refresh_if_needed(self):
        if _time.time() - self._last < CACHE_TTL:
            return
        with self._lock:
            if _time.time() - self._last < CACHE_TTL:
                return
            try:
                self._do_refresh()
            except Exception as e:
                log.error("Sheet refresh failed: %s", e)

    def _do_refresh(self):
        creds = Credentials.from_service_account_file(
            CREDS_FILE, scopes=["https://www.googleapis.com/auth/spreadsheets"]
        )
        gc = gspread.authorize(creds)
        sh = gc.open_by_key(SHEET_ID)

        # Load account rep mapping (1 API call)
        try:
            rep_rows = sh.worksheet("Account Rep").get_all_values()
            self.rep_map = {}
            for r in rep_rows[2:]:
                if r[0].strip() and len(r) > 1 and r[1].strip():
                    self.rep_map[r[0].strip()] = r[1].strip()
        except Exception:
            pass

        # --- BATCH read all master tabs (1 API call instead of ~36) ---
        all_shipments = []
        tabs = [ws.title for ws in sh.worksheets() if ws.title not in SKIP_TABS]
        ranges = [f"'{t}'!A:P" for t in tabs]
        try:
            batch_result = sh.values_batch_get(ranges)
            value_ranges = batch_result.get("valueRanges", [])
            for vr, tab_name in zip(value_ranges, tabs):
                rows = vr.get("values", [])
                if len(rows) < 2:
                    continue
                hdr_idx = 0
                if len(rows) > 1:
                    r0 = sum(1 for c in rows[0] if c.strip())
                    r1 = sum(1 for c in rows[1] if c.strip())
                    if r1 > r0:
                        hdr_idx = 1
                for row in rows[hdr_idx + 1:]:
                    efj = row[COL["efj"]].strip() if len(row) > COL["efj"] else ""
                    ctr = row[COL["container"]].strip() if len(row) > COL["container"] else ""
                    if not efj and not ctr:
                        continue
                    def cell(key, r=row):
                        idx = COL[key]
                        return r[idx].strip() if len(r) > idx else ""
                    all_shipments.append({
                        "account": tab_name,
                        "efj": efj, "move_type": cell("move_type"),
                        "container": ctr, "bol": cell("bol"),
                        "ssl": cell("ssl"), "carrier": cell("carrier"),
                        "origin": cell("origin"), "destination": cell("destination"),
                        "eta": cell("eta"), "lfd": cell("lfd"),
                        "pickup": cell("pickup"), "delivery": cell("delivery"),
                        "status": cell("status"), "notes": cell("notes"),
                        "bot_alert": cell("bot_alert"), "return_port": cell("return_port"),
                        "container_url": "",
                        "rep": self.rep_map.get(tab_name, "Unassigned"),
                    })
        except Exception as e:
            log.warning("Master batch read failed: %s", e)

        _time.sleep(2)  # breathing room between sheets

        # --- BATCH read Boviet tabs (2 API calls: metadata + batch values) ---
        try:
            bov_sh = gc.open_by_key(BOVIET_SHEET_ID)
            bov_tabs = [ws.title for ws in bov_sh.worksheets()
                        if ws.title not in BOVIET_SKIP_TABS and ws.title in BOVIET_TAB_CONFIGS]
            bov_ranges = [f"'{t}'!A:Z" for t in bov_tabs]
            bov_batch = bov_sh.values_batch_get(bov_ranges)
            bov_value_ranges = bov_batch.get("valueRanges", [])
            for vr, tab_name in zip(bov_value_ranges, bov_tabs):
                try:
                    cfg = BOVIET_TAB_CONFIGS[tab_name]
                    rows = vr.get("values", [])
                    # Get hyperlinks separately (1 call per tab — needed for Macropoint URLs)
                    bov_links = _get_sheet_hyperlinks(creds, BOVIET_SHEET_ID, tab_name)
                    bov_start = cfg.get("start_row", 1)
                    for ri, row in enumerate(rows[1:], start=1):
                        if ri + 1 < bov_start:
                            continue
                        efj = row[cfg["efj_col"]].strip() if len(row) > cfg["efj_col"] else ""
                        load_id = row[cfg["load_id_col"]].strip() if len(row) > cfg["load_id_col"] else ""
                        status = row[cfg["status_col"]].strip() if len(row) > cfg["status_col"] else ""
                        if not efj or status.lower() in BOVIET_DONE_STATUSES:
                            continue
                        bov_mp_url = ""
                        if ri < len(bov_links) and len(bov_links[ri]) > cfg["efj_col"]:
                            bov_mp_url = bov_links[ri][cfg["efj_col"]] or ""
                        # Extract optional fields from config
                        bov_pickup = ""
                        bov_delivery = ""
                        bov_phone = ""
                        bov_trailer = ""
                        bov_origin = cfg.get("default_origin", "")
                        bov_dest = cfg.get("default_dest", "")
                        if "pickup_col" in cfg:
                            bov_pickup = row[cfg["pickup_col"]].strip() if len(row) > cfg["pickup_col"] else ""
                        if "delivery_col" in cfg:
                            bov_delivery = row[cfg["delivery_col"]].strip() if len(row) > cfg["delivery_col"] else ""
                        if "phone_col" in cfg:
                            bov_phone = row[cfg["phone_col"]].strip() if len(row) > cfg["phone_col"] else ""
                        if "trailer_col" in cfg:
                            bov_trailer = row[cfg["trailer_col"]].strip() if len(row) > cfg["trailer_col"] else ""
                        bov_carrier_email = ""
                        bov_driver_name = ""
                        if "carrier_email_col" in cfg:
                            bov_carrier_email = row[cfg["carrier_email_col"]].strip() if len(row) > cfg["carrier_email_col"] else ""
                        if "driver_name_col" in cfg:
                            bov_driver_name = row[cfg["driver_name_col"]].strip() if len(row) > cfg["driver_name_col"] else ""

                        all_shipments.append({
                            "account": "Boviet", "efj": efj, "move_type": "FTL",
                            "container": load_id, "bol": "", "ssl": "",
                            "carrier": "", "origin": bov_origin, "destination": bov_dest,
                            "eta": "", "lfd": "",
                            "pickup": bov_pickup, "delivery": bov_delivery,
                            "status": status, "notes": "", "bot_alert": "",
                            "return_port": "", "rep": "Boviet",
                            "container_url": bov_mp_url,
                            "driver": bov_trailer,
                            "driver_phone": bov_phone,
                            "carrier_email": bov_carrier_email,
                            "driver_name": bov_driver_name,
                            "hub": tab_name,
                        })
                    _time.sleep(1)
                except Exception as e:
                    log.warning("Boviet tab %s: %s", tab_name, e)
        except Exception as e:
            log.warning("Boviet sheet read failed: %s", e)

        _time.sleep(2)  # breathing room before Tolead

        # ORD now handled via TOLEAD_HUB_CONFIGS (legacy block removed)

        # --- Read Tolead JFK / LAX / DFW sheets ---
        for hub_name, hub_cfg in TOLEAD_HUB_CONFIGS.items():
            try:
                _time.sleep(1)
                hub_sh = gc.open_by_key(hub_cfg["sheet_id"])
                hub_ws = hub_sh.worksheet(hub_cfg["tab"])
                hub_rows = hub_ws.get_all_values()
                # Fetch hyperlinks for MP URLs
                hub_links = _get_sheet_hyperlinks(_goog_creds, hub_cfg["sheet_id"], hub_cfg["tab"])
                cols = hub_cfg["cols"]
                hub_count = 0
                hub_start = hub_cfg.get("start_row", 1)
                for ri, row in enumerate(hub_rows[1:], start=1):
                    if ri + 1 < hub_start:
                        continue
                    def _cell(idx, r=row):
                        if idx is None:
                            return ""
                        return r[idx].strip() if len(r) > idx else ""
                    efj = _cell(cols["efj"])
                    load_id = _cell(cols["load_id"])
                    status = _cell(cols["status"])

                    # Hub-specific status derivation
                    if hub_name == "DFW":
                        if not load_id:
                            continue
                        if status and status.lower() in TOLEAD_SKIP_STATUSES:
                            continue
                        col_j = row[9].strip() if len(row) > 9 else ""
                        if col_j.lower() not in ("scheduled", "picked"):
                            status = "Needs to Cover"
                        elif not status:
                            status = col_j.capitalize()
                    elif hub_name == "ORD":
                        if not load_id:
                            continue
                        if status and status.lower() in TOLEAD_SKIP_STATUSES:
                            continue
                        if status and status.lower() == "new":
                            status = "Needs to Cover"
                        elif not status:
                            continue
                    elif hub_name == "LAX":
                        if not load_id:
                            continue
                        if status and status.lower() in TOLEAD_SKIP_STATUSES:
                            continue
                        if status and status.lower() == "unassigned":
                            status = "Needs to Cover"
                        elif not status:
                            continue
                    elif hub_name == "JFK":
                        if not load_id:
                            continue
                        if status and status.lower() in TOLEAD_SKIP_STATUSES:
                            continue
                        if status and status.lower() == "new":
                            status = "Needs to Cover"
                        elif not status:
                            continue
                    else:
                        if not efj and not load_id:
                            continue
                        if status and status in TOLEAD_SKIP_STATUSES:
                            continue
                        if not status:
                            continue

                    origin = _shorten_address(_cell(cols["origin"]) or hub_cfg["default_origin"])
                    pickup_date = _cell(cols["pickup_date"])
                    pickup_time = _cell(cols["pickup_time"])
                    pickup = f"{pickup_date} {pickup_time}".strip() if pickup_date else ""
                    delivery = _cell(cols["delivery"])
                    driver_trailer = _cell(cols.get("driver")) if cols.get("driver") is not None else ""
                    driver_phone = _cell(cols.get("phone")) if cols.get("phone") is not None else ""

                    # Extract MP URL from hyperlinks
                    _mp_url = ""
                    _efj_col = cols["efj"]
                    if hub_links and ri < len(hub_links):
                        _lr = hub_links[ri]
                        if _efj_col < len(_lr) and _lr[_efj_col]:
                            _mp_url = _lr[_efj_col]

                    all_shipments.append({
                        "account": "Tolead", "efj": efj or load_id,
                        "move_type": "FTL", "container": load_id, "bol": "",
                        "ssl": "", "carrier": "",
                        "origin": origin,
                        "destination": _shorten_address(_cell(cols["destination"])),
                        "eta": pickup_date, "lfd": "",
                        "pickup": pickup, "delivery": delivery,
                        "status": status, "notes": "", "bot_alert": "",
                        "return_port": "", "rep": "Tolead",
                        "container_url": _mp_url,
                        "hub": hub_name,
                        "driver": "",
                        "driver_phone": driver_phone,
                    })
                    hub_count += 1
                log.info("Tolead %s: %d active loads", hub_name, hub_count)
            except Exception as e:
                log.warning("Tolead %s sheet read failed: %s", hub_name, e)

        self.shipments = all_shipments

        # Enrich shipments with invoiced status from DB
        try:
            invoiced_map = db.get_invoiced_map()
        except Exception:
            invoiced_map = {}
        for s in self.shipments:
            s["_invoiced"] = invoiced_map.get(s["efj"], False)

        self._compute_stats()
        self._last = _time.time()
        log.info("Sheet cache: %d shipments", len(all_shipments))

    def _compute_stats(self):
        now = datetime.now()
        today = now.strftime("%Y-%m-%d")
        tomorrow = (now + timedelta(days=1)).strftime("%Y-%m-%d")

        active = on_sched = eta_chg = at_risk = done_today = 0
        acct_data = {}
        team_data = {}

        for s in self.shipments:
            sl = s["status"].lower()
            is_done = any(w in sl for w in ("delivered", "completed", "empty return"))

            acct = acct_data.setdefault(s["account"], {"active": 0, "alerts": 0, "done": 0})

            if is_done:
                acct["done"] += 1
                if today in s.get("delivery", ""):
                    done_today += 1
                continue

            active += 1
            acct["active"] += 1

            risk = False
            if s["lfd"] and s["lfd"][:10] <= tomorrow:
                risk = True
            if not s["eta"] and not s["status"]:
                risk = True

            if risk:
                at_risk += 1
                acct["alerts"] += 1
            else:
                on_sched += 1

            if s["bot_alert"] and today in s["bot_alert"]:
                eta_chg += 1

            rep = s.get("rep", "Unassigned")
            td = team_data.setdefault(rep, {"loads": 0, "accounts": set(), "at_risk": 0})
            td["loads"] += 1
            td["accounts"].add(s["account"])
            if risk:
                td["at_risk"] += 1

        self.stats = {
            "active": active, "on_schedule": on_sched,
            "eta_changed": eta_chg, "at_risk": at_risk,
            "completed_today": done_today,
        }
        self.accounts = sorted(
            [{"name": k, **v} for k, v in acct_data.items()],
            key=lambda x: x["active"], reverse=True,
        )
        self.team = {
            rep: {"loads": d["loads"], "accounts": sorted(d["accounts"]), "at_risk": d["at_risk"]}
            for rep, d in team_data.items()
        }


sheet_cache = SheetCache()


# ---------------------------------------------------------------------------
# Alert generation from sheet data
# ---------------------------------------------------------------------------

def _generate_alerts(shipments, limit=10):
    alerts = []
    now = datetime.now()
    today = now.strftime("%Y-%m-%d")
    tomorrow = (now + timedelta(days=1)).strftime("%Y-%m-%d")

    for s in shipments:
        sl = s["status"].lower()
        if any(w in sl for w in ("delivered", "completed", "empty return")):
            continue
        alert = None
        label = s["container"] or s["efj"]

        if s["lfd"] and s["lfd"][:10] <= today:
            alert = {
                "title": f"{s['account'].upper()} \u2014 {label}",
                "desc": f"LFD is <span style='color:var(--accent-red);font-weight:700'>TODAY ({s['lfd'][:10]})</span> \u00B7 Status: {s['status'] or 'Unknown'} \u00B7 Carrier: {s['carrier'] or 'TBD'} \u2014 <span style='color:var(--accent-red);font-weight:700'>PICK UP TODAY</span>",
                "type": "urgent", "icon": "\U0001f525",
                "move": s["move_type"], "rep": s["rep"], "efj": s["efj"], "account": s["account"],
            }
        elif s["lfd"] and s["lfd"][:10] <= tomorrow:
            alert = {
                "title": f"{s['account'].upper()} \u2014 {label}",
                "desc": f"LFD is <span style='color:var(--accent-amber);font-weight:700'>TOMORROW ({s['lfd'][:10]})</span> \u00B7 Status: {s['status'] or 'Unknown'} \u00B7 Carrier: {s['carrier'] or 'TBD'} \u2014 <span style='color:var(--accent-amber);font-weight:700'>SCHEDULE PICKUP</span>",
                "type": "warning", "icon": "\u26a0\ufe0f",
                "move": s["move_type"], "rep": s["rep"], "efj": s["efj"], "account": s["account"],
            }
        elif s["bot_alert"] and today in s["bot_alert"]:
            alert = {
                "title": f"{s['account'].upper()} \u2014 {label}",
                "desc": f"ETA: {s['eta'] or 'N/A'} \u00B7 Vessel: {s['ssl'] or 'Unknown'} \u00B7 {s['origin']} \u2192 {s['destination']}",
                "type": "info", "icon": "\U0001f916",
                "move": s["move_type"], "rep": s["rep"], "efj": s["efj"], "account": s["account"],
            }
        elif not s["eta"] and not s["status"] and s["container"]:
            alert = {
                "title": f"{s['account'].upper()} \u2014 {label}",
                "desc": f"No tracking data \u00B7 SSL: {s['ssl'] or 'Unknown'} \u00B7 Container: {s['container']}",
                "type": "info", "icon": "\U0001f4e1",
                "move": s["move_type"], "rep": s["rep"], "efj": s["efj"], "account": s["account"],
            }

        if alert:
            alerts.append(alert)

    priority = {"urgent": 0, "warning": 1, "info": 2}
    alerts.sort(key=lambda a: priority.get(a["type"], 3))
    return alerts[:limit]


# ---------------------------------------------------------------------------
# Bot status with timing from journal
# ---------------------------------------------------------------------------

def _get_service_status(unit):
    try:
        r = subprocess.run(["systemctl", "is-active", unit], capture_output=True, text=True, timeout=5)
        return r.stdout.strip()
    except Exception:
        return "unknown"


def _get_bot_status_detailed():
    results = []
    for svc in BOT_SERVICES:
        unit = svc["unit"]
        status = _get_service_status(unit)
        last_run = next_run = ""

        try:
            r = subprocess.run(
                ["journalctl", "-u", unit, "-n", "1", "--no-pager", "-o", "short"],
                capture_output=True, text=True, timeout=5,
            )
            line = r.stdout.strip().split("\n")[-1] if r.stdout.strip() else ""
            m = re.match(r"(\w+ \d+ \d+:\d+:\d+)", line)
            if m:
                ts = datetime.strptime(f"2026 {m.group(1)}", "%Y %b %d %H:%M:%S")
                mins = int((datetime.now() - ts).total_seconds() / 60)
                last_run = "just now" if mins < 1 else (f"{mins} min ago" if mins < 60 else f"{mins // 60} hr ago")
                if svc["poll_min"] > 0:
                    nm = svc["poll_min"] - mins
                    next_run = "overdue" if nm < 0 else (f"{nm} min" if nm < 60 else f"{nm // 60} hr {nm % 60} min")
        except Exception:
            pass

        results.append({
            "name": svc["name"], "unit": unit, "status": status,
            "last_run": last_run or "unknown",
            "next_run": next_run if svc["poll_min"] > 0 else "",
            "poll_min": svc["poll_min"],
        })
    return results


# ---------------------------------------------------------------------------
# Recent bot actions from journal
# ---------------------------------------------------------------------------

def _get_recent_actions(limit=8):
    actions = []
    for svc in BOT_SERVICES[:5]:
        try:
            r = subprocess.run(
                ["journalctl", "-u", svc["unit"], "-n", "40", "--no-pager", "-o", "short"],
                capture_output=True, text=True, timeout=10,
            )
            for line in r.stdout.strip().split("\n"):
                if not line.strip():
                    continue
                tm = re.match(r"\w+ \d+ (\d+:\d+)", line)
                time_str = tm.group(1) if tm else ""
                action = status_badge = None

                if "No changes in" in line:
                    tab = re.search(r"No changes in '(\w+)'", line)
                    if tab:
                        action = f"Scanned {tab.group(1)} \u2014 no updates"
                        status_badge = ("Done", "green")
                elif "Tab:" in line:
                    tab = re.search(r"Tab: (\w+)", line)
                    if tab:
                        action = f"Scanning {tab.group(1)} account"
                        status_badge = ("Active", "cyan")
                elif "[Dray Import]" in line:
                    m = re.search(r"Container: (\w+)", line)
                    if m:
                        action = f"Tracking {m.group(1)}"
                        status_badge = ("Written", "blue")
                elif "email" in line.lower() or "alert" in line.lower() or "Alert" in line:
                    action = line.split("]:")[-1].strip()[:60] if "]:" in line else line[-60:]
                    status_badge = ("Sent", "amber")

                if action and status_badge:
                    try:
                        t = datetime.strptime(time_str, "%H:%M")
                        time_str = t.strftime("%I:%M %p")
                    except Exception:
                        pass
                    actions.append({"time": time_str, "action": action, "status": status_badge[0], "color": status_badge[1]})
        except Exception:
            pass
    return actions[:limit]


# ---------------------------------------------------------------------------
# CSS
# ---------------------------------------------------------------------------

GOOGLE_FONTS = '<link href="https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@300;400;500;600;700&family=Plus+Jakarta+Sans:wght@300;400;500;600;700;800&display=swap" rel="stylesheet">'

DASHBOARD_CSS = """
:root {
  --bg-deep: #0a0d12;
  --bg-surface: #111822;
  --bg-card: #161e2c;
  --bg-card-hover: #1a2435;
  --bg-elevated: #1e2a3d;
  --accent-blue: #3b82f6;
  --accent-cyan: #06b6d4;
  --accent-green: #22c55e;
  --accent-amber: #f59e0b;
  --accent-red: #ef4444;
  --accent-purple: #8b5cf6;
  --text-primary: #e8ecf4;
  --text-secondary: #7b8ba3;
  --text-dim: #4a5568;
  --border: #1e2a3d;
  --border-subtle: #162030;
  --glow-blue: rgba(59, 130, 246, 0.15);
}
* { margin: 0; padding: 0; box-sizing: border-box; }
body {
  background: var(--bg-deep);
  color: var(--text-primary);
  font-family: 'Plus Jakarta Sans', -apple-system, BlinkMacSystemFont, sans-serif;
  min-height: 100vh;
  overflow-x: hidden;
}
body::before {
  content: '';
  position: fixed; inset: 0;
  background-image:
    linear-gradient(rgba(59,130,246,0.03) 1px, transparent 1px),
    linear-gradient(90deg, rgba(59,130,246,0.03) 1px, transparent 1px);
  background-size: 60px 60px;
  pointer-events: none; z-index: 0;
}
a { color: var(--accent-blue); text-decoration: none; }
a:hover { text-decoration: underline; }
.sidebar {
  position: fixed; left: 0; top: 0; bottom: 0; width: 72px;
  background: var(--bg-surface); border-right: 1px solid var(--border);
  display: flex; flex-direction: column; align-items: center; padding: 20px 0; z-index: 100;
}
.sidebar-logo {
  width: 42px; height: 42px;
  border-radius: 12px; display: flex; align-items: center; justify-content: center;
  margin-bottom: 32px; overflow: hidden;
  box-shadow: 0 0 20px rgba(0,222,180,0.3);
}
.sidebar-logo img { width: 42px; height: 42px; border-radius: 12px; object-fit: cover; }
.sidebar-nav { display: flex; flex-direction: column; gap: 8px; flex: 1; }
.nav-item {
  width: 44px; height: 44px; border-radius: 12px;
  display: flex; align-items: center; justify-content: center;
  color: var(--text-dim); cursor: pointer; transition: all 0.2s;
  position: relative; text-decoration: none;
}
.nav-item:hover { background: var(--bg-card); color: var(--text-secondary); text-decoration: none; }
.nav-item.active { background: var(--glow-blue); color: var(--accent-blue); }
.nav-item.active::before {
  content: ''; position: absolute; left: -14px;
  width: 3px; height: 20px; background: var(--accent-blue); border-radius: 0 3px 3px 0;
}
.nav-item svg { width: 20px; height: 20px; }
.sidebar-bottom { margin-top: auto; }
.main { margin-left: 72px; min-height: 100vh; position: relative; z-index: 1; }
.topbar {
  padding: 16px 32px; border-bottom: 1px solid var(--border);
  display: flex; align-items: center; justify-content: space-between;
  background: rgba(10,13,18,0.8); backdrop-filter: blur(12px);
  position: sticky; top: 0; z-index: 50;
}
.topbar-left { display: flex; align-items: center; gap: 20px; }
.topbar-title { font-weight: 700; font-size: 20px; letter-spacing: -0.02em; }
.topbar-title span { color: var(--accent-cyan); }
.live-badge {
  display: flex; align-items: center; gap: 6px;
  background: rgba(34,197,94,0.1); border: 1px solid rgba(34,197,94,0.2);
  padding: 4px 12px; border-radius: 20px; font-size: 11px; font-weight: 600;
  color: var(--accent-green); text-transform: uppercase; letter-spacing: 0.05em;
}
.live-dot { width: 7px; height: 7px; background: var(--accent-green); border-radius: 50%; animation: pulse 2s infinite; }
@keyframes pulse {
  0%,100% { opacity:1; box-shadow:0 0 0 0 rgba(34,197,94,0.4); }
  50% { opacity:0.7; box-shadow:0 0 0 6px rgba(34,197,94,0); }
}
.topbar-right { display: flex; align-items: center; gap: 16px; }
.topbar-time { font-family: 'JetBrains Mono', monospace; font-size: 12px; color: var(--text-secondary); }
.avatar {
  width: 34px; height: 34px; border-radius: 10px;
  background: linear-gradient(135deg, var(--accent-purple), var(--accent-blue));
  display: flex; align-items: center; justify-content: center; font-weight: 700; font-size: 13px;
}
.search-box { position: relative; }
.search-box input {
  width: 260px; background: var(--bg-card); border: 1px solid var(--border);
  border-radius: 10px; padding: 8px 14px 8px 36px; color: var(--text-primary);
  font-size: 12px; font-family: 'Plus Jakarta Sans', sans-serif;
}
.search-box input::placeholder { color: var(--text-dim); }
.search-box svg { position: absolute; left: 12px; top: 50%; transform: translateY(-50%); width: 14px; height: 14px; color: var(--text-dim); }
.content { padding: 24px 32px; }
.stats-grid { display: grid; grid-template-columns: repeat(5, 1fr); gap: 16px; margin-bottom: 24px; }
.stat-card {
  background: var(--bg-card); border: 1px solid var(--border); border-radius: 14px;
  padding: 20px; position: relative; overflow: hidden; transition: all 0.25s;
}
.stat-card:hover { border-color: rgba(59,130,246,0.3); transform: translateY(-2px); box-shadow: 0 8px 24px rgba(0,0,0,0.3); }
.stat-card::after { content: ''; position: absolute; top: 0; left: 0; right: 0; height: 2px; }
.stat-card.blue::after { background: linear-gradient(90deg, var(--accent-blue), var(--accent-cyan)); }
.stat-card.green::after { background: linear-gradient(90deg, var(--accent-green), #34d399); }
.stat-card.amber::after { background: linear-gradient(90deg, var(--accent-amber), #fbbf24); }
.stat-card.red::after { background: linear-gradient(90deg, var(--accent-red), #f87171); }
.stat-card.purple::after { background: linear-gradient(90deg, var(--accent-purple), #a78bfa); }
.stat-label { font-size: 11px; font-weight: 600; color: var(--text-dim); text-transform: uppercase; letter-spacing: 0.08em; margin-bottom: 8px; }
.stat-value { font-family: 'JetBrains Mono', monospace; font-size: 28px; font-weight: 700; letter-spacing: -0.03em; }
.stat-card.blue .stat-value { color: var(--accent-blue); }
.stat-card.green .stat-value { color: var(--accent-green); }
.stat-card.amber .stat-value { color: var(--accent-amber); }
.stat-card.red .stat-value { color: var(--accent-red); }
.stat-card.purple .stat-value { color: var(--accent-purple); }
.stat-sub { font-size: 11px; color: var(--text-dim); margin-top: 4px; }
.two-col { display: grid; grid-template-columns: 1fr 380px; gap: 20px; margin-bottom: 24px; }
.panel { background: var(--bg-card); border: 1px solid var(--border); border-radius: 14px; overflow: hidden; }
.panel-header { padding: 16px 20px; border-bottom: 1px solid var(--border); display: flex; align-items: center; justify-content: space-between; }
.panel-title { font-weight: 700; font-size: 14px; display: flex; align-items: center; gap: 8px; }
.panel-badge { background: rgba(239,68,68,0.15); color: var(--accent-red); font-family: 'JetBrains Mono', monospace; font-size: 11px; font-weight: 600; padding: 2px 8px; border-radius: 6px; }
.panel-badge.green { background: rgba(34,197,94,0.15); color: var(--accent-green); }
.filter-tabs { display: flex; gap: 6px; }
.filter-tab { padding: 5px 14px; border-radius: 8px; font-size: 12px; font-weight: 500; border: 1px solid var(--border); background: transparent; color: var(--text-secondary); cursor: pointer; transition: all 0.2s; }
.filter-tab.active { background: var(--accent-blue); color: #fff; border-color: var(--accent-blue); }
.alert-list { max-height: 420px; overflow-y: auto; }
.alert-item { padding: 14px 20px; border-bottom: 1px solid var(--border-subtle); display: flex; gap: 12px; transition: background 0.15s; }
.alert-item:hover { background: var(--bg-card-hover); }
.alert-icon { width: 36px; height: 36px; border-radius: 10px; display: flex; align-items: center; justify-content: center; font-size: 16px; flex-shrink: 0; }
.alert-icon.urgent { background: rgba(239,68,68,0.15); }
.alert-icon.warning { background: rgba(245,158,11,0.15); }
.alert-icon.info { background: rgba(59,130,246,0.12); }
.alert-body { flex: 1; min-width: 0; }
.alert-title-row { display: flex; align-items: center; justify-content: space-between; margin-bottom: 3px; }
.alert-name { font-weight: 600; font-size: 13px; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }
.alert-time { font-family: 'JetBrains Mono', monospace; font-size: 10px; color: var(--text-dim); flex-shrink: 0; }
.alert-desc { font-size: 12px; color: var(--text-secondary); line-height: 1.5; }
.alert-tags { display: flex; gap: 6px; margin-top: 6px; }
.alert-tag { font-family: 'JetBrains Mono', monospace; font-size: 10px; padding: 2px 8px; border-radius: 4px; background: var(--bg-elevated); color: var(--text-dim); }
.alert-tag.efj { background: rgba(59,130,246,0.15); color: var(--accent-blue); }
.alert-tag.import { background: rgba(6,182,212,0.15); color: var(--accent-cyan); }
.alert-tag.export { background: rgba(34,197,94,0.15); color: var(--accent-green); }
.alert-tag.ftl { background: rgba(139,92,246,0.15); color: var(--accent-purple); }
.alert-tag.rep { background: rgba(245,158,11,0.12); color: var(--accent-amber); }
.account-list { padding: 8px; }
.account-row { display: grid; grid-template-columns: 1fr 55px 55px 55px; gap: 4px; align-items: center; padding: 10px 12px; border-radius: 8px; font-size: 12px; transition: background 0.15s; }
.account-row:hover { background: var(--bg-card-hover); }
.account-row.header { font-weight: 600; color: var(--text-dim); text-transform: uppercase; font-size: 10px; letter-spacing: 0.06em; padding-bottom: 4px; border-bottom: 1px solid var(--border-subtle); }
.account-name { font-weight: 600; display: flex; align-items: center; gap: 6px; }
.account-dot { width: 6px; height: 6px; border-radius: 50%; }
.account-dot.ok { background: var(--accent-green); }
.account-dot.warn { background: var(--accent-amber); }
.account-dot.alert { background: var(--accent-red); }
.cell-count { text-align: center; font-family: 'JetBrains Mono', monospace; font-size: 12px; font-weight: 500; }
.three-col { display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 20px; }
.bot-row { display: flex; align-items: center; justify-content: space-between; padding: 12px 20px; border-bottom: 1px solid var(--border-subtle); }
.bot-info { display: flex; align-items: center; gap: 10px; }
.bot-dot { width: 8px; height: 8px; border-radius: 50%; }
.bot-dot.running { background: var(--accent-green); box-shadow: 0 0 8px rgba(34,197,94,0.5); }
.bot-dot.stopped { background: var(--accent-red); }
.bot-dot.unknown { background: var(--text-dim); }
.bot-name { font-size: 13px; font-weight: 600; }
.bot-detail { font-family: 'JetBrains Mono', monospace; font-size: 11px; color: var(--text-dim); }
.action-row { display: grid; grid-template-columns: 85px 1fr 80px; gap: 8px; padding: 10px 20px; border-bottom: 1px solid var(--border-subtle); font-size: 12px; align-items: center; }
.action-time { font-family: 'JetBrains Mono', monospace; font-size: 11px; color: var(--text-dim); }
.action-desc { color: var(--text-secondary); }
.action-status { font-family: 'JetBrains Mono', monospace; font-size: 11px; font-weight: 600; text-align: right; }
.team-row { display: flex; align-items: center; gap: 12px; padding: 12px 20px; border-bottom: 1px solid var(--border-subtle); }
.team-avatar { width: 32px; height: 32px; border-radius: 8px; display: flex; align-items: center; justify-content: center; font-weight: 700; font-size: 11px; color: #fff; flex-shrink: 0; }
.team-details { flex: 1; min-width: 0; }
.team-name { font-size: 13px; font-weight: 600; display: flex; align-items: center; justify-content: space-between; margin-bottom: 4px; }
.team-count { font-family: 'JetBrains Mono', monospace; font-size: 12px; font-weight: 600; }
.team-bar-bg { height: 6px; background: var(--bg-elevated); border-radius: 3px; overflow: hidden; margin-bottom: 4px; }
.team-bar-fill { height: 100%; border-radius: 3px; transition: width 0.6s ease; }
.team-sub { font-size: 10px; color: var(--text-dim); }
::-webkit-scrollbar { width: 6px; }
::-webkit-scrollbar-track { background: transparent; }
::-webkit-scrollbar-thumb { background: var(--border); border-radius: 3px; }
@keyframes fadeUp { from { opacity:0; transform:translateY(12px); } to { opacity:1; transform:translateY(0); } }
.stats-grid { animation: fadeUp 0.4s ease both; }
.two-col { animation: fadeUp 0.5s ease 0.1s both; }
.three-col { animation: fadeUp 0.5s ease 0.2s both; }
@media (max-width: 1200px) {
  .stats-grid { grid-template-columns: repeat(3, 1fr); }
  .two-col { grid-template-columns: 1fr; }
  .three-col { grid-template-columns: 1fr; }
}
table { width: 100%; border-collapse: collapse; font-size: 0.875rem; }
thead th { background: var(--bg-surface); color: var(--text-secondary); font-weight: 600; text-transform: uppercase; font-size: 0.7rem; letter-spacing: 0.06em; padding: 10px 14px; text-align: left; border-bottom: 1px solid var(--border); white-space: nowrap; }
tbody tr { border-bottom: 1px solid var(--border); transition: background 0.1s; }
tbody tr:hover { background: var(--bg-card-hover); }
tbody td { padding: 10px 14px; white-space: nowrap; }
.doc-yes { color: var(--accent-green); }
.doc-no { color: var(--accent-red); }
.table-wrap { padding: 16px 24px; overflow-x: auto; }
.match-form { display: inline-flex; gap: 6px; align-items: center; }
.match-form input { background: var(--bg-surface); border: 1px solid var(--border); border-radius: 6px; color: var(--text-primary); padding: 5px 10px; font-size: 0.8rem; width: 120px; }
.match-form button { background: var(--accent-blue); color: #fff; border: none; border-radius: 6px; padding: 5px 12px; font-size: 0.8rem; cursor: pointer; }
.btn-ignore { background: var(--bg-elevated) !important; color: var(--text-secondary) !important; }
/* Clickable items */
.stat-card { cursor: pointer; }
.alert-item { cursor: pointer; }
.account-name a { color: var(--text-primary); text-decoration: none; display: flex; align-items: center; gap: 6px; }
.account-name a:hover { color: var(--accent-blue); text-decoration: none; }
/* Slide-out detail panel */
.detail-overlay {
  position: fixed; inset: 0; background: rgba(0,0,0,0.5); z-index: 200;
  opacity: 0; pointer-events: none; transition: opacity 0.3s;
}
.detail-overlay.open { opacity: 1; pointer-events: auto; }
.detail-panel {
  position: fixed; top: 0; right: -460px; bottom: 0; width: 440px;
  background: var(--bg-surface); border-left: 1px solid var(--border);
  z-index: 201; transition: right 0.3s ease; overflow-y: auto;
  box-shadow: -8px 0 32px rgba(0,0,0,0.4);
}
.detail-panel.open { right: 0; }
.panel-close {
  position: absolute; top: 16px; right: 16px; width: 32px; height: 32px;
  border-radius: 8px; background: var(--bg-card); border: 1px solid var(--border);
  color: var(--text-secondary); cursor: pointer; display: flex; align-items: center;
  justify-content: center; font-size: 18px; transition: all 0.2s; z-index: 10;
}
.panel-close:hover { background: var(--bg-elevated); color: var(--text-primary); }
.panel-head {
  padding: 24px 20px; border-bottom: 1px solid var(--border);
  background: linear-gradient(180deg, var(--bg-card), var(--bg-surface));
}
.panel-head-title { font-size: 18px; font-weight: 700; margin-bottom: 4px; }
.panel-head-sub { font-size: 12px; color: var(--text-secondary); }
.panel-section { padding: 16px 20px; border-bottom: 1px solid var(--border-subtle); }
.panel-section-title {
  font-size: 11px; font-weight: 700; color: var(--text-dim); text-transform: uppercase;
  letter-spacing: 0.08em; margin-bottom: 12px;
}
.panel-field {
  display: flex; justify-content: space-between; align-items: center;
  padding: 6px 0; font-size: 13px;
}
.panel-field-label { color: var(--text-secondary); }
.panel-field-value { font-weight: 600; font-family: 'JetBrains Mono', monospace; font-size: 12px; }
.doc-item {
  display: flex; align-items: center; gap: 12px; padding: 10px 14px;
  border-radius: 10px; margin-bottom: 8px; transition: all 0.15s;
}
.doc-item.received { background: rgba(34,197,94,0.08); border: 1px solid rgba(34,197,94,0.2); }
.doc-item.missing { background: rgba(239,68,68,0.06); border: 1px solid rgba(239,68,68,0.15); }
.doc-status {
  width: 28px; height: 28px; border-radius: 8px; display: flex; align-items: center;
  justify-content: center; font-size: 14px; font-weight: 700; flex-shrink: 0;
}
.doc-item.received .doc-status { background: rgba(34,197,94,0.15); color: var(--accent-green); }
.doc-item.missing .doc-status { background: rgba(239,68,68,0.12); color: var(--accent-red); }
.doc-info { flex: 1; min-width: 0; }
.doc-type { font-weight: 600; font-size: 13px; }
.doc-filename { font-size: 11px; color: var(--text-secondary); margin-top: 2px; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }
.doc-action a, .doc-action label {
  font-size: 11px; font-weight: 600; padding: 4px 12px; border-radius: 6px; cursor: pointer; transition: all 0.15s;
}
.doc-action a { background: rgba(34,197,94,0.15); color: var(--accent-green); text-decoration: none; }
.doc-action a:hover { background: rgba(34,197,94,0.25); text-decoration: none; }
.doc-action label { background: rgba(59,130,246,0.15); color: var(--accent-blue); }
.doc-action label:hover { background: rgba(59,130,246,0.25); }
.doc-action input[type="file"] { display: none; }
.panel-loading { padding: 60px 20px; text-align: center; color: var(--text-dim); }
.highlight-flash { animation: highlightFlash 1.5s ease; }
@keyframes highlightFlash { 0% { background: rgba(59,130,246,0.15); } 100% { background: transparent; } }
"""

# ---------------------------------------------------------------------------
# Sidebar + Topbar + Page wrapper
# ---------------------------------------------------------------------------

SIDEBAR_SVG = {
    "dashboard": '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><rect x="3" y="3" width="7" height="7" rx="1"/><rect x="14" y="3" width="7" height="7" rx="1"/><rect x="3" y="14" width="7" height="7" rx="1"/><rect x="14" y="14" width="7" height="7" rx="1"/></svg>',
    "shipments": '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M21 16V8a2 2 0 00-1-1.73l-7-4a2 2 0 00-2 0l-7 4A2 2 0 003 8v8a2 2 0 001 1.73l7 4a2 2 0 002 0l7-4A2 2 0 0021 16z"/><path d="M3.27 6.96L12 12.01l8.73-5.05"/><path d="M12 22.08V12"/></svg>',
    "unmatched": '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M18 8A6 6 0 006 8c0 7-3 9-3 9h18s-3-2-3-9"/><path d="M13.73 21a2 2 0 01-3.46 0"/></svg>',
    "docs": '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M14 2H6a2 2 0 00-2 2v16a2 2 0 002 2h12a2 2 0 002-2V8z"/><path d="M14 2v6h6"/><path d="M16 13H8"/><path d="M16 17H8"/><path d="M10 9H8"/></svg>',
    "settings": '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><circle cx="12" cy="12" r="3"/><path d="M19.4 15a1.65 1.65 0 00.33 1.82l.06.06a2 2 0 010 2.83 2 2 0 01-2.83 0l-.06-.06a1.65 1.65 0 00-1.82-.33 1.65 1.65 0 00-1 1.51V21a2 2 0 01-4 0v-.09A1.65 1.65 0 009 19.4a1.65 1.65 0 00-1.82.33l-.06.06a2 2 0 01-2.83 0 2 2 0 010-2.83l.06-.06A1.65 1.65 0 004.68 15a1.65 1.65 0 00-1.51-1H3a2 2 0 010-4h.09A1.65 1.65 0 004.6 9a1.65 1.65 0 00-.33-1.82l-.06-.06a2 2 0 012.83-2.83l.06.06A1.65 1.65 0 009 4.68a1.65 1.65 0 001-1.51V3a2 2 0 014 0v.09a1.65 1.65 0 001 1.51 1.65 1.65 0 001.82-.33l.06-.06a2 2 0 012.83 2.83l-.06.06A1.65 1.65 0 0019.4 9a1.65 1.65 0 001.51 1H21a2 2 0 010 4h-.09a1.65 1.65 0 00-1.51 1z"/></svg>',
}

SEARCH_SVG = '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><circle cx="11" cy="11" r="8"/><path d="M21 21l-4.35-4.35"/></svg>'


def _sidebar(active="dashboard"):
    items = [
        ("dashboard", "/", "Dashboard"),
        ("shipments", "/shipments", "Shipments"),
        ("unmatched", "/unmatched", "Unmatched Emails"),
        ("docs", "#", "Documents"),
    ]
    nav = ""
    for key, href, title in items:
        cls = "nav-item active" if key == active else "nav-item"
        nav += f'<a class="{cls}" href="{href}" title="{title}">{SIDEBAR_SVG[key]}</a>'
    return f"""<nav class="sidebar">
  <div class="sidebar-logo"><img src="/logo.svg" alt="CSL" width="42" height="42"></div>
  <div class="sidebar-nav">{nav}</div>
  <div class="sidebar-bottom"><div class="nav-item" title="Settings">{SIDEBAR_SVG["settings"]}</div></div>
</nav>"""


def _topbar(title_main="AI", title_accent="Dispatch", search=True):
    search_html = ""
    if search:
        search_html = f'<div class="search-box">{SEARCH_SVG}<input type="text" placeholder="Search container, pro#, account..."></div>'
    return f"""<header class="topbar">
  <div class="topbar-left">
    <div class="topbar-title">{title_main} <span>{title_accent}</span></div>
    <div class="live-badge"><div class="live-dot"></div> System Live</div>
  </div>
  <div class="topbar-right">
    {search_html}
    <div class="topbar-time" id="live-clock"></div>
    <div class="avatar">JF</div>
  </div>
</header>"""


CLOCK_SCRIPT = """
function updateClock() {
  const now = new Date();
  const days = ['SUN','MON','TUE','WED','THU','FRI','SAT'];
  const months = ['JAN','FEB','MAR','APR','MAY','JUN','JUL','AUG','SEP','OCT','NOV','DEC'];
  const d = days[now.getDay()];
  const m = months[now.getMonth()];
  const day = now.getDate();
  let h = now.getHours();
  const ampm = h >= 12 ? 'PM' : 'AM';
  h = h % 12 || 12;
  const min = String(now.getMinutes()).padStart(2, '0');
  document.getElementById('live-clock').textContent = d + ' ' + m + ' ' + day + ' \\u00B7 ' + h + ':' + min + ' ' + ampm + ' ET';
}
updateClock();
setInterval(updateClock, 30000);
"""


def _page(title, body, script=""):
    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{title}</title>
{GOOGLE_FONTS}
<style>{DASHBOARD_CSS}</style>
</head>
<body>
{body}
<script>{CLOCK_SCRIPT}
{script}
</script>
</body>
</html>"""


# ---------------------------------------------------------------------------
# Dashboard HTML builders
# ---------------------------------------------------------------------------

def _build_stats_html(s):
    total = s["active"] + s["completed_today"]
    pct = round(s["on_schedule"] / s["active"] * 100, 1) if s["active"] else 0
    return f"""<div class="stats-grid">
  <div class="stat-card blue"><div class="stat-label">Active Shipments</div><div class="stat-value">{s['active']}</div><div class="stat-sub">Across {len(sheet_cache.accounts)} accounts</div></div>
  <div class="stat-card green"><div class="stat-label">On Schedule</div><div class="stat-value">{s['on_schedule']}</div><div class="stat-sub">{pct}% of active loads</div></div>
  <div class="stat-card amber"><div class="stat-label">ETA Changed</div><div class="stat-value">{s['eta_changed']}</div><div class="stat-sub">Updated today</div></div>
  <div class="stat-card red"><div class="stat-label">At Risk</div><div class="stat-value">{s['at_risk']}</div><div class="stat-sub">LFD soon or no data</div></div>
  <div class="stat-card purple"><div class="stat-label">Completed Today</div><div class="stat-value">{s['completed_today']}</div><div class="stat-sub">{total} total tracked</div></div>
</div>"""


def _build_alerts_html(alerts):
    if not alerts:
        items = '<div style="padding:40px;text-align:center;color:var(--text-dim);">All shipments on schedule</div>'
    else:
        items = ""
        for a in alerts:
            move = a.get("move", "")
            move_tag = "import" if "import" in move.lower() else ("export" if "export" in move.lower() else ("ftl" if "ftl" in move.lower() else ""))
            tags = f'<span class="alert-tag efj">{a["efj"]}</span>' if a.get("efj") else ""
            if move_tag:
                tags += f' <span class="alert-tag {move_tag}">{move}</span>'
            if a.get("rep") and a["rep"] != "Unassigned":
                tags += f' <span class="alert-tag rep">{a["rep"]}</span>'
            efj_attr = f' data-efj="{a["efj"]}"' if a.get("efj") else ""
            acct_attr = f' data-account="{a.get("account", "")}"'
            move_attr = f' data-move="{move}"' if move else ""
            items += f"""<div class="alert-item"{efj_attr}{move_attr}{acct_attr} onclick="openPanel('{a.get('efj','')}')" title="Click for details">
  <div class="alert-icon {a['type']}">{a['icon']}</div>
  <div class="alert-body">
    <div class="alert-title-row"><span class="alert-name">{a['title']}</span></div>
    <div class="alert-desc">{a['desc']}</div>
    <div class="alert-tags">{tags}</div>
  </div>
</div>"""

    badge = f'<span class="panel-badge">{len(alerts)} new</span>' if alerts else '<span class="panel-badge green">0</span>'
    filters = '<div class="filter-tabs"><button class="filter-tab active">All</button><button class="filter-tab">Imports</button><button class="filter-tab">Exports</button><button class="filter-tab">FTL</button><button class="filter-tab">Boviet</button><button class="filter-tab">Tolead</button></div>'

    return f"""<div class="panel">
  <div class="panel-header">
    <div class="panel-title">Live Alerts {badge}</div>
    {filters}
  </div>
  <div class="alert-list">{items}</div>
</div>"""


def _build_accounts_html(accounts):
    rows = ""
    for a in accounts:
        name = a["name"]
        dot_cls = "alert" if a["alerts"] > 2 else ("warn" if a["alerts"] > 0 else "ok")
        alert_style = f' style="color:var(--accent-red)"' if a["alerts"] else ""
        done_style = f' style="color:var(--accent-green)"' if a["done"] else ""
        rows += f"""<div class="account-row">
  <span class="account-name"><a href="/shipments?account={name}"><span class="account-dot {dot_cls}"></span>{name}</a></span>
  <span class="cell-count">{a['active']}</span>
  <span class="cell-count"{alert_style}>{a['alerts']}</span>
  <span class="cell-count"{done_style}>{a['done']}</span>
</div>"""

    toggle = '<div class="filter-tabs"><button class="filter-tab active">Active</button><button class="filter-tab">All</button></div>'
    return f"""<div class="panel">
  <div class="panel-header"><div class="panel-title">Account Overview</div>{toggle}</div>
  <div class="account-list">
    <div class="account-row header">
      <span>Account</span><span style="text-align:center">Active</span><span style="text-align:center">Alerts</span><span style="text-align:center">Done</span>
    </div>
    {rows}
  </div>
</div>"""


def _build_bots_html(bots):
    rows = ""
    for b in bots:
        status = b["status"]
        dot = "running" if status == "active" else ("stopped" if status in ("inactive", "failed") else "unknown")
        detail = f"Last run: {b['last_run']}"
        if b["next_run"]:
            detail += f" \u00B7 Next: {b['next_run']}"
        rows += f"""<div class="bot-row">
  <div class="bot-info">
    <div class="bot-dot {dot}"></div>
    <div><div class="bot-name">{b['name']}</div><div class="bot-detail">{detail}</div></div>
  </div>
</div>"""

    return f"""<div class="panel">
  <div class="panel-header"><div class="panel-title">Bot Status</div></div>
  {rows}
</div>"""


def _build_actions_html(actions):
    if not actions:
        rows = '<div style="padding:20px;text-align:center;color:var(--text-dim);">No recent actions</div>'
    else:
        header = '<div class="action-row" style="border-bottom:1px solid var(--border);"><span class="action-time" style="color:var(--text-dim);font-weight:600;text-transform:uppercase;font-size:10px;letter-spacing:0.06em;">Time</span><span style="color:var(--text-dim);font-weight:600;text-transform:uppercase;font-size:10px;letter-spacing:0.06em;">Action</span><span style="text-align:right;color:var(--text-dim);font-weight:600;text-transform:uppercase;font-size:10px;letter-spacing:0.06em;">Status</span></div>'
        rows = header
        color_map = {"green": "var(--accent-green)", "blue": "var(--accent-blue)", "cyan": "var(--accent-cyan)", "amber": "var(--accent-amber)", "purple": "var(--accent-purple)"}
        for a in actions:
            c = color_map.get(a["color"], "var(--text-dim)")
            rows += f"""<div class="action-row">
  <span class="action-time">{a['time']}</span>
  <span class="action-desc">{a['action']}</span>
  <span class="action-status" style="color:{c};">\u2713 {a['status']}</span>
</div>"""

    return f"""<div class="panel">
  <div class="panel-header"><div class="panel-title">Recent Bot Actions</div></div>
  {rows}
</div>"""


def _build_team_html(team_data):
    max_loads = max((d["loads"] for d in team_data.values()), default=1) or 1
    total = sum(d["loads"] for d in team_data.values())
    rows = ""
    for rep in sorted(team_data, key=lambda r: team_data[r]["loads"], reverse=True):
        d = team_data[rep]
        info = REP_STYLES.get(rep, {"color": "var(--text-dim)", "bg": "linear-gradient(135deg,#4a5568,#374151)", "initials": "??"})
        pct = int((d["loads"] / max_loads) * 100)
        acct_list = ", ".join(d["accounts"][:5])
        sub = f"{d['at_risk']} at risk \u00B7 {acct_list}"
        rows += f"""<div class="team-row">
  <div class="team-avatar" style="background:{info['bg']}">{info['initials']}</div>
  <div class="team-details">
    <div class="team-name"><span><a href="/rep/{rep}" style="color:var(--text-primary);text-decoration:none;" onmouseover="this.style.color='{info['color']}'" onmouseout="this.style.color='var(--text-primary)'">{rep}</a></span><span class="team-count" style="color:{info['color']}">{d['loads']} Loads</span></div>
    <div class="team-bar-bg"><div class="team-bar-fill" style="width:{pct}%;background:{info['color']}"></div></div>
    <div class="team-sub">{sub}</div>
  </div>
</div>"""

    return f"""<div class="panel">
  <div class="panel-header">
    <div class="panel-title">Team Load Distribution</div>
    <span style="font-family:'JetBrains Mono',monospace;font-size:11px;color:var(--text-dim);">{total} total</span>
  </div>
  {rows}
</div>"""


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------


# ---------------------------------------------------------------------------
# Auth routes
# ---------------------------------------------------------------------------
def _auth_page_style():
    return """
    <style>
    * { margin: 0; padding: 0; box-sizing: border-box; }
    body {
        background: #0a0d12; color: #e8ecf4;
        font-family: 'Plus Jakarta Sans', -apple-system, sans-serif;
        min-height: 100vh; display: flex; align-items: center; justify-content: center;
    }
    .auth-card {
        background: #161e2c; border: 1px solid #1e2a3d; border-radius: 16px;
        padding: 48px; width: 420px; max-width: 90vw;
        box-shadow: 0 20px 60px rgba(0,0,0,0.5);
    }
    .auth-card h1 { font-size: 24px; margin-bottom: 8px; }
    .auth-card p { color: #7b8ba3; margin-bottom: 28px; font-size: 14px; }
    .auth-card label { display: block; font-size: 13px; color: #7b8ba3; margin-bottom: 6px; font-weight: 500; }
    .auth-card input {
        width: 100%; padding: 12px 16px; background: #0a0d12; border: 1px solid #1e2a3d;
        border-radius: 8px; color: #e8ecf4; font-size: 15px; margin-bottom: 18px; outline: none;
    }
    .auth-card input:focus { border-color: #3b82f6; }
    .auth-card button {
        width: 100%; padding: 14px; background: linear-gradient(135deg, #3b82f6, #2563eb);
        border: none; border-radius: 8px; color: #fff; font-size: 15px; font-weight: 600; cursor: pointer;
    }
    .auth-card button:hover { transform: translateY(-1px); box-shadow: 0 8px 24px rgba(59,130,246,0.3); }
    .error-msg {
        background: rgba(239,68,68,0.1); border: 1px solid rgba(239,68,68,0.3);
        color: #ef4444; padding: 10px 14px; border-radius: 8px; font-size: 13px; margin-bottom: 18px;
    }
    .logo-row { display: flex; align-items: center; gap: 12px; margin-bottom: 28px; }
    .logo-row img { width: 36px; height: 36px; border-radius: 8px; }
    .logo-row span { font-size: 18px; font-weight: 700; color: #3b82f6; }
    </style>
    <link href="https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;500;600;700&display=swap" rel="stylesheet">
    """


@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request, error: str = Query(default="")):
    if not auth.is_configured():
        return RedirectResponse("/setup", status_code=302)
    token = request.cookies.get("csl_session")
    if auth.verify_session_token(token):
        return RedirectResponse("/", status_code=302)
    error_html = f'<div class="error-msg">{error}</div>' if error else ""
    return f"""<!DOCTYPE html><html><head><title>Login - CSL Dispatch</title>{_auth_page_style()}</head><body>
    <div class="auth-card">
        <div class="logo-row"><img src="/logo.svg" alt="logo"><span>CSL Dispatch</span></div>
        <h1>Sign In</h1>
        <p>Enter your credentials to access the dashboard.</p>
        {error_html}
        <form method="POST" action="/login">
            <label>Username</label>
            <input name="username" type="text" required autocomplete="username" autofocus>
            <label>Password</label>
            <input name="password" type="password" required autocomplete="current-password">
            <button type="submit">Sign In</button>
        </form>
    </div></body></html>"""


@app.post("/login")
def login_submit(request: Request, username: str = Form(...), password: str = Form(...)):
    ip = request.headers.get("x-real-ip", request.client.host)
    locked, remaining = auth.check_lockout(ip)
    if locked:
        return RedirectResponse(f"/login?error=Too+many+attempts.+Try+again+in+{remaining}+seconds.", status_code=302)
    if auth.verify_login(username, password):
        auth.clear_failed_attempts(ip)
        token = auth.create_session_token(username)
        resp = RedirectResponse("/", status_code=302)
        resp.set_cookie("csl_session", token, max_age=86400*7, httponly=True, secure=True, samesite="lax")
        return resp
    auth.record_failed_attempt(ip)
    return RedirectResponse("/login?error=Invalid+username+or+password.", status_code=302)


@app.get("/logout")
def logout():
    resp = RedirectResponse("/login", status_code=302)
    resp.delete_cookie("csl_session")
    return resp


@app.get("/setup", response_class=HTMLResponse)
def setup_page(error: str = Query(default="")):
    if auth.is_configured():
        return RedirectResponse("/login", status_code=302)
    error_html = f'<div class="error-msg">{error}</div>' if error else ""
    return f"""<!DOCTYPE html><html><head><title>Setup - CSL Dispatch</title>{_auth_page_style()}</head><body>
    <div class="auth-card">
        <div class="logo-row"><img src="/logo.svg" alt="logo"><span>CSL Dispatch</span></div>
        <h1>Create Admin Account</h1>
        <p>Set your username and password. This can only be done once.</p>
        {error_html}
        <form method="POST" action="/setup">
            <label>Username</label>
            <input name="username" type="text" required autofocus placeholder="e.g. admin">
            <label>Password</label>
            <input name="password" type="password" required minlength="8" placeholder="Min 8 characters">
            <label>Confirm Password</label>
            <input name="confirm" type="password" required minlength="8">
            <button type="submit">Create Account</button>
        </form>
    </div></body></html>"""


@app.post("/setup")
def setup_submit(username: str = Form(...), password: str = Form(...), confirm: str = Form(...)):
    if auth.is_configured():
        return RedirectResponse("/login", status_code=302)
    if len(password) < 8:
        return RedirectResponse("/setup?error=Password+must+be+at+least+8+characters.", status_code=302)
    if password != confirm:
        return RedirectResponse("/setup?error=Passwords+do+not+match.", status_code=302)
    auth.setup_password(username, password)
    return RedirectResponse("/login", status_code=302)


# ---------------------------------------------------------------------------
# Rep Dashboard - Account Cards with Drill-Down
# ---------------------------------------------------------------------------
@app.get("/rep/{rep_name}", response_class=HTMLResponse)
def rep_dashboard(rep_name: str):
    """Dedicated dashboard: account cards with per-account stats, click to expand load tables."""
    sheet_cache.refresh_if_needed()

    rep_shipments = [s for s in sheet_cache.shipments if s.get("rep", "Unassigned") == rep_name]
    if not rep_shipments:
        body = f"""{_sidebar("dashboard")}
<div class="main">
  {_topbar(rep_name, "Dashboard", search=False)}
  <div class="content">
    <div style="padding:48px;text-align:center;color:var(--text-dim);">No loads found for {rep_name}</div>
  </div>
</div>"""
        return HTMLResponse(_page(f"{rep_name} - CSL Dispatch", body))

    now = datetime.now()
    today = now.strftime("%Y-%m-%d")
    tomorrow = (now + timedelta(days=1)).strftime("%Y-%m-%d")

    # Get invoiced map from DB
    try:
        invoiced_map = db.get_invoiced_map()
    except Exception:
        invoiced_map = {}

    # Build per-account data
    account_data = {}
    for s in rep_shipments:
        sl = s["status"].lower()
        is_done = any(w in sl for w in ("delivered", "completed", "empty return"))
        if is_done:
            continue

        acct = s["account"]
        if acct not in account_data:
            account_data[acct] = {"active": 0, "at_risk": 0, "on_sched": 0, "unbilled": 0, "shipments": []}

        inv = invoiced_map.get(s["efj"], False)
        s["_invoiced"] = inv

        account_data[acct]["active"] += 1
        account_data[acct]["shipments"].append(s)

        risk = False
        if s["lfd"] and s["lfd"][:10] <= tomorrow:
            risk = True
        if not s["eta"] and not s["status"]:
            risk = True
        if risk:
            account_data[acct]["at_risk"] += 1
        else:
            account_data[acct]["on_sched"] += 1

        if not inv:
            account_data[acct]["unbilled"] += 1

    info = REP_STYLES.get(rep_name, {"color": "var(--accent-blue)", "bg": "linear-gradient(135deg,#3b82f6,#2563eb)", "initials": "??"})
    num_accts = len(account_data)
    min_w = "180px" if num_accts > 6 else "240px"

    # --- Account cards ---
    cards_html = ""
    for acct_name in sorted(account_data.keys(), key=lambda a: account_data[a]["active"], reverse=True):
        ad = account_data[acct_name]
        risk_color = "var(--accent-red)" if ad["at_risk"] > 0 else "var(--text-dim)"
        unbill_color = "var(--accent-amber)" if ad["unbilled"] > 0 else "var(--text-dim)"

        # Build load table rows for this account
        load_rows = ""
        for s in ad["shipments"]:
            sl = s["status"].lower()
            lfd_style = ""
            if s["lfd"] and s["lfd"][:10] <= tomorrow:
                lfd_style = ' style="color:var(--accent-red);font-weight:600"'
            status_color = "var(--accent-green)"
            if "at port" in sl or "available" in sl:
                status_color = "var(--accent-amber)"
            elif "in transit" in sl or "on vessel" in sl:
                status_color = "var(--accent-blue)"
            elif "out for" in sl or "picked up" in sl:
                status_color = "var(--accent-cyan)"

            inv_val = s.get("_invoiced", False)
            inv_sel_yes = " selected" if inv_val else ""
            inv_sel_no = "" if inv_val else " selected"
            row_bg = "background:rgba(34,197,94,0.06);" if inv_val else ""

            ctr_display = s["container"]
            if s.get("container_url"):
                ctr_display = f'<a href="{s["container_url"]}" target="_blank" style="color:var(--accent-cyan);">{s["container"]}</a>'

            efj_link = f'<a href="javascript:void(0)" onclick="openPanel(\'{s["efj"]}\')" style="color:var(--accent-blue);cursor:pointer;">{s["efj"]}</a>'

            load_rows += f"""<tr style="{row_bg}" data-efj="{s['efj']}">
  <td>{efj_link}</td>
  <td style="font-family:JetBrains Mono,monospace;font-size:12px">{ctr_display}</td>
  <td><span style="color:{status_color}">{s['status'] or '-'}</span></td>
  <td>{s['eta'] or '-'}</td>
  <td{lfd_style}>{s['lfd'] or '-'}</td>
  <td>{s['move_type'] or '-'}</td>
  <td><select class="inv-select" data-efj="{s['efj']}" onchange="toggleInvoiced(this)" style="background:var(--bg-surface);border:1px solid var(--border);border-radius:6px;color:var(--text-primary);padding:3px 6px;font-size:11px;cursor:pointer;">
    <option value="false"{inv_sel_no}>Unbilled</option>
    <option value="true"{inv_sel_yes}>Invoiced</option>
  </select></td>
</tr>"""

        safe_id = acct_name.replace(" ", "_").replace("/", "_")
        cards_html += f"""<div class="stat-card" style="cursor:pointer;padding:16px;" onclick="toggleAcct('{safe_id}')">
  <div style="font-weight:700;font-size:14px;margin-bottom:10px;">{acct_name}</div>
  <div style="display:grid;grid-template-columns:1fr 1fr;gap:4px;font-size:12px;">
    <div><span style="color:var(--text-dim);">Active</span> <span style="font-weight:700;color:var(--accent-blue);">{ad['active']}</span></div>
    <div><span style="color:var(--text-dim);">On Sched</span> <span style="font-weight:700;color:var(--accent-green);">{ad['on_sched']}</span></div>
    <div><span style="color:var(--text-dim);">At Risk</span> <span style="font-weight:700;color:{risk_color};">{ad['at_risk']}</span></div>
    <div><span style="color:var(--text-dim);">Unbilled</span> <span style="font-weight:700;color:{unbill_color};">{ad['unbilled']}</span></div>
  </div>
</div>
<div id="acct-{safe_id}" class="acct-expand" style="display:none;grid-column:1/-1;">
  <div class="panel" style="margin-bottom:12px;">
    <div class="panel-header"><div class="panel-title">{acct_name} &mdash; {ad['active']} Active Loads</div></div>
    <div class="table-wrap">
      <table>
        <thead><tr><th>EFJ #</th><th>Container/ID</th><th>Status</th><th>ETA</th><th>LFD</th><th>Type</th><th>Invoiced</th></tr></thead>
        <tbody>{load_rows}</tbody>
      </table>
    </div>
  </div>
</div>"""

    panel_html = """
<div class="detail-overlay" id="detail-overlay" onclick="closePanel()"></div>
<div class="detail-panel" id="detail-panel">
  <button class="panel-close" onclick="closePanel()">&times;</button>
  <div id="panel-content"><div class="panel-loading">Select a shipment to view details</div></div>
</div>"""

    body = f"""{_sidebar("dashboard")}
<div class="main">
  {_topbar(rep_name, "Dashboard", search=False)}
  <div class="content">
    <div style="display:flex;align-items:center;gap:16px;margin-bottom:24px;">
      <div style="width:48px;height:48px;border-radius:14px;background:{info['bg']};display:flex;align-items:center;justify-content:center;font-weight:700;font-size:18px;color:#fff;">{info['initials']}</div>
      <div>
        <h2 style="font-size:22px;font-weight:700;">{rep_name}</h2>
        <div style="font-size:13px;color:var(--text-secondary);">{num_accts} accounts &middot; {sum(d['active'] for d in account_data.values())} active loads</div>
      </div>
      <a href="/" style="margin-left:auto;font-size:12px;color:var(--text-secondary);border:1px solid var(--border);padding:6px 14px;border-radius:8px;">&larr; Back to Overview</a>
    </div>
    <div style="display:grid;grid-template-columns:repeat(auto-fill,minmax({min_w},1fr));gap:12px;margin-bottom:24px;">
      {cards_html}
    </div>
  </div>
</div>
{panel_html}"""

    interactive_js = r"""
function toggleAcct(id) {
  var el = document.getElementById('acct-' + id);
  if (!el) return;
  if (el.style.display === 'none') {
    // Close all others first
    document.querySelectorAll('.acct-expand').forEach(function(e) { e.style.display = 'none'; });
    el.style.display = 'block';
    el.scrollIntoView({behavior: 'smooth', block: 'nearest'});
  } else {
    el.style.display = 'none';
  }
}

async function toggleInvoiced(sel) {
  var efj = sel.dataset.efj;
  var val = sel.value === 'true';
  try {
    var res = await fetch('/api/load/' + encodeURIComponent(efj) + '/invoiced', {
      method: 'POST', headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({invoiced: val})
    });
    if (res.ok) {
      var row = sel.closest('tr');
      row.style.background = val ? 'rgba(34,197,94,0.06)' : '';
      row.style.transition = 'background 0.3s';
    }
  } catch(e) { console.error('Invoiced toggle failed:', e); }
}

function openPanel(efj) {
  if (!efj) return;
  document.getElementById('detail-panel').classList.add('open');
  document.getElementById('detail-overlay').classList.add('open');
  loadPanel(efj);
}
function closePanel() {
  document.getElementById('detail-panel').classList.remove('open');
  document.getElementById('detail-overlay').classList.remove('open');
}
document.addEventListener('keydown', function(e) { if (e.key === 'Escape') closePanel(); });

async function loadPanel(efj) {
  var pc = document.getElementById('panel-content');
  pc.innerHTML = '<div class="panel-loading">Loading ' + efj + '...</div>';
  try {
    var res = await fetch('/api/load/' + encodeURIComponent(efj));
    if (!res.ok) { pc.innerHTML = '<div class="panel-loading">Load not found</div>'; return; }
    var d = await res.json();
    var h = '<div class="panel-head"><div class="panel-head-title">' + d.efj + '</div>';
    h += '<div class="panel-head-sub">' + (d.account||'') + ' \u00B7 ' + (d.move_type||'') + '</div></div>';
    h += '<div class="panel-section"><div class="panel-section-title">Shipment Details</div>';
    var fields = [['Container/Load', d.container_url ? '<a href="' + d.container_url + '" target="_blank" style="color:#06b6d4">' + d.container + ' &#x2197;</a>' : d.container], ['BOL/Booking', d.bol], ['SSL/Vessel', d.ssl], ['Carrier', d.carrier], ['Origin', d.origin], ['Destination', d.destination], ['Status', d.status || 'Unknown'], ['Rep', d.rep || 'Unassigned']];
    if (d.container_url) { fields.push(['Macropoint', '<a href="' + d.container_url + '" target="_blank" style="color:#06b6d4">Track Shipment &#x2197;</a>']); }
    for (var i=0;i<fields.length;i++) { h += '<div class="panel-field"><span class="panel-field-label">' + fields[i][0] + '</span><span class="panel-field-value">' + (fields[i][1]||'-') + '</span></div>'; }
    h += '</div>';
    h += '<div class="panel-section"><div class="panel-section-title">Timeline</div>';
    var dates = [['ETA/ERD', d.eta], ['LFD/Cutoff', d.lfd], ['Pickup', d.pickup], ['Delivery', d.delivery]];
    for (var i=0;i<dates.length;i++) { h += '<div class="panel-field"><span class="panel-field-label">' + dates[i][0] + '</span><span class="panel-field-value">' + (dates[i][1]||'-') + '</span></div>'; }
    h += '</div>';
    if (d.bot_alert) { h += '<div class="panel-section"><div class="panel-section-title">Bot Notes</div><div style="font-size:12px;color:var(--text-secondary);white-space:pre-wrap;">' + d.bot_alert + '</div></div>'; }
    h += '<div class="panel-section"><div class="panel-section-title">Documents</div>';
    var docTypes = ['BOL','POD','Invoice'];
    for (var i=0;i<docTypes.length;i++) {
      var dt = docTypes[i]; var doc = null;
      if (d.documents) { for (var j=0;j<d.documents.length;j++) { if (d.documents[j].doc_type===dt) { doc=d.documents[j]; break; } } }
      if (doc && doc.filename) {
        h += '<div class="doc-item received"><div class="doc-status">\u2713</div><div class="doc-info"><div class="doc-type">' + dt + '</div><div class="doc-filename">' + doc.filename + '</div></div><div class="doc-action"><a href="/docs/' + doc.file_path + '" target="_blank">View</a></div></div>';
      } else {
        h += '<div class="doc-item missing"><div class="doc-status">\u2717</div><div class="doc-info"><div class="doc-type">' + dt + '</div><div class="doc-filename">Not received</div></div>';
        h += '<div class="doc-action"><input type="file" id="upload-' + dt + '-' + d.efj + '" accept=".pdf,.png,.jpg,.jpeg,.xlsx,.xls,.doc,.docx" onchange="uploadDoc(\'' + d.efj + '\',\'' + dt + '\',this)"><label for="upload-' + dt + '-' + d.efj + '">Upload</label></div></div>';
      }
    }
    h += '</div>';
    pc.innerHTML = h;
  } catch(e) { pc.innerHTML = '<div class="panel-loading">Error loading details</div>'; }
}

async function uploadDoc(efj, docType, input) {
  if (!input.files.length) return;
  var file = input.files[0];
  var allowed = ['.pdf','.png','.jpg','.jpeg','.xlsx','.xls','.doc','.docx'];
  var ext = file.name.substring(file.name.lastIndexOf('.')).toLowerCase();
  if (allowed.indexOf(ext)===-1) { alert('File type not allowed.'); input.value=''; return; }
  if (file.size > 25*1024*1024) { alert('File too large. Max 25 MB.'); input.value=''; return; }
  var label = input.nextElementSibling;
  var orig = label.textContent;
  label.textContent = 'Uploading...'; label.style.opacity = '0.6'; input.disabled = true;
  var fd = new FormData(); fd.append('file', file); fd.append('doc_type', docType);
  try {
    var res = await fetch('/api/load/' + encodeURIComponent(efj) + '/upload', {method:'POST', body:fd});
    if (res.ok) { label.textContent='Done!'; label.style.color='#16a34a'; setTimeout(function(){loadPanel(efj);},500); }
    else { var err = await res.json().catch(function(){return {detail:'Upload failed'};}); alert(err.detail||'Upload failed'); label.textContent=orig; label.style.opacity='1'; input.disabled=false; }
  } catch(e) { alert('Upload error: '+e.message); label.textContent=orig; label.style.opacity='1'; input.disabled=false; }
}
"""

    return HTMLResponse(_page(f"{rep_name} - CSL Dispatch", body, script=interactive_js))


@app.on_event("startup")
def startup():
    db.init_pool()

    # Create driver_contacts table if not exists
    try:
        with db.get_conn() as conn:
            with db.get_cursor(conn) as cur:
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS driver_contacts (
                        efj VARCHAR(32) PRIMARY KEY,
                        driver_name VARCHAR(120),
                        driver_phone VARCHAR(30),
                        driver_email VARCHAR(120),
                        notes TEXT,
                        created_at TIMESTAMPTZ DEFAULT NOW(),
                        updated_at TIMESTAMPTZ DEFAULT NOW()
                    )
                """)
        log.info("driver_contacts table ready")
    except Exception as e:
        log.warning("Could not create driver_contacts table: %s", e)

    # Create team_profiles table if not exists
    try:
        with db.get_conn() as conn:
            with db.get_cursor(conn) as cur:
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS team_profiles (
                        rep_name VARCHAR(64) PRIMARY KEY,
                        avatar_filename VARCHAR(256),
                        subtitle VARCHAR(256),
                        updated_at TIMESTAMPTZ DEFAULT NOW()
                    )
                """)
        log.info("team_profiles table ready")
    except Exception as e:
        log.warning("Could not create team_profiles table: %s", e)
    # Create quotes table
    try:
        db.create_quotes_table()
    except Exception as e:
        log.warning("Could not create quotes table: %s", e)

    # Create carriers table
    try:
        with db.get_conn() as conn:
            with db.get_cursor(conn) as cur:
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS carriers (
                        id              SERIAL PRIMARY KEY,
                        carrier_name    VARCHAR(256) NOT NULL,
                        mc_number       VARCHAR(32),
                        dot_number      VARCHAR(32),
                        contact_email   VARCHAR(256),
                        contact_phone   VARCHAR(64),
                        contact_name    VARCHAR(256),
                        regions         TEXT,
                        ports           TEXT,
                        rail_ramps      TEXT,
                        equipment_types TEXT,
                        notes           TEXT,
                        source          VARCHAR(32) DEFAULT 'manual',
                        pickup_area     VARCHAR(256),
                        destination_area VARCHAR(256),
                        date_quoted     DATE,
                        v_code          VARCHAR(64),
                        created_at      TIMESTAMPTZ DEFAULT NOW(),
                        updated_at      TIMESTAMPTZ DEFAULT NOW()
                    )
                """)
                cur.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_carriers_mc ON carriers(mc_number) WHERE mc_number IS NOT NULL AND mc_number != ''")
        log.info("carriers table ready")
    except Exception as e:
        log.warning("Could not create carriers table: %s", e)

    # Create warehouses + warehouse_rates tables
    try:
        with db.get_conn() as conn:
            with db.get_cursor(conn) as cur:
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS warehouses (
                        id              SERIAL PRIMARY KEY,
                        name            VARCHAR(256) NOT NULL,
                        mc_number       VARCHAR(32),
                        region          VARCHAR(64),
                        address         TEXT,
                        city            VARCHAR(128),
                        state           VARCHAR(4),
                        zip_code        VARCHAR(12),
                        contact_name    VARCHAR(256),
                        contact_email   VARCHAR(256),
                        contact_phone   VARCHAR(64),
                        services        TEXT,
                        notes           TEXT,
                        source          VARCHAR(32) DEFAULT 'manual',
                        pickup_area     VARCHAR(256),
                        destination_area VARCHAR(256),
                        date_quoted     DATE,
                        v_code          VARCHAR(64),
                        created_at      TIMESTAMPTZ DEFAULT NOW(),
                        updated_at      TIMESTAMPTZ DEFAULT NOW()
                    )
                """)
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS warehouse_rates (
                        id              SERIAL PRIMARY KEY,
                        warehouse_id    INTEGER REFERENCES warehouses(id) ON DELETE CASCADE,
                        rate_type       VARCHAR(64) NOT NULL,
                        rate_amount     DECIMAL(10,2),
                        unit            VARCHAR(32),
                        description     TEXT,
                        effective_date  DATE,
                        notes           TEXT,
                        created_at      TIMESTAMPTZ DEFAULT NOW()
                    )
                """)
                cur.execute("CREATE INDEX IF NOT EXISTS idx_warehouse_rates_wh ON warehouse_rates(warehouse_id)")
        log.info("warehouses + warehouse_rates tables ready")
    except Exception as e:
        log.warning("Could not create warehouse tables: %s", e)

    # Create lane_rates table
    try:
        with db.get_conn() as conn:
            with db.get_cursor(conn) as cur:
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS lane_rates (
                        id              SERIAL PRIMARY KEY,
                        port            VARCHAR(64),
                        destination     VARCHAR(256),
                        carrier_name    VARCHAR(256),
                        dray_rate       DECIMAL(10,2),
                        fsc             VARCHAR(32),
                        total           DECIMAL(10,2),
                        chassis_per_day DECIMAL(10,2),
                        prepull         DECIMAL(10,2),
                        storage_per_day DECIMAL(10,2),
                        detention       VARCHAR(64),
                        chassis_split   DECIMAL(10,2),
                        overweight      DECIMAL(10,2),
                        tolls           DECIMAL(10,2),
                        reefer          DECIMAL(10,2),
                        hazmat          DECIMAL(10,2),
                        all_in_total    DECIMAL(10,2),
                        rank            INTEGER,
                        equipment_type  VARCHAR(64),
                        move_type       VARCHAR(32) DEFAULT 'dray',
                        notes           TEXT,
                        source          VARCHAR(32) DEFAULT 'excel_import',
                        created_at      TIMESTAMPTZ DEFAULT NOW()
                    )
                """)
                cur.execute("CREATE INDEX IF NOT EXISTS idx_lane_rates_port ON lane_rates(port)")
                cur.execute("CREATE INDEX IF NOT EXISTS idx_lane_rates_carrier ON lane_rates(carrier_name)")
        log.info("lane_rates table ready")
    except Exception as e:
        log.warning("Could not create lane_rates table: %s", e)

    # Pre-populate sheet cache in background
    threading.Thread(target=sheet_cache.refresh_if_needed, daemon=True).start()
    log.info("Dashboard started")


@app.on_event("shutdown")
def shutdown():
    db.close_pool()


@app.get("/legacy", response_class=HTMLResponse)
def dashboard():
    sheet_cache.refresh_if_needed()

    stats = sheet_cache.stats
    accounts = sheet_cache.accounts
    alerts = _generate_alerts(sheet_cache.shipments, limit=10)
    bots = _get_bot_status_detailed()
    actions = _get_recent_actions(8)
    team = sheet_cache.team

    panel_html = """
<div class="detail-overlay" id="detail-overlay" onclick="closePanel()"></div>
<div class="detail-panel" id="detail-panel">
  <button class="panel-close" onclick="closePanel()">&times;</button>
  <div id="panel-content"><div class="panel-loading">Select a shipment to view details</div></div>
</div>"""

    body = f"""{_sidebar("dashboard")}
<div class="main">
  {_topbar()}
  <div class="content">
    {_build_stats_html(stats)}
    <div class="two-col">
      {_build_alerts_html(alerts)}
      {_build_accounts_html(accounts)}
    </div>
    <div class="three-col">
      {_build_bots_html(bots)}
      {_build_actions_html(actions)}
      {_build_team_html(team)}
    </div>
  </div>
</div>
{panel_html}"""

    interactive_js = r"""
// --- Slide-out Panel ---
function openPanel(efj) {
  if (!efj) return;
  document.getElementById('detail-panel').classList.add('open');
  document.getElementById('detail-overlay').classList.add('open');
  loadPanel(efj);
}
function closePanel() {
  document.getElementById('detail-panel').classList.remove('open');
  document.getElementById('detail-overlay').classList.remove('open');
}
document.addEventListener('keydown', function(e) { if (e.key === 'Escape') closePanel(); });

async function loadPanel(efj) {
  var pc = document.getElementById('panel-content');
  pc.innerHTML = '<div class="panel-loading">Loading ' + efj + '...</div>';
  try {
    var res = await fetch('/api/load/' + encodeURIComponent(efj));
    if (!res.ok) throw new Error('Not found');
    var data = await res.json();
    renderPanel(data);
  } catch(e) {
    pc.innerHTML = '<div class="panel-loading" style="color:var(--accent-red)">Could not load details for ' + efj + '</div>';
  }
}

function renderPanel(d) {
  var pc = document.getElementById('panel-content');
  var statusColor = 'var(--text-primary)';
  var sl = (d.status || '').toLowerCase();
  if (sl.includes('delivered') || sl.includes('completed')) statusColor = 'var(--accent-green)';
  else if (sl.includes('risk') || sl.includes('hold')) statusColor = 'var(--accent-red)';

  var h = '<div class="panel-head">';
  h += '<div class="panel-head-title">' + (d.efj || 'Unknown') + '</div>';
  h += '<div class="panel-head-sub">' + (d.account || '') + ' \u00B7 ' + (d.move_type || '') + '</div>';
  h += '</div>';

  // Shipment info
  h += '<div class="panel-section"><div class="panel-section-title">Shipment Details</div>';
  var fields = [
    ['Container / Load', d.container], ['BOL / Booking', d.bol],
    ['SSL / Vessel', d.ssl], ['Carrier', d.carrier],
    ['Origin', d.origin], ['Destination', d.destination],
    ['Status', '<span style="color:' + statusColor + ';font-weight:700">' + (d.status || 'Unknown') + '</span>'],
    ['Rep', d.rep || 'Unassigned']
  ];
  for (var i = 0; i < fields.length; i++) {
    h += '<div class="panel-field"><span class="panel-field-label">' + fields[i][0] + '</span><span class="panel-field-value">' + (fields[i][1] || '-') + '</span></div>';
  }
  h += '</div>';

  // Timeline
  h += '<div class="panel-section"><div class="panel-section-title">Timeline</div>';
  var dates = [['ETA / ERD', d.eta], ['LFD / Cutoff', d.lfd], ['Pickup', d.pickup], ['Delivery', d.delivery], ['Return to Port', d.return_port]];
  for (var i = 0; i < dates.length; i++) {
    var val = dates[i][1] || '-';
    var style = '';
    if (dates[i][0].includes('LFD') && val !== '-') {
      var today = new Date().toISOString().slice(0,10);
      if (val.slice(0,10) <= today) style = ' style="color:var(--accent-red);font-weight:700"';
    }
    h += '<div class="panel-field"><span class="panel-field-label">' + dates[i][0] + '</span><span class="panel-field-value"' + style + '>' + val + '</span></div>';
  }
  h += '</div>';

  // Bot notes
  if (d.notes || d.bot_alert) {
    h += '<div class="panel-section"><div class="panel-section-title">Notes</div>';
    if (d.notes) h += '<div style="font-size:12px;color:var(--text-secondary);line-height:1.6;margin-bottom:8px">' + d.notes + '</div>';
    if (d.bot_alert) h += '<div style="font-size:11px;color:var(--accent-cyan);font-family:JetBrains Mono,monospace">' + d.bot_alert + '</div>';
    h += '</div>';
  }

  // Document checklist
  h += '<div class="panel-section"><div class="panel-section-title">Document Checklist</div>';
  var docTypes = ['BOL', 'POD', 'Invoice'];
  for (var i = 0; i < docTypes.length; i++) {
    var dt = docTypes[i];
    var doc = null;
    if (d.documents) {
      for (var j = 0; j < d.documents.length; j++) {
        if (d.documents[j].doc_type === dt) { doc = d.documents[j]; break; }
      }
    }
    if (doc && doc.filename) {
      h += '<div class="doc-item received"><div class="doc-status">\u2713</div>';
      h += '<div class="doc-info"><div class="doc-type">' + dt + '</div><div class="doc-filename">' + doc.filename + '</div></div>';
      h += '<div class="doc-action"><a href="/docs/' + doc.file_path + '" target="_blank">Download</a></div></div>';
    } else {
      h += '<div class="doc-item missing"><div class="doc-status">\u2717</div>';
      h += '<div class="doc-info"><div class="doc-type">' + dt + '</div><div class="doc-filename">Not received</div></div>';
      h += '<div class="doc-action"><input type="file" id="upload-' + dt + '-' + d.efj + '" accept=".pdf,.png,.jpg,.jpeg,.xlsx,.xls,.doc,.docx" onchange="uploadDoc(\'' + d.efj + '\',\'' + dt + '\',this)">';
      h += '<label for="upload-' + dt + '-' + d.efj + '">Upload</label></div></div>';
    }
  }
  h += '</div>';

  pc.innerHTML = h;
}

async function uploadDoc(efj, docType, input) {
  if (!input.files.length) return;
  var file = input.files[0];
  var allowed = ['.pdf','.png','.jpg','.jpeg','.xlsx','.xls','.doc','.docx'];
  var ext = file.name.substring(file.name.lastIndexOf('.')).toLowerCase();
  if (allowed.indexOf(ext) === -1) {
    alert('File type not allowed. Allowed: ' + allowed.join(', '));
    input.value = '';
    return;
  }
  if (file.size > 25 * 1024 * 1024) {
    alert('File too large. Maximum size is 25 MB.');
    input.value = '';
    return;
  }
  var label = input.nextElementSibling;
  var origText = label.textContent;
  label.textContent = 'Uploading...';
  label.style.opacity = '0.6';
  input.disabled = true;
  var fd = new FormData();
  fd.append('file', file);
  fd.append('doc_type', docType);
  try {
    var res = await fetch('/api/load/' + encodeURIComponent(efj) + '/upload', {method:'POST', body:fd});
    if (res.ok) {
      label.textContent = 'Done!';
      label.style.color = '#16a34a';
      setTimeout(function() { loadPanel(efj); }, 500);
    } else {
      var err = await res.json().catch(function() { return {detail:'Upload failed'}; });
      alert(err.detail || 'Upload failed');
      label.textContent = origText;
      label.style.opacity = '1';
      input.disabled = false;
    }
  } catch(e) {
    alert('Upload error: ' + e.message);
    label.textContent = origText;
    label.style.opacity = '1';
    input.disabled = false;
  }
}

// --- Filter Tabs (Alerts) ---
document.querySelectorAll('.filter-tabs').forEach(function(group) {
  // Only handle the alert filter tabs (first filter-tabs group in two-col)
  if (!group.closest('.panel-header') || !group.closest('.two-col')) return;
  group.querySelectorAll('.filter-tab').forEach(function(tab) {
    tab.addEventListener('click', function() {
      group.querySelectorAll('.filter-tab').forEach(function(t){ t.classList.remove('active'); });
      this.classList.add('active');
      var filter = this.textContent.trim().toLowerCase();
      document.querySelectorAll('.alert-item').forEach(function(item) {
        if (filter === 'all') { item.style.display = ''; return; }
        var move = (item.dataset.move || '').toLowerCase();
        var acct = (item.dataset.account || '').toLowerCase();
        if (filter === 'imports' && move.includes('import')) item.style.display = '';
        else if (filter === 'exports' && move.includes('export')) item.style.display = '';
        else if (filter === 'ftl' && move.includes('ftl')) item.style.display = '';
        else if (filter === 'boviet' && acct === 'boviet') item.style.display = '';
        else if (filter === 'tolead' && acct === 'tolead') item.style.display = '';
        else item.style.display = 'none';
      });
    });
  });
});

// --- Stat Card Clicks ---
document.querySelectorAll('.stat-card').forEach(function(card) {
  card.addEventListener('click', function() {
    var label = this.querySelector('.stat-label').textContent.toLowerCase();
    if (label.includes('active')) {
      var el = document.querySelector('.account-list');
      if (el) { el.scrollIntoView({behavior:'smooth'}); el.classList.add('highlight-flash'); setTimeout(function(){el.classList.remove('highlight-flash');},1500); }
    } else if (label.includes('at risk')) {
      // Filter alerts to show urgent/warning only
      document.querySelectorAll('.alert-item').forEach(function(item) {
        var icon = item.querySelector('.alert-icon');
        item.style.display = (icon && (icon.classList.contains('urgent') || icon.classList.contains('warning'))) ? '' : 'none';
      });
      var al = document.querySelector('.alert-list');
      if (al) { al.scrollIntoView({behavior:'smooth'}); al.classList.add('highlight-flash'); setTimeout(function(){al.classList.remove('highlight-flash');},1500); }
    } else if (label.includes('eta')) {
      // Show info alerts (bot updates)
      document.querySelectorAll('.alert-item').forEach(function(item) {
        var icon = item.querySelector('.alert-icon');
        item.style.display = (icon && icon.classList.contains('info')) ? '' : 'none';
      });
      var al = document.querySelector('.alert-list');
      if (al) al.scrollIntoView({behavior:'smooth'});
    } else if (label.includes('on schedule') || label.includes('completed')) {
      var el = document.querySelector('.three-col');
      if (el) el.scrollIntoView({behavior:'smooth'});
    }
  });
});

// --- Search ---
var searchInput = document.querySelector('.search-box input');
if (searchInput) {
  searchInput.addEventListener('input', function() {
    var q = this.value.toLowerCase().trim();
    // Filter alert items
    document.querySelectorAll('.alert-item').forEach(function(item) {
      if (!q) { item.style.display = ''; return; }
      item.style.display = item.textContent.toLowerCase().includes(q) ? '' : 'none';
    });
    // Filter account rows
    document.querySelectorAll('.account-row:not(.header)').forEach(function(row) {
      if (!q) { row.style.display = ''; return; }
      row.style.display = row.textContent.toLowerCase().includes(q) ? '' : 'none';
    });
  });
}

// Auto-refresh (pause when panel is open)
setTimeout(function(){ if (!document.getElementById('detail-panel').classList.contains('open')) location.reload(); }, 60000);
"""
    return HTMLResponse(_page("CSL AI Dispatch", body, interactive_js))


@app.get("/shipments", response_class=HTMLResponse)
def shipments(account: str = Query(default=None)):
    """Document Tracker - shows all shipments with drag-drop POD upload."""
    sheet_cache.refresh_if_needed()
    account_filter = account if account else None

    # Get doc status from DB
    all_loads_db = db.get_dashboard_loads(None)
    doc_map = {}
    for ld in all_loads_db:
        doc_map[ld["load_number"]] = {
            "bol": ld.get("bol_received", False),
            "pod": ld.get("pod_received", False),
            "status": ld.get("status", "active"),
        }

    # Build list from sheet cache
    all_shipments = sheet_cache.shipments
    if account_filter:
        all_shipments = [s for s in all_shipments if s.get("account") == account_filter]

    # Filter out delivered/completed
    active = []
    ready_to_invoice = []
    for s in all_shipments:
        sl = s["status"].lower() if s.get("status") else ""
        if any(w in sl for w in ("delivered", "completed", "empty return")):
            continue
        efj = s["efj"]
        docs = doc_map.get(efj, {})
        s["_bol"] = docs.get("bol", False)
        s["_pod"] = docs.get("pod", False)
        s["_db_status"] = docs.get("status", "active")
        if s["_pod"] and s["_db_status"] == "ready_to_invoice":
            ready_to_invoice.append(s)
        else:
            active.append(s)

    # Stats
    total = len(active) + len(ready_to_invoice)
    missing_bol = sum(1 for s in active if not s["_bol"])
    missing_pod = sum(1 for s in active if not s["_pod"])

    # Build rows
    def build_rows(shipment_list, show_dismiss=False):
        rows = ""
        for s in shipment_list:
            bol = '<span class="doc-yes">&#10004;</span>' if s["_bol"] else '<span class="doc-no">&#10008;</span>'
            pod = '<span class="doc-yes">&#10004;</span>' if s["_pod"] else '<span class="doc-no drop-target">&#10008; Drop POD</span>'

            ctr = s["container"] or "-"
            if s.get("container_url"):
                ctr = f'<a href="{s["container_url"]}" target="_blank" style="color:var(--accent-cyan);">{ctr}</a>'

            status_sl = (s["status"] or "").lower()
            sc = "var(--text-secondary)"
            if "at port" in status_sl or "available" in status_sl:
                sc = "var(--accent-amber)"
            elif "in transit" in status_sl or "on vessel" in status_sl:
                sc = "var(--accent-blue)"
            elif "out for" in status_sl or "picked up" in status_sl:
                sc = "var(--accent-cyan)"

            dismiss_btn = ""
            if show_dismiss:
                dismiss_btn = f'<td><button class="btn-dismiss" onclick="dismissLoad(\'{s["efj"]}\')">&#10004; Done</button></td>'

            rows += f"""<tr class="drop-row" data-efj="{s['efj']}" ondragover="handleDragOver(event)" ondragleave="handleDragLeave(event)" ondrop="handleDrop(event, '{s['efj']}')">
  <td><strong><a href="javascript:void(0)" onclick="openPanel('{s['efj']}')" style="color:var(--accent-blue);cursor:pointer;">{s['efj']}</a></strong></td>
  <td style="font-family:JetBrains Mono,monospace;font-size:12px;">{ctr}</td>
  <td>{s['account'] or '-'}</td>
  <td>{bol}</td>
  <td class="pod-cell">{pod}</td>
  <td><span style="color:{sc}">{s['status'] or '-'}</span></td>
  {dismiss_btn}
</tr>"""
        return rows

    active_rows = build_rows(active)
    rti_rows = build_rows(ready_to_invoice, show_dismiss=True)

    rti_section = ""
    if ready_to_invoice:
        rti_section = f"""<div class="panel" style="margin-bottom:16px;border:1px solid var(--accent-amber);">
  <div class="panel-header" style="background:rgba(245,158,11,0.1);">
    <div class="panel-title" style="color:var(--accent-amber);">Ready to Invoice ({len(ready_to_invoice)})</div>
  </div>
  <div class="table-wrap">
    <table>
      <thead><tr><th>EFJ #</th><th>Container/Load ID</th><th>Account</th><th>BOL</th><th>POD</th><th>Status</th><th></th></tr></thead>
      <tbody>{rti_rows}</tbody>
    </table>
  </div>
</div>"""

    accts = [("All", None), ("EFJ-Operations", "EFJ-Operations"), ("Boviet", "Boviet"), ("Tolead", "Tolead")]
    filt = ""
    for label, val in accts:
        is_active = val == account_filter or (val is None and account_filter is None)
        href = "/shipments" if val is None else f"/shipments?account={val}"
        filt += f'<a href="{href}" style="padding:6px 14px;border-radius:8px;border:1px solid var(--border);font-size:12px;color:{"var(--accent-blue)" if is_active else "var(--text-secondary)"};background:{"var(--glow-blue)" if is_active else "var(--bg-card)"};margin-right:6px;">{label}</a>'

    body = f"""{_sidebar("shipments")}
<div class="main">
  {_topbar("Document", "Tracker", search=False)}
  <div class="content">
    <div style="margin-bottom:16px;display:flex;gap:16px;align-items:center;flex-wrap:wrap;">
      <h2 style="font-size:18px;">Shipments ({total})</h2>
      <span style="color:var(--accent-amber);font-size:12px;">Missing BOL: {missing_bol}</span>
      <span style="color:var(--accent-red);font-size:12px;">Missing POD: {missing_pod}</span>
    </div>
    <div style="margin-bottom:16px;">{filt}</div>
    {rti_section}
    <div class="panel">
      <div class="table-wrap">
        <table>
          <thead><tr><th>EFJ #</th><th>Container/Load ID</th><th>Account</th><th>BOL</th><th>POD</th><th>Status</th></tr></thead>
          <tbody>{active_rows}</tbody>
        </table>
      </div>
    </div>
    <div style="margin-top:12px;font-size:11px;color:var(--text-dim);text-align:center;">Drag &amp; drop POD files onto any row to upload</div>
  </div>
</div>
<div class="detail-overlay" id="detail-overlay" onclick="closePanel()"></div>
<div class="detail-panel" id="detail-panel">
  <button class="panel-close" onclick="closePanel()">&times;</button>
  <div id="panel-content"><div class="panel-loading">Select a shipment</div></div>
</div>"""

    dd_js = r"""
function handleDragOver(e) {
  e.preventDefault();
  e.stopPropagation();
  e.currentTarget.style.outline = '2px dashed var(--accent-blue)';
  e.currentTarget.style.background = 'rgba(59,130,246,0.05)';
}
function handleDragLeave(e) {
  e.currentTarget.style.outline = '';
  e.currentTarget.style.background = '';
}
async function handleDrop(e, efj) {
  e.preventDefault();
  e.stopPropagation();
  var row = e.currentTarget;
  row.style.outline = '';
  row.style.background = '';

  var files = e.dataTransfer.files;
  if (!files.length) return;
  var file = files[0];
  var allowed = ['.pdf','.png','.jpg','.jpeg'];
  var ext = file.name.substring(file.name.lastIndexOf('.')).toLowerCase();
  if (allowed.indexOf(ext) === -1) { alert('Only PDF/image files accepted for POD.'); return; }
  if (file.size > 25*1024*1024) { alert('File too large. Max 25 MB.'); return; }

  var podCell = row.querySelector('.pod-cell');
  podCell.innerHTML = '<span style="color:var(--accent-amber);">Uploading...</span>';
  row.style.background = 'rgba(245,158,11,0.05)';

  var fd = new FormData();
  fd.append('file', file);
  fd.append('doc_type', 'POD');
  try {
    var res = await fetch('/api/load/' + encodeURIComponent(efj) + '/upload', {method:'POST', body:fd});
    if (res.ok) {
      podCell.innerHTML = '<span class="doc-yes">&#10004;</span>';
      row.style.background = 'rgba(34,197,94,0.06)';
      row.style.transition = 'background 0.5s';
      // Mark as ready to invoice
      await fetch('/api/load/' + encodeURIComponent(efj) + '/ready-to-invoice', {method:'POST'});
      setTimeout(function() { window.location.reload(); }, 1500);
    } else {
      podCell.innerHTML = '<span class="doc-no">&#10008; Failed</span>';
      row.style.background = '';
    }
  } catch(err) {
    podCell.innerHTML = '<span class="doc-no">&#10008; Error</span>';
    row.style.background = '';
  }
}

async function dismissLoad(efj) {
  if (!confirm('Mark ' + efj + ' as invoiced/complete?')) return;
  try {
    var res = await fetch('/api/load/' + encodeURIComponent(efj) + '/dismiss', {method:'POST'});
    if (res.ok) { window.location.reload(); }
    else { alert('Failed to dismiss load.'); }
  } catch(e) { alert('Error: ' + e.message); }
}

function openPanel(efj) {
  if (!efj) return;
  document.getElementById('detail-panel').classList.add('open');
  document.getElementById('detail-overlay').classList.add('open');
  loadPanel(efj);
}
function closePanel() {
  document.getElementById('detail-panel').classList.remove('open');
  document.getElementById('detail-overlay').classList.remove('open');
}
document.addEventListener('keydown', function(e) { if (e.key === 'Escape') closePanel(); });

async function loadPanel(efj) {
  var pc = document.getElementById('panel-content');
  pc.innerHTML = '<div class="panel-loading">Loading ' + efj + '...</div>';
  try {
    var res = await fetch('/api/load/' + encodeURIComponent(efj));
    if (!res.ok) { pc.innerHTML = '<div class="panel-loading">Load not found</div>'; return; }
    var d = await res.json();
    var h = '<div class="panel-head"><div class="panel-head-title">' + d.efj + '</div>';
    h += '<div class="panel-head-sub">' + (d.account||'') + ' \u00B7 ' + (d.move_type||'') + '</div></div>';
    h += '<div class="panel-section"><div class="panel-section-title">Details</div>';
    var fields = [['Container/Load', d.container_url ? '<a href="' + d.container_url + '" target="_blank" style="color:#06b6d4">' + d.container + ' &#x2197;</a>' : d.container], ['BOL/Booking', d.bol], ['Status', d.status || 'Unknown'], ['Rep', d.rep || 'Unassigned']];
    if (d.container_url) { fields.push(['Macropoint', '<a href="' + d.container_url + '" target="_blank" style="color:#06b6d4">Track Shipment &#x2197;</a>']); }
    for (var i=0;i<fields.length;i++) { h += '<div class="panel-field"><span class="panel-field-label">' + fields[i][0] + '</span><span class="panel-field-value">' + (fields[i][1]||'-') + '</span></div>'; }
    h += '</div>';
    pc.innerHTML = h;
  } catch(e) { pc.innerHTML = '<div class="panel-loading">Error loading details</div>'; }
}
"""

    return HTMLResponse(_page("CSL Shipments", body, script=dd_js))



@app.get("/unmatched", response_class=HTMLResponse)
def unmatched_page():
    emails = db.get_unmatched_emails()
    if not emails:
        table = '<div style="padding:48px;text-align:center;color:var(--text-dim);">No unmatched emails</div>'
    else:
        rows = ""
        for em in emails[:50]:
            date_str = em["received_date"].strftime("%Y-%m-%d %H:%M") if em.get("received_date") else "-"
            rows += f"""<tr>
<td>{em.get('subject') or '-'}</td><td>{em.get('sender') or '-'}</td><td>{date_str}</td><td>{em.get('attachment_names') or '-'}</td>
<td><form class="match-form" method="POST" action="/unmatched/{em['id']}/match"><input type="text" name="load_number" placeholder="EFJ#" required><button type="submit">Match</button></form>
<form class="match-form" method="POST" action="/unmatched/{em['id']}/ignore" style="margin-top:4px;"><button type="submit" class="btn-ignore">Ignore</button></form></td>
</tr>"""
        table = f'<table><thead><tr><th>Subject</th><th>From</th><th>Date</th><th>Attachments</th><th>Action</th></tr></thead><tbody>{rows}</tbody></table>'

    body = f"""{_sidebar("unmatched")}
<div class="main">
  {_topbar("Unmatched", "Emails", search=False)}
  <div class="content">
    <div style="margin-bottom:16px;"><h2 style="font-size:18px;">Unmatched Emails ({len(emails)})</h2></div>
    <div class="panel"><div class="table-wrap">{table}</div></div>
  </div>
</div>"""
    return HTMLResponse(_page("CSL Unmatched Emails", body))


@app.get("/docs/{file_path:path}")
def serve_document(file_path: str, request: Request):
    """Serve a document, decrypting it on the fly."""
    # Defense-in-depth auth check
    token = request.cookies.get("csl_session")
    if not auth.verify_session_token(token):
        raise HTTPException(status_code=401, detail="Not authenticated")

    # Path traversal protection
    full_path = config.DOCUMENT_STORAGE_PATH / file_path
    try:
        full_path.resolve().relative_to(config.DOCUMENT_STORAGE_PATH.resolve())
    except ValueError:
        raise HTTPException(status_code=403, detail="Access denied")

    if not full_path.exists():
        raise HTTPException(status_code=404, detail="Document not found")

    # Decrypt file contents
    encrypted_data = full_path.read_bytes()
    try:
        decrypted_data = decrypt_data(encrypted_data)
    except Exception:
        # Fallback: file might be unencrypted (pre-migration)
        log.warning("Failed to decrypt %s, serving as-is", file_path)
        decrypted_data = encrypted_data

    # Determine content type from extension
    content_type, _ = mimetypes.guess_type(file_path)
    if not content_type:
        content_type = "application/octet-stream"

    return Response(
        content=decrypted_data,
        media_type=content_type,
        headers={
            "Content-Disposition": f'inline; filename="{Path(file_path).name}"',
            "Cache-Control": "no-store",
        }
    )


@app.post("/unmatched/{unmatched_id}/match")
def match_unmatched(unmatched_id: int, load_number: str = Form(...)):
    load = db.get_load_by_number(load_number.strip())
    if not load:
        raise HTTPException(status_code=404, detail=f"Load '{load_number}' not found")
    db.resolve_unmatched_email(unmatched_id, load["id"])
    return RedirectResponse(url="/unmatched", status_code=303)


@app.post("/unmatched/{unmatched_id}/ignore")
def ignore_unmatched(unmatched_id: int):
    db.ignore_unmatched_email(unmatched_id)
    return RedirectResponse(url="/unmatched", status_code=303)


@app.get("/api/stats")
def api_stats():
    sheet_cache.refresh_if_needed()
    return sheet_cache.stats


@app.get("/api/bot-status")
def api_bot_status():
    return _get_bot_status_detailed()


@app.get("/logo.svg")
def serve_logo():
    logo_path = Path(__file__).parent / "logo.svg"
    if logo_path.exists():
        return FileResponse(str(logo_path), media_type="image/svg+xml")
    # Fallback: simple colored square
    return HTMLResponse(
        '<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 42 42"><rect width="42" height="42" rx="12" fill="#00deb4"/><text x="21" y="26" text-anchor="middle" fill="#fff" font-size="14" font-weight="700">CSL</text></svg>',
        media_type="image/svg+xml",
    )


@app.get("/api/load/{efj}")
def api_load_detail(efj: str):
    """Get full load details from sheet cache + document status from DB."""
    sheet_cache.refresh_if_needed()
    shipment = None
    for s in sheet_cache.shipments:
        if s["efj"] == efj:
            shipment = s
            break
    if not shipment:
        raise HTTPException(status_code=404, detail=f"Load {efj} not found")

    # Get documents from database
    documents = []
    try:
        load_row = db.find_load_id_by_reference(efj)
        if load_row:
            docs = db.get_documents_for_load(load_row)
            for d in docs:
                documents.append({
                    "doc_type": d.get("doc_type", d.get("document_type", "")),
                    "filename": d.get("filename", d.get("original_filename", "")),
                    "file_path": d.get("file_path", d.get("stored_path", "")),
                    "received_at": str(d.get("created_at", d.get("received_at", ""))) if d.get("created_at") or d.get("received_at") else "",
                })
    except Exception as e:
        log.warning("Error fetching documents for %s: %s", efj, e)

    return {
        "efj": shipment["efj"],
        "account": shipment["account"],
        "move_type": shipment["move_type"],
        "container": shipment["container"],
        "bol": shipment["bol"],
        "ssl": shipment["ssl"],
        "carrier": shipment["carrier"],
        "origin": shipment["origin"],
        "destination": shipment["destination"],
        "eta": shipment["eta"],
        "lfd": shipment["lfd"],
        "pickup": shipment["pickup"],
        "delivery": shipment["delivery"],
        "status": shipment["status"],
        "notes": shipment["notes"],
        "bot_alert": shipment["bot_alert"],
        "return_port": shipment["return_port"],
        "rep": shipment["rep"],
        "container_url": shipment.get("container_url", ""),
        "documents": documents,
    }


@app.post("/api/load/{efj}/upload")
async def api_load_upload(efj: str, file: UploadFile = File(...), doc_type: str = Form(...)):
    """Upload a document for a load — validates, encrypts, and stores securely."""
    if doc_type not in ("BOL", "POD", "Invoice", "Other"):
        raise HTTPException(status_code=400, detail="Invalid document type")

    # Validate file extension
    original_name = file.filename or "unknown"
    ext = Path(original_name).suffix.lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail=f"File type '{ext}' not allowed. Allowed: {', '.join(sorted(ALLOWED_EXTENSIONS))}"
        )

    # Read file with size check
    contents = await file.read()
    if len(contents) > MAX_UPLOAD_SIZE:
        raise HTTPException(
            status_code=413,
            detail=f"File too large. Maximum size is {MAX_UPLOAD_SIZE // (1024*1024)} MB"
        )

    # Sanitize filename
    safe_name = _sanitize_filename(original_name)

    # Create storage directory with restricted permissions
    store_dir = config.DOCUMENT_STORAGE_PATH / efj / doc_type
    os.makedirs(store_dir, mode=0o700, exist_ok=True)

    # Handle duplicates
    stem = Path(safe_name).stem
    suffix = Path(safe_name).suffix
    dest = store_dir / safe_name
    counter = 1
    while dest.exists():
        dest = store_dir / f"{stem}_{counter}{suffix}"
        counter += 1
    final_name = dest.name

    # Encrypt and save
    encrypted_data = encrypt_data(contents)
    dest.write_bytes(encrypted_data)
    os.chmod(str(dest), 0o600)

    # Insert into database (with correct parameter order)
    rel_path = f"{efj}/{doc_type}/{final_name}"
    try:
        load_id = db.find_load_id_by_reference(efj)
        if load_id is None:
            log.warning("No load found for reference %s, skipping DB insert", efj)
        else:
            db.insert_document(load_id, doc_type, rel_path, original_name)
    except Exception as e:
        log.warning("DB insert for doc %s/%s failed: %s", efj, doc_type, e)

    return {"status": "ok", "filename": original_name, "path": rel_path}




@app.post("/api/load/{efj}/invoiced")
async def api_set_invoiced(efj: str, request: Request):
    """Toggle invoiced status for a load."""
    body = await request.json()
    invoiced = bool(body.get("invoiced", False))
    db.set_load_invoiced(efj, invoiced)
    return {"status": "ok", "invoiced": invoiced}


@app.post("/api/load/{efj}/ready-to-invoice")
async def api_ready_to_invoice(efj: str):
    """Mark a load as ready to invoice after POD is uploaded."""
    with db.get_conn() as conn:
        with db.get_cursor(conn) as cur:
            cur.execute(
                "UPDATE loads SET status = 'ready_to_invoice', updated_at = NOW() WHERE load_number = %s",
                (efj,)
            )
    return {"status": "ok"}


@app.post("/api/load/{efj}/dismiss")
async def api_dismiss_load(efj: str):
    """Dismiss a load from the ready-to-invoice list (mark as invoiced/complete)."""
    with db.get_conn() as conn:
        with db.get_cursor(conn) as cur:
            cur.execute(
                "UPDATE loads SET status = 'invoiced', invoiced = TRUE, updated_at = NOW() WHERE load_number = %s",
                (efj,)
            )
    return {"status": "ok"}


# ── React Dashboard: Shipments API ──
@app.get("/api/shipments")
async def api_shipments(request: Request, account: str = None, status: str = None):
    """Return all shipments as JSON, optionally filtered by account and/or status."""
    sheet_cache.refresh_if_needed()
    data = sheet_cache.shipments
    if account:
        data = [s for s in data if s.get("account", "").lower() == account.lower()]
    if status:
        data = [s for s in data if s.get("status", "").lower() == status.lower()]
    return {"shipments": data, "total": len(data)}


# ── React Dashboard: Alerts API ──
@app.get("/api/alerts")
async def api_alerts(request: Request):
    """Return current alerts for the dashboard."""
    sheet_cache.refresh_if_needed()
    alerts = _generate_alerts(sheet_cache.shipments)
    return {"alerts": alerts, "total": len(alerts)}


# ── React Dashboard: Accounts API ──
@app.get("/api/accounts")
async def api_accounts(request: Request):
    """Return account summaries for the dashboard."""
    sheet_cache.refresh_if_needed()
    return {"accounts": sheet_cache.accounts}


# ── React Dashboard: Team API ──
@app.get("/api/team")
async def api_team(request: Request):
    """Return team member summaries for the dashboard."""
    sheet_cache.refresh_if_needed()
    return {"team": sheet_cache.team}

# ── React SPA ──


# ═══════════════════════════════════════════════════════════════
# BATCH TABLE DATA ENDPOINTS (tracking + documents)
# ═══════════════════════════════════════════════════════════════

@app.get("/api/shipments/tracking-summary")
async def api_tracking_summary():
    """Return tracking status summary with stop timestamps for all FTL loads."""
    cache = _read_tracking_cache()
    result = {}
    for efj, entry in cache.items():
        stop_times = entry.get("stop_times") or {}
        behind = False
        for k in ("stop1_eta", "stop2_eta"):
            val = stop_times.get(k) or ""
            if "BEHIND" in val.upper():
                behind = True
        _ts_disp, _ts_detail = _classify_mp_display_status(entry)
        result[efj] = {
            "behindSchedule": behind,
            "cantMakeIt": bool(entry.get("cant_make_it")),
            "status": entry.get("status", ""),
            "lastScraped": entry.get("last_scraped", ""),
            # Stop timestamps for slide view display
            "stop1Arrived": stop_times.get("stop1_arrived"),
            "stop1Departed": stop_times.get("stop1_departed"),
            "stop2Arrived": stop_times.get("stop2_arrived"),
            "stop2Departed": stop_times.get("stop2_departed"),
            "stop1Eta": stop_times.get("stop1_eta"),
            "stop2Eta": stop_times.get("stop2_eta"),
            "mpDisplayStatus": _ts_disp,
            "mpDisplayDetail": _ts_detail,
        }
    return {"tracking": result}


@app.get("/api/shipments/document-summary")
async def api_document_summary():
    """Return document type counts + latest doc id per load for table icons and alerts."""
    with db.get_cursor() as cur:
        cur.execute("""
            SELECT efj, doc_type, COUNT(*) as cnt, MAX(id) as latest_id
            FROM load_documents
            GROUP BY efj, doc_type
            ORDER BY efj
        """)
        rows = cur.fetchall()
    result = {}
    doc_ids = {}
    for r in rows:
        efj_val = r["efj"] if isinstance(r, dict) else r[0]
        doc_type = r["doc_type"] if isinstance(r, dict) else r[1]
        cnt = r["cnt"] if isinstance(r, dict) else r[2]
        latest_id = r["latest_id"] if isinstance(r, dict) else r[3]
        if efj_val not in result:
            result[efj_val] = {}
            doc_ids[efj_val] = {}
        result[efj_val][doc_type] = cnt
        doc_ids[efj_val][doc_type] = latest_id
    return {"documents": result, "doc_ids": doc_ids}


# ═══════════════════════════════════════════════════════════════
# UNCLASSIFIED DOCUMENTS + RATE IQ
# ═══════════════════════════════════════════════════════════════

@app.get("/api/unclassified-documents")
async def api_unclassified_documents():
    """Return documents with unclassified doc_type for manual review."""
    with db.get_cursor() as cur:
        cur.execute("""
        SELECT ld.id, ld.efj, ld.doc_type, ld.original_name,
               ld.uploaded_at, ld.uploaded_by
        FROM load_documents ld
        WHERE ld.doc_type = 'unclassified'
        ORDER BY ld.uploaded_at DESC
        LIMIT 50
        """)
        rows = cur.fetchall()
    docs = []
    for r in rows:
        docs.append({
            "id": r["id"],
            "efj": r["efj"],
            "doc_type": r["doc_type"],
            "original_name": r["original_name"],
            "uploaded_at": r["uploaded_at"].isoformat() if r["uploaded_at"] else None,
            "uploaded_by": r["uploaded_by"],
        })
    return {"documents": docs, "count": len(docs)}


@app.get("/api/rate-iq")
async def api_rate_iq():
    """
    Rate IQ — return carrier rate history grouped by lane for scorecard comparison.
    Each lane shows all carrier quotes received, sorted by rate.
    """
    with db.get_cursor() as cur:
        # Get parsed rate quotes from rate_quotes table
        cur.execute("""
            SELECT rq.id, rq.email_thread_id, rq.efj, rq.lane, rq.origin,
                   rq.destination, rq.miles, rq.move_type, rq.carrier_name,
                   rq.carrier_email, rq.rate_amount, rq.rate_unit,
                   rq.quote_date, rq.indexed_at, rq.status
            FROM rate_quotes rq
            ORDER BY rq.quote_date DESC NULLS LAST
            LIMIT 500
        """)
        rate_quotes = cur.fetchall()
        # Get carrier rate emails with lane info
        cur.execute("""
            SELECT et.id, et.efj, et.sender, et.subject, et.body_preview,
                   et.lane, et.email_type, et.sent_at, et.indexed_at
            FROM email_threads et
            WHERE et.email_type = 'carrier_rate'
            ORDER BY et.sent_at DESC
            LIMIT 200
        """)
        carrier_emails = cur.fetchall()
        # Get carrier rate documents
        cur.execute("""
            SELECT ld.id, ld.efj, ld.doc_type, ld.original_name,
                   ld.uploaded_at, ld.uploaded_by
            FROM load_documents ld
            WHERE ld.doc_type = 'carrier_rate'
            ORDER BY ld.uploaded_at DESC
            LIMIT 200
        """)
        carrier_docs = cur.fetchall()
        # Get customer rate requests
        cur.execute("""
            SELECT et.id, et.efj, et.sender, et.subject, et.body_preview,
                   et.lane, et.email_type, et.sent_at
            FROM email_threads et
            WHERE et.email_type = 'customer_rate'
            ORDER BY et.sent_at DESC
            LIMIT 100
        """)
        customer_emails = cur.fetchall()

    # Group rate quotes by lane (parsed data with actual $$ amounts)
    lanes = {}
    for rq in rate_quotes:
        lane_key = rq["lane"] or "Unknown Lane"
        if lane_key not in lanes:
            lanes[lane_key] = {
                "lane": lane_key, "miles": None, "move_type": None,
                "carrier_quotes": [], "customer_requests": [],
                "cheapest": None, "avg_rate": None,
            }
        entry = lanes[lane_key]
        if rq["miles"] and not entry["miles"]:
            entry["miles"] = rq["miles"]
        if rq["move_type"] and not entry["move_type"]:
            entry["move_type"] = rq["move_type"]
        entry["carrier_quotes"].append({
            "id": rq["id"],
            "efj": rq["efj"],
            "carrier": rq["carrier_name"] or rq["carrier_email"] or "Unknown",
            "carrier_email": rq["carrier_email"],
            "rate": float(rq["rate_amount"]) if rq["rate_amount"] else None,
            "rate_unit": rq["rate_unit"],
            "date": rq["quote_date"].isoformat() if rq["quote_date"] else None,
            "status": rq["status"],
            "move_type": rq["move_type"],
        })

    # Also add carrier emails that might not have parsed rates
    for e in carrier_emails:
        lane_key = e["lane"] or "Unknown Lane"
        if lane_key not in lanes:
            lanes[lane_key] = {
                "lane": lane_key, "miles": None, "move_type": None,
                "carrier_quotes": [], "customer_requests": [],
                "cheapest": None, "avg_rate": None,
            }
        # Only add if not already represented by a rate_quote
        existing_ids = {q.get("efj") for q in lanes[lane_key]["carrier_quotes"]}
        if e["efj"] not in existing_ids:
            lanes[lane_key]["carrier_quotes"].append({
                "id": e["id"],
                "efj": e["efj"],
                "carrier": e["sender"],
                "carrier_email": e["sender"],
                "rate": None,
                "rate_unit": None,
                "date": e["sent_at"].isoformat() if e["sent_at"] else None,
                "status": "pending",
                "move_type": None,
                "source": "email",
            })

    # Add customer requests to their lanes
    for e in customer_emails:
        lane_key = e["lane"] or "Unknown Lane"
        if lane_key not in lanes:
            lanes[lane_key] = {
                "lane": lane_key, "miles": None, "move_type": None,
                "carrier_quotes": [], "customer_requests": [],
                "cheapest": None, "avg_rate": None,
            }
        lanes[lane_key]["customer_requests"].append({
            "id": e["id"],
            "efj": e["efj"],
            "sender": e["sender"],
            "subject": e["subject"],
            "sent_at": e["sent_at"].isoformat() if e["sent_at"] else None,
        })

    # Compute cheapest + avg per lane
    for lane_data in lanes.values():
        rates = [q["rate"] for q in lane_data["carrier_quotes"] if q.get("rate")]
        if rates:
            min_rate = min(rates)
            cheapest_q = next(q for q in lane_data["carrier_quotes"] if q.get("rate") == min_rate)
            lane_data["cheapest"] = {"carrier": cheapest_q["carrier"], "rate": min_rate}
            lane_data["avg_rate"] = round(sum(rates) / len(rates), 2)

    # Carrier scorecard — frequency, win rate, avg rate
    carrier_scores = {}
    for rq in rate_quotes:
        carrier_key = rq["carrier_email"] or rq["carrier_name"] or "Unknown"
        if carrier_key not in carrier_scores:
            carrier_scores[carrier_key] = {
                "carrier": rq["carrier_name"] or carrier_key,
                "quote_count": 0, "win_count": 0,
                "total_rate": 0, "rated_count": 0,
                "lanes_covered": set(),
            }
        cs = carrier_scores[carrier_key]
        cs["quote_count"] += 1
        if rq["status"] == "accepted":
            cs["win_count"] += 1
        if rq["rate_amount"]:
            cs["total_rate"] += float(rq["rate_amount"])
            cs["rated_count"] += 1
        if rq["lane"]:
            cs["lanes_covered"].add(rq["lane"])

    # Fallback: also count carrier emails not in rate_quotes
    for e in carrier_emails:
        sender = e["sender"] or "Unknown"
        sender_key = sender.split("<")[-1].replace(">", "").strip() if "<" in sender else sender
        if sender_key not in carrier_scores:
            carrier_scores[sender_key] = {
                "carrier": sender,
                "quote_count": 0, "win_count": 0,
                "total_rate": 0, "rated_count": 0,
                "lanes_covered": set(),
            }
            carrier_scores[sender_key]["quote_count"] += 1
            if e["lane"]:
                carrier_scores[sender_key]["lanes_covered"].add(e["lane"])

    scorecard = []
    for key, data in carrier_scores.items():
        scorecard.append({
            "carrier": data["carrier"],
            "quote_count": data["quote_count"],
            "win_count": data["win_count"],
            "avg_rate": round(data["total_rate"] / data["rated_count"], 2) if data["rated_count"] else None,
            "lanes_covered": len(data["lanes_covered"]),
            "lane_list": list(data["lanes_covered"]),
        })
    scorecard.sort(key=lambda x: x["quote_count"], reverse=True)

    return {
        "lanes": list(lanes.values()),
        "scorecard": scorecard,
        "carrier_docs": [
            {
                "id": d["id"], "efj": d["efj"], "doc_type": d["doc_type"],
                "original_name": d["original_name"],
                "uploaded_at": d["uploaded_at"].isoformat() if d["uploaded_at"] else None,
            }
            for d in carrier_docs
        ],
        "total_carrier_quotes": len(carrier_emails),
        "total_customer_requests": len(customer_emails),
        "total_rate_quotes": len(rate_quotes),
    }


@app.get("/api/rate-iq/lane/{lane}")
async def api_rate_iq_lane(lane: str):
    """Get all quotes for a specific lane."""
    from urllib.parse import unquote
    lane = unquote(lane)
    with db.get_cursor() as cur:
        cur.execute("""
        SELECT et.id, et.efj, et.sender, et.subject, et.body_preview,
               et.lane, et.email_type, et.sent_at
        FROM email_threads et
        WHERE et.lane = %s
        ORDER BY et.sent_at DESC
        """, (lane,))
        emails = cur.fetchall()
    results = []
    for e in emails:
        results.append({
            "id": e["id"],
            "efj": e["efj"],
            "sender": e["sender"],
            "subject": e["subject"],
            "body_preview": e["body_preview"],
            "lane": e["lane"],
            "email_type": e["email_type"],
            "sent_at": e["sent_at"].isoformat() if e["sent_at"] else None,
        })
    return {"lane": lane, "emails": results}


@app.patch("/api/rate-iq/{quote_id}")
async def update_rate_quote(quote_id: int, request: Request):
    """Accept or reject a rate quote."""
    body = await request.json()
    new_status = body.get("status", "").strip()
    if new_status not in ("accepted", "rejected", "pending"):
        return JSONResponse(status_code=400, content={"error": "status must be accepted, rejected, or pending"})
    with db.get_conn() as conn:
        with db.get_cursor(conn) as cur:
            cur.execute(
                "UPDATE rate_quotes SET status = %s WHERE id = %s RETURNING id, lane, carrier_name, rate_amount",
                (new_status, quote_id),
            )
            row = cur.fetchone()
    if not row:
        return JSONResponse(status_code=404, content={"error": "quote not found"})
    return {"ok": True, "id": row["id"], "status": new_status,
            "carrier": row["carrier_name"], "rate": float(row["rate_amount"]) if row["rate_amount"] else None}


@app.get("/api/rate-iq/search-lane")
async def api_search_lane(origin: str = Query(""), destination: str = Query("")):
    """
    Unified lane rate search: rate_quotes + lane_rates + won quotes.
    Groups results by normalized lane. Returns lane_groups[] with
    floor/avg/ceiling per lane, plus flat matches[] for backwards compat.
    """
    if not origin and not destination:
        return {"matches": [], "lane_groups": [], "carriers": [], "stats": {}}

    origin_q = origin.strip().lower()
    dest_q = destination.strip().lower()

    # Build WHERE conditions for each source
    def _build_conditions(o_col, d_col, lane_col=None):
        conds, params = [], []
        if origin_q:
            if lane_col:
                conds.append(f"(LOWER({o_col}) LIKE %s OR LOWER({lane_col}) LIKE %s)")
                params.extend([f"%{origin_q}%", f"%{origin_q}%"])
            else:
                conds.append(f"LOWER({o_col}) LIKE %s")
                params.append(f"%{origin_q}%")
        if dest_q:
            if lane_col:
                conds.append(f"(LOWER({d_col}) LIKE %s OR LOWER({lane_col}) LIKE %s)")
                params.extend([f"%{dest_q}%", f"%{dest_q}%"])
            else:
                conds.append(f"LOWER({d_col}) LIKE %s")
                params.append(f"%{dest_q}%")
        return (" AND ".join(conds) if conds else "TRUE"), params

    rq_where, rq_params = _build_conditions("rq.origin", "rq.destination", "rq.lane")
    lr_where, lr_params = _build_conditions("lr.port", "lr.destination")
    # won quotes use pod/final_delivery
    wq_where, wq_params = _build_conditions("q.pod", "q.final_delivery")

    with db.get_cursor() as cur:
        # ── Unified UNION: rate_quotes + lane_rates + won quotes ──
        cur.execute(f"""
            SELECT origin, destination,
                   COALESCE(NULLIF(TRIM(origin),'') || ' \u2192 ' || NULLIF(TRIM(destination),''), lane, '') AS norm_lane,
                   carrier_name, carrier_email, rate_amount, rate_unit,
                   quote_date, source, move_type, id, miles, status
            FROM (
                -- Email-extracted carrier rates
                SELECT rq.origin, rq.destination, rq.lane,
                       rq.carrier_name, rq.carrier_email, rq.rate_amount, rq.rate_unit,
                       rq.quote_date, 'email' AS source, rq.move_type, rq.id, rq.miles, rq.status
                FROM rate_quotes rq
                WHERE {rq_where} AND rq.rate_amount IS NOT NULL

                UNION ALL

                -- Excel-imported lane rates
                SELECT lr.port AS origin, lr.destination,
                       lr.port || ' \u2192 ' || lr.destination AS lane,
                       lr.carrier_name, NULL AS carrier_email,
                       lr.dray_rate AS rate_amount, 'flat' AS rate_unit,
                       lr.created_at AS quote_date, 'import' AS source,
                       lr.move_type, lr.id, NULL AS miles, 'import' AS status
                FROM lane_rates lr
                WHERE {lr_where} AND lr.dray_rate IS NOT NULL

                UNION ALL

                -- Won/accepted customer quotes (carrier cost only)
                SELECT q.pod AS origin, q.final_delivery AS destination,
                       q.pod || ' \u2192 ' || q.final_delivery AS lane,
                       q.carrier_name, NULL AS carrier_email,
                       q.carrier_total AS rate_amount, 'flat' AS rate_unit,
                       q.created_at AS quote_date, 'quote' AS source,
                       q.shipment_type AS move_type, q.id, NULL AS miles, 'accepted' AS status
                FROM quotes q
                WHERE {wq_where} AND q.status = 'accepted' AND q.carrier_total > 0
            ) combined
            ORDER BY rate_amount ASC NULLS LAST, quote_date DESC NULLS LAST
            LIMIT 100
        """, rq_params + lr_params + wq_params)
        all_rows = cur.fetchall()

        # ── Directory carriers ──
        c_conds, c_params = [], []
        if origin_q:
            c_conds.append("LOWER(c.pickup_area) LIKE %s")
            c_params.append(f"%{origin_q}%")
        if dest_q:
            c_conds.append("LOWER(c.destination_area) LIKE %s")
            c_params.append(f"%{dest_q}%")
        c_where = " OR ".join(c_conds) if c_conds else "FALSE"
        try:
            cur.execute(f"""
                SELECT c.id, c.name, c.mc_number, c.email, c.phone,
                       c.pickup_area, c.destination_area,
                       c.can_dray, c.hazmat, c.overweight, c.date_quoted
                FROM carriers c WHERE {c_where}
                ORDER BY c.date_quoted DESC NULLS LAST LIMIT 20
            """, c_params)
            matching_carriers = cur.fetchall()
        except Exception:
            matching_carriers = []

    # ── Build flat matches (backwards compat) ──
    matches = []
    for r in all_rows:
        matches.append({
            "id": r["id"],
            "lane": r["norm_lane"],
            "origin": r["origin"],
            "destination": r["destination"],
            "miles": r["miles"],
            "carrier": r["carrier_name"] or "Unknown",
            "carrier_email": r["carrier_email"],
            "rate": float(r["rate_amount"]) if r["rate_amount"] else None,
            "rate_unit": r["rate_unit"],
            "date": r["quote_date"].isoformat() if r["quote_date"] else None,
            "status": r["status"],
            "source": r["source"],
            "move_type": r["move_type"],
        })

    # ── Group by normalized lane ──
    from collections import defaultdict
    lane_map = defaultdict(list)
    for m in matches:
        key = (
            (m["origin"] or "").strip().lower(),
            (m["destination"] or "").strip().lower(),
        )
        lane_map[key].append(m)

    lane_groups = []
    for (orig_key, dest_key), group in sorted(
        lane_map.items(),
        key=lambda kv: min((q["rate"] or 9e9) for q in kv[1])
    ):
        rates = [q["rate"] for q in group if q["rate"]]
        source_counts = {}
        for q in group:
            source_counts[q["source"]] = source_counts.get(q["source"], 0) + 1
        last_date = max((q["date"] or "") for q in group) or None
        lane_groups.append({
            "lane": group[0]["lane"],
            "origin": group[0]["origin"],
            "destination": group[0]["destination"],
            "count": len(group),
            "rated_count": len(rates),
            "floor": min(rates) if rates else None,
            "avg": round(sum(rates) / len(rates), 2) if rates else None,
            "ceiling": max(rates) if rates else None,
            "carriers": len(set(q["carrier"] for q in group)),
            "last_quoted": last_date,
            "sources": source_counts,  # {"email": 3, "import": 2}
            "quotes": sorted(group, key=lambda q: q["rate"] or 9e9),
        })

    # Sort lane_groups: most quotes first
    lane_groups.sort(key=lambda g: g["rated_count"], reverse=True)

    # ── Global stats ──
    all_rates = [m["rate"] for m in matches if m["rate"]]
    stats = {}
    if all_rates:
        stats = {
            "floor": min(all_rates),
            "ceiling": max(all_rates),
            "avg": round(sum(all_rates) / len(all_rates), 2),
            "count": len(all_rates),
            "total_carriers": len(set(m["carrier"] for m in matches)),
            "total_lanes": len(lane_groups),
            "sources": {
                "email": sum(1 for m in matches if m["source"] == "email"),
                "import": sum(1 for m in matches if m["source"] == "import"),
                "quote": sum(1 for m in matches if m["source"] == "quote"),
            },
        }

    carriers = [
        {
            "id": c["id"], "name": c["name"], "mc": c["mc_number"],
            "email": c["email"], "phone": c["phone"],
            "pickup": c["pickup_area"], "destination": c["destination_area"],
            "can_dray": c["can_dray"], "hazmat": c["hazmat"],
            "overweight": c["overweight"],
            "date_quoted": c["date_quoted"].isoformat() if c["date_quoted"] else None,
        }
        for c in matching_carriers
    ]

    return {"matches": matches, "lane_groups": lane_groups, "carriers": carriers, "stats": stats}


@app.get("/api/customer-reply-alerts")
async def api_customer_reply_alerts():
    """Get active customer reply alerts (unreplied for 15+ min)."""
    with db.get_cursor() as cur:
        cur.execute("""
        SELECT cra.id, cra.email_thread_id, cra.efj, cra.sender,
               cra.subject, cra.alerted_at, cra.dismissed
        FROM customer_reply_alerts cra
        WHERE cra.dismissed = FALSE
        ORDER BY cra.alerted_at DESC
        LIMIT 50
        """)
        alerts = cur.fetchall()
    return [
        {
            "id": a["id"], "email_thread_id": a["email_thread_id"],
            "efj": a["efj"], "sender": a["sender"], "subject": a["subject"],
            "alerted_at": a["alerted_at"].isoformat() if a["alerted_at"] else None,
        }
        for a in alerts
    ]


@app.post("/api/customer-reply-alerts/{alert_id}/dismiss")
async def dismiss_customer_reply_alert(alert_id: int):
    """Dismiss a customer reply alert."""
    with db.get_conn() as conn:
        with db.get_cursor(conn) as cur:
            cur.execute(
                "UPDATE customer_reply_alerts SET dismissed = TRUE WHERE id = %s RETURNING id",
                (alert_id,),
            )
            row = cur.fetchone()
    if not row:
        return JSONResponse(status_code=404, content={"error": "alert not found"})
    return {"ok": True}


# ═══════════════════════════════════════════════════════════════
# TEAM PROFILE ENDPOINTS (avatars + subtitles)
# ═══════════════════════════════════════════════════════════════

@app.get("/api/team/profiles")
async def get_team_profiles():
    """Return all team profile data (avatars + subtitles)."""
    with db.get_cursor() as cur:
        cur.execute("SELECT rep_name, avatar_filename, subtitle FROM team_profiles")
        rows = cur.fetchall()
    profiles = {}
    for row in rows:
        profiles[row["rep_name"]] = {
            "avatar_url": f"/api/team/avatar/{row['rep_name']}" if row.get("avatar_filename") else None,
            "subtitle": row.get("subtitle"),
        }
    return {"profiles": profiles}


@app.post("/api/team/{rep_name}/avatar")
async def upload_avatar(rep_name: str, file: UploadFile = File(...)):
    """Upload a profile picture for a team rep."""
    import uuid as _uuid
    allowed = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".pdf"}
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in allowed:
        return JSONResponse(status_code=400, content={"error": f"Invalid file type. Allowed: {', '.join(allowed)}"})

    os.makedirs(os.path.join("uploads", "avatars"), exist_ok=True)

    # If PDF, convert first page to PNG
    if ext == ".pdf":
        import tempfile
        from pdf2image import convert_from_bytes
        pdf_bytes = await file.read()
        try:
            images = convert_from_bytes(pdf_bytes, first_page=1, last_page=1, dpi=150)
            ext = ".png"
            unique_name = f"{_uuid.uuid4().hex[:8]}_{rep_name}{ext}"
            save_path = os.path.join("uploads", "avatars", unique_name)
            images[0].save(save_path, "PNG")
        except Exception as e:
            return JSONResponse(status_code=400, content={"error": f"Could not convert PDF: {str(e)}"})
    else:
        unique_name = f"{_uuid.uuid4().hex[:8]}_{rep_name}{ext}"
        save_path = os.path.join("uploads", "avatars", unique_name)

    # Remove old avatar if exists
    with db.get_conn() as conn:
        with db.get_cursor(conn) as cur:
            cur.execute("SELECT avatar_filename FROM team_profiles WHERE rep_name = %s", (rep_name,))
            old = cur.fetchone()
            if old and old.get("avatar_filename"):
                old_path = os.path.join("uploads", "avatars", old["avatar_filename"])
                if os.path.exists(old_path):
                    os.remove(old_path)

            if not os.path.exists(save_path):
                content = await file.read()
                with open(save_path, "wb") as f:
                    f.write(content)

            # Auto-resize + compress avatar to 200x200
            try:
                from PIL import Image as PILImage
                img = PILImage.open(save_path)
                img = img.convert("RGB")
                # Center crop to square
                w, h = img.size
                side = min(w, h)
                left = (w - side) // 2
                top = (h - side) // 2
                img = img.crop((left, top, left + side, top + side))
                img = img.resize((200, 200), PILImage.LANCZOS)
                # Save as JPEG for smaller file size
                compressed_name = os.path.splitext(unique_name)[0] + ".jpg"
                compressed_path = os.path.join("uploads", "avatars", compressed_name)
                img.save(compressed_path, "JPEG", quality=85, optimize=True)
                # Remove original if different
                if compressed_path != save_path and os.path.exists(save_path):
                    os.remove(save_path)
                unique_name = compressed_name
                save_path = compressed_path
            except Exception as resize_err:
                import traceback; traceback.print_exc()
                pass  # keep original if resize fails

            cur.execute("""
                INSERT INTO team_profiles (rep_name, avatar_filename, updated_at)
                VALUES (%s, %s, NOW())
                ON CONFLICT (rep_name)
                DO UPDATE SET avatar_filename = EXCLUDED.avatar_filename, updated_at = NOW()
            """, (rep_name, unique_name))

    return {"ok": True, "avatar_url": f"/api/team/avatar/{rep_name}"}


@app.delete("/api/team/{rep_name}/avatar")
async def delete_avatar(rep_name: str):
    """Remove a rep's profile picture."""
    with db.get_conn() as conn:
        with db.get_cursor(conn) as cur:
            cur.execute("SELECT avatar_filename FROM team_profiles WHERE rep_name = %s", (rep_name,))
            row = cur.fetchone()
            if row and row.get("avatar_filename"):
                path = os.path.join("uploads", "avatars", row["avatar_filename"])
                if os.path.exists(path):
                    os.remove(path)
                cur.execute("UPDATE team_profiles SET avatar_filename = NULL, updated_at = NOW() WHERE rep_name = %s", (rep_name,))
    return {"ok": True}


@app.get("/api/team/avatar/{rep_name}")
async def serve_avatar(rep_name: str):
    """Serve a rep's profile picture."""
    with db.get_cursor() as cur:
        cur.execute("SELECT avatar_filename FROM team_profiles WHERE rep_name = %s", (rep_name,))
        row = cur.fetchone()
    if not row or not row.get("avatar_filename"):
        return JSONResponse(status_code=404, content={"error": "No avatar found"})
    path = os.path.join("uploads", "avatars", row["avatar_filename"])
    if not os.path.exists(path):
        return JSONResponse(status_code=404, content={"error": "File not found"})
    import mimetypes
    mime = mimetypes.guess_type(path)[0] or "image/png"
    return FileResponse(path, media_type=mime, headers={"Cache-Control": "public, max-age=3600"})


_react_index = Path(__file__).parent / "static" / "dist" / "index.html"

@app.get("/")
async def react_root():
    """Redirect root to React app."""
    return RedirectResponse("/app", status_code=302)



# ═══════════════════════════════════════════════════════════════
# UNBILLED ORDERS API
# ═══════════════════════════════════════════════════════════════
import xlrd
from io import BytesIO

def _parse_unbilled_excel(file_bytes: bytes, filename: str) -> list:
    """Parse .xls or .xlsx unbilled orders file. Returns list of dicts."""
    rows = []
    if filename.lower().endswith(".xls"):
        wb = xlrd.open_workbook(file_contents=file_bytes)
        ws = wb.sheet_by_index(0)
        # Find header row: look for row containing "Order#"
        header_row = 0
        for r in range(min(5, ws.nrows)):
            vals = [str(ws.cell_value(r, c)).strip().lower() for c in range(ws.ncols)]
            if any("order#" in v for v in vals):
                header_row = r
                break
        headers = [str(ws.cell_value(header_row, c)).strip() for c in range(ws.ncols)]
        for r in range(header_row + 1, ws.nrows):
            row = {}
            for c, h in enumerate(headers):
                val = ws.cell_value(r, c)
                if ws.cell_type(r, c) == xlrd.XL_CELL_DATE:
                    try:
                        dt = xlrd.xldate_as_datetime(val, wb.datemode)
                        val = dt.strftime("%Y-%m-%d")
                    except Exception:
                        val = str(val)
                row[h] = val
            rows.append(row)
    else:
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            return []
        # Find header row (look for "Order")
        header_idx = 0
        for idx, r in enumerate(all_rows[:5]):
            if any("order#" in str(v).lower() for v in r if v):
                header_idx = idx
                break
        headers = [str(h).strip() if h else f"col{i}" for i, h in enumerate(all_rows[header_idx])]
        for vals in all_rows[header_idx + 1:]:
            row = {}
            for h, v in zip(headers, vals):
                if hasattr(v, "strftime"):
                    v = v.strftime("%Y-%m-%d")
                row[h] = v
            rows.append(row)
        wb.close()
    return rows


def _map_unbilled_row(row: dict) -> dict:
    """Map Excel columns to DB columns. Exact header matching with fallbacks."""
    def find(keys):
        for k in keys:
            # Try exact match first
            if k in row:
                return row[k]
            # Then case-insensitive exact
            for h in row:
                if h.lower().strip() == k.lower():
                    return row[h]
            # Then substring
            for h in row:
                if k.lower() in h.lower():
                    return row[h]
        return None

    def safe_date(val):
        if not val:
            return None
        s = str(val).strip()
        if not s or s == "None":
            return None
        # Already YYYY-MM-DD from parser
        if len(s) == 10 and s[4] == "-":
            return s
        # Try MM/DD/YYYY
        for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
            try:
                from datetime import datetime
                return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
        return None

    return {
        "order_num": str(find(["Order"]) or "").strip(),
        "container": str(find(["Container"]) or "").strip(),
        "bill_to": str(find(["Bill"]) or "").strip(),
        "tractor": str(find(["Tractor"]) or "").strip(),
        "entered": safe_date(find(["Entered"])),
        "appt_date": safe_date(find(["Appt"])),
        "dliv_dt": safe_date(find(["DLIV"])),
        "act_dt": safe_date(find(["ACT"])),
    }


def _calc_age(entered_str):
    """Days since entered date."""
    if not entered_str:
        return 0
    try:
        from datetime import datetime, date
        d = datetime.strptime(str(entered_str), "%Y-%m-%d").date()
        return (date.today() - d).days
    except Exception:
        return 0


@app.post("/api/unbilled/upload")
async def api_unbilled_upload(request: Request):
    """Upload .xls/.xlsx file of unbilled orders. Smart UPSERT reconciliation."""
    form = await request.form()
    file = form.get("file")
    if not file:
        return JSONResponse({"error": "No file"}, status_code=400)

    filename = file.filename or "upload.xls"
    contents = await file.read()
    try:
        rows = _parse_unbilled_excel(contents, filename)
    except Exception as e:
        return JSONResponse({"error": f"Parse error: {e}"}, status_code=400)

    if not rows:
        return JSONResponse({"error": "No data rows found"}, status_code=400)

    mapped = [_map_unbilled_row(r) for r in rows]
    mapped = [m for m in mapped if m["order_num"]]  # skip empty rows

    from datetime import datetime
    batch_id = datetime.utcnow().strftime("%Y%m%d_%H%M%S")

    inserted = 0
    updated = 0
    reconciled = 0

    with db.get_conn() as conn:
        with db.get_cursor(conn) as cur:
            # UPSERT each row — preserves billing_status for existing orders
            for m in mapped:
                age = _calc_age(m["entered"])
                cur.execute(
                    """INSERT INTO unbilled_orders
                       (order_num, container, bill_to, tractor, entered, appt_date, dliv_dt, act_dt, age_days, upload_batch)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                       ON CONFLICT (order_num) WHERE dismissed = FALSE
                       DO UPDATE SET
                           container = EXCLUDED.container,
                           bill_to = EXCLUDED.bill_to,
                           tractor = EXCLUDED.tractor,
                           entered = EXCLUDED.entered,
                           appt_date = EXCLUDED.appt_date,
                           dliv_dt = EXCLUDED.dliv_dt,
                           act_dt = EXCLUDED.act_dt,
                           age_days = EXCLUDED.age_days,
                           upload_batch = EXCLUDED.upload_batch""",
                    (m["order_num"], m["container"], m["bill_to"], m["tractor"],
                     m["entered"], m["appt_date"], m["dliv_dt"], m["act_dt"], age, batch_id)
                )
                if cur.statusmessage.startswith("INSERT"):
                    inserted += 1
                else:
                    updated += 1

            # Auto-dismiss orders NOT in this upload (they dropped off the report)
            cur.execute(
                """UPDATE unbilled_orders
                   SET dismissed = TRUE, dismissed_at = NOW(), dismissed_reason = 'reconciled'
                   WHERE dismissed = FALSE AND upload_batch != %s""",
                (batch_id,)
            )
            reconciled = cur.rowcount

    # Cross-reference: how many active unbilled orders have delivered shipments?
    delivered_count = 0
    try:
        with db.get_cursor() as _xref_cur:
            _xref_cur.execute("""
                SELECT COUNT(*) as cnt
                FROM unbilled_orders u
                JOIN shipments s ON REPLACE(u.order_num, ' ', '') = s.efj
                WHERE u.dismissed = FALSE
                  AND LOWER(s.status) IN ('delivered','completed','empty returned','billed_closed','returned to port')
            """)
            delivered_count = _xref_cur.fetchone()["cnt"]
    except Exception:
        pass

    log.info("Unbilled upload: %d inserted, %d updated, %d reconciled, %d delivered (batch=%s)",
             inserted, updated, reconciled, delivered_count, batch_id)
    return JSONResponse({
        "ok": True, "imported": inserted + updated,
        "inserted": inserted, "updated": updated, "reconciled": reconciled,
        "delivered_count": delivered_count,
        "batch": batch_id
    })


@app.get("/api/unbilled")
def api_unbilled_list():
    """List all non-dismissed unbilled orders, enriched with shipment status."""
    with db.get_cursor() as cur:
        cur.execute(
        """SELECT u.id, u.order_num, u.container, u.bill_to, u.tractor,
                  u.entered::text, u.appt_date::text, u.dliv_dt::text, u.act_dt::text,
                  u.age_days, u.upload_batch, u.created_at::text,
                  COALESCE(u.billing_status, 'ready_to_bill') as billing_status,
                  s.status as shipment_status,
                  s.account as shipment_account,
                  s.delivery_date::text as shipment_delivery,
                  s.archived as shipment_archived,
                  CASE WHEN LOWER(COALESCE(s.status,'')) IN
                       ('delivered','completed','empty returned','billed_closed','returned to port')
                       THEN TRUE ELSE FALSE END as shipment_delivered
           FROM unbilled_orders u
           LEFT JOIN shipments s ON REPLACE(u.order_num, ' ', '') = s.efj
           WHERE u.dismissed = FALSE
           ORDER BY u.age_days DESC"""
        )
        rows = cur.fetchall()
    return JSONResponse({"orders": [dict(r) for r in rows]})


@app.get("/api/unbilled/stats")
def api_unbilled_stats():
    """Unbilled orders summary stats."""
    with db.get_cursor() as cur:
        cur.execute("SELECT COUNT(*) as count FROM unbilled_orders WHERE dismissed = FALSE")
        count = cur.fetchone()["count"]
        cur.execute(
        """SELECT bill_to, COUNT(*) as cnt, MAX(age_days) as max_age
           FROM unbilled_orders WHERE dismissed = FALSE
           GROUP BY bill_to ORDER BY cnt DESC"""
        )
        by_customer = [dict(r) for r in cur.fetchall()]
        cur.execute(
        "SELECT COALESCE(SUM(age_days), 0) as total_age FROM unbilled_orders WHERE dismissed = FALSE"
        )
        total_age = cur.fetchone()["total_age"]
        avg_age = round(total_age / count, 1) if count > 0 else 0
    # Count unbilled orders with delivered shipments
    delivered_count = 0
    try:
        with db.get_cursor() as _dc:
            _dc.execute("""
                SELECT COUNT(*) as cnt
                FROM unbilled_orders u
                JOIN shipments s ON REPLACE(u.order_num, ' ', '') = s.efj
                WHERE u.dismissed = FALSE
                  AND LOWER(s.status) IN ('delivered','completed','empty returned','billed_closed','returned to port')
            """)
            delivered_count = _dc.fetchone()["cnt"]
    except Exception:
        pass
    return JSONResponse({"count": count, "avg_age": avg_age, "by_customer": by_customer, "delivered_count": delivered_count})


@app.post("/api/unbilled/{order_id}/dismiss")
def api_unbilled_dismiss(order_id: int):
    """Dismiss (hide) an unbilled order."""
    with db.get_conn() as conn:
        with db.get_cursor(conn) as cur:
            cur.execute(
                "UPDATE unbilled_orders SET dismissed = TRUE, dismissed_at = NOW(), dismissed_reason = 'manual' WHERE id = %s",
                (order_id,)
            )
    return JSONResponse({"ok": True})


@app.post("/api/unbilled/{order_id}/status")
async def api_unbilled_update_status(order_id: int, request: Request):
    """Update billing status of an unbilled order. Auto-archives shipment on 'closed'."""
    body = await request.json()
    new_status = body.get("billing_status", "").strip()
    if not new_status:
        raise HTTPException(400, "Missing billing_status")
    valid = ("ready_to_bill", "billed_cx", "driver_paid", "closed")
    if new_status not in valid:
        raise HTTPException(400, f"Invalid billing_status: {new_status}")

    # Update billing status + get order_num for cross-reference
    order_num = None
    with db.get_conn() as conn:
        with db.get_cursor(conn) as cur:
            cur.execute(
                "UPDATE unbilled_orders SET billing_status = %s WHERE id = %s AND dismissed = FALSE RETURNING order_num",
                (new_status, order_id)
            )
            row = cur.fetchone()
            if not row:
                raise HTTPException(404, f"Order {order_id} not found or dismissed")
            order_num = row["order_num"]

    # Auto-dismiss + archive on closed
    if new_status == "closed" and order_num:
        with db.get_conn() as conn:
            with db.get_cursor(conn) as cur:
                cur.execute(
                    "UPDATE unbilled_orders SET dismissed = TRUE, dismissed_at = NOW(), dismissed_reason = 'closed' WHERE id = %s",
                    (order_id,)
                )
        # Archive matching shipment
        efj = order_num.replace(" ", "").strip()
        if efj:
            _archive_shipment_on_close(efj)

    return JSONResponse({"ok": True, "billing_status": new_status})





# (duplicate endpoint removed — consolidated into single handler above)


# ===================================================================
# UNBILLED → SHIPMENT ARCHIVE HELPERS
# ===================================================================

def _archive_shipment_on_close(efj: str):
    """Archive a shipment in PG + Google Sheet when its unbilled order is closed."""
    import sys as _sys
    _csl_bot_dir = "/root/csl-bot"
    if _csl_bot_dir not in _sys.path:
        _sys.path.insert(0, _csl_bot_dir)
    # PG archive
    try:
        from csl_pg_writer import pg_archive_shipment
        pg_archive_shipment(efj)
        log.info("Unbilled close → PG archived %s", efj)
    except Exception as e:
        log.warning("Unbilled close → PG archive failed for %s: %s", efj, e)
    # Sheet archive (fire-and-forget)
    try:
        with db.get_cursor() as cur:
            cur.execute("SELECT account, rep FROM shipments WHERE efj = %s", (efj,))
            ship = cur.fetchone()
        if ship and ship.get("account"):
            from csl_sheet_writer import sheet_archive_row
            sheet_archive_row(efj, ship["account"], ship.get("rep"))
    except Exception as e:
        log.warning("Unbilled close → sheet archive failed for %s: %s", efj, e)
    # Invalidate completed cache
    _completed_cache["ts"] = 0


@app.post("/api/unbilled/bulk-close-delivered")
async def api_unbilled_bulk_close_delivered():
    """Close all unbilled orders whose matching shipment is delivered/completed."""
    closed_efjs = []
    with db.get_conn() as conn:
        with db.get_cursor(conn) as cur:
            cur.execute("""
                UPDATE unbilled_orders u
                SET billing_status = 'closed',
                    dismissed = TRUE,
                    dismissed_at = NOW(),
                    dismissed_reason = 'auto_delivered'
                FROM shipments s
                WHERE REPLACE(u.order_num, ' ', '') = s.efj
                  AND u.dismissed = FALSE
                  AND LOWER(s.status) IN ('delivered','completed','empty returned','billed_closed','returned to port')
                RETURNING s.efj
            """)
            closed_efjs = [r["efj"] for r in cur.fetchall()]

    # Archive each in PG + sheet (best-effort)
    for efj in closed_efjs:
        _archive_shipment_on_close(efj)

    log.info("Bulk close delivered: %d orders closed + archived", len(closed_efjs))
    return JSONResponse({"ok": True, "closed_count": len(closed_efjs)})


# ===================================================================
# DOCUMENT HUB API
# ===================================================================

@app.get("/api/load/{efj}/documents")
async def get_load_documents(efj: str):
    """List all documents for a load."""
    with db.get_cursor() as cur:
        cur.execute(
        "SELECT id, doc_type, original_name, size_bytes, uploaded_at "
        "FROM load_documents WHERE efj = %s ORDER BY uploaded_at DESC",
        (efj,)
        )
        rows = cur.fetchall()
    docs = [
        {
            "id": r["id"],
            "doc_type": r["doc_type"],
            "original_name": r["original_name"],
            "size_bytes": r["size_bytes"],
            "uploaded_at": r["uploaded_at"].isoformat() if r["uploaded_at"] else None,
        }
        for r in rows
    ]
    return JSONResponse({"documents": docs})


@app.post("/api/load/{efj}/documents")
async def upload_load_document(efj: str, file: UploadFile = File(...), doc_type: str = Form("other")):
    """Upload a document for a load."""
    import uuid
    upload_dir = f"/root/csl-bot/csl-doc-tracker/uploads/{efj}"
    os.makedirs(upload_dir, exist_ok=True)
    safe_name = f"{uuid.uuid4().hex[:8]}_{file.filename}"
    file_path = os.path.join(upload_dir, safe_name)
    contents = await file.read()
    with open(file_path, "wb") as f:
        f.write(contents)
    with db.get_conn() as conn:
        with db.get_cursor(conn) as cur:
            cur.execute(
                "INSERT INTO load_documents (efj, doc_type, filename, original_name, size_bytes) "
                "VALUES (%s, %s, %s, %s, %s) RETURNING id",
                (efj, doc_type, safe_name, file.filename, len(contents))
            )
            doc_id = cur.fetchone()["id"]
    return JSONResponse({"ok": True, "id": doc_id, "original_name": file.filename})


@app.delete("/api/load/{efj}/documents/{doc_id}")
async def delete_load_document(efj: str, doc_id: int):
    """Delete a document for a load."""
    with db.get_cursor() as cur:
        cur.execute("SELECT filename FROM load_documents WHERE id = %s AND efj = %s", (doc_id, efj))
        row = cur.fetchone()
    if row:
        file_path = f"/root/csl-bot/csl-doc-tracker/uploads/{efj}/{row['filename']}"
        if os.path.exists(file_path):
            os.remove(file_path)
        with db.get_conn() as conn:
            with db.get_cursor(conn) as cur:
                cur.execute("DELETE FROM load_documents WHERE id = %s AND efj = %s", (doc_id, efj))
    return JSONResponse({"ok": True})


@app.get("/api/load/{efj}/documents/{doc_id}/download")
async def download_load_document(efj: str, doc_id: int, inline: bool = False):
    """Download or inline-preview a specific document."""
    with db.get_cursor() as cur:
        cur.execute(
        "SELECT filename, original_name FROM load_documents WHERE id = %s AND efj = %s",
        (doc_id, efj)
        )
        row = cur.fetchone()
    if not row:
        return JSONResponse(status_code=404, content={"error": "not found"})
    file_path = f"/root/csl-bot/csl-doc-tracker/uploads/{efj}/{row['filename']}"
    if not os.path.exists(file_path):
        return JSONResponse(status_code=404, content={"error": "file missing"})
    if inline:
        import mimetypes
        media_type = mimetypes.guess_type(row["original_name"])[0] or "application/octet-stream"
        return FileResponse(
            file_path,
            media_type=media_type,
            headers={"Content-Disposition": f'inline; filename="{row["original_name"]}"'}
        )
    return FileResponse(file_path, filename=row["original_name"])


@app.patch("/api/load/{efj}/documents/{doc_id}")
async def update_load_document(efj: str, doc_id: int, request: Request):
    """Update document metadata (doc_type reclassification)."""
    body = await request.json()
    new_type = body.get("doc_type", "").strip()
    valid_types = [
        "customer_rate", "carrier_rate", "rate", "unclassified",
        "pod", "bol", "carrier_invoice", "screenshot", "email", "other",
    ]
    if new_type not in valid_types:
        return JSONResponse(status_code=400, content={"error": f"Invalid doc_type. Must be one of: {valid_types}"})
    with db.get_conn() as conn:
        with db.get_cursor(conn) as cur:
            cur.execute(
                "UPDATE load_documents SET doc_type = %s WHERE id = %s AND efj = %s RETURNING id",
                (new_type, doc_id, efj)
            )
            row = cur.fetchone()
    if not row:
        return JSONResponse(status_code=404, content={"error": "not found"})
    return JSONResponse({"ok": True, "doc_type": new_type})


# ===================================================================
# MACROPOINT SCREENSHOT API
# ===================================================================


# ── Driver contact endpoints ──────────────────────────────────────────────

@app.get("/api/load/{efj}/driver")
async def api_get_driver(efj: str):
    """Return driver contact info for a load."""
    contact = _get_driver_contact(efj)

    # Fall back to shipments table for driver/phone/container_url
    # (Tolead sync + PG-native loads populate these but not driver_contacts)
    pg_driver = ""
    pg_phone = ""
    pg_mp_url = ""
    try:
        with db.get_cursor() as cur:
            cur.execute(
                "SELECT driver, driver_phone, container_url FROM shipments WHERE efj = %s",
                (efj,),
            )
            pg_row = cur.fetchone()
            if pg_row:
                pg_driver = (pg_row["driver"] or "").strip()
                pg_phone = (pg_row["driver_phone"] or "").strip()
                pg_mp_url = (pg_row["container_url"] or "").strip()
    except Exception:
        pass

    # Also check Column N (driver/truck) from sheet cache as fallback for name
    sheet_driver = ""
    for s in sheet_cache.shipments:
        if s["efj"] == efj:
            sheet_driver = s.get("notes", "")  # COL index 13 = Column N = Driver/Truck
            break

    # Fall back to tracking cache for driver_phone (scraped from MP portal)
    cached_phone = ""
    cached_mp_url = ""
    tracking_cache = _read_tracking_cache()
    cached = _find_tracking_entry(tracking_cache, efj)
    if cached.get("driver_phone"):
        cached_phone = cached["driver_phone"]
        # Auto-save scraped phone to DB if DB doesn't have one yet
        if not contact.get("driver_phone"):
            try:
                _upsert_driver_contact(efj, phone=cached_phone)
            except Exception:
                pass
    if cached.get("macropoint_url"):
        cached_mp_url = cached["macropoint_url"]

    # Priority: driver_contacts > shipments table > sheet cache > tracking cache
    final_name = contact.get("driver_name") or pg_driver or sheet_driver or ""
    final_phone = contact.get("driver_phone") or pg_phone or cached_phone or ""
    final_mp_url = contact.get("macropoint_url") or pg_mp_url or cached_mp_url or ""
    phone_source = "db" if contact.get("driver_phone") else ("pg" if pg_phone else ("macropoint" if cached_phone else ""))

    return {
        "efj": efj,
        "driverName": final_name,
        "driverPhone": final_phone,
        "driverEmail": contact.get("driver_email") or "",
        "carrierEmail": contact.get("carrier_email") or "",
        "trailerNumber": contact.get("trailer_number") or "",
        "macropointUrl": final_mp_url,
        "notes": contact.get("notes") or "",
        "phoneSource": phone_source,
    }


@app.post("/api/load/{efj}/driver")
async def api_update_driver(efj: str, request: Request):
    """Create or update driver contact info."""
    body = await request.json()
    _upsert_driver_contact(
        efj,
        name=body.get("driverName"),
        phone=body.get("driverPhone"),
        email=body.get("driverEmail"),
        notes=body.get("notes"),
        carrier_email=body.get("carrierEmail"),
        trailer_number=body.get("trailerNumber"),
        macropoint_url=body.get("macropointUrl"),
    )
    log.info("Driver contact updated for %s", efj)
    return {"status": "ok", "efj": efj}


# -- Timestamped Notes Log ---------------------------------------------------

@app.get("/api/load/{efj}/notes")
def api_load_notes_list(efj: str):
    """List all timestamped notes for a load, newest first."""
    with db.get_cursor() as cur:
        cur.execute(
        """SELECT id, efj, note_text, created_by, created_at::text
           FROM load_notes
           WHERE efj = %s
           ORDER BY created_at DESC""",
        (efj,)
        )
        rows = cur.fetchall()
    return JSONResponse({"notes": [dict(r) for r in rows]})


@app.post("/api/load/{efj}/notes")
async def api_load_notes_add(efj: str, request: Request):
    """Add a timestamped note to a load."""
    body = await request.json()
    text = (body.get("text") or "").strip()
    if not text:
        raise HTTPException(400, "Missing note text")
    created_by = (body.get("created_by") or "dashboard").strip()
    with db.get_conn() as conn:
        with db.get_cursor(conn) as cur:
            cur.execute(
                """INSERT INTO load_notes (efj, note_text, created_by)
                   VALUES (%s, %s, %s)
                   RETURNING id, efj, note_text, created_by, created_at::text""",
                (efj, text, created_by)
            )
            row = cur.fetchone()
    log.info("Note added for %s by %s", efj, created_by)
    return JSONResponse({"ok": True, "note": dict(row)})


@app.get("/api/macropoint/{efj}/screenshot")
async def get_macropoint_screenshot(efj: str):
    """Serve a cached Macropoint tracking screenshot."""
    screenshot_path = f"/root/csl-bot/csl-doc-tracker/uploads/mp_screenshots/{efj}.png"
    if not os.path.exists(screenshot_path):
        return JSONResponse(status_code=404, content={"error": "no screenshot"})
    import json as _json
    meta_path = f"/root/csl-bot/csl-doc-tracker/uploads/mp_screenshots/{efj}.json"
    headers = {"Cache-Control": "max-age=300"}
    if os.path.exists(meta_path):
        with open(meta_path) as f:
            meta = _json.load(f)
        headers["X-Captured-At"] = meta.get("captured_at", "")
    return FileResponse(screenshot_path, media_type="image/png", headers=headers)


@app.get("/app")
@app.get("/app/{path:path}")
async def react_spa(path: str = ""):
    """Serve React SPA for all /app/* routes."""
    # Serve actual static files from dist/ if they exist (images, etc.)
    if path:
        import mimetypes as _mt
        static_file = _react_dist / path
        if static_file.is_file():
            mime, _ = _mt.guess_type(str(static_file))
            return FileResponse(str(static_file), media_type=mime or "application/octet-stream")
    if _react_index.exists():
        return FileResponse(str(_react_index), media_type="text/html")
    return RedirectResponse("/legacy", status_code=302)

@app.get("/api/macropoint/{efj}")
async def api_macropoint(efj: str):
    """Return Macropoint tracking data for an FTL load."""
    sheet_cache.refresh_if_needed()
    # Find the shipment — check sheet cache first, then Postgres
    shipment = None
    for s in sheet_cache.shipments:
        if s["efj"] == efj:
            shipment = s
            break
    if not shipment:
        # Fall back to Postgres (PG-migrated loads aren't in sheet cache)
        try:
            with db.get_cursor() as cur:
                cur.execute(
                    "SELECT efj, container, status, origin, destination, pickup_date, "
                    "delivery_date, eta, account, move_type, container_url, carrier "
                    "FROM shipments WHERE efj = %s",
                    (efj,),
                )
                pg_row = cur.fetchone()
                if pg_row:
                    shipment = {
                        "efj": pg_row["efj"],
                        "container": pg_row["container"] or "",
                        "status": pg_row["status"] or "",
                        "origin": pg_row["origin"] or "",
                        "destination": pg_row["destination"] or "",
                        "pickup": pg_row["pickup_date"] or "",
                        "delivery": pg_row["delivery_date"] or "",
                        "eta": pg_row["eta"] or "",
                        "account": pg_row["account"] or "",
                        "move_type": pg_row["move_type"] or "",
                        "container_url": pg_row["container_url"] or "",
                        "carrier": pg_row["carrier"] or "",
                    }
        except Exception:
            pass
    if not shipment:
        raise HTTPException(404, f"Load {efj} not found")
    # Also check tracking cache
    _tracking_cache = _read_tracking_cache()
    _cached_entry = _find_tracking_entry(_tracking_cache, efj)
    _cached_url = _cached_entry.get("macropoint_url", "")

    status = shipment.get("status", "")
    progress = _build_macropoint_progress(status)

    # Format phone
    phone_raw = TRACKING_PHONE
    if len(phone_raw) == 10:
        phone_fmt = f"({phone_raw[:3]}) {phone_raw[3:6]}-{phone_raw[6:]}"
    else:
        phone_fmt = phone_raw

    # ── Tracking cache (stop timeline from ftl_monitor) ──
    tracking_cache = _read_tracking_cache()
    cached = _find_tracking_entry(tracking_cache, efj)

    # ── Driver contact info (from DB, with cache fallback) ──
    contact = _get_driver_contact(efj)
    driver_name = contact.get("driver_name") or ""
    driver_phone = contact.get("driver_phone") or cached.get("driver_phone") or ""
    driver_email = contact.get("driver_email") or ""
    stop_times = cached.get("stop_times", {})
    cant_make_it = cached.get("cant_make_it")
    last_scraped = cached.get("last_scraped")
    mp_load_id_cached = cached.get("mp_load_id")

    # Build stop timeline from PG tracking_events (durable) + cache supplement
    timeline = _build_timeline_from_pg(efj)

    # Supplement: merge cache stop_times for any events missing from PG.
    # PG might only have Tracking Started (AF) while cache has arrival/departure
    # from GPS inference that failed to persist (e.g. "ET" timestamp bug).
    _pg_events = {e.get("event", "") for e in timeline}
    _cache_supplements = []
    if stop_times.get("stop1_arrived") and "Arrived at Pickup" not in _pg_events:
        _cache_supplements.append({"event": "Arrived at Pickup", "time": stop_times["stop1_arrived"], "type": "arrived"})
    if stop_times.get("stop1_departed") and "Departed Pickup" not in _pg_events:
        _cache_supplements.append({"event": "Departed Pickup", "time": stop_times["stop1_departed"], "type": "departed"})
    if stop_times.get("stop2_arrived") and "Arrived at Delivery" not in _pg_events:
        _cache_supplements.append({"event": "Arrived at Delivery", "time": stop_times["stop2_arrived"], "type": "arrived"})
    if stop_times.get("stop2_departed") and "Departed Delivery" not in _pg_events:
        _cache_supplements.append({"event": "Departed Delivery", "time": stop_times["stop2_departed"], "type": "departed"})
    if not timeline and not _cache_supplements:
        # Pure fallback for ETA-only loads
        if stop_times.get("stop1_eta"):
            _cache_supplements.append({"event": "Pickup ETA", "time": stop_times["stop1_eta"], "type": "eta"})
        if stop_times.get("stop2_eta"):
            _cache_supplements.append({"event": "Delivery ETA", "time": stop_times["stop2_eta"], "type": "eta"})
    timeline.extend(_cache_supplements)

    # Detect behind schedule from ETA strings
    behind_schedule = False
    for k in ("stop1_eta", "stop2_eta"):
        if stop_times.get(k) and "BEHIND" in stop_times[k].upper():
            behind_schedule = True

    return {
        "loadId": shipment.get("container", "") or shipment.get("efj", ""),
        "carrier": "Evans Delivery Company, Inc.",
        "driver": driver_name,
        "phone": phone_fmt,
        "email": DISPATCH_EMAIL,
        "trackingStatus": status or "Unknown",
        "macropointUrl": shipment.get("container_url", "") or cached.get("macropoint_url", ""),
        "progress": progress,
        "origin": shipment.get("origin", ""),
        "destination": shipment.get("destination", ""),
        "pickup": shipment.get("pickup", ""),
        "delivery": shipment.get("delivery", ""),
        "eta": shipment.get("eta", ""),
        "account": shipment.get("account", ""),
        "moveType": shipment.get("move_type", ""),
        # ── New fields ──
        "driverName": driver_name,
        "driverPhone": driver_phone,
        "driverEmail": driver_email,
        "mpLoadId": mp_load_id_cached,
        "timeline": timeline,
        "behindSchedule": behind_schedule,
        "cantMakeIt": cant_make_it,
        "lastScraped": last_scraped,
        "mpLastUpdated": cached.get("last_event_at") or cached.get("last_ping_at") or last_scraped or None,
        "lastLocation": cached.get("last_location") or None,
        "mpDisplayStatus": _classify_mp_display_status(cached, shipment)[0] if cached else "",
        "mpDisplayDetail": _classify_mp_display_status(cached, shipment)[1] if cached else "",
        "scheduleAlert": cached.get("schedule_alert", "") if cached else "",
        "distanceToStop": cached.get("distance_to_stop", "") if cached else "",
    }


@app.post("/api/load/{efj}/status")
async def api_update_status(efj: str, request: Request):
    """Update a load's status in Google Sheet."""
    body = await request.json()
    new_status = body.get("status", "").strip()
    if not new_status:
        raise HTTPException(400, "Missing status")

    # Find the shipment in cache to get its tab and row
    shipment = None
    for s in sheet_cache.shipments:
        if s["efj"] == efj:
            shipment = s
            break
    if not shipment:
        raise HTTPException(404, f"Load {efj} not found")

    tab = shipment.get("account", "")
    if not tab:
        raise HTTPException(400, "Cannot determine sheet tab for this load")

    # Write to Google Sheet
    try:
        creds = Credentials.from_service_account_file(
            CREDS_FILE, scopes=["https://www.googleapis.com/auth/spreadsheets"]
        )
        gc = gspread.authorize(creds)

        if tab == "Tolead":
            sh = gc.open_by_key(TOLEAD_SHEET_ID)
            ws = sh.worksheet(TOLEAD_TAB)
            rows = ws.get_all_values()
            target_row = None
            for i, row in enumerate(rows):
                if len(row) > TOLEAD_COL_EFJ and row[TOLEAD_COL_EFJ].strip() == efj:
                    target_row = i + 1
                    break
            if not target_row:
                raise HTTPException(404, f"Row for {efj} not found in Tolead")
            ws.update_cell(target_row, TOLEAD_COL_STATUS + 1, new_status)

        elif tab == "Boviet":
            # Search all Boviet tabs for the EFJ
            sh = gc.open_by_key(BOVIET_SHEET_ID)
            found = False
            for bov_tab, cfg in BOVIET_TAB_CONFIGS.items():
                try:
                    ws = sh.worksheet(bov_tab)
                    rows = ws.get_all_values()
                    for i, row in enumerate(rows):
                        if len(row) > cfg["efj_col"] and row[cfg["efj_col"]].strip() == efj:
                            ws.update_cell(i + 1, cfg["status_col"] + 1, new_status)
                            found = True
                            log.info("Status updated: %s → %s (Boviet/%s, row=%d)", efj, new_status, bov_tab, i + 1)
                            break
                    if found:
                        break
                except Exception:
                    continue
            if not found:
                raise HTTPException(404, f"Row for {efj} not found in Boviet tabs")

        else:
            # Master sheet — tab name is the account name
            sh = gc.open_by_key(SHEET_ID)
            ws = sh.worksheet(tab)
            rows = ws.get_all_values()
            target_row = None
            efj_col = COL.get("efj", 0)
            status_col = COL.get("status", 12)
            for i, row in enumerate(rows):
                if len(row) > efj_col and row[efj_col].strip() == efj:
                    target_row = i + 1
                    break
            if not target_row:
                raise HTTPException(404, f"Row for {efj} not found in {tab}")
            ws.update_cell(target_row, status_col + 1, new_status)

            # ── Auto-archive to Completed tab on billed_closed ──
            if new_status == "billed_closed" and target_row:
                try:
                    rep_name = sheet_cache.rep_map.get(tab, "")
                    completed_tab = f"Completed {rep_name}" if rep_name else ""
                    if completed_tab and completed_tab in _COMPLETED_TABS:
                        row_data = rows[target_row - 1] if target_row - 1 < len(rows) else []
                        if row_data:
                            # Write billed_closed into the row data before archiving
                            while len(row_data) <= status_col:
                                row_data.append("")
                            row_data[status_col] = new_status
                            # Append to Completed tab
                            cws = sh.worksheet(completed_tab)
                            cws.append_row(row_data, value_input_option="USER_ENTERED")
                            # Delete from active tab
                            ws.delete_rows(target_row)
                            log.info("Archived %s from %s → %s (row %d)", efj, tab, completed_tab, target_row)
                            # Invalidate completed cache so it refreshes
                            _completed_cache["ts"] = 0
                    else:
                        log.warning("No completed tab for rep=%r (account=%s) — skipping archive", rep_name, tab)
                except Exception as archive_err:
                    log.error("Auto-archive failed for %s: %s (status still updated)", efj, archive_err)

        # Update cache
        shipment["status"] = new_status
        log.info("Status updated: %s → %s (tab=%s)", efj, new_status, tab)

        # Send delivery email for Master loads
        _normalized = new_status.strip().lower()
        if _normalized == "delivered" and tab not in ("Tolead", "Boviet"):
            send_delivery_email(shipment)

        return {"status": "ok", "efj": efj, "new_status": new_status}

    except HTTPException:
        raise
    except Exception as e:
        log.error("Failed to update status for %s: %s", efj, e)
        raise HTTPException(500, f"Failed to update status: {e}")




# ── SeaRates Vessel Schedules + Port Codes ──

import requests as _requests
from datetime import timedelta as _timedelta

_CARRIER_SCAC_MAP = {
    "maersk": "MAEU", "msc": "MSCU", "cosco": "COSU",
    "evergreen": "EGLV", "yang ming": "YMLU", "hmm": "HDMU",
    "hyundai": "HDMU", "oocl": "OOLU", "one": "ONEY",
    "ocean network": "ONEY", "cma": "CMDU", "cma cgm": "CMDU",
    "hapag": "HLCU", "hapag-lloyd": "HLCU", "zim": "ZIMU",
    "wan hai": "WHLC", "apl": "CMDU",
}

def _resolve_locode(city_name):
    """Resolve a city name to a UN/LOCODE via DB lookup."""
    if not city_name:
        return None
    clean = city_name.strip().lower()
    # Strip state/country suffixes like ", NJ" or ", OH"
    for sep in [",", " - "]:
        if sep in clean:
            clean = clean.split(sep)[0].strip()
    # Remove common prefixes
    for prefix in ["port ", "port of "]:
        if clean.startswith(prefix):
            clean = clean[len(prefix):]
    try:
        with db.get_conn() as conn:
            with db.get_cursor(conn) as cur:
                cur.execute("SELECT locode FROM port_locode_map WHERE city_name = %s", (clean,))
                row = cur.fetchone()
                if row:
                    return row[0] if isinstance(row, tuple) else row.get("locode")
    except Exception:
        pass
    return None

def _resolve_scac(ssl_field):
    """Extract SCAC code from carrier/vessel name."""
    if not ssl_field:
        return None
    lower = ssl_field.lower()
    for key, scac in _CARRIER_SCAC_MAP.items():
        if key in lower:
            return scac
    return None

def _searates_schedule_lookup(origin_locode, dest_locode, carrier_scac=None, from_date=None):
    """Query SeaRates Ship Schedules API for sailing schedules."""
    api_key = os.environ.get("SEARATES_SCHEDULES_API_KEY") or os.environ.get("SEARATES_API_KEY")
    if not api_key:
        return []
    if not from_date:
        from_date = datetime.now().strftime("%Y-%m-%d")
    params = {
        "cargo_type": "GC",
        "origin": origin_locode,
        "destination": dest_locode,
        "from_date": from_date,
        "weeks": 4,
        "sort": "DEP",
    }
    if carrier_scac:
        params["carriers"] = carrier_scac
    try:
        resp = _requests.get(
            "https://schedules.searates.com/api/v2/schedules/by-points",
            params=params,
            headers={"X-API-KEY": api_key},
            timeout=30
        )
        if resp.status_code == 200:
            data = resp.json()
            return data.get("data", {}).get("schedules", [])
    except Exception as e:
        log.warning("SeaRates schedules lookup failed: %s", e)
    return []

def _searates_container_lookup(number):
    """Query SeaRates Container Tracking API."""
    api_key = os.environ.get("SEARATES_API_KEY")
    if not api_key:
        return {}
    try:
        resp = _requests.get(
            "https://tracking.searates.com/tracking",
            params={"api_key": api_key, "number": number, "sealine": "auto"},
            timeout=30
        )
        if resp.status_code == 200:
            return resp.json()
    except Exception as e:
        log.warning("SeaRates container lookup failed: %s", e)
    return {}

def _extract_tracking_data(raw):
    """Extract ETA, vessel, carrier, LFD from SeaRates tracking response."""
    result = {"eta": None, "vessel": None, "carrier": None, "lfd": None, "status": None}
    if not raw or not raw.get("data"):
        return result
    data = raw.get("data", {})
    # Carrier info
    metadata = data.get("metadata", {})
    if metadata.get("sealine_name"):
        result["carrier"] = metadata["sealine_name"]
    # Parse route/events for ETA
    route = data.get("route", {})
    pod = route.get("pod", {})
    if pod.get("date"):
        result["eta"] = pod["date"][:10] if len(pod.get("date", "")) >= 10 else None
    # Vessel from route
    prepol = route.get("prepol", {})
    if prepol.get("name") and "vessel" in prepol.get("transport_type", "").lower():
        result["vessel"] = prepol.get("name")
    # Parse containers for status
    containers = data.get("containers", [])
    if containers:
        events = containers[0].get("events", [])
        if events:
            last = events[-1]
            result["status"] = last.get("description", "")
            # Look for LFD in event descriptions
            for ev in events:
                desc = (ev.get("description") or "").lower()
                if "last free" in desc or "lfd" in desc:
                    result["lfd"] = ev.get("date", "")[:10] if ev.get("date") else None
    return result


@app.post("/api/searates/lookup")
async def api_searates_lookup(request: Request):
    """Auto-fetch vessel schedule data from SeaRates APIs."""
    body = await request.json()
    move_type = body.get("moveType", "")
    number = body.get("number", "").strip()
    origin = body.get("origin", "")
    destination = body.get("destination", "")

    result = {
        "eta": None, "lfd": None, "cutoff": None, "erd": None,
        "vessel": None, "carrier": None, "terminal": None,
        "voyage": None, "transitDays": None
    }

    # Step 1: Container/booking tracking
    if number:
        raw = _searates_container_lookup(number)
        tracking = _extract_tracking_data(raw)
        result["eta"] = tracking.get("eta")
        result["lfd"] = tracking.get("lfd")
        result["vessel"] = tracking.get("vessel")
        result["carrier"] = tracking.get("carrier")

    # Step 2: Ship Schedules (if ports resolve)
    origin_locode = _resolve_locode(origin)
    dest_locode = _resolve_locode(destination)
    carrier_scac = _resolve_scac(result.get("carrier") or "")

    if origin_locode and dest_locode:
        schedules = _searates_schedule_lookup(origin_locode, dest_locode, carrier_scac)
        if schedules:
            best = schedules[0]  # Already sorted by departure
            is_export = "Export" in move_type
            if is_export:
                dep = best.get("origin", {}).get("estimated_date", "")
                result["erd"] = dep[:10] if dep else None
                # Cut-off from legs or schedule data
                legs = best.get("legs", [])
                if legs:
                    cut = legs[0].get("departure", {}).get("estimated_date", "")
                    result["cutoff"] = cut[:10] if cut else result["erd"]
                result["terminal"] = best.get("origin", {}).get("terminal_name")
            else:
                arr = best.get("destination", {}).get("estimated_date", "")
                if not result["eta"] and arr:
                    result["eta"] = arr[:10]
                result["terminal"] = best.get("destination", {}).get("terminal_name")

            result["vessel"] = result["vessel"] or (best.get("legs", [{}])[0].get("vessel_name") if best.get("legs") else None)
            result["voyage"] = (best.get("legs", [{}])[0].get("voyages", [{}])[0].get("voyage") if best.get("legs") and best["legs"][0].get("voyages") else None)
            result["transitDays"] = best.get("transit_time")
            result["carrier"] = result["carrier"] or best.get("carrier_name")

    return JSONResponse(result)


@app.get("/api/port-codes")
async def api_port_codes():
    """Return all known port code mappings."""
    try:
        with db.get_conn() as conn:
            with db.get_cursor(conn) as cur:
                cur.execute("SELECT city_name, locode, port_name, country, region FROM port_locode_map ORDER BY city_name")
                rows = cur.fetchall()
        ports = []
        for r in rows:
            if isinstance(r, dict):
                ports.append(r)
            else:
                ports.append({"city_name": r[0], "locode": r[1], "port_name": r[2], "country": r[3], "region": r[4]})
        return {"ports": ports, "count": len(ports)}
    except Exception as e:
        log.error("Failed to fetch port codes: %s", e)
        raise HTTPException(500, str(e))


@app.get("/api/reps")
async def api_reps():
    """Return list of account reps."""
    return {"reps": ["Eli", "Radka", "John F", "Janice"]}


@app.post("/api/accounts/add")
async def api_add_account(request: Request):
    """Add a new account tab to the Master Tracker."""
    body = await request.json()
    name = body.get("name", "").strip()
    rep = body.get("rep", "").strip()
    if not name:
        raise HTTPException(400, "Account name is required")
    # For now, just return success — actual sheet tab creation would need gspread
    log.info("New account requested: %s (rep: %s)", name, rep)
    return {"ok": True, "account": name, "rep": rep}


# ── Add New Load ──

# Master Tracker account tabs (anything NOT in SKIP_TABS, Tolead, or Boviet)
_TOLEAD_ACCOUNTS = {"Tolead", "Tolead ORD", "Tolead JFK", "Tolead LAX", "Tolead DFW"}
_BOVIET_ACCOUNTS = set(BOVIET_TAB_CONFIGS.keys()) | {"Boviet"}

@app.post("/api/load/add")
async def api_add_load(request: Request):
    """Add a new load to the appropriate Google Sheet tab."""
    body = await request.json()
    efj = body.get("efj", "").strip()
    account = body.get("account", "").strip()
    if not efj:
        raise HTTPException(400, "EFJ Pro # is required")
    if not account:
        raise HTTPException(400, "Account is required")

    move_type = body.get("moveType", "Dray Import")
    container = body.get("container", "").strip()
    carrier = body.get("carrier", "").strip()
    origin = body.get("origin", "").strip()
    destination = body.get("destination", "").strip()
    eta = body.get("eta", "")
    lfd = body.get("lfd", "")
    pickup = body.get("pickupDate", "")
    delivery = body.get("deliveryDate", "")
    status = body.get("status", "")
    notes = body.get("notes", "")
    bol = body.get("bol", "").strip()
    customer_ref = body.get("customerRef", "").strip()
    equipment_type = body.get("equipmentType", "").strip()
    rep = body.get("rep", "").strip()
    driver_phone = body.get("driverPhone", "")
    trailer = body.get("trailerNumber", "")
    carrier_email = body.get("carrierEmail", "")
    mp_url = body.get("macropointUrl", "")

    try:
        creds = Credentials.from_service_account_file(
            CREDS_FILE, scopes=["https://www.googleapis.com/auth/spreadsheets"]
        )
        gc = gspread.authorize(creds)

        # --- Determine which sheet to write to ---
        target_sheet = "master"
        hub_key = None
        boviet_tab = None

        # Check if Tolead hub
        if account in _TOLEAD_ACCOUNTS:
            hub_key = account.replace("Tolead ", "").upper() if account != "Tolead" else "ORD"
            if hub_key not in TOLEAD_HUB_CONFIGS:
                raise HTTPException(400, f"Unknown Tolead hub: {hub_key}")
            target_sheet = "tolead"

        # Check if Boviet tab
        elif account in _BOVIET_ACCOUNTS:
            boviet_tab = account if account != "Boviet" else "Piedra"
            if boviet_tab not in BOVIET_TAB_CONFIGS:
                raise HTTPException(400, f"Unknown Boviet tab: {boviet_tab}")
            target_sheet = "boviet"

        if target_sheet == "master":
            # --- Master Tracker: columns A-P ---
            sh = gc.open_by_key(SHEET_ID)
            try:
                ws = sh.worksheet(account)
            except gspread.WorksheetNotFound:
                raise HTTPException(400, f"Account tab '{account}' not found in Master Tracker")
            status_display = status.replace("_", " ").title() if status else ""
            row = [""] * 16
            row[COL["efj"]] = efj
            row[COL["move_type"]] = move_type
            row[COL["container"]] = container
            row[COL["carrier"]] = carrier
            if bol:
                row[3] = bol  # Column D: BOL/Booking
            row[COL["origin"]] = origin
            row[COL["destination"]] = destination
            row[COL["eta"]] = eta
            row[COL["lfd"]] = lfd
            row[COL["pickup"]] = pickup
            row[COL["delivery"]] = delivery
            row[COL["status"]] = status_display
            row[COL["notes"]] = notes
            ws.append_row(row, value_input_option="USER_ENTERED")
            log.info("Added load %s to Master/%s", efj, account)
            result_tab = account

        elif target_sheet == "tolead":
            # --- Tolead hub sheet ---
            hub_cfg = TOLEAD_HUB_CONFIGS[hub_key]
            sh = gc.open_by_key(hub_cfg["sheet_id"])
            ws = sh.worksheet(hub_cfg["tab"])
            cols = hub_cfg["cols"]
            max_col = max(v for v in cols.values() if v is not None) + 1
            row = [""] * max_col
            if cols.get("efj") is not None:
                row[cols["efj"]] = efj
            if cols.get("load_id") is not None:
                row[cols["load_id"]] = container or efj
            if cols.get("status") is not None:
                row[cols["status"]] = status.replace("_", " ").title() if status else ""
            if cols.get("origin") is not None:
                row[cols["origin"]] = origin or hub_cfg.get("default_origin", "")
            if cols.get("destination") is not None:
                row[cols["destination"]] = destination
            if cols.get("pickup_date") is not None:
                row[cols["pickup_date"]] = pickup
            if cols.get("delivery") is not None:
                row[cols["delivery"]] = delivery
            if cols.get("driver") is not None:
                row[cols["driver"]] = trailer
            if cols.get("phone") is not None:
                row[cols["phone"]] = driver_phone
            ws.append_row(row, value_input_option="USER_ENTERED")
            log.info("Added load %s to Tolead/%s", efj, hub_key)
            result_tab = f"Tolead {hub_key}"

        elif target_sheet == "boviet":
            # --- Boviet tab ---
            cfg = BOVIET_TAB_CONFIGS[boviet_tab]
            sh = gc.open_by_key(BOVIET_SHEET_ID)
            ws = sh.worksheet(boviet_tab)
            max_col = max(v for v in cfg.values() if isinstance(v, int)) + 1
            row = [""] * max_col
            row[cfg["efj_col"]] = efj
            row[cfg["load_id_col"]] = container or efj
            row[cfg["status_col"]] = status.replace("_", " ").title() if status else ""
            if cfg.get("pickup_col") is not None:
                row[cfg["pickup_col"]] = pickup
            if cfg.get("delivery_col") is not None:
                row[cfg["delivery_col"]] = delivery
            if cfg.get("phone_col") is not None:
                row[cfg["phone_col"]] = driver_phone
            if cfg.get("trailer_col") is not None:
                row[cfg["trailer_col"]] = trailer
            ws.append_row(row, value_input_option="USER_ENTERED")
            log.info("Added load %s to Boviet/%s", efj, boviet_tab)
            result_tab = f"Boviet {boviet_tab}"

        # Invalidate cache so next fetch picks up the new row
        sheet_cache._last = 0

        # Store FTL driver info in DB if provided
        if move_type == "FTL" and (driver_phone or carrier_email):
            try:
                # Build notes with trailer/MP URL info
                extra_notes = []
                if trailer:
                    extra_notes.append(f"Trailer: {trailer}")
                if mp_url:
                    extra_notes.append(f"MP: {mp_url}")
                notes_str = " | ".join(extra_notes) or None
                with db.get_conn() as conn:
                    with db.get_cursor(conn) as cur:
                        cur.execute(
                            """INSERT INTO driver_contacts (efj, driver_phone, driver_email, notes, updated_at)
                               VALUES (%s, %s, %s, %s, NOW())
                               ON CONFLICT (efj) DO UPDATE SET
                                 driver_phone = COALESCE(EXCLUDED.driver_phone, driver_contacts.driver_phone),
                                 driver_email = COALESCE(EXCLUDED.driver_email, driver_contacts.driver_email),
                                 notes        = COALESCE(EXCLUDED.notes, driver_contacts.notes),
                                 updated_at = NOW()""",
                            (efj, driver_phone or None, carrier_email or None, notes_str),
                        )
            except Exception as db_err:
                log.warning("Driver contact save failed for %s: %s", efj, db_err)

        return {"ok": True, "efj": efj, "tab": result_tab, "sheet": target_sheet}

    except HTTPException:
        raise
    except Exception as e:
        log.error("Failed to add load %s: %s", efj, e)
        raise HTTPException(500, f"Failed to add load: {e}")


# ── Email History & Unmatched Inbox Endpoints ──

@app.get("/api/load/{efj}/emails")
async def get_load_emails(efj: str):
    """Return indexed emails for a load, newest first."""
    with db.get_cursor() as cur:
        cur.execute(
        """SELECT id, gmail_message_id, gmail_thread_id, subject, sender,
                  recipients, body_preview, has_attachments, attachment_names,
                  sent_at, indexed_at, email_type, lane, priority, ai_summary
           FROM email_threads
           WHERE efj = %s
           ORDER BY sent_at DESC NULLS LAST""",
        (efj,)
        )
        rows = cur.fetchall()
    emails = [
        {
            "id": r["id"],
            "gmail_message_id": r["gmail_message_id"],
            "subject": r["subject"],
            "sender": r["sender"],
            "recipients": r["recipients"],
            "body_preview": r["body_preview"],
            "has_attachments": r["has_attachments"],
            "attachment_names": r["attachment_names"],
            "sent_at": r["sent_at"].isoformat() if r["sent_at"] else None,
            "indexed_at": r["indexed_at"].isoformat() if r["indexed_at"] else None,
            "email_type": r.get("email_type"),
            "lane": r.get("lane"),
            "priority": r.get("priority"),
            "ai_summary": r.get("ai_summary"),
        }
        for r in rows
    ]
    return JSONResponse({"emails": emails, "count": len(emails)})




@app.post("/api/load/{efj}/summary")
async def api_load_summary(efj: str, request: Request):
    """Generate an AI-powered operational summary for a load using Claude."""
    if not config.ANTHROPIC_API_KEY:
        raise HTTPException(422, "ANTHROPIC_API_KEY not configured")

    body = await request.json()
    shipment = body.get("shipment", {})
    emails = body.get("emails", [])
    documents = body.get("documents", [])
    driver = body.get("driver", {})
    tracking = body.get("tracking")
    today = datetime.now().strftime("%Y-%m-%d")

    lines = [
        f"Load: {shipment.get('efj', efj)}",
        f"Move Type: {shipment.get('moveType', 'Unknown')}",
        f"Account: {shipment.get('account', 'Unknown')}",
        f"Status: {shipment.get('rawStatus', shipment.get('status', 'Unknown'))}",
        f"Container/Load#: {shipment.get('container', 'N/A')}",
        f"Carrier: {shipment.get('carrier', 'N/A')}",
        f"Origin: {shipment.get('origin', 'N/A')} -> Destination: {shipment.get('destination', 'N/A')}",
        f"ETA: {shipment.get('eta', 'N/A')}",
        f"LFD/Cutoff: {shipment.get('lfd', 'N/A')}",
        f"Pickup: {shipment.get('pickupDate', 'N/A')}",
        f"Delivery: {shipment.get('deliveryDate', 'N/A')}",
        f"BOL: {shipment.get('bol', 'N/A')}",
        f"SSL/Vessel: {shipment.get('ssl', 'N/A')}",
        f"Return Port: {shipment.get('returnPort', 'N/A')}",
        f"Notes: {shipment.get('notes', 'None')}",
        f"Bot Alert: {shipment.get('botAlert', 'None')}",
        f"Rep: {shipment.get('rep', 'N/A')}",
    ]
    if shipment.get('hub'):
        lines.append(f"Hub: {shipment['hub']}")
    if shipment.get('project'):
        lines.append(f"Project: {shipment['project']}")

    if any(driver.get(k) for k in ("driverName", "driverPhone", "driverEmail", "trailerNumber")):
        lines.append("")
        lines.append("--- Driver/Carrier Contact ---")
        if driver.get("driverName"):
            lines.append(f"Driver: {driver['driverName']}")
        if driver.get("driverPhone"):
            lines.append(f"Phone: {driver['driverPhone']}")
        if driver.get("driverEmail"):
            lines.append(f"Email: {driver['driverEmail']}")
        if driver.get("carrierEmail"):
            lines.append(f"Carrier Email: {driver['carrierEmail']}")
        if driver.get("trailerNumber"):
            lines.append(f"Trailer: {driver['trailerNumber']}")

    if tracking:
        lines.append("")
        lines.append("--- Tracking Status ---")
        lines.append(f"Tracking Status: {tracking.get('trackingStatus', 'N/A')}")
        if tracking.get('eta'):
            lines.append(f"Tracking ETA: {tracking['eta']}")
        if tracking.get('behindSchedule'):
            lines.append("WARNING: Behind Schedule")
        if tracking.get('cantMakeIt'):
            lines.append(f"CRITICAL: {tracking['cantMakeIt']}")

    lines.append("")
    lines.append("--- Documents on File ---")
    if documents:
        doc_types = {}
        for d in documents:
            dt = d.get("doc_type", "other")
            doc_types.setdefault(dt, []).append(d.get("original_name", "unknown"))
        for dt, names in doc_types.items():
            lines.append(f"  {dt}: {len(names)} file(s) - {', '.join(names[:3])}")
    else:
        lines.append("  No documents uploaded")

    doc_type_set = {d.get("doc_type") for d in documents}
    missing_docs = []
    if "bol" not in doc_type_set:
        missing_docs.append("BOL")
    if "pod" not in doc_type_set:
        missing_docs.append("POD")
    if "customer_rate" not in doc_type_set:
        missing_docs.append("Customer Rate Con")
    if "carrier_rate" not in doc_type_set:
        missing_docs.append("Carrier Rate Con")
    if missing_docs:
        lines.append(f"  MISSING: {', '.join(missing_docs)}")

    lines.append("")
    lines.append("--- Recent Email Activity ---")
    if emails:
        lines.append(f"Total emails: {len(emails)}")
        for e in emails[:5]:
            sent = e.get('sent_at', '')[:10] if e.get('sent_at') else 'N/A'
            lines.append(f"  [{sent}] From: {e.get('sender', 'Unknown')}")
            lines.append(f"    Subject: {e.get('subject', 'No subject')}")
            if e.get('body_preview'):
                lines.append(f"    Preview: {e['body_preview'][:120]}")
    else:
        lines.append("  No emails indexed for this load")

    context_str = "\n".join(lines)

    system_prompt = (
        "You are a logistics operations assistant for Evans Delivery (EFJ Operations). "
        "You produce concise, actionable load summaries for dispatchers.\n\n"
        "Rules:\n"
        "- Output exactly 3-5 bullet points using the bullet character\n"
        "- Each bullet should be one sentence, max 20 words\n"
        "- First bullet: Current status and location context\n"
        "- Flag any issues: behind schedule, missing documents, approaching LFD, no driver, no tracking\n"
        "- Note document completeness (what is present vs missing)\n"
        "- Summarize recent email activity if any\n"
        "- If everything looks good, say so\n"
        "- Today is: " + today + "\n"
        "- Use plain text only, no markdown, no bold, no headers\n"
        "- Be direct and operational for experienced dispatchers"
    )

    try:
        import anthropic
        client = anthropic.Anthropic(api_key=config.ANTHROPIC_API_KEY)
        message = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=300,
            system=system_prompt,
            messages=[{"role": "user", "content": f"Generate an operational summary for this load:\n\n{context_str}"}],
        )
        summary_text = message.content[0].text.strip()
        return JSONResponse({"summary": summary_text})
    except Exception as e:
        log.error("AI summary generation failed for %s: %s", efj, e)
        raise HTTPException(500, f"Summary generation failed: {str(e)}")


@app.get("/api/unmatched-emails")
async def get_unmatched_emails():
    """List unmatched inbox emails pending review."""
    with db.get_cursor() as cur:
        cur.execute(
        """SELECT id, gmail_message_id, subject, sender, recipients,
                  body_preview, has_attachments, attachment_names,
                  sent_at, indexed_at, review_status,
                  email_type, lane, priority, ai_summary, suggested_rep
           FROM unmatched_inbox_emails
           WHERE review_status = 'pending'
           ORDER BY COALESCE(priority, 0) DESC, sent_at DESC NULLS LAST
           LIMIT 100""",
        )
        rows = cur.fetchall()
    emails = [
        {
            "id": r["id"],
            "subject": r["subject"],
            "sender": r["sender"],
            "body_preview": r["body_preview"],
            "has_attachments": r["has_attachments"],
            "attachment_names": r["attachment_names"],
            "sent_at": r["sent_at"].isoformat() if r["sent_at"] else None,
            "review_status": r["review_status"],
            "email_type": r.get("email_type"),
            "lane": r.get("lane"),
            "priority": r.get("priority"),
            "ai_summary": r.get("ai_summary"),
            "suggested_rep": r.get("suggested_rep"),
        }
        for r in rows
    ]
    return JSONResponse({"emails": emails, "count": len(emails)})


@app.post("/api/unmatched-emails/{email_id}/assign")
async def assign_unmatched_email(email_id: int, request: Request):
    """Assign an unmatched email to a load by EFJ#."""
    body = await request.json()
    efj = body.get("efj", "").strip().upper()
    if not efj:
        raise HTTPException(400, "efj is required")

    with db.get_conn() as conn:
        with db.get_cursor(conn) as cur:
            # Get the unmatched email
            cur.execute(
                "SELECT * FROM unmatched_inbox_emails WHERE id = %s AND review_status = 'pending'",
                (email_id,)
            )
            row = cur.fetchone()
            if not row:
                raise HTTPException(404, "Email not found or already processed")

            # Move to email_threads
            cur.execute(
                """INSERT INTO email_threads
                   (efj, gmail_thread_id, gmail_message_id, subject, sender,
                    recipients, body_preview, has_attachments, attachment_names, sent_at)
                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                   ON CONFLICT (gmail_message_id) DO NOTHING""",
                (efj, row["gmail_thread_id"], row["gmail_message_id"],
                 row["subject"], row["sender"], row["recipients"],
                 row["body_preview"], row["has_attachments"],
                 row["attachment_names"], row["sent_at"]),
            )

            # Mark as assigned
            cur.execute(
                "UPDATE unmatched_inbox_emails SET review_status = 'assigned', assigned_efj = %s WHERE id = %s",
                (efj, email_id),
            )

    log.info("Unmatched email %d assigned to %s", email_id, efj)
    return JSONResponse({"status": "ok", "efj": efj})


@app.post("/api/unmatched-emails/{email_id}/dismiss")
async def dismiss_unmatched_email(email_id: int):
    """Dismiss an unmatched email (mark as not relevant)."""
    with db.get_conn() as conn:
        with db.get_cursor(conn) as cur:
            cur.execute(
                "UPDATE unmatched_inbox_emails SET review_status = 'dismissed' WHERE id = %s",
                (email_id,),
            )
    return JSONResponse({"status": "ok"})


# ===================================================================
# RATE IQ — Quote Builder API
# ===================================================================

# ── Settings (MUST be before /api/quotes/{id} to avoid FastAPI matching "settings" as an id) ──

@app.get("/api/quotes/settings")
async def api_quote_settings():
    """Return default quote builder settings."""
    return JSONResponse({
        "default_margin_pct": 15,
        "default_terms": [
            "Rates valid for 7 days from quote date",
            "Subject to carrier availability at time of booking",
            "Accessorial charges may vary based on actual services required",
            "Payment terms: Net 30 days from invoice date",
        ],
        "default_accessorials": [
            {"charge": "Storage", "rate": "45.00", "frequency": "per day", "checked": False, "amount": "45.00"},
            {"charge": "Pre-Pull", "rate": "150.00", "frequency": "flat", "checked": False, "amount": "150.00"},
            {"charge": "Chassis (2 days)", "rate": "45.00", "frequency": "per day", "checked": False, "amount": "45.00"},
            {"charge": "Overweight", "rate": "150.00", "frequency": "flat", "checked": False, "amount": "150.00"},
            {"charge": "Detention", "rate": "85.00", "frequency": "per hour", "checked": False, "amount": "85.00"},
        ],
    })


# ── Distance lookup (geocode via Nominatim, route via OSRM) ──

@app.get("/api/quotes/distance")
async def api_quote_distance(origin: str = Query(...), destination: str = Query(...)):
    """Calculate mileage and transit time between origin and destination."""
    import requests as _req

    def _geocode(place: str):
        r = _req.get(
            "https://nominatim.openstreetmap.org/search",
            params={"q": place, "format": "json", "limit": 1, "countrycodes": "us"},
            headers={"User-Agent": "CSLogix-Dashboard/1.0"},
            timeout=10,
        )
        r.raise_for_status()
        results = r.json()
        if not results:
            return None
        return float(results[0]["lon"]), float(results[0]["lat"])

    try:
        orig = _geocode(origin)
        dest = _geocode(destination)
        if not orig or not dest:
            return JSONResponse({"error": "Could not geocode one or both locations"}, status_code=400)

        # OSRM route
        coords = f"{orig[0]},{orig[1]};{dest[0]},{dest[1]}"
        r = _req.get(
            f"https://router.project-osrm.org/route/v1/driving/{coords}",
            params={"overview": "false"},
            headers={"User-Agent": "CSLogix-Dashboard/1.0"},
            timeout=10,
        )
        r.raise_for_status()
        data = r.json()
        if data.get("code") != "Ok" or not data.get("routes"):
            return JSONResponse({"error": "No route found"}, status_code=400)

        route = data["routes"][0]
        dist_meters = route["distance"]
        dur_seconds = route["duration"]

        one_way_miles = round(dist_meters / 1609.34)
        round_trip_miles = one_way_miles * 2
        duration_hours = round(dur_seconds / 3600, 2)

        # Transit time string: for short distances show hours, for long show days
        if duration_hours < 8:
            transit_time = f"{round(duration_hours)} hours"
        else:
            days_low = max(1, int(duration_hours / 10))  # ~10hr driving days
            days_high = days_low + 1
            transit_time = f"{days_low}-{days_high} days"

        return JSONResponse({
            "one_way_miles": one_way_miles,
            "round_trip_miles": round_trip_miles,
            "duration_hours": duration_hours,
            "transit_time": transit_time,
        })

    except _req.exceptions.Timeout:
        return JSONResponse({"error": "Geocoding service timeout"}, status_code=504)
    except Exception as e:
        log.warning("Distance lookup failed: %s", e)
        return JSONResponse({"error": str(e)}, status_code=500)


# ── Extract rates from file or text ──

def _parse_rate_text(text: str) -> dict:
    """Parse carrier rate info from plain text (email body, rate confirmation).
    Returns structured fields that the frontend can populate."""
    import re as _re
    result = {}

    # Try to find carrier name (first line or "Carrier: X" pattern)
    carrier_m = _re.search(r'(?:carrier|trucking|transport|logistics|freight)\s*[:\-]?\s*(.+)', text, _re.IGNORECASE)
    if carrier_m:
        result["carrier_name"] = carrier_m.group(1).strip()[:120]

    # MC number — usually in email signature (MC#123456, MC-123456, MC 123456, MC:123456)
    mc_m = _re.search(r'MC\s*[#:\-]?\s*(\d{4,7})', text, _re.IGNORECASE)
    if mc_m:
        result["carrier_mc"] = mc_m.group(1)

    # DOT number — often near MC#
    dot_m = _re.search(r'(?:DOT|USDOT)\s*[#:\-]?\s*(\d{4,8})', text, _re.IGNORECASE)
    if dot_m:
        result["carrier_dot"] = dot_m.group(1)

    # Origin / destination patterns
    orig_m = _re.search(r'(?:origin|pick\s*up|from|shipper)\s*[:\-]\s*(.+)', text, _re.IGNORECASE)
    if orig_m:
        result["origin"] = orig_m.group(1).strip().split("\n")[0][:200]
    dest_m = _re.search(r'(?:destination|deliver(?:y)?|to|consignee)\s*[:\-]\s*(.+)', text, _re.IGNORECASE)
    if dest_m:
        result["destination"] = dest_m.group(1).strip().split("\n")[0][:200]

    # Shipment type
    text_lower = text.lower()
    if "dray" in text_lower or "container" in text_lower or "chassis" in text_lower:
        result["shipment_type"] = "Dray"
    elif "ftl" in text_lower or "full truckload" in text_lower or "53" in text_lower:
        result["shipment_type"] = "FTL"
    elif "ltl" in text_lower:
        result["shipment_type"] = "LTL"
    elif "transload" in text_lower:
        result["shipment_type"] = "Transload"

    # Mileage
    miles_m = _re.search(r'(\d[\d,]*)\s*(?:miles|mi\b)', text, _re.IGNORECASE)
    if miles_m:
        result["one_way_miles"] = miles_m.group(1).replace(",", "")

    # Dollar amounts — collect all as potential linehaul items
    dollar_matches = _re.findall(
        r'(.{0,60}?)\$\s*([\d,]+(?:\.\d{2})?)', text
    )
    linehaul_items = []
    for context, amount in dollar_matches:
        # Clean up context to use as description
        desc = context.strip().rstrip(":-–—").strip()
        # Remove leading junk
        desc = _re.sub(r'^.*?(?:rate|charge|fee|cost|price|total)\s*[:\-]?\s*', '', desc, flags=_re.IGNORECASE).strip()
        if not desc:
            # Try to label from nearby keywords
            ctx_lower = context.lower()
            if "line" in ctx_lower or "haul" in ctx_lower:
                desc = "Linehaul"
            elif "fuel" in ctx_lower or "fsc" in ctx_lower:
                desc = "Fuel Surcharge"
            elif "stop" in ctx_lower:
                desc = "Stop Charge"
            else:
                desc = "Linehaul"
        linehaul_items.append({
            "description": desc[:80],
            "rate": amount.replace(",", ""),
        })

    if linehaul_items:
        result["linehaul_items"] = linehaul_items

    # Accessorials — look for common terms
    acc_patterns = [
        (r'storage\s*[:\-$]*\s*\$?([\d,.]+)', "Storage"),
        (r'pre[\-\s]?pull\s*[:\-$]*\s*\$?([\d,.]+)', "Pre-Pull"),
        (r'chassis\s*[:\-$]*\s*\$?([\d,.]+)', "Chassis (2 days)"),
        (r'over\s*weight\s*[:\-$]*\s*\$?([\d,.]+)', "Overweight"),
        (r'detention\s*[:\-$]*\s*\$?([\d,.]+)', "Detention"),
        (r'demurrage\s*[:\-$]*\s*\$?([\d,.]+)', "Demurrage"),
        (r'layover\s*[:\-$]*\s*\$?([\d,.]+)', "Layover"),
    ]
    accessorials = []
    for pat, name in acc_patterns:
        m = _re.search(pat, text, _re.IGNORECASE)
        if m:
            accessorials.append({
                "charge": name,
                "rate": m.group(1).replace(",", ""),
                "frequency": "flat",
                "amount": m.group(1).replace(",", ""),
            })
    if accessorials:
        result["accessorials"] = accessorials

    return result


_EXTRACT_PROMPT = """Extract rate/quote information from this carrier rate confirmation or email.
Look carefully for the carrier's MC# and DOT# — these are often in the email signature block at the bottom.
Return ONLY valid JSON (no markdown, no explanation) with these fields:
{
  "carrier_name": "string or null",
  "carrier_mc": "MC number (digits only) or null — look in email signature for MC#, MC:, MC-, etc",
  "carrier_dot": "DOT/USDOT number (digits only) or null — look in email signature",
  "origin": "city, state or full address or null",
  "destination": "city, state or full address or null",
  "shipment_type": "Dray|FTL|LTL|Transload|OTR or null",
  "round_trip_miles": "string or null",
  "one_way_miles": "string or null",
  "transit_time": "string or null",
  "linehaul_items": [{"description": "string", "rate": "numeric string", "section": "Charges|OTR|Dray/Transload|Transload|LTL"}],
  "accessorials": [{"charge": "name", "rate": "numeric string", "frequency": "flat|per day|per hour", "amount": "numeric string"}]
}
Only include fields you can confidently extract. For linehaul_items, list each charge line separately (linehaul, fuel surcharge, stop charges, etc). Omit null fields."""


def _extract_with_claude(content: list) -> dict:
    """Call Claude API with content blocks (text or image), return parsed rate data."""
    import anthropic
    client = anthropic.Anthropic(api_key=config.ANTHROPIC_API_KEY)
    message = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=1024,
        messages=[{"role": "user", "content": content}],
    )
    # Parse the JSON from Claude's response
    response_text = message.content[0].text.strip()
    # Strip markdown code fences if present
    if response_text.startswith("```"):
        response_text = response_text.split("\n", 1)[-1].rsplit("```", 1)[0].strip()
    return json.loads(response_text)


@app.post("/api/quotes/extract")
async def api_quote_extract(request: Request):
    """Extract rate info from uploaded file (image/PDF) or pasted text.
    Uses Claude Vision when ANTHROPIC_API_KEY is configured, falls back to regex."""
    content_type = request.headers.get("content-type", "")
    has_claude = bool(config.ANTHROPIC_API_KEY)

    if "multipart/form-data" not in content_type:
        raise HTTPException(400, "No file or text provided")

    form = await request.form()
    text = form.get("text")
    file = form.get("file")

    # ── Text extraction ──
    if text:
        text_str = str(text)
        if has_claude:
            try:
                content = [{"type": "text", "text": _EXTRACT_PROMPT + "\n\n" + text_str}]
                result = _extract_with_claude(content)
                return JSONResponse(result)
            except Exception as e:
                log.warning("Claude text extraction failed, falling back to regex: %s", e)
        result = _parse_rate_text(text_str)
        if not result:
            raise HTTPException(400, "Could not extract any rate information from the text")
        return JSONResponse(result)

    # ── File extraction ──
    if file:
        file_bytes = await file.read()
        filename = getattr(file, "filename", "") or ""
        ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""

        # Text-based files
        if ext in ("txt", "csv", "eml"):
            try:
                text_content = file_bytes.decode("utf-8", errors="replace")
            except Exception:
                text_content = file_bytes.decode("latin-1", errors="replace")
            if has_claude:
                try:
                    content = [{"type": "text", "text": _EXTRACT_PROMPT + "\n\n" + text_content}]
                    result = _extract_with_claude(content)
                    return JSONResponse(result)
                except Exception as e:
                    log.warning("Claude text extraction failed, falling back to regex: %s", e)
            result = _parse_rate_text(text_content)
            if not result:
                raise HTTPException(400, "Could not extract rate info from file")
            return JSONResponse(result)

        # Images — send to Claude Vision
        if ext in ("png", "jpg", "jpeg", "gif", "webp"):
            if not has_claude:
                raise HTTPException(422, "Image extraction requires ANTHROPIC_API_KEY to be configured")
            import base64
            media_map = {"png": "image/png", "jpg": "image/jpeg", "jpeg": "image/jpeg",
                         "gif": "image/gif", "webp": "image/webp"}
            media_type = media_map.get(ext, "image/png")
            try:
                content = [
                    {"type": "image", "source": {
                        "type": "base64",
                        "media_type": media_type,
                        "data": base64.b64encode(file_bytes).decode(),
                    }},
                    {"type": "text", "text": _EXTRACT_PROMPT},
                ]
                result = _extract_with_claude(content)
                return JSONResponse(result)
            except Exception as e:
                log.error("Claude Vision extraction failed: %s", e)
                raise HTTPException(500, f"AI extraction failed: {e}")

        # PDFs — send to Claude as document
        if ext == "pdf":
            if not has_claude:
                raise HTTPException(422, "PDF extraction requires ANTHROPIC_API_KEY to be configured")
            import base64
            try:
                content = [
                    {"type": "document", "source": {
                        "type": "base64",
                        "media_type": "application/pdf",
                        "data": base64.b64encode(file_bytes).decode(),
                    }},
                    {"type": "text", "text": _EXTRACT_PROMPT},
                ]
                result = _extract_with_claude(content)
                return JSONResponse(result)
            except Exception as e:
                log.error("Claude PDF extraction failed: %s", e)
                raise HTTPException(500, f"AI extraction failed: {e}")


    # .msg (Outlook) - use extract-msg for proper parsing
    if ext in ('msg',):
        try:
            import extract_msg, io as _io
            msg = extract_msg.openMsg(_io.BytesIO(file_bytes))
            parts = []
            if msg.subject:
                parts.append('Subject: ' + str(msg.subject))
            if msg.sender:
                parts.append('From: ' + str(msg.sender))
            if msg.date:
                parts.append('Date: ' + str(msg.date))
            if msg.body:
                parts.append(msg.body)
            text_content = chr(10).join(parts)
            # Extract image attachments for Claude Vision
            attachment_images = []
            for att in (msg.attachments or []):
                att_name = (att.longFilename or att.shortFilename or '').lower()
                if att_name.endswith(('.png', '.jpg', '.jpeg', '.gif', '.webp')):
                    att_data = att.data
                    if att_data:
                        import base64 as _b64
                        ext2 = att_name.rsplit('.', 1)[-1]
                        mt = dict(png='image/png', jpg='image/jpeg', jpeg='image/jpeg', gif='image/gif', webp='image/webp').get(ext2, 'image/png')
                        attachment_images.append(dict(type='image', source=dict(type='base64', media_type=mt, data=_b64.b64encode(att_data).decode())))
                elif att_name.endswith('.pdf') and att.data:
                    try:
                        import fitz
                        pdf_doc = fitz.open(stream=att.data, filetype='pdf')
                        pdf_text = chr(10).join(page.get_text() for page in pdf_doc)
                        if pdf_text.strip():
                            text_content += chr(10)*2 + '--- Attached PDF: ' + att_name + ' ---' + chr(10) + pdf_text[:4000]
                    except Exception:
                        pass
            if not text_content.strip():
                raise HTTPException(400, 'Could not extract readable text from .msg file')
            if has_claude:
                try:
                    content_msg = [dict(type='text', text=_EXTRACT_PROMPT + chr(10)*2 + text_content[:8000])]
                    content_msg = attachment_images + content_msg
                    result = _extract_with_claude(content_msg)
                    return JSONResponse(result)
                except Exception as e:
                    log.warning('Claude .msg extraction failed: %s', e)
            result = _parse_rate_text(text_content)
            if not result:
                raise HTTPException(400, 'Could not extract rate info from .msg file')
            return JSONResponse(result)
        except HTTPException:
            raise
        except Exception as e:
            log.error('.msg extraction failed: %s', e)
            # Fallback: brute-force ASCII extraction
            try:
                raw = file_bytes.decode('latin-1', errors='replace')
                blocks = re.findall(r'[ -~]{20,}', raw)
                fallback_text = chr(10).join(blocks)
                if fallback_text.strip() and has_claude:
                    content_msg = [dict(type='text', text=_EXTRACT_PROMPT + chr(10)*2 + fallback_text[:8000])]
                    result = _extract_with_claude(content_msg)
                    return JSONResponse(result)
            except Exception:
                pass
            raise HTTPException(500, f'Failed to process .msg file: {e}')

    # .htm/.html files
    if ext in ('htm', 'html'):
        try:
            raw = file_bytes.decode('utf-8', errors='replace')
            text_content = re.sub(r'<[^>]+>', ' ', raw)
            text_content = re.sub(chr(92) + 's+', ' ', text_content).strip()
            if not text_content:
                raise HTTPException(400, 'Could not extract text from HTML file')
            if has_claude:
                try:
                    content_msg = [dict(type='text', text=_EXTRACT_PROMPT + chr(10)*2 + text_content[:8000])]
                    result = _extract_with_claude(content_msg)
                    return JSONResponse(result)
                except Exception as e:
                    log.warning('Claude HTML extraction failed: %s', e)
            result = _parse_rate_text(text_content)
            if not result:
                raise HTTPException(400, 'Could not extract rate info from HTML')
            return JSONResponse(result)
        except HTTPException:
            raise
        except Exception as e:
            log.error('HTML extraction failed: %s', e)
            raise HTTPException(500, f'Failed to process HTML file: {e}')

        raise HTTPException(400, f"Unsupported file type: .{ext}")

    raise HTTPException(400, "No file or text provided")


# ── CRUD (list, create, get, update) — register AFTER /settings and /distance ──

def _index_quote_to_rate_iq(row: dict, quote_id: int = None):
    """
    Write carrier cost from a saved quote into rate_quotes for Rate IQ lane intelligence.
    Only indexes dray-type quotes with a valid carrier_total and origin+dest.
    """
    DRAY_TYPES = {"Dray", "Dray+Transload", "OTR", "Transload"}
    shipment_type = (row.get("shipment_type") or "").strip()
    if shipment_type not in DRAY_TYPES:
        return
    origin = (row.get("pod") or "").strip()
    destination = (row.get("final_delivery") or "").strip()
    carrier_name = (row.get("carrier_name") or "").strip()
    try:
        carrier_total = float(row.get("carrier_total") or 0)
    except (TypeError, ValueError):
        carrier_total = None
    if not origin or not destination or not carrier_total:
        return
    lane = f"{origin} → {destination}"
    quote_date = row.get("created_at") or row.get("updated_at")
    qid = quote_id or row.get("id")
    try:
        with db.get_cursor() as cur:
            if qid:
                # Update existing rate_quotes entry for this quote
                cur.execute(
                    "UPDATE rate_quotes SET origin=%s, destination=%s, lane=%s, "
                    "carrier_name=%s, rate_amount=%s, move_type=%s, quote_date=%s, status='quoted' "
                    "WHERE source_quote_id=%s",
                    (origin, destination, lane, carrier_name, carrier_total,
                     shipment_type.lower(), quote_date, qid)
                )
                if cur.rowcount == 0:
                    # No existing row — insert
                    cur.execute(
                        "INSERT INTO rate_quotes "
                        "(origin, destination, lane, carrier_name, rate_amount, rate_unit, "
                        " move_type, quote_date, status, source_quote_id) "
                        "VALUES (%s,%s,%s,%s,%s,'flat',%s,%s,'quoted',%s) "
                        "ON CONFLICT DO NOTHING",
                        (origin, destination, lane, carrier_name, carrier_total,
                         shipment_type.lower(), quote_date, qid)
                    )
            else:
                cur.execute(
                    "INSERT INTO rate_quotes "
                    "(origin, destination, lane, carrier_name, rate_amount, rate_unit, "
                    " move_type, quote_date, status) "
                    "VALUES (%s,%s,%s,%s,%s,'flat',%s,%s,'quoted') "
                    "ON CONFLICT DO NOTHING",
                    (origin, destination, lane, carrier_name, carrier_total,
                     shipment_type.lower(), quote_date)
                )
    except Exception as e:
        import logging
        logging.getLogger("csl-dashboard").warning("rate_quotes index failed: %s", e)


@app.get("/api/quotes")
async def api_list_quotes(
    status: str = Query(default=None),
    search: str = Query(default=None),
    move_types: str = Query(default=None),
    limit: int = Query(default=50),
    offset: int = Query(default=0),
):
    """List quotes with optional filters. move_types is comma-separated shipment types."""
    mt_list = [t.strip() for t in move_types.split(",")] if move_types else None
    rows = db.list_quotes(status=status, search=search, move_types=mt_list, limit=limit, offset=offset)
    rows = [_sanitize_row(r) for r in rows]
    return JSONResponse({"quotes": rows})



def _sanitize_row(row: dict) -> dict:
    """Convert Decimal->float, datetime->str for JSON serialization."""
    from decimal import Decimal as _Dec
    out = {}
    for k, v in row.items():
        if isinstance(v, _Dec):
            out[k] = float(v)
        elif hasattr(v, "isoformat"):
            out[k] = v.isoformat()
        else:
            out[k] = v
    return out

@app.post("/api/quotes")
async def api_create_quote(request: Request):
    """Create a new quote."""
    content_type = request.headers.get("content-type", "")
    if "multipart/form-data" in content_type:
        form = await request.form()
        raw = form.get("quote_data")
        if not raw:
            raise HTTPException(400, "Missing quote_data")
        data = json.loads(raw)
    else:
        data = await request.json()

    row = db.insert_quote(data)
    # Index carrier cost to rate_quotes for Rate IQ lane intelligence
    _index_quote_to_rate_iq(row)
    return JSONResponse(_sanitize_row(row), status_code=201)


@app.get("/api/quotes/{quote_id}")
async def api_get_quote(quote_id: int):
    """Get a single quote by ID."""
    row = db.get_quote(quote_id)
    if not row:
        raise HTTPException(404, "Quote not found")
    return JSONResponse(_sanitize_row(row))



@app.patch("/api/quotes/{quote_id}/status")
async def update_quote_status(quote_id: int, request: Request):
    """Quick status update for a quote (won/lost/expired/sent)"""
    body = await request.json()
    new_status = body.get("status")
    if new_status not in ("draft", "sent", "accepted", "lost", "expired"):
        return JSONResponse({"error": "Invalid status"}, status_code=400)
    conn = db.get_conn()
    try:
        cur = conn.cursor()
        cur.execute(
            "UPDATE quotes SET status=%s, updated_at=NOW() WHERE id=%s RETURNING id, quote_number, status",
            (new_status, quote_id)
        )
        row = cur.fetchone()
        conn.commit()
        if not row:
            return JSONResponse({"error": "Quote not found"}, status_code=404)
        return {"id": row[0], "quote_number": row[1], "status": row[2]}
    except Exception as e:
        conn.rollback()
        return JSONResponse({"error": str(e)}, status_code=500)
    finally:
        db.put_conn(conn)


@app.put("/api/quotes/{quote_id}")
async def api_update_quote(quote_id: int, request: Request):
    """Update an existing quote."""
    content_type = request.headers.get("content-type", "")
    if "multipart/form-data" in content_type:
        form = await request.form()
        raw = form.get("quote_data")
        if not raw:
            raise HTTPException(400, "Missing quote_data")
        data = json.loads(raw)
    else:
        data = await request.json()

    row = db.update_quote(quote_id, data)
    if not row:
        raise HTTPException(404, "Quote not found")
    # Re-index carrier cost to rate_quotes for Rate IQ
    _index_quote_to_rate_iq(row, quote_id=quote_id)
    return JSONResponse(_sanitize_row(row))


# ═══════════════════════════════════════════════════════════
# CARRIER DIRECTORY API
# ═══════════════════════════════════════════════════════════

def _serialize_row(row):
    """Convert a DB row dict to JSON-safe dict."""
    d = dict(row)
    for k in ("created_at", "updated_at", "effective_date", "quote_date", "indexed_at", "date_quoted"):
        if d.get(k) and hasattr(d[k], "isoformat"):
            d[k] = d[k].isoformat()
    # Decimal -> float
    for k, v in d.items():
        if hasattr(v, "as_tuple"):
            d[k] = float(v)
    return d


@app.get("/api/carriers")
async def api_list_carriers(search: str = Query(default=None), region: str = Query(default=None)):
    with db.get_cursor() as cur:
        clauses, params = [], []
        if search:
            clauses.append("(carrier_name ILIKE %s OR mc_number ILIKE %s OR contact_email ILIKE %s)")
            s = f"%{search}%"
            params.extend([s, s, s])
        if region:
            clauses.append("regions ILIKE %s")
            params.append(f"%{region}%")
        where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
        cur.execute(f"SELECT * FROM carriers {where} ORDER BY carrier_name ASC", params)
        rows = [_serialize_row(r) for r in cur.fetchall()]
    return JSONResponse({"carriers": rows})


@app.post("/api/carriers")
async def api_create_carrier(request: Request):
    body = await request.json()
    fields = ["carrier_name", "mc_number", "dot_number", "contact_email", "contact_phone",
              "contact_name", "regions", "ports", "rail_ramps", "equipment_types", "notes", "source", "pickup_area", "destination_area", "date_quoted", "v_code",
              "can_dray", "can_hazmat", "can_overweight", "can_transload"]
    bool_fields = {"can_dray", "can_hazmat", "can_overweight", "can_transload"}
    vals = []
    for f in fields:
        if f in bool_fields:
            vals.append(bool(body.get(f, False)))
        elif f == "carrier_name":
            vals.append(body.get(f, ""))
        else:
            vals.append(body.get(f, None))
    if not vals[0]:
        vals[0] = "Unknown Carrier"
    with db.get_conn() as conn:
        with db.get_cursor(conn) as cur:
            cur.execute(f"""
                INSERT INTO carriers ({', '.join(fields)})
                VALUES ({', '.join(['%s'] * len(fields))}) RETURNING *
            """, vals)
            row = _serialize_row(cur.fetchone())
    return JSONResponse(row)


@app.put("/api/carriers/{carrier_id}")
async def api_update_carrier(carrier_id: int, request: Request):
    body = await request.json()
    allowed = {"carrier_name", "mc_number", "dot_number", "contact_email", "contact_phone",
               "contact_name", "regions", "ports", "rail_ramps", "equipment_types", "notes", "pickup_area", "destination_area", "date_quoted", "v_code",
               "can_dray", "can_hazmat", "can_overweight", "can_transload"}
    bool_fields = {"can_dray", "can_hazmat", "can_overweight", "can_transload"}
    sets, params = [], []
    for k, v in body.items():
        if k in allowed:
            sets.append(f"{k} = %s")
            params.append(bool(v) if k in bool_fields else v)
    if not sets:
        raise HTTPException(400, "No valid fields")
    sets.append("updated_at = NOW()")
    params.append(carrier_id)
    with db.get_conn() as conn:
        with db.get_cursor(conn) as cur:
            cur.execute(f"UPDATE carriers SET {', '.join(sets)} WHERE id = %s RETURNING *", params)
            row = cur.fetchone()
    if not row:
        raise HTTPException(404, "Carrier not found")
    return JSONResponse(_serialize_row(row))


@app.delete("/api/carriers/{carrier_id}")
async def api_delete_carrier(carrier_id: int):
    with db.get_conn() as conn:
        with db.get_cursor(conn) as cur:
            cur.execute("DELETE FROM carriers WHERE id = %s", (carrier_id,))
    return JSONResponse({"ok": True})


_CARRIER_EXTRACT_PROMPT = """Extract carrier directory information from this document (rate sheet, carrier packet, or screenshot).
Return ONLY valid JSON (no markdown, no explanation) with these fields:
{
  "carrier_name": "string",
  "mc_number": "MC number digits only, or null",
  "dot_number": "DOT/USDOT number digits only, or null",
  "contact_email": "string or null",
  "contact_phone": "string or null",
  "contact_name": "string or null",
  "regions": "comma-separated service regions or null",
  "ports": "comma-separated ports served or null",
  "equipment_types": "comma-separated equipment types (Dry Van, Flatbed, Reefer, etc.) or null",
  "rates": [{"lane": "origin to destination", "rate": "dollar amount", "equipment": "type"}]
}
Extract ALL lanes/rates if this is a rate sheet. Omit null fields."""


@app.post("/api/carriers/extract")
async def api_carrier_extract(request: Request):
    """Extract carrier info from uploaded file via Claude Vision."""
    if not config.ANTHROPIC_API_KEY:
        raise HTTPException(422, "ANTHROPIC_API_KEY not configured")
    form = await request.form()
    file = form.get("file")
    if not file:
        raise HTTPException(400, "No file provided")
    file_bytes = await file.read()
    filename = getattr(file, "filename", "") or ""
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""

    import base64
    if ext in ("png", "jpg", "jpeg", "gif", "webp"):
        media_map = {"png": "image/png", "jpg": "image/jpeg", "jpeg": "image/jpeg",
                     "gif": "image/gif", "webp": "image/webp"}
        content = [
            {"type": "image", "source": {"type": "base64", "media_type": media_map.get(ext, "image/png"),
                                         "data": base64.b64encode(file_bytes).decode()}},
            {"type": "text", "text": _CARRIER_EXTRACT_PROMPT},
        ]
    elif ext == "pdf":
        content = [
            {"type": "document", "source": {"type": "base64", "media_type": "application/pdf",
                                            "data": base64.b64encode(file_bytes).decode()}},
            {"type": "text", "text": _CARRIER_EXTRACT_PROMPT},
        ]
    elif ext in ("txt", "csv", "eml"):
        text_content = file_bytes.decode("utf-8", errors="replace")
        content = [{"type": "text", "text": _CARRIER_EXTRACT_PROMPT + "\n\n" + text_content}]
    else:
        raise HTTPException(400, f"Unsupported file type: {ext}")

    try:
        result = _extract_with_claude(content)
        return JSONResponse(result)
    except Exception as e:
        log.error("Carrier extraction failed: %s", e)
        raise HTTPException(500, f"AI extraction failed: {e}")


# ═══════════════════════════════════════════════════════════
# WAREHOUSE DIRECTORY API
# ═══════════════════════════════════════════════════════════

@app.get("/api/warehouses")
async def api_list_warehouses(search: str = Query(default=None), region: str = Query(default=None),
                              state: str = Query(default=None)):
    with db.get_cursor() as cur:
        clauses, params = [], []
        if search:
            clauses.append("(name ILIKE %s OR mc_number ILIKE %s OR contact_email ILIKE %s)")
            s = f"%{search}%"
            params.extend([s, s, s])
        if region:
            clauses.append("region ILIKE %s")
            params.append(f"%{region}%")
        if state:
            clauses.append("state = %s")
            params.append(state.upper())
        where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
        cur.execute(f"SELECT * FROM warehouses {where} ORDER BY name ASC", params)
        warehouses = [_serialize_row(r) for r in cur.fetchall()]
        # Attach rate summaries
        wh_ids = [w["id"] for w in warehouses]
        if wh_ids:
            cur.execute("""SELECT warehouse_id, rate_type, rate_amount, unit, description
                           FROM warehouse_rates WHERE warehouse_id = ANY(%s) ORDER BY warehouse_id, rate_type""",
                        (wh_ids,))
            rates_map = {}
            for r in cur.fetchall():
                rates_map.setdefault(r["warehouse_id"], []).append(_serialize_row(r))
            for w in warehouses:
                w["rates"] = rates_map.get(w["id"], [])
    return JSONResponse({"warehouses": warehouses})


@app.post("/api/warehouses")
async def api_create_warehouse(request: Request):
    body = await request.json()
    fields = ["name", "mc_number", "region", "address", "city", "state", "zip_code",
              "contact_name", "contact_email", "contact_phone", "services", "notes", "source"]
    vals = [body.get(f, "" if f == "name" else None) for f in fields]
    if not vals[0]:
        vals[0] = "Unknown Warehouse"
    with db.get_conn() as conn:
        with db.get_cursor(conn) as cur:
            cur.execute(f"""INSERT INTO warehouses ({', '.join(fields)})
                VALUES ({', '.join(['%s'] * len(fields))}) RETURNING *""", vals)
            row = _serialize_row(cur.fetchone())
    return JSONResponse(row)


@app.put("/api/warehouses/{wh_id}")
async def api_update_warehouse(wh_id: int, request: Request):
    body = await request.json()
    allowed = {"name", "mc_number", "region", "address", "city", "state", "zip_code",
               "contact_name", "contact_email", "contact_phone", "services", "notes"}
    sets, params = [], []
    for k, v in body.items():
        if k in allowed:
            sets.append(f"{k} = %s")
            params.append(v)
    if not sets:
        raise HTTPException(400, "No valid fields")
    sets.append("updated_at = NOW()")
    params.append(wh_id)
    with db.get_conn() as conn:
        with db.get_cursor(conn) as cur:
            cur.execute(f"UPDATE warehouses SET {', '.join(sets)} WHERE id = %s RETURNING *", params)
            row = cur.fetchone()
    if not row:
        raise HTTPException(404, "Warehouse not found")
    return JSONResponse(_serialize_row(row))


@app.delete("/api/warehouses/{wh_id}")
async def api_delete_warehouse(wh_id: int):
    with db.get_conn() as conn:
        with db.get_cursor(conn) as cur:
            cur.execute("DELETE FROM warehouses WHERE id = %s", (wh_id,))
    return JSONResponse({"ok": True})


@app.get("/api/warehouses/{wh_id}/rates")
async def api_warehouse_rates(wh_id: int):
    with db.get_cursor() as cur:
        cur.execute("SELECT * FROM warehouse_rates WHERE warehouse_id = %s ORDER BY rate_type", (wh_id,))
        rows = [_serialize_row(r) for r in cur.fetchall()]
    return JSONResponse({"rates": rows})


@app.post("/api/warehouses/{wh_id}/rates")
async def api_add_warehouse_rate(wh_id: int, request: Request):
    body = await request.json()
    with db.get_conn() as conn:
        with db.get_cursor(conn) as cur:
            cur.execute("""INSERT INTO warehouse_rates (warehouse_id, rate_type, rate_amount, unit, description, notes)
                VALUES (%s, %s, %s, %s, %s, %s) RETURNING *""",
                (wh_id, body.get("rate_type", "flat"), body.get("rate_amount"),
                 body.get("unit"), body.get("description"), body.get("notes")))
            row = _serialize_row(cur.fetchone())
    return JSONResponse(row)


@app.delete("/api/warehouses/{wh_id}/rates/{rate_id}")
async def api_delete_warehouse_rate(wh_id: int, rate_id: int):
    with db.get_conn() as conn:
        with db.get_cursor(conn) as cur:
            cur.execute("DELETE FROM warehouse_rates WHERE id = %s AND warehouse_id = %s", (rate_id, wh_id))
    return JSONResponse({"ok": True})


_WAREHOUSE_EXTRACT_PROMPT = """Extract warehouse/transloading facility information from this rate card or document.
Return ONLY valid JSON (no markdown, no explanation) with these fields:
{
  "name": "facility name",
  "mc_number": "MC number digits or null",
  "address": "full address or null",
  "city": "string or null",
  "state": "2-letter state code or null",
  "contact_name": "string or null",
  "contact_email": "string or null",
  "contact_phone": "string or null",
  "services": "comma-separated services (Transload, Cross-dock, Storage, etc.) or null",
  "rates": [
    {"rate_type": "per_pallet|per_day|per_container|monthly_min|per_hour|flat|per_case|per_label",
     "rate_amount": numeric, "unit": "string", "description": "what this covers"}
  ]
}
Extract ALL rate line items. Look for per-pallet, per-day storage, container handling, labeling, monthly minimums. Omit null fields."""


@app.post("/api/warehouses/extract")
async def api_warehouse_extract(request: Request):
    """Extract warehouse info from uploaded file via Claude Vision."""
    if not config.ANTHROPIC_API_KEY:
        raise HTTPException(422, "ANTHROPIC_API_KEY not configured")
    form = await request.form()
    file = form.get("file")
    if not file:
        raise HTTPException(400, "No file provided")
    file_bytes = await file.read()
    filename = getattr(file, "filename", "") or ""
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""

    import base64
    if ext in ("png", "jpg", "jpeg", "gif", "webp"):
        media_map = {"png": "image/png", "jpg": "image/jpeg", "jpeg": "image/jpeg",
                     "gif": "image/gif", "webp": "image/webp"}
        content = [
            {"type": "image", "source": {"type": "base64", "media_type": media_map.get(ext, "image/png"),
                                         "data": base64.b64encode(file_bytes).decode()}},
            {"type": "text", "text": _WAREHOUSE_EXTRACT_PROMPT},
        ]
    elif ext == "pdf":
        content = [
            {"type": "document", "source": {"type": "base64", "media_type": "application/pdf",
                                            "data": base64.b64encode(file_bytes).decode()}},
            {"type": "text", "text": _WAREHOUSE_EXTRACT_PROMPT},
        ]
    elif ext in ("txt", "csv", "eml"):
        text_content = file_bytes.decode("utf-8", errors="replace")
        content = [{"type": "text", "text": _WAREHOUSE_EXTRACT_PROMPT + "\n\n" + text_content}]
    else:
        raise HTTPException(400, f"Unsupported file type: {ext}")

    try:
        result = _extract_with_claude(content)
        return JSONResponse(result)
    except Exception as e:
        log.error("Warehouse extraction failed: %s", e)
        raise HTTPException(500, f"AI extraction failed: {e}")


# ═══════════════════════════════════════════════════════════
# LANE RATES API
# ═══════════════════════════════════════════════════════════

@app.get("/api/lane-rates")
async def api_list_lane_rates(port: str = Query(default=None), carrier: str = Query(default=None),
                              destination: str = Query(default=None)):
    with db.get_cursor() as cur:
        clauses, params = [], []
        if port:
            clauses.append("port ILIKE %s")
            params.append(f"%{port}%")
        if carrier:
            clauses.append("carrier_name ILIKE %s")
            params.append(f"%{carrier}%")
        if destination:
            clauses.append("destination ILIKE %s")
            params.append(f"%{destination}%")
        where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
        cur.execute(f"SELECT * FROM lane_rates {where} ORDER BY port, destination, total ASC NULLS LAST LIMIT 1000", params)
        rows = [_serialize_row(r) for r in cur.fetchall()]
    return JSONResponse({"lane_rates": rows, "total": len(rows)})


# ═══════════════════════════════════════════════════════════
# EXCEL BULK IMPORT
# ═══════════════════════════════════════════════════════════

_CITY_TABS = {
    "Atlanta", "Baltimore", "Birmingham", "Buffalo", "Boston", "Charleston", "Charlotte",
    "Chicago", "Cincinnati", "Cleveland", "Columbus", "Dallas", "Denver", "Detroit",
    "El Paso, TX", "Houston", "Indianapolis", "Jacksonville", "Kansas City", "Louisville",
    "Los Angeles", "Memphis", "Miami", "Mobile", "Nashville", "New York", "NOLA",
    "Norfolk", "Oakland", "Omaha", "Phildadelphia", "Pittsburgh", "Portland",
    "Salt Lake City", "Seattle", "Savannah", "St Louis", "Wilmington", "Tacoma", "Tampa",
    "Toronto",
}

_SKIP_TABS = {
    "Sheet1", "Sheet2", "Sheet3", "Sheet4", "Rate Quote Sheet-John (1)",
    "Priority 1", "PostMaster", "Oversize", "Timberlab",
}


def _safe_float(val):
    """Try to convert a value to float, return None on failure."""
    if val is None:
        return None
    try:
        s = str(val).replace(",", "").replace("$", "").strip()
        if not s:
            return None
        return float(s)
    except (ValueError, TypeError):
        return None


def _parse_city_tab(ws, tab_name):
    """Parse a city tab from the Excel workbook. Returns (carriers, lane_rates)."""
    carriers = []
    lane_rates = []
    max_row = ws.max_row or 1

    # Detect the rate grid header row (cols E-R) and carrier directory section
    rate_grid_header_row = None
    carrier_section_row = None

    for row_idx in range(1, min(max_row + 1, 5)):
        cell_e = ws.cell(row=row_idx, column=5).value  # Col E
        cell_f = ws.cell(row=row_idx, column=6).value  # Col F
        if cell_e and cell_f:
            e_str = str(cell_e).lower().strip()
            f_str = str(cell_f).lower().strip()
            if any(kw in e_str for kw in ("destination", "lane", "carrier")) or any(kw in f_str for kw in ("carrier", "dray", "lane")):
                rate_grid_header_row = row_idx
                break

    # Find carrier directory section (look for "Carrier" header in col B)
    for row_idx in range(1, max_row + 1):
        cell_b = ws.cell(row=row_idx, column=2).value
        if cell_b and str(cell_b).strip().lower() == "carrier":
            cell_c = ws.cell(row=row_idx, column=3).value
            if cell_c and "email" in str(cell_c).lower():
                carrier_section_row = row_idx
                break

    # Parse rate grid (rows after header, cols E-R)
    if rate_grid_header_row:
        # Determine column mapping from header
        hdr = {}
        for col in range(5, 25):
            v = ws.cell(row=rate_grid_header_row, column=col).value
            if v:
                hdr[str(v).strip().lower()] = col

        dest_col = hdr.get("destination", hdr.get("lane", 5))
        carrier_col = hdr.get("carrier", 6)

        for row_idx in range(rate_grid_header_row + 1, max_row + 1):
            carrier_val = ws.cell(row=row_idx, column=carrier_col).value
            dest_val = ws.cell(row=row_idx, column=dest_col).value
            if not carrier_val and not dest_val:
                # Check if we've hit the quote template or carrier section
                cell_b = ws.cell(row=row_idx, column=2).value
                if cell_b and str(cell_b).strip().lower() in ("pod", "carrier", "warehousing"):
                    break
                continue
            if not carrier_val:
                continue

            lr = {
                "port": tab_name,
                "destination": str(dest_val).strip() if dest_val else None,
                "carrier_name": str(carrier_val).strip(),
                "dray_rate": _safe_float(ws.cell(row=row_idx, column=hdr.get("dray", hdr.get("rate", 7))).value),
                "fsc": str(ws.cell(row=row_idx, column=hdr.get("fsc", 8)).value or "").strip() or None,
                "total": _safe_float(ws.cell(row=row_idx, column=hdr.get("total", 9)).value),
                "chassis_per_day": _safe_float(ws.cell(row=row_idx, column=hdr.get("chassis", 10)).value),
                "prepull": _safe_float(ws.cell(row=row_idx, column=hdr.get("prepull", hdr.get("drop", 11))).value),
                "storage_per_day": _safe_float(ws.cell(row=row_idx, column=hdr.get("storage", hdr.get("reefer", 12))).value),
                "detention": str(ws.cell(row=row_idx, column=hdr.get("detention", 14)).value or "").strip() or None,
                "overweight": _safe_float(ws.cell(row=row_idx, column=hdr.get("ow", 15)).value),
                "tolls": _safe_float(ws.cell(row=row_idx, column=hdr.get("tolls", hdr.get("toll", 16))).value),
                "all_in_total": _safe_float(ws.cell(row=row_idx, column=hdr.get("total", 17)).value),
                "rank": int(_safe_float(ws.cell(row=row_idx, column=hdr.get("rank", 18)).value) or 0) or None,
                "move_type": "dray",
                "source": "excel_import",
            }
            # Use the last "total" column for all_in
            for k2 in sorted(hdr.keys()):
                if k2 == "total" and hdr[k2] > 9:
                    lr["all_in_total"] = _safe_float(ws.cell(row=row_idx, column=hdr[k2]).value)
            if lr["dray_rate"] or lr["total"]:
                lane_rates.append(lr)

    # Parse carrier directory (rows after "Carrier | Email | MC" header)
    if carrier_section_row:
        for row_idx in range(carrier_section_row + 1, max_row + 1):
            cell_b = ws.cell(row=row_idx, column=2).value
            cell_c = ws.cell(row=row_idx, column=3).value
            if not cell_b and not cell_c:
                # Check for end of carrier section
                continue
            name = str(cell_b).strip() if cell_b else None
            email = str(cell_c).strip() if cell_c else None
            if not name or name == "\xa0":
                continue
            if name.lower() in ("warehousing", "warehouse charges", ""):
                break
            mc = ws.cell(row=row_idx, column=4).value
            notes_val = ws.cell(row=row_idx, column=5).value
            carriers.append({
                "carrier_name": name,
                "contact_email": email if email and "@" in str(email) else None,
                "mc_number": str(mc).strip() if mc else None,
                "regions": tab_name,
                "notes": str(notes_val).strip() if notes_val else None,
                "source": "excel_import",
            })

    return carriers, lane_rates


@app.post("/api/directory/import-excel")
async def api_import_excel(file: UploadFile = File(...)):
    """Bulk import carriers, lane rates, and warehouses from the rate quote Excel."""
    import tempfile
    file_bytes = await file.read()
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(file_bytes)
        tmp_path = tmp.name

    try:
        import openpyxl
        wb = openpyxl.load_workbook(tmp_path, data_only=True, read_only=True)
    except Exception as e:
        os.unlink(tmp_path)
        raise HTTPException(400, f"Failed to read Excel file: {e}")

    all_carriers = []
    all_lane_rates = []
    all_warehouses = []
    summary = {"sheets_processed": [], "carriers": 0, "lane_rates": 0, "warehouses": 0}

    for name in wb.sheetnames:
        if name in _SKIP_TABS:
            continue
        ws = wb[name]

        if name in _CITY_TABS:
            carriers, lane_rates = _parse_city_tab(ws, name)
            all_carriers.extend(carriers)
            all_lane_rates.extend(lane_rates)
            if carriers or lane_rates:
                summary["sheets_processed"].append(name)
            continue

        # Specialty tabs
        if name == "Tolead Rates ORD":
            for row_idx in range(2, (ws.max_row or 1) + 1):
                city = ws.cell(row=row_idx, column=1).value
                rate = _safe_float(ws.cell(row=row_idx, column=3).value)
                if city and rate:
                    all_lane_rates.append({
                        "port": "Chicago (ORD)", "destination": str(city).strip(),
                        "carrier_name": "Tolead", "total": rate, "move_type": "ftl",
                        "source": "excel_import",
                    })
            summary["sheets_processed"].append(name)

        elif name == "LTL":
            for row_idx in range(2, (ws.max_row or 1) + 1):
                pallets = ws.cell(row=row_idx, column=1).value
                dest = ws.cell(row=row_idx, column=2).value
                rate = _safe_float(ws.cell(row=row_idx, column=4).value)
                if rate:
                    all_lane_rates.append({
                        "port": "LTL", "destination": str(dest or "").strip() or None,
                        "carrier_name": f"LTL ({pallets} pallet{'s' if str(pallets) != '1' else ''})",
                        "total": rate, "move_type": "ltl", "source": "excel_import",
                        "notes": f"{ws.cell(row=row_idx, column=3).value} transit days" if ws.cell(row=row_idx, column=3).value else None,
                    })
            summary["sheets_processed"].append(name)

        elif name == "Step Deck":
            for row_idx in range(2, (ws.max_row or 1) + 1):
                pickup = ws.cell(row=row_idx, column=1).value
                lane = ws.cell(row=row_idx, column=2).value
                rate = _safe_float(ws.cell(row=row_idx, column=3).value)
                if lane and rate:
                    all_lane_rates.append({
                        "port": str(pickup or "").strip() or "Step Deck",
                        "destination": str(lane).strip(),
                        "carrier_name": "Step Deck",
                        "total": rate, "move_type": "step_deck", "equipment_type": "Step Deck",
                        "source": "excel_import",
                    })
            summary["sheets_processed"].append(name)

        elif name == "Sutton WH - Inventory (Meiborg)":
            wh = {"name": "Sutton Warehouse (Meiborg)", "source": "excel_import", "services": "Storage, Labeling, Palletizing"}
            rates = []
            for row_idx in range(2, (ws.max_row or 1) + 1):
                desc = ws.cell(row=row_idx, column=1).value
                rate = _safe_float(ws.cell(row=row_idx, column=3).value)
                if desc and rate:
                    desc_str = str(desc).strip()
                    if "total" in desc_str.lower():
                        continue
                    rates.append({"rate_type": "flat", "rate_amount": rate, "unit": "each", "description": desc_str})
            if rates:
                all_warehouses.append({"warehouse": wh, "rates": rates})
            summary["sheets_processed"].append(name)

        elif name == "Tolead Box Rates":
            # Similar to Tolead Rates ORD but for box rates
            for row_idx in range(2, (ws.max_row or 1) + 1):
                city = ws.cell(row=row_idx, column=1).value
                rate = _safe_float(ws.cell(row=row_idx, column=3).value)
                if city and rate:
                    all_lane_rates.append({
                        "port": "Tolead Box", "destination": str(city).strip(),
                        "carrier_name": "Tolead", "total": rate, "move_type": "ftl",
                        "source": "excel_import",
                    })
            summary["sheets_processed"].append(name)

        elif name == "Heavy Haul":
            # Heavy haul has truck assignments with rates
            for row_idx in range(2, (ws.max_row or 1) + 1):
                truck = ws.cell(row=row_idx, column=13).value  # M
                rate = _safe_float(ws.cell(row=row_idx, column=14).value)  # N
                equip = ws.cell(row=row_idx, column=16).value  # P
                if rate:
                    all_lane_rates.append({
                        "port": "Heavy Haul", "destination": "Project Site",
                        "carrier_name": f"Truck {truck}" if truck else "Heavy Haul",
                        "total": rate, "move_type": "heavy_haul",
                        "equipment_type": str(equip).strip() if equip else None,
                        "source": "excel_import",
                    })
            summary["sheets_processed"].append(name)

    wb.close()
    os.unlink(tmp_path)

    # Bulk insert into database
    with db.get_conn() as conn:
        with db.get_cursor(conn) as cur:
            # Insert carriers (deduplicate by name+email within import)
            seen_carriers = set()
            for c in all_carriers:
                key = (c["carrier_name"].lower(), (c.get("contact_email") or "").lower())
                if key in seen_carriers:
                    # Append region to existing
                    continue
                seen_carriers.add(key)
                cur.execute("""
                    INSERT INTO carriers (carrier_name, mc_number, contact_email, regions, notes, source)
                    VALUES (%s, %s, %s, %s, %s, %s)
                    ON CONFLICT DO NOTHING
                """, (c["carrier_name"], c.get("mc_number"), c.get("contact_email"),
                      c.get("regions"), c.get("notes"), "excel_import"))
                summary["carriers"] += 1

            # Insert lane rates
            for lr in all_lane_rates:
                cur.execute("""
                    INSERT INTO lane_rates (port, destination, carrier_name, dray_rate, fsc, total,
                        chassis_per_day, prepull, storage_per_day, detention, chassis_split,
                        overweight, tolls, reefer, hazmat, all_in_total, rank,
                        equipment_type, move_type, notes, source)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (lr.get("port"), lr.get("destination"), lr.get("carrier_name"),
                      lr.get("dray_rate"), lr.get("fsc"), lr.get("total"),
                      lr.get("chassis_per_day"), lr.get("prepull"), lr.get("storage_per_day"),
                      lr.get("detention"), lr.get("chassis_split"), lr.get("overweight"),
                      lr.get("tolls"), lr.get("reefer"), lr.get("hazmat"),
                      lr.get("all_in_total"), lr.get("rank"),
                      lr.get("equipment_type"), lr.get("move_type", "dray"),
                      lr.get("notes"), lr.get("source", "excel_import")))
                summary["lane_rates"] += 1

            # Insert warehouses + rates
            for wh_data in all_warehouses:
                wh = wh_data["warehouse"]
                cur.execute("""INSERT INTO warehouses (name, services, source)
                    VALUES (%s, %s, %s) RETURNING id""",
                    (wh["name"], wh.get("services"), "excel_import"))
                wh_id = cur.fetchone()["id"]
                for rate in wh_data["rates"]:
                    cur.execute("""INSERT INTO warehouse_rates (warehouse_id, rate_type, rate_amount, unit, description)
                        VALUES (%s, %s, %s, %s, %s)""",
                        (wh_id, rate["rate_type"], rate["rate_amount"], rate["unit"], rate.get("description")))
                summary["warehouses"] += 1

    return JSONResponse(summary)




# ── Inbox Command Center Endpoints ──────────────────────────────────

@app.get("/api/inbox")
async def get_inbox(request: Request):
    """Unified inbox: thread-grouped emails with sent-reply detection."""
    from urllib.parse import parse_qs
    params = dict(request.query_params)
    days = int(params.get("days", "7"))
    tab = params.get("tab", "all")
    email_type_filter = params.get("type", "")
    priority_min = int(params.get("priority_min", "0"))
    rep_filter = params.get("rep", "")
    limit = min(int(params.get("limit", "200")), 500)

    with db.get_cursor() as cur:
        # Fetch all inbound emails (matched + unmatched) from the lookback window
        cur.execute("""
            SELECT
                e.id, e.gmail_thread_id, e.gmail_message_id, e.subject, e.sender,
                e.body_preview, e.has_attachments, e.attachment_names,
                e.sent_at, e.email_type, e.lane, e.priority, e.ai_summary,
                e.classification_feedback, e.corrected_type,
                e.efj, NULL as review_status, NULL as suggested_rep,
                'matched' as source
            FROM email_threads e
            WHERE e.sent_at >= NOW() - INTERVAL '%s days'
            UNION ALL
            SELECT
                u.id, u.gmail_thread_id, u.gmail_message_id, u.subject, u.sender,
                u.body_preview, u.has_attachments, u.attachment_names,
                u.sent_at, u.email_type, u.lane, u.priority, u.ai_summary,
                u.classification_feedback, u.corrected_type,
                NULL as efj, u.review_status, u.suggested_rep,
                'unmatched' as source
            FROM unmatched_inbox_emails u
            WHERE u.sent_at >= NOW() - INTERVAL '%s days'
              AND u.review_status = 'pending'
            ORDER BY sent_at DESC
        """ % (days, days))
        inbound_rows = cur.fetchall()

        # No sent_messages query needed — reply detection uses sender patterns in email_threads
        pass  # Sent messages are detected by sender domain in the thread messages below

    # CSL team domains for reply detection
    _CSL_DOMAINS = ('evansdelivery', 'commonsenselogistics')

    # Group inbound by thread_id
    threads_map = {}
    for row in inbound_rows:
        tid = row["gmail_thread_id"] or row["gmail_message_id"]  # fallback if no thread
        if tid not in threads_map:
            threads_map[tid] = {
                "thread_id": tid,
                "efj": row["efj"],
                "messages": [],
                "max_priority": 0,
                "email_type": None,
                "ai_summary": None,
                "lane": None,
                "has_attachments": False,
                "source": row["source"],
                "review_status": row["review_status"],
                "suggested_rep": row["suggested_rep"],
                "classification_feedback": row["classification_feedback"],
            }

        thread = threads_map[tid]
        _sender_lower = (row["sender"] or "").lower()
        _is_csl = any(d in _sender_lower for d in _CSL_DOMAINS)
        thread["messages"].append({
            "id": row["id"],
            "sender": row["sender"],
            "subject": row["subject"],
            "sent_at": row["sent_at"].isoformat() if row["sent_at"] else None,
            "body_preview": row["body_preview"],
            "direction": "sent" if _is_csl else "inbound",
            "has_attachments": row["has_attachments"],
            "attachment_names": row["attachment_names"],
            "email_type": row["email_type"],
            "priority": row["priority"],
            "ai_summary": row["ai_summary"],
        })

        # Track thread-level aggregates
        p = row["priority"] or 0
        if p > thread["max_priority"]:
            thread["max_priority"] = p
            thread["email_type"] = row["email_type"]
            thread["ai_summary"] = row["ai_summary"]
        if row["lane"]:
            thread["lane"] = row["lane"]
        if row["has_attachments"]:
            thread["has_attachments"] = True
        if row["efj"]:
            thread["efj"] = row["efj"]
        if row["source"] == "matched":
            thread["source"] = "matched"
        if row["suggested_rep"]:
            thread["suggested_rep"] = row["suggested_rep"]

    # Build final thread list with reply detection
    threads = []
    for tid, thread in threads_map.items():
        # Sort messages chronologically (sent/inbound already tagged by direction)
        all_msgs = thread["messages"][:]
        all_msgs.sort(key=lambda m: m["sent_at"] or "")

        # Determine latest message and reply status
        latest_msg = all_msgs[-1] if all_msgs else None
        # Reply detection is done via external/CSL sender matching below

        # has_csl_reply: any CSL team message AFTER the latest external inbound
        external_msgs = [m for m in all_msgs if m.get("direction") == "inbound"]
        csl_msgs = [m for m in all_msgs if m.get("direction") == "sent"]
        latest_external = external_msgs[-1] if external_msgs else None
        has_csl_reply = False
        if latest_external and csl_msgs:
            latest_ext_time = latest_external.get("sent_at") or ""
            has_csl_reply = any(
                (m.get("sent_at") or "") > latest_ext_time
                for m in csl_msgs
            )

        # needs_reply: latest external message exists + no CSL reply after it
        needs_reply = (
            latest_external is not None
            and not has_csl_reply
        )

        thread_start = thread["messages"][0]["sent_at"] if thread["messages"] else None

        threads.append({
            "thread_id": tid,
            "efj": thread["efj"],
            "message_count": len(all_msgs),
            "latest_subject": latest_msg["subject"] if latest_msg else None,
            "latest_sender": latest_msg["sender"] if latest_msg else None,
            "latest_sent_at": latest_msg["sent_at"] if latest_msg else None,
            "thread_start": thread_start,
            "max_priority": thread["max_priority"],
            "email_type": thread["email_type"],
            "ai_summary": thread["ai_summary"],
            "lane": thread["lane"],
            "has_attachments": thread["has_attachments"],
            "has_csl_reply": has_csl_reply,
            "needs_reply": needs_reply,
            "suggested_rep": thread["suggested_rep"],
            "source": thread["source"],
            "review_status": thread["review_status"],
            "classification_feedback": thread["classification_feedback"],
            "messages": all_msgs,
        })

    # Sort by priority DESC, then latest_sent_at DESC
    threads.sort(key=lambda t: (-(t["max_priority"] or 0), t["latest_sent_at"] or ""), reverse=False)
    threads.sort(key=lambda t: -(t["max_priority"] or 0))

    # Apply tab filter
    if tab == "needs_reply":
        threads = [t for t in threads if t["needs_reply"]]
    elif tab == "unmatched":
        threads = [t for t in threads if t["source"] == "unmatched"]
    elif tab == "rates":
        threads = [t for t in threads if t["email_type"] in (
            "carrier_rate", "customer_rate", "carrier_rate_response",
            "carrier_rate_confirmation", "rate_outreach")]

    # Apply additional filters
    if email_type_filter:
        types = set(email_type_filter.split(","))
        threads = [t for t in threads if t["email_type"] in types]
    if priority_min > 0:
        threads = [t for t in threads if (t["max_priority"] or 0) >= priority_min]
    if rep_filter:
        threads = [t for t in threads if t.get("suggested_rep") == rep_filter]

    # Stats
    all_threads = list(threads_map.values())
    stats = {
        "total_threads": len(threads_map),
        "needs_reply": sum(1 for t in threads if t["needs_reply"]),
        "unmatched": sum(1 for t in threads_map.values() if t["source"] == "unmatched"),
        "high_priority": sum(1 for t in threads_map.values() if t["max_priority"] >= 4),
    }

    return JSONResponse({
        "threads": threads[:limit],
        "stats": stats,
    })


@app.post("/api/inbox/{email_id}/feedback")
async def inbox_classification_feedback(email_id: int, request: Request):
    """Store classification feedback (correct/incorrect) for an email."""
    body = await request.json()
    feedback = body.get("feedback")  # "correct" or "incorrect"
    corrected_type = body.get("corrected_type")  # only if incorrect

    if feedback not in ("correct", "incorrect"):
        raise HTTPException(400, "feedback must be 'correct' or 'incorrect'")

    # Try email_threads first, then unmatched
    updated = False
    with db.get_conn() as conn:
        with db.get_cursor(conn) as cur:
            cur.execute(
                """UPDATE email_threads
                   SET classification_feedback = %s, corrected_type = %s
                   WHERE id = %s""",
                (feedback, corrected_type if feedback == "incorrect" else None, email_id),
            )
            if cur.rowcount > 0:
                updated = True
            else:
                cur.execute(
                    """UPDATE unmatched_inbox_emails
                       SET classification_feedback = %s, corrected_type = %s
                       WHERE id = %s""",
                    (feedback, corrected_type if feedback == "incorrect" else None, email_id),
                )
                if cur.rowcount > 0:
                    updated = True

    if not updated:
        raise HTTPException(404, "Email not found")

    return JSONResponse({"status": "ok", "feedback": feedback})


@app.get("/api/inbox/reply-alerts")
async def get_inbox_reply_alerts():
    """Return unreplied customer quote emails (threads needing reply)."""
    with db.get_cursor() as cur:
        cur.execute("""
        SELECT ca.id, ca.email_thread_id, ca.efj, ca.sender, ca.subject,
               ca.alerted_at, ca.dismissed
        FROM customer_reply_alerts ca
        WHERE ca.dismissed = false
        ORDER BY ca.alerted_at DESC
        LIMIT 50
        """)
        rows = cur.fetchall()

    alerts = []
    for r in rows:
        alerts.append({
            "id": r["id"],
            "email_thread_id": r["email_thread_id"],
            "efj": r["efj"],
            "sender": r["sender"],
            "subject": r["subject"],
            "alerted_at": r["alerted_at"].isoformat() if r["alerted_at"] else None,
        })

    return JSONResponse({"alerts": alerts, "count": len(alerts)})




@app.get("/api/rate-response-alerts")
async def rate_response_alerts():
    """Return recent carrier_rate_response emails for live alert consumption."""
    with db.get_cursor() as cur:
        cur.execute("""
            SELECT id, efj, gmail_thread_id, sender, subject, lane,
                   ai_summary, sent_at, suggested_rep
            FROM email_threads
            WHERE email_type IN ('carrier_rate_response', 'payment_escalation',
                                 'carrier_invoice', 'carrier_rate_confirmation')
              AND sent_at > NOW() - INTERVAL '24 hours'
            ORDER BY sent_at DESC
        """)
        rows = cur.fetchall()
    alerts = []
    for r in rows:
        alerts.append({
            "id": r["id"],
            "efj": r["efj"],
            "email_type": r["email_type"],
            "sender": r["sender"],
            "subject": r["subject"],
            "lane": r["lane"],
            "summary": r["ai_summary"],
            "sent_at": r["sent_at"].isoformat() if r["sent_at"] else None,
            "rep": r["suggested_rep"],
        })
    return JSONResponse({"alerts": alerts})


@app.get("/health")
def health():
    return {"status": "ok"}



# ── BOL Generator Proxy Endpoints ──────────────────────────────────────────

_BOL_ACCOUNTS = {
    "accounts": [
        {
            "name": "Boviet",
            "columns": ["PO Number", "Quantity", "Weight", "Description",
                        "Consignee", "Ship From", "Ship To"],
        },
        {
            "name": "General",
            "columns": ["PO Number", "Quantity", "Weight", "Description",
                        "Consignee", "Ship From", "Ship To"],
        },
    ]
}

BOL_WEBAPP_URL = "http://localhost:5002"
BOL_WEBAPP_PASSWORD = os.getenv("BOL_PASSWORD", "evans2026")


@app.get("/api/bol/accounts")
async def api_bol_accounts():
    """Return static BOL account configuration."""
    return JSONResponse(_BOL_ACCOUNTS)


@app.post("/api/bol/generate")
async def api_bol_generate(request: Request):
    """Proxy file upload to BOL webapp on localhost:5002.
    Authenticates with the BOL webapp, forwards the uploaded file,
    and streams back the generated ZIP."""
    import requests as _bol_req

    form = await request.form()
    file = form.get("file") or form.get("datafile")
    if not file:
        raise HTTPException(400, "No file uploaded")

    file_bytes = await file.read()
    filename = getattr(file, "filename", "data.xlsx") or "data.xlsx"

    try:
        sess = _bol_req.Session()
        # Authenticate with the BOL webapp
        sess.post(
            f"{BOL_WEBAPP_URL}/login",
            data={"password": BOL_WEBAPP_PASSWORD},
            allow_redirects=False,
            timeout=10,
        )
        # Forward the file to /generate
        resp = sess.post(
            f"{BOL_WEBAPP_URL}/generate",
            files={"datafile": (filename, file_bytes)},
            allow_redirects=False,
            timeout=300,
        )
    except _bol_req.ConnectionError:
        raise HTTPException(503, "BOL webapp is unreachable (port 5002)")
    except _bol_req.Timeout:
        raise HTTPException(504, "BOL webapp timed out")
    except Exception as e:
        log.error("BOL proxy error: %s", e)
        raise HTTPException(502, f"BOL proxy error: {e}")

    # If the BOL webapp redirected (flash error), extract the error
    if resp.status_code in (301, 302, 303):
        raise HTTPException(422, "BOL generation failed — check file format and columns")

    content_type = resp.headers.get("content-type", "application/zip")
    content_disp = resp.headers.get("content-disposition", "attachment; filename=BOLs.zip")

    return Response(
        content=resp.content,
        status_code=resp.status_code,
        headers={
            "Content-Type": content_type,
            "Content-Disposition": content_disp,
        },
    )


_BOL_EXTRACT_PROMPT = """You are a logistics data extraction assistant. Extract BOL (Bill of Lading) data from this image/document.

Return a JSON object with this structure:
{
  "rows": [
    {
      "po_number": "string or null",
      "quantity": "string or null",
      "weight": "string or null",
      "description": "string or null",
      "consignee": "string or null",
      "ship_from": "string or null",
      "ship_to": "string or null"
    }
  ],
  "shipper": "string or null",
  "consignee": "string or null",
  "carrier": "string or null",
  "bol_number": "string or null",
  "date": "string or null",
  "notes": "string or null"
}

Extract as many line-item rows as you can find. Include addresses, PO numbers, weights, quantities, and descriptions.
If a field is not present, use null. Return ONLY valid JSON, no markdown fences."""


@app.post("/api/bol/extract")
async def api_bol_extract(request: Request):
    """Extract BOL fields from an uploaded image/PDF using Claude Vision."""
    if not getattr(config, "ANTHROPIC_API_KEY", None):
        raise HTTPException(422, "ANTHROPIC_API_KEY not configured — OCR extraction unavailable")

    content_type = request.headers.get("content-type", "")
    if "multipart/form-data" not in content_type:
        raise HTTPException(400, "Expected multipart/form-data with a file upload")

    form = await request.form()
    file = form.get("file")
    if not file:
        raise HTTPException(400, "No file uploaded")

    file_bytes = await file.read()
    filename = getattr(file, "filename", "") or ""
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""

    import anthropic
    import base64

    client = anthropic.Anthropic(api_key=config.ANTHROPIC_API_KEY)

    # Build content blocks based on file type
    if ext in ("png", "jpg", "jpeg", "gif", "webp"):
        media_map = {
            "png": "image/png", "jpg": "image/jpeg", "jpeg": "image/jpeg",
            "gif": "image/gif", "webp": "image/webp",
        }
        content = [
            {"type": "image", "source": {
                "type": "base64",
                "media_type": media_map.get(ext, "image/png"),
                "data": base64.b64encode(file_bytes).decode(),
            }},
            {"type": "text", "text": _BOL_EXTRACT_PROMPT},
        ]
    elif ext == "pdf":
        content = [
            {"type": "document", "source": {
                "type": "base64",
                "media_type": "application/pdf",
                "data": base64.b64encode(file_bytes).decode(),
            }},
            {"type": "text", "text": _BOL_EXTRACT_PROMPT},
        ]
    else:
        raise HTTPException(400, f"Unsupported file type for BOL extraction: .{ext}. Use PNG, JPG, or PDF.")

    try:
        message = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=2048,
            messages=[{"role": "user", "content": content}],
        )
        response_text = message.content[0].text.strip()
        # Strip markdown code fences if present
        if response_text.startswith("```"):
            response_text = response_text.split("\n", 1)[-1].rsplit("```", 1)[0].strip()
        result = json.loads(response_text)
        return JSONResponse(result)
    except json.JSONDecodeError as e:
        log.error("BOL extract: Claude returned invalid JSON: %s", e)
        raise HTTPException(500, "AI extraction returned invalid JSON")
    except Exception as e:
        log.error("BOL extract failed: %s", e)
        raise HTTPException(500, f"AI extraction failed: {e}")


# ---------------------------------------------------------------------------
# Completed loads cache + API — reads "Completed Eli/Radka/John F" tabs
# ---------------------------------------------------------------------------

_completed_cache = {"data": [], "ts": 0}
_completed_lock = threading.Lock()
_COMPLETED_TABS = {
    "Completed Eli": "Eli",
    "Completed Radka": "Radka",
    "Completed John F": "John F",
}


def _refresh_completed_cache():
    """Fetch completed loads from all 3 Completed tabs in Master Tracker.
    Enriches account from PG shipments table and sorts newest-first."""
    now = _time.time()
    if now - _completed_cache["ts"] < CACHE_TTL:
        return
    with _completed_lock:
        if _time.time() - _completed_cache["ts"] < CACHE_TTL:
            return
        try:
            creds = Credentials.from_service_account_file(
                CREDS_FILE, scopes=["https://www.googleapis.com/auth/spreadsheets"]
            )
            gc_local = gspread.authorize(creds)
            sh = gc_local.open_by_key(SHEET_ID)
            tab_names = list(_COMPLETED_TABS.keys())
            ranges = [f"\'{t}\'!A:P" for t in tab_names]
            batch_result = sh.values_batch_get(ranges)
            value_ranges = batch_result.get("valueRanges", [])
            loads = []
            for vr, tab_name in zip(value_ranges, tab_names):
                rep = _COMPLETED_TABS[tab_name]
                rows = vr.get("values", [])
                if len(rows) < 2:
                    continue
                # Detect header row (skip title rows)
                hdr_idx = 0
                if len(rows) > 1:
                    r0 = sum(1 for c in rows[0] if c.strip())
                    r1 = sum(1 for c in rows[1] if c.strip())
                    if r1 > r0:
                        hdr_idx = 1
                for row in rows[hdr_idx + 1:]:
                    efj = row[COL["efj"]].strip() if len(row) > COL["efj"] else ""
                    ctr = row[COL["container"]].strip() if len(row) > COL["container"] else ""
                    if not efj and not ctr:
                        continue
                    def cell(key, r=row):
                        idx = COL[key]
                        return r[idx].strip() if len(r) > idx else ""
                    loads.append({
                        "efj": efj,
                        "move_type": cell("move_type"),
                        "container": ctr,
                        "bol": cell("bol"),
                        "ssl": cell("ssl"),
                        "carrier": cell("carrier"),
                        "origin": cell("origin"),
                        "destination": cell("destination"),
                        "eta": cell("eta"),
                        "lfd": cell("lfd"),
                        "pickup": cell("pickup"),
                        "delivery": cell("delivery"),
                        "status": cell("status"),
                        "notes": cell("notes"),
                        "bot_alert": cell("bot_alert"),
                        "return_port": cell("return_port"),
                        "rep": rep,
                        "account": "",  # enriched below from PG
                    })

            # --- Enrich account from PG shipments table ---
            efj_list = [l["efj"] for l in loads if l["efj"]]
            pg_accounts = {}
            if efj_list:
                try:
                    with db.get_cursor() as cursor:
                        cursor.execute(
                            "SELECT efj, account FROM shipments WHERE efj = ANY(%s)",
                            (efj_list,)
                        )
                        pg_accounts = {r["efj"]: r["account"] for r in cursor.fetchall() if r.get("account")}
                except Exception as pg_err:
                    log.warning("Completed cache PG account lookup failed: %s", pg_err)

            # Fallback: reverse rep_map from sheet_cache (account -> rep)
            rep_map = getattr(sheet_cache, "rep_map", {})
            # Build reverse: rep -> list of accounts
            rev_rep = {}
            for acct, r in rep_map.items():
                rev_rep.setdefault(r, []).append(acct)

            for load in loads:
                efj = load["efj"]
                if efj in pg_accounts:
                    load["account"] = pg_accounts[efj]
                else:
                    # If only one account for this rep, use it; otherwise leave blank
                    accts = rev_rep.get(load["rep"], [])
                    load["account"] = accts[0] if len(accts) == 1 else ""

            # --- Sort by delivery date descending (newest first) ---
            import re as _re
            def _delivery_sort_key(load):
                """Parse delivery date for sorting. Returns sortable string, empty last."""
                d = load.get("delivery", "") or load.get("pickup", "") or ""
                # Try to extract date portion (handles "MM-DD HH:MM", "MM/DD", "YYYY-MM-DD", etc.)
                # Normalize to comparable form
                d = d.strip().split()[0] if d.strip() else ""  # take date part only
                if not d:
                    return "0000-00-00"
                # Handle MM-DD or MM/DD format
                m = _re.match(r"(\d{1,2})[/-](\d{1,2})", d)
                if m:
                    return f"2026-{int(m.group(1)):02d}-{int(m.group(2)):02d}"
                # Handle YYYY-MM-DD
                m2 = _re.match(r"(\d{4})[/-](\d{1,2})[/-](\d{1,2})", d)
                if m2:
                    return f"{m2.group(1)}-{int(m2.group(2)):02d}-{int(m2.group(3)):02d}"
                return d

            loads.sort(key=_delivery_sort_key, reverse=True)

            # --- Merge PG-archived shipments (Tolead, Boviet, and any PG-only archives) ---
            try:
                with db.get_cursor() as _pg_cur:
                    _pg_cur.execute("""
                        SELECT efj, move_type, container, bol, vessel, carrier,
                               origin, destination, eta, lfd,
                               pickup_date::text, delivery_date::text, status,
                               notes, bot_notes, return_date::text,
                               rep, account, hub
                        FROM shipments
                        WHERE archived = TRUE
                        ORDER BY archived_at DESC NULLS LAST
                    """)
                    pg_archived = _pg_cur.fetchall()

                existing_efjs = {l["efj"] for l in loads}
                pg_added = 0
                for row in pg_archived:
                    if row["efj"] not in existing_efjs:
                        loads.append({
                            "efj": row["efj"] or "",
                            "move_type": row["move_type"] or "",
                            "container": row["container"] or "",
                            "bol": row["bol"] or "",
                            "ssl": row["vessel"] or "",
                            "carrier": row["carrier"] or "",
                            "origin": row["origin"] or "",
                            "destination": row["destination"] or "",
                            "eta": row["eta"] or "",
                            "lfd": row["lfd"] or "",
                            "pickup": row["pickup_date"] or "",
                            "delivery": row["delivery_date"] or "",
                            "status": row["status"] or "",
                            "notes": row["notes"] or "",
                            "bot_alert": row["bot_notes"] or "",
                            "return_port": row["return_date"] or "",
                            "rep": row["rep"] or "",
                            "account": row["account"] or "",
                        })
                        existing_efjs.add(row["efj"])
                        pg_added += 1
                if pg_added:
                    loads.sort(key=_delivery_sort_key, reverse=True)
                    log.info("Completed cache: merged %d PG-archived shipments", pg_added)
            except Exception as pg_merge_err:
                log.warning("Completed cache PG merge failed: %s", pg_merge_err)

            _completed_cache["data"] = loads
            _completed_cache["ts"] = _time.time()
            log.info("Completed cache: %d total loads (sheets + PG)", len(loads))
        except Exception as e:
            log.error("Completed cache refresh failed: %s", e)




# -- Carrier Performance Scorecard ─────────────────────────────────────

@app.get("/api/carriers/scorecard")
async def api_carrier_scorecard():
    """Aggregate carrier delivery performance from completed loads."""
    from datetime import datetime as _dt
    from collections import defaultdict

    _refresh_completed_cache()
    loads = _completed_cache.get("data", [])

    carriers = defaultdict(lambda: {
        "loads": 0, "on_time": 0, "total_transit": 0, "transit_count": 0,
        "lanes": defaultdict(int), "last_delivery": None, "move_types": defaultdict(int),
    })

    def _parse_date(s):
        if not s:
            return None
        for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y", "%m-%d-%Y", "%m-%d-%y"):
            try:
                return _dt.strptime(s.strip(), fmt)
            except ValueError:
                continue
        return None

    for load in loads:
        carrier = (load.get("carrier") or "").strip()
        if not carrier or carrier.lower() in ("", "tbd", "tba", "n/a", "none"):
            continue

        carriers[carrier]["loads"] += 1
        move = load.get("move_type", "").strip()
        if move:
            carriers[carrier]["move_types"][move] += 1

        origin = load.get("origin", "").strip()
        dest = load.get("destination", "").strip()
        if origin and dest:
            lane = f"{origin} → {dest}"
            carriers[carrier]["lanes"][lane] += 1

        pickup_dt = _parse_date(load.get("pickup"))
        delivery_dt = _parse_date(load.get("delivery"))
        lfd_dt = _parse_date(load.get("lfd"))

        # Transit time
        if pickup_dt and delivery_dt and delivery_dt > pickup_dt:
            delta = (delivery_dt - pickup_dt).days
            if 0 < delta < 60:
                carriers[carrier]["total_transit"] += delta
                carriers[carrier]["transit_count"] += 1

        # On-time: delivery <= LFD (for dray) or delivery exists (for FTL)
        if delivery_dt:
            if lfd_dt:
                if delivery_dt <= lfd_dt:
                    carriers[carrier]["on_time"] += 1
            else:
                carriers[carrier]["on_time"] += 1

        # Track most recent delivery
        if delivery_dt:
            if not carriers[carrier]["last_delivery"] or delivery_dt > carriers[carrier]["last_delivery"]:
                carriers[carrier]["last_delivery"] = delivery_dt

    # Build response
    results = []
    for name, data in carriers.items():
        total = data["loads"]
        on_time_pct = round(data["on_time"] / total * 100) if total > 0 else 0
        avg_transit = round(data["total_transit"] / data["transit_count"], 1) if data["transit_count"] > 0 else None
        top_lanes = sorted(data["lanes"].items(), key=lambda x: -x[1])[:5]
        primary_move = max(data["move_types"].items(), key=lambda x: x[1])[0] if data["move_types"] else None

        results.append({
            "carrier": name,
            "total_loads": total,
            "on_time_pct": on_time_pct,
            "avg_transit_days": avg_transit,
            "lanes_served": len(data["lanes"]),
            "top_lanes": [{"lane": l, "count": c} for l, c in top_lanes],
            "primary_move_type": primary_move,
            "last_delivery": data["last_delivery"].strftime("%Y-%m-%d") if data["last_delivery"] else None,
        })

    results.sort(key=lambda x: -x["total_loads"])
    return JSONResponse({"carriers": results, "total": len(results)})


@app.get("/api/completed")
async def api_completed(
    request: Request,
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=200),
    search: str = Query(""),
    rep: str = Query(""),
    account: str = Query(""),
):
    """Return paginated completed shipment loads from Completed tabs."""
    _refresh_completed_cache()
    data = list(_completed_cache["data"])

    # --- Assign account from rep_map if available ---
    rep_map = getattr(sheet_cache, "rep_map", {})

    # --- Filter by rep ---
    if rep:
        rep_lower = rep.lower()
        data = [s for s in data if s.get("rep", "").lower() == rep_lower]

    # --- Filter by account ---
    if account:
        acct_lower = account.lower()
        data = [s for s in data if s.get("account", "").lower() == acct_lower]

    # --- Search filter (EFJ, container, carrier, origin, destination, bol) ---
    if search:
        q = search.lower()
        def matches(s):
            for field in ("efj", "container", "carrier", "origin", "destination", "bol", "account"):
                if q in s.get(field, "").lower():
                    return True
            return False
        data = [s for s in data if matches(s)]

    total = len(data)
    start = (page - 1) * limit
    end = start + limit
    page_data = data[start:end]
    return {"loads": page_data, "total": total, "hasMore": end < total}


# ---------------------------------------------------------------------------
# Bot Health â deep per-service diagnostics from journalctl
# ---------------------------------------------------------------------------

def _analyze_service_health(unit: str, name: str, poll_min: int) -> dict:
    """Analyze 24h of journalctl logs for a single service."""
    import subprocess as _sp

    # 1. Active state
    try:
        r = _sp.run(["systemctl", "is-active", unit], capture_output=True, text=True, timeout=5)
        active_state = r.stdout.strip()
    except Exception:
        active_state = "unknown"

    # 2. Pull 24h of journal logs
    try:
        r = _sp.run(
            ["journalctl", "-u", unit, "--since", "24 hours ago", "--no-pager", "-o", "short"],
            capture_output=True, text=True, timeout=30,
        )
        raw = r.stdout.strip()
        lines = raw.split(chr(10)) if raw else []
    except Exception:
        lines = []

    # 3. Count ACTUAL crashes = systemd restart events (not just any error line)
    crash_pattern = re.compile(
        r"(Failed with result|Main process exited, code=exited, status=1|"
        r"Scheduled restart job, restart counter|"
        r"systemd\[\d+\]: .+: Main process exited)",
        re.IGNORECASE,
    )
    crash_count = sum(1 for l in lines if crash_pattern.search(l))

    # 4. Operational errors (for display, not health determination)
    error_pattern = re.compile(r"error|traceback|failed|exception", re.IGNORECASE)
    error_lines = [l for l in lines if error_pattern.search(l) and not crash_pattern.search(l)]
    recent_errors = []
    for el in (error_lines[-5:] if error_lines else []):
        tm = re.match(r"\w+ \d+ (\d+:\d+:\d+)", el)
        time_str = tm.group(1) if tm else ""
        msg = el.split(":", 3)[-1].strip() if ":" in el else el
        recent_errors.append({"time": time_str, "level": "error", "msg": msg[:120]})

    # Also add crash lines to recent_errors
    crash_lines = [l for l in lines if crash_pattern.search(l)]
    for cl in (crash_lines[-3:] if crash_lines else []):
        tm = re.match(r"\w+ \d+ (\d+:\d+:\d+)", cl)
        time_str = tm.group(1) if tm else ""
        msg = cl.split(":", 3)[-1].strip() if ":" in cl else cl
        recent_errors.append({"time": time_str, "level": "crash", "msg": msg[:120]})
    recent_errors.sort(key=lambda e: e.get("time", ""), reverse=True)
    recent_errors = recent_errors[:8]

    # 5. Email count
    email_pattern = re.compile(r"Sent alert|SMTP|email sent", re.IGNORECASE)
    email_count = sum(1 for l in lines if email_pattern.search(l))

    # 6. Cycle count
    cycle_pattern = re.compile(
        r"\[Dray Import\]|\[Dray Export\]|\[FTL\]|\[Boviet\]|\[Tolead\]|Tab:|Checking |Starting cycle|--- Cycle"
    )
    cycle_count = sum(1 for l in lines if cycle_pattern.search(l))

    # 7. Loads tracked
    loads_pattern = re.compile(r"Tracking|Scraping|Container:|Row \d+:")
    loads_count = sum(1 for l in lines if loads_pattern.search(l))

    # 8. Last successful cycle timestamp
    last_cycle_ts = None
    cycle_end_pattern = re.compile(r"Run complete|poll complete|Done|Sleeping")
    for l in reversed(lines):
        if cycle_end_pattern.search(l):
            tm = re.match(r"(\w+ \d+ \d+:\d+:\d+)", l)
            if tm:
                try:
                    from datetime import datetime as _dt
                    last_cycle_ts = _dt.strptime(f"2026 {tm.group(1)}", "%Y %b %d %H:%M:%S").isoformat()
                except Exception:
                    pass
            break

    # 9. Health status — based on real crashes, not operational errors
    if active_state not in ("active", "activating"):
        health = "down"
    elif crash_count > 10:
        health = "crash_loop"
    elif crash_count > 3:
        health = "degraded"
    else:
        health = "healthy"

    # 10. Uptime
    uptime_str = ""
    try:
        r = _sp.run(
            ["systemctl", "show", unit, "--property=ActiveEnterTimestamp"],
            capture_output=True, text=True, timeout=5,
        )
        ts_line = r.stdout.strip()
        if "=" in ts_line:
            ts_val = ts_line.split("=", 1)[1].strip()
            if ts_val:
                from datetime import datetime as _dt
                try:
                    started = _dt.strptime(ts_val, "%a %Y-%m-%d %H:%M:%S %Z")
                    delta = _dt.now() - started
                    hours = int(delta.total_seconds() // 3600)
                    mins = int((delta.total_seconds() % 3600) // 60)
                    if hours >= 24:
                        uptime_str = f"{hours // 24}d {hours % 24}h"
                    elif hours > 0:
                        uptime_str = f"{hours}h {mins}m"
                    else:
                        uptime_str = f"{mins}m"
                except Exception:
                    pass
    except Exception:
        pass

    # 11. Last run / next run
    last_run = ""
    next_run = ""
    for l in reversed(lines):
        tm = re.match(r"(\w+ \d+ \d+:\d+:\d+)", l)
        if tm:
            try:
                from datetime import datetime as _dt
                ts = _dt.strptime(f"2026 {tm.group(1)}", "%Y %b %d %H:%M:%S")
                mins = int((_dt.now() - ts).total_seconds() / 60)
                last_run = "just now" if mins < 1 else (f"{mins}m ago" if mins < 60 else f"{mins // 60}h {mins % 60}m ago")
                if poll_min > 0:
                    nm = poll_min - mins
                    next_run = "overdue" if nm < 0 else (f"{nm} min" if nm < 60 else f"{nm // 60}h {nm % 60}m")
            except Exception:
                pass
            break

    return {
        "unit": unit,
        "name": name,
        "active_state": active_state,
        "health": health,
        "uptime": uptime_str,
        "poll_min": poll_min,
        "crashes_24h": crash_count,
        "emails_24h": email_count,
        "cycles_24h": cycle_count,
        "loads_24h": loads_count,
        "last_run": last_run or "unknown",
        "next_run": next_run,
        "last_successful_cycle": last_cycle_ts,
        "recent_errors": recent_errors,
        "journal_24h": {
            "crashes": crash_count,
            "emails_sent": email_count,
            "cycles_completed": cycle_count,
            "loads_tracked": loads_count,
        },
    }



@app.get("/api/bot-health")
def api_bot_health():
    """Deep health check for all bot services - 24h window."""
    services = {}
    total_crashes = 0
    total_emails = 0
    total_cycles = 0
    healthy_count = 0

    for svc in BOT_SERVICES:
        info = _analyze_service_health(svc["unit"], svc["name"], svc["poll_min"])
        services[svc["unit"]] = info
        total_crashes += info["crashes_24h"]
        total_emails += info["emails_24h"]
        total_cycles += info["cycles_24h"]
        if info["health"] == "healthy":
            healthy_count += 1

    summary = {
        "total_crashes_24h": total_crashes,
        "total_emails_24h": total_emails,
        "total_cycles_24h": total_cycles,
        "services_healthy": healthy_count,
        "services_total": len(services),
    }

    return {"services": services, "summary": summary, "generated_at": __import__("datetime").datetime.now().isoformat()}




@app.get("/api/health")
async def api_health():
    """Comprehensive health check: PG, sheets, services, crons."""
    import subprocess, datetime
    checks = {}
    overall = "healthy"

    # 1. Postgres check
    try:
        with db.get_cursor() as cur:
            cur.execute("SELECT COUNT(*) as total, COUNT(*) FILTER (WHERE archived=false) as active FROM shipments")
            row = cur.fetchone()
            pg_total = row["total"] if isinstance(row, dict) else row[0]
            pg_active = row["active"] if isinstance(row, dict) else row[1]
        checks["postgres"] = {"status": "ok", "total_shipments": pg_total, "active_shipments": pg_active}
    except Exception as e:
        checks["postgres"] = {"status": "error", "error": str(e)}
        overall = "degraded"

    # 2. Google Sheets cache check
    try:
        age = _time.time() - sheet_cache._last
        sheet_count = len(sheet_cache.shipments)
        checks["sheets_cache"] = {
            "status": "ok" if age < 900 else "stale",
            "shipment_count": sheet_count,
            "cache_age_seconds": round(age, 1),
        }
        if age > 900:
            overall = "degraded"
    except Exception as e:
        checks["sheets_cache"] = {"status": "error", "error": str(e)}
        overall = "degraded"

    # 3. Systemd services check
    svc_names = ["csl-dashboard", "csl-boviet", "csl-tolead", "csl-inbox"]
    svc_status = {}
    for svc in svc_names:
        try:
            r = subprocess.run(["systemctl", "is-active", svc], capture_output=True, text=True, timeout=5)
            st = r.stdout.strip()
            svc_status[svc] = st
            if st != "active":
                overall = "degraded"
        except Exception:
            svc_status[svc] = "unknown"
    checks["services"] = svc_status

    # 4. Cron check (most recent logs)
    cron_status = {}
    cron_files = {
        "dray_import": "/tmp/csl_import.log",
        "dray_export": "/tmp/export_monitor.log",
        "ftl_monitor": "/tmp/ftl_monitor.log",
        "sheet_sync": "/tmp/sheet_sync.log",
    }
    for name, logfile in cron_files.items():
        try:
            r = subprocess.run(["tail", "-3", logfile], capture_output=True, text=True, timeout=5)
            lines = r.stdout.strip().split("\n")
            has_error = any("error" in l.lower() or "traceback" in l.lower() for l in lines)
            cron_status[name] = "error" if has_error else "ok"
            if has_error:
                overall = "degraded"
        except Exception:
            cron_status[name] = "unknown"
    checks["cron_jobs"] = cron_status

    # 5. Disk check
    try:
        r = subprocess.run(["df", "-h", "/"], capture_output=True, text=True, timeout=5)
        lines = r.stdout.strip().split("\n")
        if len(lines) >= 2:
            parts = lines[1].split()
            checks["disk"] = {"status": "ok", "used": parts[2], "available": parts[3], "use_pct": parts[4]}
            pct = int(parts[4].replace("%", ""))
            if pct > 90:
                overall = "critical"
                checks["disk"]["status"] = "critical"
    except Exception:
        checks["disk"] = {"status": "unknown"}

    return {
        "overall": overall,
        "checks": checks,
        "timestamp": datetime.datetime.now().isoformat(),
    }

@app.get("/api/cron-status")
def api_cron_status():
    """Status of cron-based monitors (dray import/export)."""
    import sys
    if "/root/csl-bot" not in sys.path:
        sys.path.insert(0, "/root/csl-bot")
    try:
        from cron_log_parser import get_all_cron_status
        return {"cron_jobs": get_all_cron_status()}
    except Exception as exc:
        return {"cron_jobs": {}, "error": str(exc)}


def main():

    logging.basicConfig(
        level=getattr(logging, config.LOG_LEVEL),
        format="%(asctime)s [%(name)s] %(levelname)s: %(message)s",
        handlers=[
            logging.StreamHandler(sys.stdout),
            logging.FileHandler(config.LOG_FILE),
        ],
    )
    uvicorn.run(
        "app:app",
        host=config.DASHBOARD_HOST,
        port=config.DASHBOARD_PORT,
        log_level=config.LOG_LEVEL.lower(),
    )



# ═══════════════════════════════════════════════════════════════════════════
# V2 API ENDPOINTS — Postgres-backed (Phase 1 migration)
# Coexist with existing sheet-based /api/ endpoints.
# ═══════════════════════════════════════════════════════════════════════════

# Shared accounts that still need Google Sheet writes
_SHARED_SHEET_ACCOUNTS = {"Tolead", "Boviet"}


def _shipment_row_to_dict(row: dict) -> dict:
    """Convert a Postgres shipments row to the same JSON shape as sheet_cache."""
    return {
        "efj": row["efj"] or "",
        "move_type": row["move_type"] or "",
        "container": row["container"] or "",
        "bol": row["bol"] or "",
        "ssl": row["vessel"] or "",
        "carrier": row["carrier"] or "",
        "origin": row["origin"] or "",
        "destination": row["destination"] or "",
        "eta": row["eta"] or "",
        "lfd": row["lfd"] or "",
        "pickup": row["pickup_date"] or "",
        "delivery": row["delivery_date"] or "",
        "status": row["status"] or "",
        "notes": row["notes"] or "",
        "bot_alert": row["bot_notes"] or "",
        "return_port": row["return_date"] or "",
        "container_url": row["container_url"] or "",
        "rep": row["rep"] or "Unassigned",
        "account": row["account"] or "",
        "hub": row["hub"] or "",
        "driver": row["driver"] or "",
        "driver_phone": row["driver_phone"] or "",
        "carrier_email": row.get("carrier_email") or "",
        "trailer": row.get("trailer") or row.get("driver") or "",
        "driver_name": row.get("driver_name") or "",
        "source": row["source"] or "sheet",
        "archived": row.get("archived", False),
        "updated_at": row["updated_at"].isoformat() if row.get("updated_at") else "",
    }


@app.get("/api/v2/shipments")
async def api_v2_shipments(request: Request, account: str = None, status: str = None,
                            hub: str = None, rep: str = None, archived: bool = False):
    """Return shipments from Postgres, same shape as /api/shipments."""
    where_clauses = ["archived = %s"]
    params = [archived]

    if account:
        where_clauses.append("LOWER(account) = LOWER(%s)")
        params.append(account)
    if status:
        where_clauses.append("LOWER(status) = LOWER(%s)")
        params.append(status)
    if hub:
        where_clauses.append("LOWER(hub) = LOWER(%s)")
        params.append(hub)
    if rep:
        where_clauses.append("LOWER(rep) = LOWER(%s)")
        params.append(rep)

    where = " AND ".join(where_clauses)

    with db.get_cursor() as cur:
        cur.execute(
        f"SELECT * FROM shipments WHERE {where} ORDER BY created_at DESC",
        params,
        )
        rows = cur.fetchall()

    shipments = [_shipment_row_to_dict(r) for r in rows]

    # Enrich with invoiced status from DB
    try:
        invoiced_map = db.get_invoiced_map()
    except Exception:
        invoiced_map = {}
    for s in shipments:
        s["_invoiced"] = invoiced_map.get(s["efj"], False)

    # Enrich with email thread stats (count + max priority)
    try:
        with db.get_cursor() as cur:
            cur.execute("""
                SELECT efj, COUNT(*) as email_count, COALESCE(MAX(priority), 0) as email_max_priority
                FROM email_threads
                GROUP BY efj
            """)
            email_stats = {r["efj"]: r for r in cur.fetchall()}
        for s in shipments:
            es = email_stats.get(s["efj"])
            s["email_count"] = es["email_count"] if es else 0
            s["email_max_priority"] = es["email_max_priority"] if es else 0
    except Exception:
        for s in shipments:
            s["email_count"] = 0
            s["email_max_priority"] = 0

    # Enrich with mp_status + mp_display_status from tracking cache
    try:
        _tc = _read_tracking_cache()
        for s in shipments:
            _ce = _find_tracking_entry(_tc, s["efj"])
            if _ce:
                s["mp_status"] = _ce.get("status", "")
                _disp, _detail = _classify_mp_display_status(_ce, s)
                s["mp_display_status"] = _disp
                s["mp_display_detail"] = _detail
                s["mp_last_updated"] = _ce.get("last_event_at") or _ce.get("last_ping_at") or _ce.get("last_scraped") or ""
                if not s.get("container_url") and _ce.get("macropoint_url"):
                    s["container_url"] = _ce["macropoint_url"]
            else:
                s["mp_display_status"] = ""
                s["mp_display_detail"] = ""
                s["mp_last_updated"] = ""
    except Exception:
        pass

    
    # Enrich with driver_contacts (carrier_email, trailer_number, driver_name)
    try:
        with db.get_cursor() as cur:
            cur.execute("SELECT efj, carrier_email, trailer_number, driver_name, driver_phone FROM driver_contacts")
            dc_rows = cur.fetchall()
        dc_map = {r["efj"]: r for r in dc_rows}
        for s in shipments:
            dc = dc_map.get(s["efj"])
            if dc:
                if dc.get("carrier_email") and not s.get("carrier_email"):
                    s["carrier_email"] = dc["carrier_email"]
                if dc.get("trailer_number"):
                    s["trailer"] = dc["trailer_number"]
                if dc.get("driver_name"):
                    s["driver_name"] = dc["driver_name"]
                if dc.get("driver_phone") and not s.get("driver_phone"):
                    s["driver_phone"] = dc["driver_phone"]
    except Exception as e:
        import logging
        logging.getLogger("csl").warning("driver_contacts enrichment failed: %s", e)

    return {"shipments": shipments, "total": len(shipments)}


@app.get("/api/v2/shipments/{efj}")
async def api_v2_shipment_detail(efj: str, request: Request):
    """Return a single shipment from Postgres."""
    with db.get_cursor() as cur:
        cur.execute("SELECT * FROM shipments WHERE efj = %s", (efj,))
        row = cur.fetchone()
    if not row:
        raise HTTPException(404, f"Shipment {efj} not found")
    return _shipment_row_to_dict(row)


@app.get("/api/v2/stats")
async def api_v2_stats(request: Request):
    """Dashboard stats computed from Postgres."""
    from datetime import datetime as _dt, timedelta as _td
    today = _dt.now().strftime("%Y-%m-%d")
    tomorrow = (_dt.now() + _td(days=1)).strftime("%Y-%m-%d")

    with db.get_cursor() as cur:
        # Active count (not archived, not delivered/completed)
        cur.execute("""
            SELECT COUNT(*) as cnt FROM shipments
            WHERE archived = FALSE
              AND LOWER(status) NOT IN ('delivered', 'completed', 'empty returned', 'billed_closed')
        """)
        active = cur.fetchone()["cnt"]

        # At risk (LFD is today or tomorrow)
        cur.execute("""
            SELECT COUNT(*) as cnt FROM shipments
            WHERE archived = FALSE
              AND LOWER(status) NOT IN ('delivered', 'completed', 'empty returned', 'billed_closed')
              AND lfd != '' AND lfd IS NOT NULL
              AND LEFT(lfd, 10) <= %s
        """, (tomorrow,))
        at_risk = cur.fetchone()["cnt"]

        # Completed today
        cur.execute("""
            SELECT COUNT(*) as cnt FROM shipments
            WHERE archived = FALSE
              AND LOWER(status) IN ('delivered', 'completed')
              AND delivery_date LIKE %s
        """, (f"%{today}%",))
        completed_today = cur.fetchone()["cnt"]

        # ETA changed (bot_notes mentions today)
        cur.execute("""
            SELECT COUNT(*) as cnt FROM shipments
            WHERE archived = FALSE
              AND bot_notes LIKE %s
        """, (f"%{today}%",))
        eta_changed = cur.fetchone()["cnt"]

    on_schedule = max(0, active - at_risk)

    return {
        "active": active,
        "on_schedule": on_schedule,
        "eta_changed": eta_changed,
        "at_risk": at_risk,
        "completed_today": completed_today,
    }


@app.get("/api/v2/accounts")
async def api_v2_accounts(request: Request):
    from datetime import datetime as _dt, timedelta as _td
    """Account list with counts from Postgres."""
    with db.get_cursor() as cur:
        cur.execute("""
        SELECT
            account,
            COUNT(*) FILTER (WHERE LOWER(status) NOT IN ('delivered','completed','empty returned','billed_closed') AND archived = FALSE) as active,
            COUNT(*) FILTER (WHERE LOWER(status) IN ('delivered','completed','empty returned','billed_closed') AND archived = FALSE) as done,
            COUNT(*) FILTER (
                WHERE archived = FALSE
                  AND LOWER(status) NOT IN ('delivered','completed','empty returned','billed_closed')
                  AND lfd != '' AND lfd IS NOT NULL
                  AND LEFT(lfd, 10) <= %s
            ) as alerts
        FROM shipments
        WHERE archived = FALSE
        GROUP BY account
        ORDER BY active DESC
        """, ((_dt.now() + _td(days=1)).strftime("%Y-%m-%d"),))
        rows = cur.fetchall()

    accounts = [{"name": r["account"], "active": r["active"], "done": r["done"], "alerts": r["alerts"]} for r in rows]
    return {"accounts": accounts}


@app.get("/api/v2/team")
async def api_v2_team(request: Request):
    """Team member summaries from Postgres."""
    with db.get_cursor() as cur:
        cur.execute("""
        SELECT rep,
               COUNT(*) FILTER (WHERE LOWER(status) NOT IN ('delivered','completed','empty returned','billed_closed') AND archived = FALSE) as loads,
               array_agg(DISTINCT account) FILTER (WHERE LOWER(status) NOT IN ('delivered','completed','empty returned','billed_closed') AND archived = FALSE) as accounts,
               COUNT(*) FILTER (
                   WHERE archived = FALSE
                     AND LOWER(status) NOT IN ('delivered','completed','empty returned','billed_closed')
                     AND lfd != '' AND lfd IS NOT NULL
                     AND LEFT(lfd, 10) <= to_char(NOW() + interval '1 day', 'YYYY-MM-DD')
               ) as at_risk
        FROM shipments
        WHERE archived = FALSE AND rep IS NOT NULL AND rep != ''
        GROUP BY rep
        """)
        rows = cur.fetchall()
    team = {}
    for r in rows:
        accts = r["accounts"] if r["accounts"] else []
        accts = [a for a in accts if a]
        team[r["rep"]] = {"loads": r["loads"], "accounts": sorted(accts), "at_risk": r["at_risk"]}
    return {"team": team}


@app.post("/api/v2/load/{efj}/status")
async def api_v2_update_status(efj: str, request: Request):
    """Update status in Postgres. Write back to Google Sheet if shared account."""
    body = await request.json()
    new_status = body.get("status", "").strip()
    if not new_status:
        raise HTTPException(400, "Missing status")

    # Update Postgres
    with db.get_conn() as conn:
        with db.get_cursor(conn) as cur:
            cur.execute(
                "UPDATE shipments SET status = %s, updated_at = NOW() WHERE efj = %s RETURNING account, hub",
                (new_status, efj),
            )
            row = cur.fetchone()
            if not row:
                raise HTTPException(404, f"Shipment {efj} not found")

    account = row["account"]

    # Write back to Google Sheet for shared accounts
    if account in _SHARED_SHEET_ACCOUNTS:
        try:
            _v2_write_status_to_sheet(efj, new_status, account, row.get("hub"))
        except Exception as e:
            log.warning("Sheet write-back failed for %s: %s (Postgres updated OK)", efj, e)

    return {"ok": True, "efj": efj, "status": new_status}


def _v2_write_status_to_sheet(efj: str, new_status: str, account: str, hub: str = None):
    """Write status back to Google Sheet for shared accounts (Tolead/Boviet)."""
    creds = Credentials.from_service_account_file(
        CREDS_FILE, scopes=["https://www.googleapis.com/auth/spreadsheets"]
    )
    gc = gspread.authorize(creds)

    if account == "Tolead" and hub and hub in TOLEAD_HUB_CONFIGS:
        cfg = TOLEAD_HUB_CONFIGS[hub]
        sh = gc.open_by_key(cfg["sheet_id"])
        ws = sh.worksheet(cfg["tab"])
        rows = ws.get_all_values()
        cols = cfg["cols"]
        for i, row in enumerate(rows):
            efj_val = row[cols["efj"]].strip() if len(row) > cols["efj"] else ""
            load_val = row[cols["load_id"]].strip() if len(row) > cols["load_id"] else ""
            if efj_val == efj or load_val == efj:
                ws.update_cell(i + 1, cols["status"] + 1, new_status)
                log.info("Sheet write-back: Tolead %s %s → %s (row %d)", hub, efj, new_status, i + 1)
                return
        log.warning("Sheet write-back: %s not found in Tolead %s", efj, hub)

    elif account == "Boviet":
        sh = gc.open_by_key(BOVIET_SHEET_ID)
        for bov_tab, cfg in BOVIET_TAB_CONFIGS.items():
            try:
                ws = sh.worksheet(bov_tab)
                rows = ws.get_all_values()
                for i, row in enumerate(rows):
                    if len(row) > cfg["efj_col"] and row[cfg["efj_col"]].strip() == efj:
                        ws.update_cell(i + 1, cfg["status_col"] + 1, new_status)
                        log.info("Sheet write-back: Boviet/%s %s → %s (row %d)", bov_tab, efj, new_status, i + 1)
                        return
            except Exception:
                continue
        log.warning("Sheet write-back: %s not found in Boviet tabs", efj)


@app.post("/api/v2/load/{efj}/update")
async def api_v2_update_field(efj: str, request: Request, background_tasks: BackgroundTasks):
    """Update any field(s) on a shipment in Postgres + write back to Master Sheet."""
    body = await request.json()

    # Allowed fields to update
    ALLOWED = {
        "move_type", "container", "bol", "vessel", "carrier",
        "origin", "destination", "eta", "lfd", "pickup_date", "delivery_date",
        "status", "notes", "driver", "bot_notes", "return_date",
        "rep", "customer_ref", "equipment_type", "container_url",
        "driver_phone", "hub", "archived", "customer_rate",
    }
    updates = {k: v for k, v in body.items() if k in ALLOWED}
    if not updates:
        raise HTTPException(400, "No valid fields to update")

    set_clauses = [f"{k} = %({k})s" for k in updates]
    set_clauses.append("updated_at = NOW()")
    updates["efj"] = efj

    with db.get_conn() as conn:
        with db.get_cursor(conn) as cur:
            cur.execute(
                f"UPDATE shipments SET {', '.join(set_clauses)} WHERE efj = %(efj)s RETURNING *",
                updates,
            )
            row = cur.fetchone()
            if not row:
                raise HTTPException(404, f"Shipment {efj} not found")

    # Fire-and-forget: write changed fields back to Master Sheet
    account = row["account"] or ""
    if account and account not in _SHARED_SHEET_ACCOUNTS:
        sheet_fields = {k: v for k, v in body.items() if k in ALLOWED and k != "archived"}
        if sheet_fields:
            background_tasks.add_task(_write_fields_to_master_sheet, efj, account, sheet_fields)

    return {"ok": True, "shipment": _shipment_row_to_dict(row)}


@app.post("/api/v2/load/add")
async def api_v2_add_shipment(request: Request, background_tasks: BackgroundTasks):
    """Insert a new shipment into Postgres. Write to Google Sheet."""
    body = await request.json()
    efj = body.get("efj", "").strip()
    account = body.get("account", "").strip()
    if not efj:
        raise HTTPException(400, "Missing EFJ #")
    if not account:
        raise HTTPException(400, "Missing account")

    with db.get_conn() as conn:
        with db.get_cursor(conn) as cur:
            cur.execute("""
                INSERT INTO shipments (
                    efj, move_type, container, bol, vessel, carrier,
                    origin, destination, eta, lfd, pickup_date, delivery_date,
                    status, notes, driver, bot_notes, return_date,
                    account, hub, rep, source
                ) VALUES (
                    %(efj)s, %(move_type)s, %(container)s, %(bol)s, %(vessel)s, %(carrier)s,
                    %(origin)s, %(destination)s, %(eta)s, %(lfd)s, %(pickup_date)s, %(delivery_date)s,
                    %(status)s, %(notes)s, %(driver)s, %(bot_notes)s, %(return_date)s,
                    %(account)s, %(hub)s, %(rep)s, 'dashboard'
                )
                ON CONFLICT (efj) DO NOTHING
                RETURNING *
            """, {
                "efj": efj,
                "move_type": body.get("move_type", ""),
                "container": body.get("container", ""),
                "bol": body.get("bol", ""),
                "vessel": body.get("vessel", ""),
                "carrier": body.get("carrier", ""),
                "origin": body.get("origin", ""),
                "destination": body.get("destination", ""),
                "eta": body.get("eta", ""),
                "lfd": body.get("lfd", ""),
                "pickup_date": body.get("pickup_date", ""),
                "delivery_date": body.get("delivery_date", ""),
                "status": body.get("status", ""),
                "notes": body.get("notes", ""),
                "driver": body.get("driver", ""),
                "bot_notes": body.get("bot_notes", ""),
                "return_date": body.get("return_date", ""),
                "account": account,
                "hub": body.get("hub", ""),
                "rep": body.get("rep", "Unassigned"),
            })
            row = cur.fetchone()

    if not row:
        raise HTTPException(409, f"Shipment {efj} already exists")

    # Fire-and-forget: add new row to Master Sheet
    if account and account not in _SHARED_SHEET_ACCOUNTS:
        background_tasks.add_task(_add_load_to_master_sheet, efj, account, body)

    return {"ok": True, "shipment": _shipment_row_to_dict(row)}




# ═══ Tracking Events Persistence ═══════════════════════════════════════════

def _persist_tracking_event(efj: str, load_ref: str, event_code: str, stop_name: str,
                            stop_type: str, status_mapped: str, lat: str, lon: str,
                            city: str, state: str, event_time: str, mp_order_id: str,
                            raw_params: dict = None):
    """Persist a webhook event to tracking_events PG table. Fire-and-forget."""
    try:
        import json as _json
        # Fix "ET" timestamps — PG doesn't recognize "ET" as a timezone.
        # Convert "2026-03-09 10:33:18 ET" → "2026-03-09 10:33:18-04:00" (or -05:00 for EST)
        _et = event_time
        if _et and isinstance(_et, str) and _et.rstrip().endswith(" ET"):
            from zoneinfo import ZoneInfo
            from datetime import datetime as _dt
            try:
                _naive_str = _et.rstrip()[:-3]  # strip " ET"
                _naive = _dt.strptime(_naive_str, "%Y-%m-%d %H:%M:%S")
                _aware = _naive.replace(tzinfo=ZoneInfo("America/New_York"))
                _et = _aware.isoformat()
            except Exception:
                pass  # fall through with original value
        with db.get_conn() as conn:
            with db.get_cursor(conn) as cur:
                cur.execute("""
                    INSERT INTO tracking_events (
                        efj, load_ref, event_code, event_type, stop_name, stop_type,
                        status_mapped, lat, lon, city, state, event_time,
                        mp_order_id, raw_params
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (efj, load_ref, event_code, event_code, stop_name, stop_type,
                      status_mapped, lat, lon, city, state, _et or None,
                      mp_order_id, _json.dumps(raw_params) if raw_params else None))
    except Exception as e:
        log.warning(f"Failed to persist tracking event for {efj}: {e}")


def _build_timeline_from_pg(efj: str) -> list:
    """Build a chronological timeline from tracking_events PG table.

    Uses stop_name + shipment origin/destination to determine pickup vs delivery
    instead of counting X1 occurrences (which fails when pickup events are missing).
    """
    try:
        with db.get_cursor() as cur:
            cur.execute("""
                SELECT event_code, status_mapped, stop_name, stop_type,
                       event_time, city, state
                FROM tracking_events
                WHERE efj = %s AND event_code IN ('AF', 'X1', 'X2', 'X4', 'X6', 'D1')
                ORDER BY event_time ASC
            """, (efj,))
            rows = cur.fetchall()
    except Exception:
        return []

    if not rows:
        return []

    # Get shipment origin/destination for stop matching
    origin = ""
    destination = ""
    try:
        with db.get_cursor() as cur:
            cur.execute("SELECT origin, destination FROM shipments WHERE efj = %s", (efj,))
            ship = cur.fetchone()
            if ship:
                origin = (ship["origin"] or "").upper()
                destination = (ship["destination"] or "").upper()
    except Exception:
        pass

    def _is_pickup_stop(stop_name, stop_type, city, state):
        """Determine if an event is at the pickup stop."""
        sn = (stop_name or "").upper()
        st = (stop_type or "").upper()
        c = (city or "").upper()
        s = (state or "").upper()
        # Explicit stop type from Macropoint
        if st in ("PICKUP", "ORIGIN"):
            return True
        if st in ("DROPOFF", "DELIVERY", "DESTINATION"):
            return False
        # Match against shipment origin/destination
        if origin and (origin in sn or c in origin or sn in origin):
            return True
        if destination and (destination in sn or c in destination or sn in destination):
            return False
        # Fallback: if city matches destination, it's delivery
        if destination and c and c in destination:
            return False
        if origin and c and c in origin:
            return True
        return None  # Unknown

    timeline = []
    for row in rows:
        code = row["event_code"]
        stop = row["stop_name"] or ""
        stop_type = row["stop_type"] or ""
        city = row["city"] or ""
        state = row["state"] or ""
        location = f"{city}, {state}".strip(", ") if city or state else stop
        event_time = row["event_time"].isoformat() if row["event_time"] else ""

        if code == "AF":
            timeline.append({
                "event": "Tracking Started",
                "time": event_time,
                "type": "info",
                "location": location,
            })
        elif code == "X1":
            is_pickup = _is_pickup_stop(stop, stop_type, city, state)
            if is_pickup:
                timeline.append({
                    "event": "Arrived at Pickup",
                    "time": event_time,
                    "type": "arrived",
                    "location": location,
                })
            else:
                timeline.append({
                    "event": "Arrived at Delivery",
                    "time": event_time,
                    "type": "arrived",
                    "location": location,
                })
        elif code == "X2":
            is_pickup = _is_pickup_stop(stop, stop_type, city, state)
            if is_pickup:
                timeline.append({
                    "event": "Departed Pickup",
                    "time": event_time,
                    "type": "departed",
                    "location": location,
                })
            else:
                timeline.append({
                    "event": "Departed Delivery",
                    "time": event_time,
                    "type": "departed",
                    "location": location,
                })
        elif code == "X4":
            timeline.append({
                "event": "Departed Pickup - En Route",
                "time": event_time,
                "type": "departed",
                "location": location,
            })
        elif code in ("X6", "D1"):
            timeline.append({
                "event": "Delivered",
                "time": event_time,
                "type": "delivered",
                "location": location,
            })

    return timeline

# ═══ Sheet Write-Back Helpers (fire-and-forget) ═══════════════════════════

def _write_fields_to_master_sheet(efj: str, account: str, fields: dict):
    """Background task: write dashboard field edits back to Master Sheet."""
    try:
        from csl_sheet_writer import sheet_update_field
        sheet_update_field(efj, account, fields)
    except Exception as e:
        log.warning("Sheet write-back failed for %s [%s]: %s (PG update succeeded)", efj, account, e)


def _add_load_to_master_sheet(efj: str, account: str, body: dict):
    """Background task: append new dashboard-created load to Master Sheet."""
    try:
        from csl_sheet_writer import sheet_add_row
        sheet_add_row(efj, account, body)
    except Exception as e:
        log.warning("Sheet add-row failed for %s [%s]: %s (PG insert succeeded)", efj, account, e)


# ═══ Macropoint Webhook (migrated from standalone webhook.py) ═══════════

import hmac as _hmac

# ── Webhook → Postgres direct write ──────────────────────────────────────
_WEBHOOK_STATUS_TO_DROPDOWN = {
    "Driver Arrived at Pickup": "At Pickup",
    "Departed Pickup - En Route": "In Transit",
    "Arrived at Delivery": "At Delivery",
    "Departed Delivery": "Departed Delivery",
    "Running Late": "Running Behind",
    "Delivered": "Delivered",
    "Driver Phone Unresponsive": "Driver Phone Unresponsive",
}

def _webhook_pg_write(load_ref: str, mapped_status: str, payload: dict):
    """Write FTL status update directly to Postgres from webhook."""
    try:
        dropdown = _WEBHOOK_STATUS_TO_DROPDOWN.get(mapped_status)
        if not dropdown:
            return  # Status not mappable to dropdown — skip PG write

        efj_clean = load_ref.replace("EFJ", "").strip()
        # Try both formats: "EFJ107230" and "107230"
        efj_key = f"EFJ{efj_clean}" if not load_ref.startswith("EFJ") else load_ref

        # Extract stop times from payload
        event_upper = (payload.get("eventType") or payload.get("event_type") or "").upper()
        event_time = payload.get("eventTime") or payload.get("timestamp") or ""

        kwargs = {"status": dropdown}
        if "ARRIVED" in event_upper and "PICKUP" in event_upper and event_time:
            kwargs["pickup_date"] = event_time
        elif "ARRIVED" in event_upper and "DELIVERY" in event_upper and event_time:
            kwargs["delivery_date"] = event_time

        # Use pg_update_shipment from csl_pg_writer
        import sys
        if "/root/csl-bot" not in sys.path:
            sys.path.insert(0, "/root/csl-bot")
        from csl_pg_writer import pg_update_shipment


        pg_update_shipment(efj_key, **kwargs)
        log.info(f"Webhook PG write: {efj_key} -> {dropdown}")
    except Exception as exc:
        log.warning(f"Webhook PG write failed (non-fatal): {exc}")



# Real-time webhook email alerts
import sys as _sys
if "/root/csl-bot" not in _sys.path:
    _sys.path.insert(0, "/root/csl-bot")
from csl_ftl_alerts import send_webhook_alert, STATUS_TO_DROPDOWN as _ALERT_STATUS_MAP

# ── Status hierarchy: higher number = further along in lifecycle ──
_STATUS_RANK = {
    "Tracking Started": 1,
    "Tracking Waiting for Update": 2,
    "Driver Phone Unresponsive": 2,
    "Driver Arrived at Pickup": 3,
    "At Pickup": 3,
    "Departed Pickup - En Route": 4,
    "In Transit": 4,
    "Running Late": 4,
    "Tracking Behind Schedule": 4,
    "Arrived at Delivery": 5,
    "At Delivery": 5,
    "Departed Delivery": 6,
    "Delivered": 7,
    "Tracking Completed Successfully": 8,
}


# Statuses that are "done" — ignore all further webhook/monitor events
_TERMINAL_STATUSES = {"Delivered", "Tracking Completed Successfully", "Billed/Closed", "billed_closed"}

def _status_is_regression(old_status, new_status):
    """Return True if new_status is a regression from old_status."""
    old_rank = _STATUS_RANK.get(old_status, 0)
    new_rank = _STATUS_RANK.get(new_status, 0)
    return new_rank > 0 and old_rank > 0 and new_rank < old_rank


_WEBHOOK_STATUS_MAP = {
    "ARRIVED_PICKUP": "Driver Arrived at Pickup",
    "DEPARTED_PICKUP": "Departed Pickup - En Route",
    "IN_TRANSIT": "Departed Pickup - En Route",
    "ARRIVED_DELIVERY": "Arrived at Delivery",
    "DEPARTED_DELIVERY": "Departed Delivery",
    "DELIVERED": "Delivered",
    "TRACKING_STARTED": "Tracking Started",
    "DRIVER_UNRESPONSIVE": "Driver Phone Unresponsive",
    "RUNNING_LATE": "Running Late",
    "CANT_MAKE_IT": "Can\'t Make It",
    "TRACKING_COMPLETED_SUCCESSFULLY": "Delivered",
    "TRACKING_COMPLETED": "Delivered",
    "COMPLETED": "Delivered",
}

_WEBHOOK_LOG = "/root/csl-bot/webhook_payloads.log"
_WEBHOOK_EVENTS_LOG = "/root/csl-bot/webhook_events.log"
_WH_USER = os.getenv("WEBHOOK_AUTH_USERNAME", "")
_WH_PASS = os.getenv("WEBHOOK_AUTH_PASSWORD", "")


def _webhook_basic_auth(request: Request) -> bool:
    """Validate HTTP Basic Auth for Macropoint webhook."""
    auth_header = request.headers.get("authorization", "")
    if not auth_header.startswith("Basic "):
        return False
    import base64 as _b64
    try:
        decoded = _b64.b64decode(auth_header[6:]).decode()
        user, passwd = decoded.split(":", 1)
    except Exception:
        return False
    return (_hmac.compare_digest(user, _WH_USER)
            and _hmac.compare_digest(passwd, _WH_PASS))


def _update_tracking_cache_webhook(load_ref: str, status: str, now: str, payload: dict):
    """Update ftl_tracking_cache.json when a webhook event arrives."""
    try:
        with open(TRACKING_CACHE_FILE, "r") as f:
            cache = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return False, None
    efj_clean = load_ref.replace("EFJ", "").strip()
    matched_key = None
    for key, entry in cache.items():
        efj = entry.get("efj", "")
        load_num = entry.get("load_num", "")
        mp_load_id = entry.get("mp_load_id", "")
        if (efj == load_ref or efj == efj_clean or
                load_num == load_ref or mp_load_id == load_ref or
                key == efj_clean):
            matched_key = key
            break

    # Fallback: look up by container field in Postgres
    if not matched_key:
        try:
            with db.get_cursor() as cur:
                # Try exact match, then LIKE for partial (PO # 336840 vs PO336840)
                cur.execute(
                    "SELECT efj FROM shipments WHERE container = %s OR container LIKE %s LIMIT 1",
                    (load_ref, f"%{load_ref}%"),
                )
                row = cur.fetchone()
                if row:
                    pg_efj = row["efj"]
                    efj_num = pg_efj.replace("EFJ", "").strip()
                    # Check if this EFJ is in cache
                    if efj_num in cache:
                        matched_key = efj_num
                    elif pg_efj in cache:
                        matched_key = pg_efj
                    else:
                        # Create a new cache entry for this load
                        cache[efj_num] = {
                            "efj": pg_efj,
                            "load_num": load_ref,
                            "status": "",
                            "mp_load_id": load_ref,
                            "cant_make_it": None,
                            "stop_times": {},
                            "macropoint_url": "",
                            "last_scraped": "",
                            "driver_phone": "",
                        }
                        matched_key = efj_num
                        log.info(f"Webhook: created cache entry {efj_num} for container {load_ref}")
        except Exception as exc:
            log.debug(f"Webhook PG container lookup failed: {exc}")

    if not matched_key:
        return False, None
    entry = cache[matched_key]
    old_status = entry.get("status", "")
    # ── Status hierarchy guard: block regressions ──
    if _status_is_regression(old_status, status):
        import logging
        logging.getLogger("uvicorn.error").info(
            f"Webhook: BLOCKED status regression {matched_key} [{old_status}] -> [{status}]"
        )
        return False, None
    entry["status"] = status
    entry["last_scraped"] = now
    entry["webhook_updated"] = True
    entry["last_event_at"] = now

    # Auto-populate macropoint_url from MPOrderID if available
    mp_order_id = payload.get("mp_order_id") or payload.get("MPOrderID") or ""
    if mp_order_id and not entry.get("macropoint_url"):
        entry["macropoint_url"] = f"https://visibility.macropoint.com/shipments?l={mp_order_id}"

    stop_times = entry.get("stop_times", {})
    event_upper = (payload.get("eventType") or payload.get("event_type") or "").upper()
    event_time = payload.get("eventTime") or payload.get("timestamp") or now

    if "ARRIVED" in event_upper and "PICKUP" in event_upper:
        stop_times["stop1_arrived"] = event_time
    elif "DEPARTED" in event_upper and "PICKUP" in event_upper:
        stop_times["stop1_departed"] = event_time
    elif "ARRIVED" in event_upper and "DELIVERY" in event_upper:
        stop_times["stop2_arrived"] = event_time
    elif "DEPARTED" in event_upper and "DELIVERY" in event_upper:
        stop_times["stop2_departed"] = event_time
    elif "DELIVERED" in event_upper:
        stop_times["stop2_departed"] = event_time

    entry["stop_times"] = stop_times

    tmp_path = TRACKING_CACHE_FILE + ".tmp"
    with open(tmp_path, "w") as f:
        json.dump(cache, f, indent=2)
    os.replace(tmp_path, TRACKING_CACHE_FILE)

    log.info(f"Webhook cache updated: {matched_key} [{old_status}] -> [{status}]")
    _match_info = {
        "efj": entry.get("efj", ""),
        "load_num": entry.get("load_num", ""),
        "stop_times": entry.get("stop_times", {}),
        "mp_load_id": entry.get("mp_load_id", ""),
        "matched_key": matched_key,
    }
    return True, _match_info


@app.get("/webhook-test")
async def webhook_health():
    now = datetime.now(tz=__import__("zoneinfo").ZoneInfo("America/New_York")).strftime("%Y-%m-%d %H:%M:%S ET")
    return {"status": "ok", "service": "csl-webhook (integrated)", "time": now}




def _webhook_send_alert_background(efj: str, load_num: str, status: str,
                                    stop_times: dict, mp_load_id: str,
                                    matched_key: str):
    """Background task: resolve account from PG, write PG status, send email."""
    try:
        # Resolve account from Postgres
        account = ""
        dest = ""
        existing_status = ""
        try:
            with db.get_cursor() as cur:
                cur.execute(
                    "SELECT account, destination, status FROM shipments WHERE efj = %s",
                    (efj,)
                )
                row = cur.fetchone()
                if row:
                    account = row.get("account", "") or ""
                    dest = row.get("destination", "") or ""
                    existing_status = row.get("status", "") or ""
        except Exception as exc:
            log.warning(f"Webhook bg: PG lookup failed for {efj}: {exc}")

        if not account:
            log.warning(f"Webhook bg: no account found for {efj} — skipping email")
            return

        # Map webhook status to dropdown value for PG write
        dropdown = _ALERT_STATUS_MAP.get(status)

        # Write PG status update
        # ── Block status regression (skip PG write AND email) ──
        if _status_is_regression(existing_status, status):
            log.info(f"Webhook BG: BLOCKED regression {efj} [{existing_status}] -> [{status}] — skipping entirely")
            return
        # ── Skip if load is already in terminal status ──
        if existing_status in _TERMINAL_STATUSES:
            log.info(f"Webhook BG: SKIP {efj} — already in terminal status [{existing_status}]")
            return
        if dropdown and dropdown != existing_status:
            try:
                from csl_pg_writer import pg_update_shipment as _pg_up
                kwargs = {"status": dropdown}
                # Extract dates from stop_times
                if stop_times.get("stop1_arrived"):
                    kwargs["pickup_date"] = stop_times["stop1_arrived"]
                if stop_times.get("stop2_arrived") or stop_times.get("stop2_departed"):
                    kwargs["delivery_date"] = stop_times.get("stop2_departed") or stop_times.get("stop2_arrived")
                _pg_up(efj, **kwargs)
                log.info(f"Webhook bg: PG updated {efj} -> {dropdown}")
            except Exception as exc:
                log.warning(f"Webhook bg: PG write failed for {efj}: {exc}")

        # Write container_url to PG if available from cache
        try:
            with open(TRACKING_CACHE_FILE, "r") as _cf:
                _cache_tmp = json.load(_cf)
            _entry_tmp = _cache_tmp.get(matched_key, {})
            _mp_url = _entry_tmp.get("macropoint_url", "")
            if _mp_url:
                with db.get_cursor() as cur:
                    cur.execute(
                        "UPDATE shipments SET container_url = %s WHERE efj = %s AND (container_url IS NULL OR container_url = '')",
                        (_mp_url, efj)
                    )
                    log.info(f"Webhook bg: wrote container_url for {efj}")
        except Exception as _url_exc:
            log.debug(f"Webhook bg: container_url write skipped for {efj}: {_url_exc}")

        # Send real-time email alert (with dedup)
        sent = send_webhook_alert(
            efj=efj,
            load_num=load_num,
            status=status,
            account=account,
            stop_times=stop_times,
            mp_load_id=mp_load_id,
        )
        log.info(f"Webhook bg: alert {'sent' if sent else 'skipped (dedup)'} for {efj} [{status}]")

    except Exception as exc:
        log.error(f"Webhook bg: unhandled error for {efj}: {exc}")

@app.get("/macropoint-webhook")
async def macropoint_webhook_get(request: Request, background_tasks: BackgroundTasks):
    """Handle Macropoint native protocol callbacks (GET with query params)."""
    from urllib.parse import unquote
    params = dict(request.query_params)
    now = datetime.now(tz=__import__("zoneinfo").ZoneInfo("America/New_York")).strftime("%Y-%m-%d %H:%M:%S ET")

    load_ref = params.get("ID", "").strip()
    data_source = params.get("DataSource", "").strip()
    mp_order_id = params.get("MPOrderID", "")
    lat = params.get("Latitude", "")
    lon = params.get("Longitude", "")
    city = params.get("City", "")
    state = params.get("State", "")
    street = params.get("Street1", "")
    timestamp = params.get("LocationDateTimeUTC", "") or params.get("ApproxLocationDateTimeInLocalTime", "")

    # Log all incoming events
    event_record = {
        "time": now,
        "load_ref": load_ref,
        "data_source": data_source,
        "mp_order_id": mp_order_id,
        "params": {k: v for k, v in params.items() if k not in ("MPOrderID",)},
    }
    try:
        with open(_WEBHOOK_EVENTS_LOG, "a") as f:
            f.write(json.dumps(event_record) + "\n")
    except Exception:
        pass

    if not load_ref:
        return {"status": "ok", "processed": False, "reason": "no load ID"}

    # ── Location Pings ────────────────────────────────────────────────
    if data_source == "Ping" and lat and lon:
        # Update tracking cache with latest location
        try:
            with open(TRACKING_CACHE_FILE, "r") as f:
                cache = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            cache = {}

        # Find matching cache entry by load_ref (could be ORD1260305010 format)
        matched_key = None
        for key, entry in cache.items():
            efj = entry.get("efj", "")
            load_num = entry.get("load_num", "")
            mp_load_id = entry.get("mp_load_id", "")
            if (load_ref == efj or load_ref == load_num or
                    load_ref == mp_load_id or load_ref == key):
                matched_key = key
                break

        if matched_key:
            entry = cache[matched_key]
            entry["last_location"] = {
                "lat": lat, "lon": lon,
                "city": city, "state": state, "street": street,
                "timestamp": timestamp,
            }
            entry["last_scraped"] = now
            entry["last_ping_at"] = now
            if not entry.get("last_event_at"):
                entry["last_event_at"] = now
            tmp_path = TRACKING_CACHE_FILE + ".tmp"
            with open(tmp_path, "w") as f:
                json.dump(cache, f, indent=2)
            os.replace(tmp_path, TRACKING_CACHE_FILE)

        log.debug(f"Webhook ping: {load_ref} @ {city}, {state}")
        return {"status": "ok", "type": "location_ping", "load": load_ref}

    # ── Trip Events (Event field: X1, X2, X3, X6, AG, AF, etc.) ─────
    event_code = params.get("Event", "").strip()
    stop_name = params.get("Stop", "").strip()
    stop_type = params.get("StopType", "").strip()  # Pickup / DropOff

    # Macropoint event code → internal status mapping
    
    _MP_EVENT_MAP = {
        "AF": "Tracking Started",
        "X1": None,   # Arrived — need stop context
        "X2": None,   # Departed — need stop context
        "X3": None,   # Position near stop — skip (too noisy)
        "X4": "Departed Pickup - En Route",
        "X6": "Delivered",
        "D1": "Delivered",   # Delivery confirmation
        "AG": "Driver Phone Unresponsive",
    }

    if event_code:
        mapped = _MP_EVENT_MAP.get(event_code)

        # X1 (Arrived) / X2 (Departed) — need to infer pickup vs delivery
        if event_code == "X1":
            # If it's the first stop (pickup) or stop_type=Pickup
            if stop_type.lower() == "pickup" or "pickup" in stop_name.lower():
                mapped = "Driver Arrived at Pickup"
            else:
                mapped = "Arrived at Delivery"
        elif event_code == "X2":
            if stop_type.lower() == "pickup" or "pickup" in stop_name.lower():
                mapped = "Departed Pickup - En Route"
            else:
                mapped = "Departed Delivery"
        elif event_code == "X3":
            # Position update near stop — don't treat as status change
            log.debug(f"Webhook X3 (position near stop): {load_ref} @ {stop_name}")
            return {"status": "ok", "type": "position_near_stop", "load": load_ref}

        if mapped:
            # Map human-readable status back to structured event type for stop_times parsing
            _MAPPED_TO_EVENT_TYPE = {
                "Driver Arrived at Pickup": "ARRIVED_PICKUP",
                "Departed Pickup - En Route": "DEPARTED_PICKUP",
                "Arrived at Delivery": "ARRIVED_DELIVERY",
                "Departed Delivery": "DEPARTED_DELIVERY",
                "Delivered": "DELIVERED",
            }
            structured_event = _MAPPED_TO_EVENT_TYPE.get(mapped, event_code)

            cache_updated, _match_info = _update_tracking_cache_webhook(load_ref, mapped, now, {
                "loadNumber": load_ref,
                "eventType": structured_event,
                "event": event_code,
                "stop": stop_name,
                "timestamp": timestamp,
                "mp_order_id": mp_order_id,
            })
            if cache_updated and _match_info:
                background_tasks.add_task(
                    _webhook_send_alert_background,
                    efj=_match_info["efj"],
                    load_num=_match_info["load_num"],
                    status=mapped,
                    stop_times=_match_info["stop_times"],
                    mp_load_id=_match_info["mp_load_id"],
                    matched_key=_match_info["matched_key"],
                )
            # Persist event to PG for historical timeline
            _persist_efj = _match_info["efj"] if cache_updated and _match_info else None
            if _persist_efj:
                _persist_tracking_event(
                    efj=_persist_efj,
                    load_ref=load_ref,
                    event_code=event_code,
                    stop_name=stop_name,
                    stop_type=stop_type,
                    status_mapped=mapped,
                    lat=params.get("Latitude", ""),
                    lon=params.get("Longitude", ""),
                    city=params.get("City", ""),
                    state=params.get("State", ""),
                    event_time=timestamp,
                    mp_order_id=mp_order_id,
                    raw_params=dict(params),
                )
            log.info(f"Webhook trip event: {load_ref} [{event_code}] -> {mapped} (cache={cache_updated})")
            return {"status": "ok", "type": "trip_event", "load": load_ref, "event": event_code, "status_mapped": mapped}

    # ── Schedule Alerts (ScheduleAlertText field) ─────────────────────
    schedule_alert = params.get("ScheduleAlertText", "").strip()
    distance_str = params.get("DistanceToStopInMiles", "").strip()
    if schedule_alert or (stop_type and distance_str):
        # ── Infer stop arrival/departure from GPS proximity ──
        # Macropoint only sends X1 for delivery stops.  For pickup stops,
        # we infer arrival when DistanceToStopInMiles < 0.5 and departure
        # when distance increases after a recorded arrival.
        ARRIVAL_THRESHOLD = 0.5   # miles — "at the stop"
        DEPARTURE_THRESHOLD = 2.0  # miles — "left the stop"
        try:
            distance_mi = float(distance_str) if distance_str else None
        except (ValueError, TypeError):
            distance_mi = None

        if distance_mi is not None and stop_type:
            stop_key = "stop1" if stop_type.lower() in ("pickup", "origin") else "stop2"
            arrived_key = f"{stop_key}_arrived"
            departed_key = f"{stop_key}_departed"

            # Read current cache to check existing stop_times
            try:
                with open(TRACKING_CACHE_FILE, "r") as f:
                    _sc_cache = json.load(f)
            except (FileNotFoundError, json.JSONDecodeError):
                _sc_cache = {}

            _sc_matched = None
            for _sk, _sv in _sc_cache.items():
                _efj = _sv.get("efj", "")
                _ln = _sv.get("load_num", "")
                _ml = _sv.get("mp_load_id", "")
                if load_ref in (_efj, _ln, _ml, _sk):
                    _sc_matched = _sk
                    break

            if _sc_matched:
                _sc_entry = _sc_cache[_sc_matched]
                _sc_stops = _sc_entry.get("stop_times", {})
                _changed = False

                # Use the handler's 'now' timestamp (ET) — schedule alerts
                # don't carry ApproxLocationDateTimeInLocalTime (that's on Pings).
                # EtaToStop is the predicted arrival, not the current time.
                event_local_time = now

                # Arrival detection: close to stop and not yet recorded
                if distance_mi < ARRIVAL_THRESHOLD and not _sc_stops.get(arrived_key):
                    _sc_stops[arrived_key] = event_local_time
                    _changed = True
                    _inferred_event = "Arrived at Pickup" if stop_key == "stop1" else "Arrived at Delivery"
                    _inferred_code = "X1"
                    log.info(f"Webhook INFERRED {_inferred_event}: {load_ref} dist={distance_mi:.2f}mi at {stop_type}")

                    # Persist to PG tracking_events
                    _persist_efj = _sc_entry.get("efj")
                    if _persist_efj:
                        _persist_tracking_event(
                            efj=_persist_efj,
                            load_ref=load_ref,
                            event_code="X1",
                            stop_name=params.get("StopName", ""),
                            stop_type=stop_type,
                            status_mapped=_inferred_event,
                            lat=params.get("Latitude", ""),
                            lon=params.get("Longitude", ""),
                            city=params.get("City", ""),
                            state=params.get("State", ""),
                            event_time=event_local_time,
                            mp_order_id=mp_order_id,
                            raw_params=dict(params),
                        )

                # Departure detection: was at stop, now moved away
                if (distance_mi > DEPARTURE_THRESHOLD and
                        _sc_stops.get(arrived_key) and
                        not _sc_stops.get(departed_key)):
                    _sc_stops[departed_key] = event_local_time
                    _changed = True
                    _inferred_event = "Departed Pickup" if stop_key == "stop1" else "Departed Delivery"
                    log.info(f"Webhook INFERRED {_inferred_event}: {load_ref} dist={distance_mi:.2f}mi from {stop_type}")

                    # Persist to PG tracking_events
                    _persist_efj = _sc_entry.get("efj")
                    if _persist_efj:
                        _persist_tracking_event(
                            efj=_persist_efj,
                            load_ref=load_ref,
                            event_code="X2",
                            stop_name=params.get("StopName", ""),
                            stop_type=stop_type,
                            status_mapped=_inferred_event,
                            lat=params.get("Latitude", ""),
                            lon=params.get("Longitude", ""),
                            city=params.get("City", ""),
                            state=params.get("State", ""),
                            event_time=event_local_time,
                            mp_order_id=mp_order_id,
                            raw_params=dict(params),
                        )

                if _changed:
                    _sc_entry["stop_times"] = _sc_stops
                    tmp_path = TRACKING_CACHE_FILE + ".tmp"
                    with open(tmp_path, "w") as f:
                        json.dump(_sc_cache, f, indent=2)
                    os.replace(tmp_path, TRACKING_CACHE_FILE)

        # ── Store schedule alert intelligence in tracking cache ──
        if schedule_alert:
            try:
                with open(TRACKING_CACHE_FILE, "r") as f:
                    _sa_cache = json.load(f)
                _sa_key = None
                for _sak, _sav in _sa_cache.items():
                    if load_ref in (_sav.get("efj", ""), _sav.get("load_num", ""),
                                    _sav.get("mp_load_id", ""), _sak):
                        _sa_key = _sak
                        break
                if _sa_key:
                    _sa_entry = _sa_cache[_sa_key]
                    _old_alert = _sa_entry.get("schedule_alert", "")
                    if schedule_alert != _old_alert:
                        _sa_entry["schedule_alert"] = schedule_alert
                        _sa_entry["schedule_alert_code"] = params.get("ScheduleAlertCode", "")
                        _sa_entry["distance_to_stop"] = distance_str
                        _sa_entry["eta_to_stop"] = params.get("EtaToStop", "")
                        _sa_entry["schedule_stop_type"] = stop_type
                        tmp_path = TRACKING_CACHE_FILE + ".tmp"
                        with open(tmp_path, "w") as f:
                            json.dump(_sa_cache, f, indent=2)
                        os.replace(tmp_path, TRACKING_CACHE_FILE)
                        log.debug(f"Webhook: stored schedule alert for {_sa_key}: {schedule_alert}")
            except Exception as _sa_exc:
                log.debug(f"Webhook: schedule alert storage failed: {_sa_exc}")

        log.info(f"Webhook schedule alert: {load_ref} [{stop_type}] {schedule_alert} (dist={distance_str} mi)")
        return {"status": "ok", "type": "schedule_alert", "load": load_ref, "alert": schedule_alert}

    # ── Fallback: check other event fields ────────────────────────────
    event_type = (
        params.get("EventType", "") or params.get("Status", "") or
        params.get("OrderStatus", "") or data_source
    ).strip()

    if event_type and event_type != "Ping":
        mapped_status = _WEBHOOK_STATUS_MAP.get(event_type.upper().replace(" ", "_"), "")
        if not mapped_status:
            mapped_status = event_type
        if mapped_status:
            cache_updated, _match_info = _update_tracking_cache_webhook(load_ref, mapped_status, now, {
                "loadNumber": load_ref,
                "eventType": event_type.upper().replace(" ", "_"),
                "timestamp": timestamp,
                "mp_order_id": mp_order_id,
            })
            if cache_updated and _match_info:
                background_tasks.add_task(
                    _webhook_send_alert_background,
                    efj=_match_info["efj"],
                    load_num=_match_info["load_num"],
                    status=mapped_status,
                    stop_times=_match_info["stop_times"],
                    mp_load_id=_match_info["mp_load_id"],
                    matched_key=_match_info["matched_key"],
                )
            log.info(f"Webhook GET event: {load_ref} -> {mapped_status} (cache={cache_updated})")
            return {"status": "ok", "type": "status_event", "load": load_ref, "status_mapped": mapped_status}

    return {"status": "ok", "type": data_source or "unknown", "load": load_ref}


@app.post("/macropoint-webhook")
async def macropoint_webhook(request: Request, background_tasks: BackgroundTasks):
    # Basic Auth check
    if not _webhook_basic_auth(request):
        raise HTTPException(status_code=401, detail="Unauthorized",
                            headers={"WWW-Authenticate": 'Basic realm="csl-bot"'})

    payload = await request.json() if request.headers.get("content-type", "").startswith("application/json") else {}
    now = datetime.now(tz=__import__("zoneinfo").ZoneInfo("America/New_York")).strftime("%Y-%m-%d %H:%M:%S ET")

    # Raw payload log
    log_entry = f"\n{'=' * 60}\n[{now}]\n{json.dumps(payload, indent=2)}\n"
    try:
        with open(_WEBHOOK_LOG, "a") as f:
            f.write(log_entry)
    except Exception:
        pass

    log.info(f"Webhook received: {json.dumps(payload)[:200]}")

    # Skip test payloads
    if payload.get("test"):
        log.info("Test payload — skipping processing")
        return {"status": "ok", "processed": False, "reason": "test payload"}

    # Extract load identifier
    load_ref = (
        payload.get("loadNumber")
        or payload.get("pro")
        or payload.get("referenceNumber")
        or payload.get("load_number")
        or payload.get("shipmentId")
        or payload.get("reference_number")
        or ""
    ).strip()

    event_type = (
        payload.get("eventType")
        or payload.get("status")
        or payload.get("event_type")
        or payload.get("event")
        or ""
    ).strip()

    # Map event to internal status
    mapped_status = _WEBHOOK_STATUS_MAP.get(event_type.upper().replace(" ", "_"), "")
    if not mapped_status and event_type:
        mapped_status = event_type

    # Structured event log
    event_record = {
        "time": now,
        "load_ref": load_ref,
        "raw_event": event_type,
        "mapped_status": mapped_status,
        "payload_keys": list(payload.keys()),
    }
    try:
        with open(_WEBHOOK_EVENTS_LOG, "a") as f:
            f.write(json.dumps(event_record) + "\n")
    except Exception:
        pass

    # Update tracking cache
    cache_updated = False
    if load_ref and mapped_status:
        cache_updated, _match_info = _update_tracking_cache_webhook(load_ref, mapped_status, now, payload)

        # Persist event to PG for historical timeline
        if cache_updated and _match_info:
            _persist_tracking_event(
                efj=_match_info["efj"],
                load_ref=load_ref,
                event_code=payload.get("event", payload.get("eventCode", "")),
                stop_name=payload.get("stop", ""),
                stop_type=payload.get("stopType", ""),
                status_mapped=mapped_status,
                lat=payload.get("latitude", ""),
                lon=payload.get("longitude", ""),
                city=payload.get("city", ""),
                state=payload.get("state", ""),
                event_time=payload.get("eventTime", payload.get("timestamp", "")),
                mp_order_id=payload.get("MPOrderID", payload.get("mp_order_id", "")),
                raw_params=payload,
            )

        # Real-time: write PG + send email in background task
        if cache_updated and _match_info:
            background_tasks.add_task(
                _webhook_send_alert_background,
                efj=_match_info["efj"],
                load_num=_match_info["load_num"],
                status=mapped_status,
                stop_times=_match_info["stop_times"],
                mp_load_id=_match_info["mp_load_id"],
                matched_key=_match_info["matched_key"],
            )

    if not load_ref:
        log.warning(f"Webhook: unknown payload format — keys: {list(payload.keys())}")
        return {"status": "ok", "processed": False, "reason": "unknown format"}

    log.info(f"Webhook processed: {load_ref} -> {mapped_status or '(no status)'} (cache_updated={cache_updated})")
    return {
        "status": "ok",
        "processed": True,
        "load_ref": load_ref,
        "mapped_status": mapped_status,
        "cache_updated": cache_updated,
    }


# ═══ End of v2 endpoints ═══

if __name__ == "__main__":
    main()
