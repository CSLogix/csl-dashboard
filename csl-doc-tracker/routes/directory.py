import logging
import os
from decimal import Decimal
from fastapi import APIRouter, Query, Request, HTTPException, UploadFile, File
from fastapi.responses import JSONResponse
import database as db

log = logging.getLogger(__name__)
router = APIRouter()


def _serialize_row(row):
    """Convert a DB row dict to JSON-safe dict."""
    d = dict(row)
    for k in ("created_at", "updated_at", "effective_date", "quote_date", "indexed_at", "date_quoted"):
        if d.get(k) and hasattr(d[k], "isoformat"):
            d[k] = d[k].isoformat()
    # Decimal -> float
    for k, v in d.items():
        if hasattr(v, "as_tuple"):
            d[k] = float(v)
    return d


@router.get("/api/carriers")
async def api_list_carriers(
    search: str = Query(default=None),
    region: str = Query(default=None),
    market: str = Query(default=None),
    capability: str = Query(default=None),
    exclude_dnu: bool = Query(default=False),
    include_lanes: bool = Query(default=False),
):
    with db.get_cursor() as cur:
        clauses, params = [], []
        if search:
            clauses.append("(carrier_name ILIKE %s OR mc_number ILIKE %s OR contact_email ILIKE %s)")
            s = f"%{search}%"
            params.extend([s, s, s])
        if region:
            clauses.append("regions ILIKE %s")
            params.append(f"%{region}%")
        if market:
            clauses.append("%s = ANY(markets)")
            params.append(market)
        if exclude_dnu:
            clauses.append("(dnu IS NOT TRUE)")
        if capability:
            caps = [c.strip().lower() for c in capability.split(",") if c.strip()]
            cap_map = {
                "hazmat": "can_hazmat", "dray": "can_dray", "overweight": "can_overweight",
                "transload": "can_transload", "reefer": "can_reefer", "bonded": "can_bonded",
                "oog": "can_oog", "warehousing": "can_warehousing",
            }
            for cap in caps:
                col = cap_map.get(cap)
                if col:
                    clauses.append(f"{col} IS TRUE")
        where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
        cur.execute(f"SELECT * FROM carriers {where} ORDER BY tier_rank ASC NULLS LAST, carrier_name ASC", params)
        rows = [_serialize_row(r) for r in cur.fetchall()]

        # Optionally nest lane_rates per carrier
        if include_lanes and rows:
            carrier_names = list({r.get("carrier_name") for r in rows if r.get("carrier_name")})
            if carrier_names:
                placeholders = ",".join(["%s"] * len(carrier_names))
                cur.execute(
                    f"SELECT * FROM lane_rates WHERE carrier_name IN ({placeholders}) ORDER BY port, destination, total ASC NULLS LAST",
                    carrier_names,
                )
                lane_rows = [_serialize_row(r) for r in cur.fetchall()]
                lanes_by_carrier = {}
                for lr in lane_rows:
                    cn = lr.get("carrier_name", "")
                    lanes_by_carrier.setdefault(cn, []).append(lr)
                for row in rows:
                    row["lane_rates"] = lanes_by_carrier.get(row.get("carrier_name", ""), [])

    return JSONResponse({"carriers": rows, "total": len(rows)})


@router.post("/api/carriers")
async def api_create_carrier(request: Request):
    body = await request.json()
    fields = ["carrier_name", "mc_number", "dot_number", "contact_email", "contact_phone",
              "contact_name", "regions", "ports", "rail_ramps", "equipment_types", "notes", "source",
              "pickup_area", "destination_area", "date_quoted", "v_code",
              "can_dray", "can_hazmat", "can_overweight", "can_transload",
              "can_reefer", "can_bonded", "can_oog", "can_warehousing",
              "tier_rank", "dnu", "trucks",
              "service_feedback", "service_notes", "service_record", "comments", "insurance_info",
              "markets", "haz_classes"]
    bool_fields = {"can_dray", "can_hazmat", "can_overweight", "can_transload",
                   "can_reefer", "can_bonded", "can_oog", "can_warehousing", "dnu"}
    int_fields = {"tier_rank", "trucks"}
    array_fields = {"markets", "haz_classes"}
    vals = []
    for f in fields:
        if f in bool_fields:
            vals.append(bool(body.get(f, False)))
        elif f in int_fields:
            v = body.get(f)
            vals.append(int(v) if v is not None else None)
        elif f in array_fields:
            v = body.get(f)
            if isinstance(v, list):
                vals.append(v)
            elif isinstance(v, str) and v:
                vals.append([x.strip() for x in v.split(",") if x.strip()])
            else:
                vals.append(None)
        elif f == "carrier_name":
            vals.append(body.get(f, ""))
        else:
            vals.append(body.get(f, None))
    if not vals[0]:
        vals[0] = "Unknown Carrier"
    with db.get_conn() as conn:
        with db.get_cursor(conn) as cur:
            cur.execute(f"""
                INSERT INTO carriers ({', '.join(fields)})
                VALUES ({', '.join(['%s'] * len(fields))}) RETURNING *
            """, vals)
            row = _serialize_row(cur.fetchone())
    return JSONResponse(row)


@router.put("/api/carriers/{carrier_id}")
async def api_update_carrier(carrier_id: int, request: Request):
    body = await request.json()
    allowed = {"carrier_name", "mc_number", "dot_number", "contact_email", "contact_phone",
               "contact_name", "regions", "ports", "rail_ramps", "equipment_types", "notes",
               "pickup_area", "destination_area", "date_quoted", "v_code",
               "can_dray", "can_hazmat", "can_overweight", "can_transload",
               "can_reefer", "can_bonded", "can_oog", "can_warehousing",
               "tier_rank", "dnu", "trucks",
               "service_feedback", "service_notes", "service_record", "comments", "insurance_info",
               "markets", "haz_classes"}
    bool_fields = {"can_dray", "can_hazmat", "can_overweight", "can_transload",
                   "can_reefer", "can_bonded", "can_oog", "can_warehousing", "dnu"}
    int_fields = {"tier_rank", "trucks"}
    array_fields = {"markets", "haz_classes"}
    # Sync transload/warehousing — treat as synonyms
    if "can_transload" in body and "can_warehousing" not in body:
        body["can_warehousing"] = body["can_transload"]
    elif "can_warehousing" in body and "can_transload" not in body:
        body["can_transload"] = body["can_warehousing"]
    sets, params = [], []
    for k, v in body.items():
        if k in allowed:
            sets.append(f"{k} = %s")
            if k in bool_fields:
                params.append(bool(v))
            elif k in int_fields:
                params.append(int(v) if v is not None else None)
            elif k in array_fields:
                if isinstance(v, list):
                    params.append(v)
                elif isinstance(v, str) and v:
                    params.append([x.strip() for x in v.split(",") if x.strip()])
                else:
                    params.append(None)
            else:
                params.append(v)
    if not sets:
        raise HTTPException(400, "No valid fields")
    sets.append("updated_at = NOW()")
    params.append(carrier_id)
    with db.get_conn() as conn:
        with db.get_cursor(conn) as cur:
            cur.execute(f"UPDATE carriers SET {', '.join(sets)} WHERE id = %s RETURNING *", params)
            row = cur.fetchone()
    if not row:
        raise HTTPException(404, "Carrier not found")
    return JSONResponse(_serialize_row(row))


@router.delete("/api/carriers/{carrier_id}")
async def api_delete_carrier(carrier_id: int):
    with db.get_conn() as conn:
        with db.get_cursor(conn) as cur:
            cur.execute("DELETE FROM carriers WHERE id = %s", (carrier_id,))
    return JSONResponse({"ok": True})


_CARRIER_EXTRACT_PROMPT = """Extract carrier directory information from this document (rate sheet, carrier packet, or directory screenshot).

If this is a DIRECTORY LISTING (like LoadMatch/Drayage Directory) with MULTIPLE carriers in a table:
Return ONLY valid JSON (no markdown): {"carriers": [array of carrier objects]}

If this is a SINGLE carrier rate sheet or packet:
Return ONLY valid JSON (no markdown): a single carrier object (not wrapped in array)

Each carrier object should have:
{
  "carrier_name": "string",
  "mc_number": "MC number digits only, or null",
  "dot_number": "DOT/USDOT number digits only, or null",
  "contact_email": "string or null",
  "contact_phone": "string or null",
  "contact_name": "string or null",
  "city": "city name or null",
  "state": "state abbreviation or null",
  "regions": "comma-separated service regions or null",
  "ports": "comma-separated ports served or null",
  "equipment_types": "comma-separated equipment types or null",
  "can_hazmat": true/false (HZ abbreviation in directory),
  "can_bonded": true/false (BD abbreviation),
  "can_transload": true/false (TR or TR-WH abbreviation),
  "can_warehousing": true/false (TR-WH or WH abbreviation, same as transload),
  "can_overweight": true/false (Overweight label),
  "can_oog": true/false (OOG or oversized),
  "can_reefer": true/false (Refrigerated),
  "rates": [{"lane": "origin to destination", "rate": "dollar amount", "equipment": "type"}]
}
Extract ALL carriers visible. For directory screenshots, focus on carrier name, location, and capability flags. Omit null fields."""


@router.post("/api/carriers/extract")
async def api_carrier_extract(request: Request):
    """Extract carrier info from uploaded file via Claude Vision."""
    import config
    if not config.ANTHROPIC_API_KEY:
        raise HTTPException(422, "ANTHROPIC_API_KEY not configured")
    form = await request.form()
    file = form.get("file")
    if not file:
        raise HTTPException(400, "No file provided")
    file_bytes = await file.read()
    filename = getattr(file, "filename", "") or ""
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""

    import base64
    if ext in ("png", "jpg", "jpeg", "gif", "webp"):
        media_map = {"png": "image/png", "jpg": "image/jpeg", "jpeg": "image/jpeg",
                     "gif": "image/gif", "webp": "image/webp"}
        content = [
            {"type": "image", "source": {"type": "base64", "media_type": media_map.get(ext, "image/png"),
                                         "data": base64.b64encode(file_bytes).decode()}},
            {"type": "text", "text": _CARRIER_EXTRACT_PROMPT},
        ]
    elif ext == "pdf":
        content = [
            {"type": "document", "source": {"type": "base64", "media_type": "application/pdf",
                                            "data": base64.b64encode(file_bytes).decode()}},
            {"type": "text", "text": _CARRIER_EXTRACT_PROMPT},
        ]
    elif ext in ("txt", "csv", "eml"):
        text_content = file_bytes.decode("utf-8", errors="replace")
        content = [{"type": "text", "text": _CARRIER_EXTRACT_PROMPT + "\n\n" + text_content}]
    else:
        raise HTTPException(400, f"Unsupported file type: {ext}")

    try:
        from shared import _extract_with_claude
        result = _extract_with_claude(content)

        # If bulk carriers extracted (directory screenshot), save them
        carriers_list = result.get("carriers") if isinstance(result, dict) else None
        if not carriers_list and isinstance(result, dict) and result.get("carrier_name"):
            carriers_list = [result]

        if carriers_list and isinstance(carriers_list, list) and len(carriers_list) > 0:
            saved = []
            for c in carriers_list:
                name = (c.get("carrier_name") or "").strip()
                if not name:
                    continue
                mc = (c.get("mc_number") or "").strip() or None
                # Sync transload/warehousing
                can_tl = bool(c.get("can_transload", False))
                can_wh = bool(c.get("can_warehousing", False))
                if can_tl or can_wh:
                    can_tl = can_wh = True
                area = c.get("city", "")
                if c.get("state"):
                    area = f"{area}, {c['state']}" if area else c["state"]
                with db.get_conn() as conn:
                    with db.get_cursor(conn) as cur:
                        # Upsert by mc_number or carrier_name
                        if mc:
                            cur.execute("SELECT id FROM carriers WHERE mc_number = %s", (mc,))
                        else:
                            cur.execute("SELECT id FROM carriers WHERE LOWER(carrier_name) = LOWER(%s)", (name,))
                        existing = cur.fetchone()
                        if existing:
                            cur.execute("""UPDATE carriers SET
                                can_hazmat = COALESCE(%s, can_hazmat), can_overweight = COALESCE(%s, can_overweight),
                                can_transload = COALESCE(%s, can_transload), can_warehousing = COALESCE(%s, can_warehousing),
                                can_reefer = COALESCE(%s, can_reefer), can_bonded = COALESCE(%s, can_bonded),
                                can_oog = COALESCE(%s, can_oog),
                                pickup_area = COALESCE(NULLIF(%s, ''), pickup_area),
                                updated_at = NOW()
                                WHERE id = %s RETURNING id, carrier_name""",
                                (bool(c.get("can_hazmat")), bool(c.get("can_overweight")),
                                 can_tl, can_wh,
                                 bool(c.get("can_reefer")), bool(c.get("can_bonded")),
                                 bool(c.get("can_oog")), area or "", existing["id"]))
                            saved.append({"id": existing["id"], "name": name, "action": "updated"})
                        else:
                            cur.execute("""INSERT INTO carriers (carrier_name, mc_number, pickup_area,
                                can_hazmat, can_overweight, can_transload, can_warehousing,
                                can_reefer, can_bonded, can_oog, source)
                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'loadmatch')
                                RETURNING id, carrier_name""",
                                (name, mc, area or None,
                                 bool(c.get("can_hazmat")), bool(c.get("can_overweight")),
                                 can_tl, can_wh,
                                 bool(c.get("can_reefer")), bool(c.get("can_bonded")),
                                 bool(c.get("can_oog"))))
                            row = cur.fetchone()
                            saved.append({"id": row["id"], "name": name, "action": "created"})
            return JSONResponse({"carriers": carriers_list, "saved": saved})

        return JSONResponse(result)
    except Exception as e:
        log.error("Carrier extraction failed: %s", e)
        raise HTTPException(500, f"AI extraction failed: {e}")


# ═══════════════════════════════════════════════════════════
# WAREHOUSE DIRECTORY API
# ═══════════════════════════════════════════════════════════

@router.get("/api/warehouses")
async def api_list_warehouses(search: str = Query(default=None), region: str = Query(default=None),
                              state: str = Query(default=None)):
    with db.get_cursor() as cur:
        clauses, params = [], []
        if search:
            clauses.append("(name ILIKE %s OR mc_number ILIKE %s OR contact_email ILIKE %s)")
            s = f"%{search}%"
            params.extend([s, s, s])
        if region:
            clauses.append("region ILIKE %s")
            params.append(f"%{region}%")
        if state:
            clauses.append("state = %s")
            params.append(state.upper())
        where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
        cur.execute(f"SELECT * FROM warehouses {where} ORDER BY name ASC", params)
        warehouses = [_serialize_row(r) for r in cur.fetchall()]
        # Attach rate summaries
        wh_ids = [w["id"] for w in warehouses]
        if wh_ids:
            cur.execute("""SELECT warehouse_id, rate_type, rate_amount, unit, description
                           FROM warehouse_rates WHERE warehouse_id = ANY(%s) ORDER BY warehouse_id, rate_type""",
                        (wh_ids,))
            rates_map = {}
            for r in cur.fetchall():
                rates_map.setdefault(r["warehouse_id"], []).append(_serialize_row(r))
            for w in warehouses:
                w["rates"] = rates_map.get(w["id"], [])
    return JSONResponse({"warehouses": warehouses})


@router.post("/api/warehouses")
async def api_create_warehouse(request: Request):
    body = await request.json()
    fields = ["name", "mc_number", "region", "address", "city", "state", "zip_code",
              "contact_name", "contact_email", "contact_phone", "services", "notes", "source"]
    vals = [body.get(f, "" if f == "name" else None) for f in fields]
    if not vals[0]:
        vals[0] = "Unknown Warehouse"
    with db.get_conn() as conn:
        with db.get_cursor(conn) as cur:
            cur.execute(f"""INSERT INTO warehouses ({', '.join(fields)})
                VALUES ({', '.join(['%s'] * len(fields))}) RETURNING *""", vals)
            row = _serialize_row(cur.fetchone())
    return JSONResponse(row)


@router.put("/api/warehouses/{wh_id}")
async def api_update_warehouse(wh_id: int, request: Request):
    body = await request.json()
    allowed = {"name", "mc_number", "region", "address", "city", "state", "zip_code",
               "contact_name", "contact_email", "contact_phone", "services", "notes"}
    sets, params = [], []
    for k, v in body.items():
        if k in allowed:
            sets.append(f"{k} = %s")
            params.append(v)
    if not sets:
        raise HTTPException(400, "No valid fields")
    sets.append("updated_at = NOW()")
    params.append(wh_id)
    with db.get_conn() as conn:
        with db.get_cursor(conn) as cur:
            cur.execute(f"UPDATE warehouses SET {', '.join(sets)} WHERE id = %s RETURNING *", params)
            row = cur.fetchone()
    if not row:
        raise HTTPException(404, "Warehouse not found")
    return JSONResponse(_serialize_row(row))


@router.delete("/api/warehouses/{wh_id}")
async def api_delete_warehouse(wh_id: int):
    with db.get_conn() as conn:
        with db.get_cursor(conn) as cur:
            cur.execute("DELETE FROM warehouses WHERE id = %s", (wh_id,))
    return JSONResponse({"ok": True})


@router.get("/api/warehouses/{wh_id}/rates")
async def api_warehouse_rates(wh_id: int):
    with db.get_cursor() as cur:
        cur.execute("SELECT * FROM warehouse_rates WHERE warehouse_id = %s ORDER BY rate_type", (wh_id,))
        rows = [_serialize_row(r) for r in cur.fetchall()]
    return JSONResponse({"rates": rows})


@router.post("/api/warehouses/{wh_id}/rates")
async def api_add_warehouse_rate(wh_id: int, request: Request):
    body = await request.json()
    with db.get_conn() as conn:
        with db.get_cursor(conn) as cur:
            cur.execute("""INSERT INTO warehouse_rates (warehouse_id, rate_type, rate_amount, unit, description, notes)
                VALUES (%s, %s, %s, %s, %s, %s) RETURNING *""",
                (wh_id, body.get("rate_type", "flat"), body.get("rate_amount"),
                 body.get("unit"), body.get("description"), body.get("notes")))
            row = _serialize_row(cur.fetchone())
    return JSONResponse(row)


@router.delete("/api/warehouses/{wh_id}/rates/{rate_id}")
async def api_delete_warehouse_rate(wh_id: int, rate_id: int):
    with db.get_conn() as conn:
        with db.get_cursor(conn) as cur:
            cur.execute("DELETE FROM warehouse_rates WHERE id = %s AND warehouse_id = %s", (rate_id, wh_id))
    return JSONResponse({"ok": True})


_WAREHOUSE_EXTRACT_PROMPT = """Extract warehouse/transloading facility information from this rate card or document.
Return ONLY valid JSON (no markdown, no explanation) with these fields:
{
  "name": "facility name",
  "mc_number": "MC number digits or null",
  "address": "full address or null",
  "city": "string or null",
  "state": "2-letter state code or null",
  "contact_name": "string or null",
  "contact_email": "string or null",
  "contact_phone": "string or null",
  "services": "comma-separated services (Transload, Cross-dock, Storage, etc.) or null",
  "rates": [
    {"rate_type": "per_pallet|per_day|per_container|monthly_min|per_hour|flat|per_case|per_label",
     "rate_amount": numeric, "unit": "string", "description": "what this covers"}
  ]
}
Extract ALL rate line items. Look for per-pallet, per-day storage, container handling, labeling, monthly minimums. Omit null fields."""


@router.post("/api/warehouses/extract")
async def api_warehouse_extract(request: Request):
    """Extract warehouse info from uploaded file via Claude Vision."""
    import config
    if not config.ANTHROPIC_API_KEY:
        raise HTTPException(422, "ANTHROPIC_API_KEY not configured")
    form = await request.form()
    file = form.get("file")
    if not file:
        raise HTTPException(400, "No file provided")
    file_bytes = await file.read()
    filename = getattr(file, "filename", "") or ""
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""

    import base64
    if ext in ("png", "jpg", "jpeg", "gif", "webp"):
        media_map = {"png": "image/png", "jpg": "image/jpeg", "jpeg": "image/jpeg",
                     "gif": "image/gif", "webp": "image/webp"}
        content = [
            {"type": "image", "source": {"type": "base64", "media_type": media_map.get(ext, "image/png"),
                                         "data": base64.b64encode(file_bytes).decode()}},
            {"type": "text", "text": _WAREHOUSE_EXTRACT_PROMPT},
        ]
    elif ext == "pdf":
        content = [
            {"type": "document", "source": {"type": "base64", "media_type": "application/pdf",
                                            "data": base64.b64encode(file_bytes).decode()}},
            {"type": "text", "text": _WAREHOUSE_EXTRACT_PROMPT},
        ]
    elif ext in ("txt", "csv", "eml"):
        text_content = file_bytes.decode("utf-8", errors="replace")
        content = [{"type": "text", "text": _WAREHOUSE_EXTRACT_PROMPT + "\n\n" + text_content}]
    else:
        raise HTTPException(400, f"Unsupported file type: {ext}")

    try:
        from shared import _extract_with_claude
        result = _extract_with_claude(content)
        return JSONResponse(result)
    except Exception as e:
        log.error("Warehouse extraction failed: %s", e)
        raise HTTPException(500, f"AI extraction failed: {e}")


# ═══════════════════════════════════════════════════════════
# LANE RATES API
# ═══════════════════════════════════════════════════════════

@router.get("/api/lane-rates")
async def api_list_lane_rates(port: str = Query(default=None), carrier: str = Query(default=None),
                              destination: str = Query(default=None),
                              move_type: str = Query(default=None)):
    """
                              List lane rates and recent rate quotes filtered by optional criteria.
                              
                              Returns a combined list of lane rate records and recent rate quotes matching the provided filters, with normalized fields and source-prefixed IDs.
                              
                              Parameters:
                                  port (str | None): Partial match against lane_rates.port or rate_quotes.origin.
                                  carrier (str | None): Partial match against carrier_name.
                                  destination (str | None): Partial match against destination.
                                  move_type (str | None): Exact match for move_type; pass "all" to disable move_type filtering.
                              
                              Returns:
                                  dict: JSON object with:
                                      - lane_rates (list): Array of rate objects (both lane_rates and rate_quotes) including derived fields
                                        `dest_city`, `dest_state`, `origin_city`, and `origin_state`. Each lane_rates record has an `id`
                                        prefixed with "lr-<id>" and each rate_quote has `id` prefixed with "rq-<id>". Rate quote entries
                                        include an `_source` field indicating their origin.
                                      - total (int): Number of returned records.
                              """
                              with db.get_cursor() as cur:
        # Query lane_rates table
        clauses, params = [], []
        if port:
            clauses.append("port ILIKE %s")
            params.append(f"%{port}%")
        if carrier:
            clauses.append("carrier_name ILIKE %s")
            params.append(f"%{carrier}%")
        if destination:
            clauses.append("destination ILIKE %s")
            params.append(f"%{destination}%")
        if move_type and move_type != "all":
            clauses.append("move_type = %s")
            params.append(move_type)
        where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
        cur.execute(f"""SELECT *,
            TRIM(SPLIT_PART(destination, ',', 1)) AS dest_city,
            TRIM(SPLIT_PART(TRIM(SPLIT_PART(destination, ',', 2)), ' ', 1)) AS dest_state,
            TRIM(SPLIT_PART(port, ',', 1)) AS origin_city,
            TRIM(SPLIT_PART(TRIM(SPLIT_PART(port, ',', 2)), ' ', 1)) AS origin_state
            FROM lane_rates {where} ORDER BY port, destination, total ASC NULLS LAST LIMIT 1000""", params)
        rows = [_serialize_row(r) for r in cur.fetchall()]

        # Also include rate_quotes (from manual intake, emails, etc.)
        rq_clauses, rq_params = [], []
        if port:
            rq_clauses.append("origin ILIKE %s")
            rq_params.append(f"%{port}%")
        if carrier:
            rq_clauses.append("carrier_name ILIKE %s")
            rq_params.append(f"%{carrier}%")
        if destination:
            rq_clauses.append("destination ILIKE %s")
            rq_params.append(f"%{destination}%")
        if move_type and move_type != "all":
            rq_clauses.append("move_type = %s")
            rq_params.append(move_type)
        rq_clauses.append("rate_amount IS NOT NULL")
        rq_where = "WHERE " + " AND ".join(rq_clauses) if rq_clauses else ""
        cur.execute(f"""
            SELECT id, origin AS port, destination, carrier_name, carrier_email,
                   rate_amount AS dray_rate, rate_amount AS total,
                   NULL::numeric AS fsc, NULL::numeric AS chassis_per_day,
                   NULL::numeric AS prepull, NULL::numeric AS storage_per_day,
                   NULL::numeric AS detention, NULL::numeric AS chassis_split,
                   NULL::numeric AS overweight, NULL::numeric AS tolls,
                   NULL::numeric AS reefer, NULL::numeric AS hazmat,
                   NULL::numeric AS triaxle, NULL::numeric AS bond_fee,
                   NULL::numeric AS residential, NULL::numeric AS all_in_total,
                   move_type, miles, quote_date AS created_at,
                   source, status, NULL::int AS rank,
                   NULL AS equipment_type, NULL AS notes,
                   TRIM(SPLIT_PART(destination, ',', 1)) AS dest_city,
                   TRIM(SPLIT_PART(TRIM(SPLIT_PART(destination, ',', 2)), ' ', 1)) AS dest_state,
                   TRIM(SPLIT_PART(origin, ',', 1)) AS origin_city,
                   TRIM(SPLIT_PART(TRIM(SPLIT_PART(origin, ',', 2)), ' ', 1)) AS origin_state
            FROM rate_quotes {rq_where}
            ORDER BY quote_date DESC NULLS LAST LIMIT 500
        """, rq_params)
        rq_rows = [_serialize_row(r) for r in cur.fetchall()]
        # Mark source and prefix IDs to avoid collisions with lane_rates
        for r in rq_rows:
            r["_source"] = r.get("source") or "rate_quote"
            r["id"] = f"rq-{r['id']}"
        for r in rows:
            r["id"] = f"lr-{r['id']}"
        rows.extend(rq_rows)

    return JSONResponse({"lane_rates": rows, "total": len(rows)})

@router.put("/api/lane-rates/{rate_id}")
async def api_update_lane_rate(rate_id: str, request: Request):
    body = await request.json()
    allowed = {"port", "destination", "carrier_name", "dray_rate", "fsc", "total",
               "chassis_per_day", "prepull", "storage_per_day", "detention",
               "chassis_split", "overweight", "tolls", "reefer", "hazmat",
               "triaxle", "bond_fee", "residential", "all_in_total",
               "rank", "equipment_type", "move_type", "notes"}
    decimal_fields = {"dray_rate", "total", "chassis_per_day", "prepull", "storage_per_day",
                      "chassis_split", "overweight", "tolls", "reefer", "hazmat",
                      "triaxle", "bond_fee", "residential", "all_in_total"}
    int_fields = {"rank"}
    sets, params = [], []
    for k, v in body.items():
        if k in allowed:
            sets.append(f"{k} = %s")
            if k in decimal_fields:
                from decimal import InvalidOperation
                try:
                    params.append(Decimal(str(v)) if v is not None and str(v).strip() != "" else None)
                except (InvalidOperation, ValueError):
                    params.append(None)
            elif k in int_fields:
                params.append(int(v) if v is not None else None)
            else:
                params.append(str(v).strip() if v is not None else None)
    if not sets:
        raise HTTPException(400, "No valid fields")
    # Auto-recalculate total if dray_rate or fsc changed but total not explicitly sent
    if ("dray_rate" in body or "fsc" in body) and "total" not in body:
        sets.append("total = COALESCE(dray_rate, 0) + CASE WHEN fsc ~ '^[0-9.]+$' THEN fsc::numeric ELSE 0 END")
    # Parse namespaced ID (lr-123 or rq-456 or bare integer)
    if str(rate_id).startswith("lr-"):
        table, real_id = "lane_rates", int(rate_id[3:])
    elif str(rate_id).startswith("rq-"):
        table, real_id = "rate_quotes", int(rate_id[3:])
    else:
        table, real_id = "lane_rates", int(rate_id)
    params.append(real_id)
    with db.get_conn() as conn:
        with db.get_cursor(conn) as cur:
            cur.execute(f"UPDATE {table} SET {', '.join(sets)} WHERE id = %s RETURNING *", params)
            row = cur.fetchone()
    if not row:
        raise HTTPException(404, "Lane rate not found")
    return JSONResponse(_serialize_row(row))


@router.delete("/api/lane-rates/{rate_id}")
async def api_delete_lane_rate(rate_id: str):
    # IDs are namespaced: "lr-123" for lane_rates, "rq-456" for rate_quotes
    if rate_id.startswith("rq-"):
        table, real_id = "rate_quotes", int(rate_id[3:])
    elif rate_id.startswith("lr-"):
        table, real_id = "lane_rates", int(rate_id[3:])
    else:
        # Legacy: bare integer ID — assume lane_rates
        table, real_id = "lane_rates", int(rate_id)
    with db.get_conn() as conn:
        with db.get_cursor(conn) as cur:
            cur.execute(f"DELETE FROM {table} WHERE id = %s RETURNING id", (real_id,))
            row = cur.fetchone()
    if not row:
        raise HTTPException(404, "Lane rate not found")
    return {"ok": True, "deleted_id": rate_id}


# ═══════════════════════════════════════════════════════════
# EXCEL BULK IMPORT
# ═══════════════════════════════════════════════════════════

_CITY_TABS = {
    "Atlanta", "Baltimore", "Birmingham", "Buffalo", "Boston", "Charleston", "Charlotte",
    "Chicago", "Cincinnati", "Cleveland", "Columbus", "Dallas", "Denver", "Detroit",
    "El Paso, TX", "Houston", "Indianapolis", "Jacksonville", "Kansas City", "Louisville",
    "Los Angeles", "Memphis", "Miami", "Mobile", "Nashville", "New York", "NOLA",
    "Norfolk", "Oakland", "Omaha", "Phildadelphia", "Pittsburgh", "Portland",
    "Salt Lake City", "Seattle", "Savannah", "St Louis", "Wilmington", "Tacoma", "Tampa",
    "Toronto",
}

_SKIP_TABS = {
    "Sheet1", "Sheet2", "Sheet3", "Sheet4", "Rate Quote Sheet-John (1)",
    "Priority 1", "PostMaster", "Oversize", "Timberlab",
}


def _safe_float(val):
    """Try to convert a value to float, return None on failure."""
    if val is None:
        return None
    try:
        s = str(val).replace(",", "").replace("$", "").strip()
        if not s:
            return None
        return float(s)
    except (ValueError, TypeError):
        return None


def _parse_city_tab(ws, tab_name):
    """Parse a city tab from the Excel workbook. Returns (carriers, lane_rates)."""
    carriers = []
    lane_rates = []
    max_row = ws.max_row or 1

    # Detect the rate grid header row (cols E-R) and carrier directory section
    rate_grid_header_row = None
    carrier_section_row = None

    for row_idx in range(1, min(max_row + 1, 5)):
        cell_e = ws.cell(row=row_idx, column=5).value  # Col E
        cell_f = ws.cell(row=row_idx, column=6).value  # Col F
        if cell_e and cell_f:
            e_str = str(cell_e).lower().strip()
            f_str = str(cell_f).lower().strip()
            if any(kw in e_str for kw in ("destination", "lane", "carrier")) or any(kw in f_str for kw in ("carrier", "dray", "lane")):
                rate_grid_header_row = row_idx
                break

    # Find carrier directory section (look for "Carrier" header in col B)
    for row_idx in range(1, max_row + 1):
        cell_b = ws.cell(row=row_idx, column=2).value
        if cell_b and str(cell_b).strip().lower() == "carrier":
            cell_c = ws.cell(row=row_idx, column=3).value
            if cell_c and "email" in str(cell_c).lower():
                carrier_section_row = row_idx
                break

    # Parse rate grid (rows after header, cols E-R)
    if rate_grid_header_row:
        # Determine column mapping from header
        hdr = {}
        for col in range(5, 25):
            v = ws.cell(row=rate_grid_header_row, column=col).value
            if v:
                hdr[str(v).strip().lower()] = col

        dest_col = hdr.get("destination", hdr.get("lane", 5))
        carrier_col = hdr.get("carrier", 6)

        for row_idx in range(rate_grid_header_row + 1, max_row + 1):
            carrier_val = ws.cell(row=row_idx, column=carrier_col).value
            dest_val = ws.cell(row=row_idx, column=dest_col).value
            if not carrier_val and not dest_val:
                # Check if we've hit the quote template or carrier section
                cell_b = ws.cell(row=row_idx, column=2).value
                if cell_b and str(cell_b).strip().lower() in ("pod", "carrier", "warehousing"):
                    break
                continue
            if not carrier_val:
                continue

            lr = {
                "port": tab_name,
                "destination": str(dest_val).strip() if dest_val else None,
                "carrier_name": str(carrier_val).strip(),
                "dray_rate": _safe_float(ws.cell(row=row_idx, column=hdr.get("dray", hdr.get("rate", 7))).value),
                "fsc": str(ws.cell(row=row_idx, column=hdr.get("fsc", 8)).value or "").strip() or None,
                "total": _safe_float(ws.cell(row=row_idx, column=hdr.get("total", 9)).value),
                "chassis_per_day": _safe_float(ws.cell(row=row_idx, column=hdr.get("chassis", 10)).value),
                "prepull": _safe_float(ws.cell(row=row_idx, column=hdr.get("prepull", hdr.get("drop", 11))).value),
                "storage_per_day": _safe_float(ws.cell(row=row_idx, column=hdr.get("storage", hdr.get("reefer", 12))).value),
                "detention": str(ws.cell(row=row_idx, column=hdr.get("detention", 14)).value or "").strip() or None,
                "overweight": _safe_float(ws.cell(row=row_idx, column=hdr.get("ow", 15)).value),
                "tolls": _safe_float(ws.cell(row=row_idx, column=hdr.get("tolls", hdr.get("toll", 16))).value),
                "all_in_total": _safe_float(ws.cell(row=row_idx, column=hdr.get("total", 17)).value),
                "rank": int(_safe_float(ws.cell(row=row_idx, column=hdr.get("rank", 18)).value) or 0) or None,
                "move_type": "dray",
                "source": "excel_import",
            }
            # Use the last "total" column for all_in
            for k2 in sorted(hdr.keys()):
                if k2 == "total" and hdr[k2] > 9:
                    lr["all_in_total"] = _safe_float(ws.cell(row=row_idx, column=hdr[k2]).value)
            if lr["dray_rate"] or lr["total"]:
                lane_rates.append(lr)

    # Parse carrier directory (rows after "Carrier | Email | MC" header)
    if carrier_section_row:
        for row_idx in range(carrier_section_row + 1, max_row + 1):
            cell_b = ws.cell(row=row_idx, column=2).value
            cell_c = ws.cell(row=row_idx, column=3).value
            if not cell_b and not cell_c:
                # Check for end of carrier section
                continue
            name = str(cell_b).strip() if cell_b else None
            email = str(cell_c).strip() if cell_c else None
            if not name or name == "\xa0":
                continue
            if name.lower() in ("warehousing", "warehouse charges", ""):
                break
            mc = ws.cell(row=row_idx, column=4).value
            notes_val = ws.cell(row=row_idx, column=5).value
            carriers.append({
                "carrier_name": name,
                "contact_email": email if email and "@" in str(email) else None,
                "mc_number": str(mc).strip() if mc else None,
                "regions": tab_name,
                "notes": str(notes_val).strip() if notes_val else None,
                "source": "excel_import",
            })

    return carriers, lane_rates


@router.post("/api/directory/import-excel")
async def api_import_excel(file: UploadFile = File(...)):
    """Bulk import carriers, lane rates, and warehouses from the rate quote Excel."""
    import tempfile
    file_bytes = await file.read()
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(file_bytes)
        tmp_path = tmp.name

    try:
        import openpyxl
        wb = openpyxl.load_workbook(tmp_path, data_only=True, read_only=True)
    except Exception as e:
        os.unlink(tmp_path)
        raise HTTPException(400, f"Failed to read Excel file: {e}")

    all_carriers = []
    all_lane_rates = []
    all_warehouses = []
    summary = {"sheets_processed": [], "carriers": 0, "lane_rates": 0, "warehouses": 0}

    for name in wb.sheetnames:
        if name in _SKIP_TABS:
            continue
        ws = wb[name]

        if name in _CITY_TABS:
            carriers, lane_rates = _parse_city_tab(ws, name)
            all_carriers.extend(carriers)
            all_lane_rates.extend(lane_rates)
            if carriers or lane_rates:
                summary["sheets_processed"].append(name)
            continue

        # Specialty tabs
        if name == "Tolead Rates ORD":
            for row_idx in range(2, (ws.max_row or 1) + 1):
                city = ws.cell(row=row_idx, column=1).value
                rate = _safe_float(ws.cell(row=row_idx, column=3).value)
                if city and rate:
                    all_lane_rates.append({
                        "port": "Chicago (ORD)", "destination": str(city).strip(),
                        "carrier_name": "Tolead", "total": rate, "move_type": "ftl",
                        "source": "excel_import",
                    })
            summary["sheets_processed"].append(name)

        elif name == "LTL":
            for row_idx in range(2, (ws.max_row or 1) + 1):
                pallets = ws.cell(row=row_idx, column=1).value
                dest = ws.cell(row=row_idx, column=2).value
                rate = _safe_float(ws.cell(row=row_idx, column=4).value)
                if rate:
                    all_lane_rates.append({
                        "port": "LTL", "destination": str(dest or "").strip() or None,
                        "carrier_name": f"LTL ({pallets} pallet{'s' if str(pallets) != '1' else ''})",
                        "total": rate, "move_type": "ltl", "source": "excel_import",
                        "notes": f"{ws.cell(row=row_idx, column=3).value} transit days" if ws.cell(row=row_idx, column=3).value else None,
                    })
            summary["sheets_processed"].append(name)

        elif name == "Step Deck":
            for row_idx in range(2, (ws.max_row or 1) + 1):
                pickup = ws.cell(row=row_idx, column=1).value
                lane = ws.cell(row=row_idx, column=2).value
                rate = _safe_float(ws.cell(row=row_idx, column=3).value)
                if lane and rate:
                    all_lane_rates.append({
                        "port": str(pickup or "").strip() or "Step Deck",
                        "destination": str(lane).strip(),
                        "carrier_name": "Step Deck",
                        "total": rate, "move_type": "step_deck", "equipment_type": "Step Deck",
                        "source": "excel_import",
                    })
            summary["sheets_processed"].append(name)

        elif name == "Sutton WH - Inventory (Meiborg)":
            wh = {"name": "Sutton Warehouse (Meiborg)", "source": "excel_import", "services": "Storage, Labeling, Palletizing"}
            rates = []
            for row_idx in range(2, (ws.max_row or 1) + 1):
                desc = ws.cell(row=row_idx, column=1).value
                rate = _safe_float(ws.cell(row=row_idx, column=3).value)
                if desc and rate:
                    desc_str = str(desc).strip()
                    if "total" in desc_str.lower():
                        continue
                    rates.append({"rate_type": "flat", "rate_amount": rate, "unit": "each", "description": desc_str})
            if rates:
                all_warehouses.append({"warehouse": wh, "rates": rates})
            summary["sheets_processed"].append(name)

        elif name == "Tolead Box Rates":
            # Similar to Tolead Rates ORD but for box rates
            for row_idx in range(2, (ws.max_row or 1) + 1):
                city = ws.cell(row=row_idx, column=1).value
                rate = _safe_float(ws.cell(row=row_idx, column=3).value)
                if city and rate:
                    all_lane_rates.append({
                        "port": "Tolead Box", "destination": str(city).strip(),
                        "carrier_name": "Tolead", "total": rate, "move_type": "ftl",
                        "source": "excel_import",
                    })
            summary["sheets_processed"].append(name)

        elif name == "Heavy Haul":
            # Heavy haul has truck assignments with rates
            for row_idx in range(2, (ws.max_row or 1) + 1):
                truck = ws.cell(row=row_idx, column=13).value  # M
                rate = _safe_float(ws.cell(row=row_idx, column=14).value)  # N
                equip = ws.cell(row=row_idx, column=16).value  # P
                if rate:
                    all_lane_rates.append({
                        "port": "Heavy Haul", "destination": "Project Site",
                        "carrier_name": f"Truck {truck}" if truck else "Heavy Haul",
                        "total": rate, "move_type": "heavy_haul",
                        "equipment_type": str(equip).strip() if equip else None,
                        "source": "excel_import",
                    })
            summary["sheets_processed"].append(name)

    wb.close()
    os.unlink(tmp_path)

    # Bulk insert into database
    with db.get_conn() as conn:
        with db.get_cursor(conn) as cur:
            # Insert carriers (deduplicate by name+email within import)
            seen_carriers = set()
            for c in all_carriers:
                key = (c["carrier_name"].lower(), (c.get("contact_email") or "").lower())
                if key in seen_carriers:
                    # Append region to existing
                    continue
                seen_carriers.add(key)
                cur.execute("""
                    INSERT INTO carriers (carrier_name, mc_number, contact_email, regions, notes, source)
                    VALUES (%s, %s, %s, %s, %s, %s)
                    ON CONFLICT DO NOTHING
                """, (c["carrier_name"], c.get("mc_number"), c.get("contact_email"),
                      c.get("regions"), c.get("notes"), "excel_import"))
                summary["carriers"] += 1

            # Insert lane rates
            for lr in all_lane_rates:
                cur.execute("""
                    INSERT INTO lane_rates (port, destination, carrier_name, dray_rate, fsc, total,
                        chassis_per_day, prepull, storage_per_day, detention, chassis_split,
                        overweight, tolls, reefer, hazmat, all_in_total, rank,
                        equipment_type, move_type, notes, source)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (lr.get("port"), lr.get("destination"), lr.get("carrier_name"),
                      lr.get("dray_rate"), lr.get("fsc"), lr.get("total"),
                      lr.get("chassis_per_day"), lr.get("prepull"), lr.get("storage_per_day"),
                      lr.get("detention"), lr.get("chassis_split"), lr.get("overweight"),
                      lr.get("tolls"), lr.get("reefer"), lr.get("hazmat"),
                      lr.get("all_in_total"), lr.get("rank"),
                      lr.get("equipment_type"), lr.get("move_type", "dray"),
                      lr.get("notes"), lr.get("source", "excel_import")))
                summary["lane_rates"] += 1

            # Insert warehouses + rates
            for wh_data in all_warehouses:
                wh = wh_data["warehouse"]
                cur.execute("""INSERT INTO warehouses (name, services, source)
                    VALUES (%s, %s, %s) RETURNING id""",
                    (wh["name"], wh.get("services"), "excel_import"))
                wh_id = cur.fetchone()["id"]
                for rate in wh_data["rates"]:
                    cur.execute("""INSERT INTO warehouse_rates (warehouse_id, rate_type, rate_amount, unit, description)
                        VALUES (%s, %s, %s, %s, %s)""",
                        (wh_id, rate["rate_type"], rate["rate_amount"], rate["unit"], rate.get("description")))
                summary["warehouses"] += 1

    return JSONResponse(summary)


# ═══════════════════════════════════════════════════════════════
# ─── Carrier Suggestion Endpoint (Quote Builder integration) ──
# ═══════════════════════════════════════════════════════════════

@router.get("/api/directory/suggest")
async def api_directory_suggest(
    port_code: str = Query(..., min_length=3),
    caps: str = Query(default=None),
    destination: str = Query(default=None),
):
    """Return ranked carrier suggestions for a port, filtered by capabilities."""
    cap_map = {
        "hazmat": "can_hazmat", "dray": "can_dray", "overweight": "can_overweight",
        "transload": "can_transload", "reefer": "can_reefer", "bonded": "can_bonded",
        "oog": "can_oog", "warehousing": "can_warehousing",
    }
    cap_filters = []
    if caps:
        for c in caps.split(","):
            col = cap_map.get(c.strip().lower())
            if col:
                cap_filters.append(col)

    pc = port_code.strip()
    pc_like = f"%{pc}%"

    with db.get_cursor() as cur:
        # Build WHERE: two-tier port match + DNU exclusion + capability filters
        cap_sql = "".join(f" AND {col} IS TRUE" for col in cap_filters)
        cur.execute(f"""
            SELECT *,
                CASE
                    WHEN ports ILIKE %s THEN 1
                    WHEN pickup_area ILIKE %s OR regions ILIKE %s THEN 2
                    ELSE 3
                END AS port_match_tier
            FROM carriers
            WHERE (dnu IS NOT TRUE)
              AND (ports ILIKE %s OR pickup_area ILIKE %s OR regions ILIKE %s)
              {cap_sql}
            ORDER BY port_match_tier ASC, tier_rank ASC NULLS LAST, carrier_name ASC
        """, (pc_like, pc_like, pc_like, pc_like, pc_like, pc_like))
        rows = cur.fetchall()

        if not rows:
            return JSONResponse({"carriers": [], "port_code": pc, "total": 0})

        # Gather lane_rates for matched carriers
        carrier_names = list({r["carrier_name"] for r in rows if r.get("carrier_name")})
        lanes_by_carrier = {}
        if carrier_names:
            ph = ",".join(["%s"] * len(carrier_names))
            lane_params = list(carrier_names) + [pc_like]
            dest_sql = ""
            if destination and destination.strip():
                dest_sql = " AND destination ILIKE %s"
                lane_params.append(f"%{destination.strip()}%")
            cur.execute(f"""
                SELECT carrier_name, total, dray_rate, destination, created_at
                FROM lane_rates
                WHERE carrier_name IN ({ph}) AND port ILIKE %s {dest_sql}
                ORDER BY created_at DESC
            """, lane_params)
            for lr in cur.fetchall():
                cn = lr["carrier_name"]
                lanes_by_carrier.setdefault(cn, []).append(lr)

        # Build response
        result = []
        for r in rows:
            cn = r.get("carrier_name", "")
            lanes = lanes_by_carrier.get(cn, [])
            rates = [float(l["total"] or l.get("dray_rate") or 0) for l in lanes if (l.get("total") or l.get("dray_rate"))]

            # Capabilities list
            capabilities = []
            for cap_key, col in cap_map.items():
                if r.get(col):
                    capabilities.append(cap_key)

            rate_min = min(rates) if rates else None
            rate_max = max(rates) if rates else None
            rate_range = f"${rate_min:,.0f} – ${rate_max:,.0f}" if rate_min is not None and rate_max is not None else None

            # Lane match: did they run this exact destination?
            lane_match = False
            lane_rate = None
            if destination and destination.strip():
                dest_lower = destination.strip().lower()
                for l in lanes:
                    if dest_lower in (l.get("destination") or "").lower():
                        lane_match = True
                        lane_rate = float(l["total"] or l.get("dray_rate") or 0)
                        break

            result.append({
                "carrier_id": r["id"],
                "name": cn,
                "capabilities": capabilities,
                "last_quoted": float(rates[0]) if rates else None,
                "rate_range": rate_range,
                "rate_min": rate_min,
                "rate_max": rate_max,
                "lane_match": lane_match,
                "lane_rate": lane_rate,
                "contact": {
                    "name": r.get("contact_name") or "",
                    "phone": r.get("contact_phone") or "",
                    "email": r.get("contact_email") or "",
                },
                "tier_rank": r.get("tier_rank"),
                "port_match": "exact" if r["port_match_tier"] == 1 else "fuzzy",
            })

        # Final sort: lane_match first, then exact port, then tier
        result.sort(key=lambda c: (
            0 if c["lane_match"] else 1,
            0 if c["port_match"] == "exact" else 1,
            c["tier_rank"] if c["tier_rank"] is not None else 999,
        ))

    return JSONResponse({"carriers": result, "port_code": pc, "total": len(result)})


# ═══════════════════════════════════════════════════════════════
# ─── Feedback Endpoint (Quote → Directory) ────────────────────
# ═══════════════════════════════════════════════════════════════

@router.post("/api/directory/feedback")
async def api_directory_feedback(request: Request):
    """Update directory when a quote is saved. Creates carrier if unknown."""
    body = await request.json()
    carrier_id = body.get("carrier_id")  # int PK from suggestions, or None
    carrier_name = (body.get("carrier_name") or "").strip()
    port_code = (body.get("port_code") or "").strip()
    destination = (body.get("destination") or "").strip()
    rate = body.get("rate")
    quote_id = body.get("quote_id")

    if not carrier_name or not port_code:
        raise HTTPException(400, "carrier_name and port_code required")

    with db.get_conn() as conn:
        with db.get_cursor(conn) as cur:
            # Find carrier: by ID first, then name fallback
            carrier_row = None
            if carrier_id:
                cur.execute("SELECT id, carrier_name FROM carriers WHERE id = %s", (carrier_id,))
                carrier_row = cur.fetchone()
            if not carrier_row:
                cur.execute("SELECT id, carrier_name FROM carriers WHERE carrier_name ILIKE %s LIMIT 1",
                            (carrier_name,))
                carrier_row = cur.fetchone()

            action = "updated"
            if not carrier_row:
                # Auto-create with needs_review flag
                cur.execute("""
                    INSERT INTO carriers (carrier_name, ports, needs_review, date_quoted)
                    VALUES (%s, %s, TRUE, NOW()) RETURNING id, carrier_name
                """, (carrier_name, port_code))
                carrier_row = cur.fetchone()
                action = "added_for_review"
            else:
                # Update date_quoted
                cur.execute("UPDATE carriers SET date_quoted = NOW(), updated_at = NOW() WHERE id = %s",
                            (carrier_row["id"],))

            # Insert lane_rate if we have rate + destination
            if rate and destination:
                cur.execute("""
                    INSERT INTO lane_rates (carrier_name, port, destination, total, source)
                    VALUES (%s, %s, %s, %s, %s)
                """, (carrier_row["carrier_name"], port_code, destination, rate, "quote"))

    return JSONResponse({"ok": True, "action": action, "carrier_id": carrier_row["id"]})
